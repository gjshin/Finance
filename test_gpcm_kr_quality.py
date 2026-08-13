"""gpcm_kr.py Data_Quality 회귀 테스트.

DART 조회는 회사·기간별로 조용히 실패한다. 실패한 자리에 0이 남고 나머지 계산은
그대로 돌아가, 엑셀만 받아 본 사람은 그 0이 '정말 0'인지 '못 가져온 것'인지 알 수 없었다.
특히 LTM은 (당기누계 + 전기연간 - 전년동기)라서 셋 중 하나만 빠져도 숫자가
그럴듯하게 틀린다 — 오류도 안 나고 눈에도 안 띈다.

여기서는 일부러 실패시켜 두고, 그 실패가 Data_Quality 시트까지 실제로 도달하는지 본다.
"""
import sys, os
import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import gpcm_kr as M

BS = pd.DataFrame([
    {'sj_nm': '연결재무상태표', 'account_nm': '현금및현금성자산', 'account_id': 'ifrs-full_CashAndCashEquivalents', 'thstrm_amount': '100000000000'},
    {'sj_nm': '연결재무상태표', 'account_nm': '자본총계', 'account_id': 'ifrs-full_Equity', 'thstrm_amount': '600000000000'},
])
PL = pd.DataFrame([
    {'sj_nm': '연결포괄손익계산서', 'account_nm': '매출액', 'account_id': 'ifrs-full_Revenue',
     'thstrm_amount': '800000000000', 'thstrm_add_amount': '800000000000'},
    {'sj_nm': '연결포괄손익계산서', 'account_nm': '영업이익', 'account_id': 'dart_OperatingIncomeLoss',
     'thstrm_amount': '80000000000', 'thstrm_add_amount': '80000000000'},
])


class Dart:
    """전년동기(2023 1Q) 손익계산서만 못 가져오는 DART."""
    corp_codes = pd.DataFrame({'corp_code': ['00126380'], 'corp_name': ['테스트'], 'stock_code': ['005930']})

    def find_corp_code(self, c): return '00126380'
    def company(self, c): return {'corp_name': '테스트'}
    def finstate_all(self, corp, y, reprt_code='11011', fs_div='CFS'):
        return BS.copy() if fs_div == 'CFS' else pd.DataFrame()


def fake_pl(dart_instance, corp_code, year, reprt_code):
    if (year, reprt_code) == (2023, '11013'):     # 전년동기 — 조회 실패
        return None, 'N/A', 'NO_DATA'
    return PL.copy(), 'CFS', 'OK'


class Silent:
    def write(self, *a, **k): pass
    def update(self, *a, **k): pass
    def progress(self, *a, **k): pass


def run(shares, pl=fake_pl):
    M.get_krx_listing = lambda: pd.DataFrame({'Code': ['005930'], 'Name': ['테스트'], 'Stocks': [100]})
    M.get_stock_price = lambda t, d: (70000, d)
    M.get_outstanding_shares = lambda *a, **k: (shares, 'KRX' if shares else 'N/A', {'status': '013', 'message': '조회된 데이타가 없습니다.'})
    M.fetch_pl_df = pl
    return M.fetch_financial_data('KEY', ['005930'], ['2024.1Q'], Dart(), Silent(), Silent())


def items(quality, level=None):
    return [r for r in quality.rows if level is None or r['Level'] == level]


fails = []


def check(label, ok, detail=''):
    print(f"{'OK  ' if ok else '실패'}  {label}{('  — ' + detail) if detail and not ok else ''}")
    if not ok:
        fails.append(label)


# --- 1. LTM 조각이 빠지면 ERROR 로 남는가 -----------------------------------
*_, quality = run(shares=1_000_000)
ltm = [r for r in items(quality, M.SEV_ERROR) if '전년동기' in r['Message']]
check('LTM 전년동기 누락이 ERROR 로 기록된다', len(ltm) == 1, f'{len(ltm)}건')
check('메시지에 합산이 틀어진다는 사실이 적혀 있다',
      bool(ltm) and '실제와 다릅니다' in ltm[0]['Message'])

# --- 2. 시가총액이 0이면 ERROR 인가 -----------------------------------------
*_, quality_no_shares = run(shares=None)
mc = [r for r in items(quality_no_shares, M.SEV_ERROR) if '시가총액' in r['Item']]
check('발행주식수를 못 가져오면 시가총액 ERROR 가 남는다', len(mc) >= 1, f'{len(mc)}건')
check('DART 가 돌려준 사유가 메시지에 실린다',
      bool(mc) and '조회된 데이타가 없습니다' in mc[0]['Message'])

# --- 3. 정상 조회는 조용한가 (허위 경보 방지) -------------------------------
*_, quality_clean = run(shares=1_000_000, pl=lambda d, c, y, r: (PL.copy(), 'CFS', 'OK'))
noise = [r for r in items(quality_clean, M.SEV_ERROR)]
check('전부 정상이면 ERROR 가 하나도 없다', not noise, f'{[r["Item"] for r in noise]}')

# --- 4. 엑셀에 Data_Quality 시트가 실제로 만들어지는가 -----------------------
wacc = {'Rf': 0.03, 'MRP': 0.06, 'Size_Premium': 0.0, 'Avg_Unlevered_Beta': 0.8,
        'Target_Tax_Rate': 0.22, 'Avg_Debt_Ratio': 0.3, 'Target_DE_Ratio': 0.43,
        'Target_Relevered_Beta': 0.9, 'Target_Ke': 0.09, 'Kd_Pretax': 0.05,
        'Kd_Aftertax': 0.04, 'Equity_Weight': 0.7, 'Debt_Weight': 0.3, 'Target_WACC': 0.08}
bs_rows, pl_rows, mkt, names, summary, byear, bqtr, bdate, multiples, quality = run(shares=None)
book = M.export_gpcm_excel('2024.1Q', '1Q', ['005930'], summary, bs_rows, pl_rows, mkt, names,
                           wacc, '5Y Monthly', [], 0.3, bdate,
                           pd.DataFrame(multiples), ['2024.1Q'], quality)
wb = load_workbook(book)
check('Data_Quality 시트가 만들어진다', 'Data_Quality' in wb.sheetnames, str(wb.sheetnames))
if 'Data_Quality' in wb.sheetnames:
    ws = wb['Data_Quality']
    body = [ws.cell(r, 1).value for r in range(5, ws.max_row + 1)]
    check('빈 시트가 아니라 실제 항목이 적혀 있다', M.SEV_ERROR in body, str(body))
    check('GPCM 바로 뒤에 놓여 눈에 걸린다',
          wb.sheetnames.index('Data_Quality') == 1, str(wb.sheetnames))


# --- 5. 업종 후보 목록 (KRX 차단 환경이라 응답을 흉내 내 검증) ---------------
IND = pd.DataFrame([
    {'Code': '005930', 'Name': '삼성전자', 'Market': 'KOSPI', 'Sector': '반도체 제조업',
     'Industry': 'DRAM, NAND Flash', 'SettleMonth': '12월'},
    {'Code': '000660', 'Name': 'SK하이닉스', 'Market': 'KOSPI', 'Sector': '반도체 제조업',
     'Industry': 'DRAM', 'SettleMonth': '12월'},
    {'Code': '111111', 'Name': '삼월결산㈜', 'Market': 'KOSDAQ', 'Sector': '반도체 제조업',
     'Industry': '반도체 장비', 'SettleMonth': '3월'},
    {'Code': '222222', 'Name': '딴업종㈜', 'Market': 'KOSDAQ', 'Sector': '음식료품',
     'Industry': '라면', 'SettleMonth': '12월'},
])

rows = M.peer_candidate_rows(IND, '반도체 제조업')
check('고른 업종만 후보로 나온다', [r['Code'] for r in rows] == ['000660', '005930', '111111'],
      str([r['Code'] for r in rows]))
check('종목코드 오름차순으로 고정된다', [r['Code'] for r in rows] == sorted(r['Code'] for r in rows))
check('12월 결산이 아닌 회사가 표시된다',
      [r['FiscalNot12'] for r in rows] == [False, False, True], str([r['FiscalNot12'] for r in rows]))
check('주요제품이 함께 실린다', rows[1]['Product'] == 'DRAM, NAND Flash', rows[1]['Product'])
check('업종 목록을 못 받으면 빈 목록', M.peer_candidate_rows(pd.DataFrame(), '반도체 제조업') == [])

# --- 6. 고른 결과가 엑셀에 남는가 -------------------------------------------
sel = {'sector': '반도체 제조업', 'candidates': rows, 'picked': ['005930', '000660']}
bs2, pl2, mkt2, nm2, sm2, _, _, bd2, mul2, q2 = run(shares=1_000_000)
wb2 = load_workbook(M.export_gpcm_excel('2024.1Q', '1Q', ['005930'], sm2, bs2, pl2, mkt2, nm2,
                                        wacc, '5Y Monthly', [], 0.3, bd2,
                                        pd.DataFrame(mul2), ['2024.1Q'], q2, sel))
check('Peer_Selection 시트가 만들어진다', 'Peer_Selection' in wb2.sheetnames, str(wb2.sheetnames))
if 'Peer_Selection' in wb2.sheetnames:
    ws2 = wb2['Peer_Selection']
    body = [(ws2.cell(r, 1).value, ws2.cell(r, 2).value) for r in range(5, ws2.max_row + 1)]
    check('고르지 않은 후보까지 전부 남는다', len(body) == 3, str(body))
    check('고른 회사에만 표시가 붙는다',
          {c for m, c in body if m == '●'} == {'005930', '000660'}, str(body))

# --- 7. 업종에서 고르지 않았으면 시트를 만들지 않는다 -------------------------
wb3 = load_workbook(M.export_gpcm_excel('2024.1Q', '1Q', ['005930'], sm2, bs2, pl2, mkt2, nm2,
                                        wacc, '5Y Monthly', [], 0.3, bd2,
                                        pd.DataFrame(mul2), ['2024.1Q'], q2))
check('직접 입력이면 Peer_Selection 시트가 없다', 'Peer_Selection' not in wb3.sheetnames)

print()
print(f"잘못된 항목 {len(fails)}건")
sys.exit(1 if fails else 0)
