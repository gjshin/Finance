# 최신 수정: 2026-02-16 16:00 KST
# 주요 변경사항:
# - D/E Ratio 컬럼 추가 (IBD / (시총 + NCI))
# - Unlevered Beta 계산에 D/E Ratio 사용 (정확한 Hamada 공식)
# - 국내 베타 계산 수정: 시장 지수는 yfinance 사용 (^KS11, ^KQ11)
# - Debt Ratio 수식 수정: IBD/(시총+IBD+NCI) [총부채/총자산]
# - 노트 업데이트: 주가 데이터 소스 명시 (yfinance 통합)

import streamlit as st
import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
import io
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import ChartLines
# import FinanceDataReader as fdr (Removed as per user request to use yfinance only)
from scipy import stats
import requests
from bs4 import BeautifulSoup

# ==========================================
# Module-level constants & small helpers
# ==========================================

# 국가 코드 → KPMG 테이블 국가명 매핑 (두 곳에서 공통 사용)
COUNTRY_NAME_MAP = {
    'US': 'United States', 'JP': 'Japan', 'CN': 'China', 'IN': 'India',
    'CA': 'Canada', 'GB': 'United Kingdom', 'DE': 'Germany', 'FR': 'France',
    'AU': 'Australia', 'AT': 'Austria', 'NL': 'Netherlands', 'BE': 'Belgium',
    'IT': 'Italy', 'ES': 'Spain', 'SE': 'Sweden', 'NO': 'Norway',
    'DK': 'Denmark', 'FI': 'Finland', 'IE': 'Ireland', 'CH': 'Switzerland',
    'MX': 'Mexico',
}


def label_sort_key(x):
    """'Recent' → -1, 'Y' → 0, 'Y-N' → N."""
    if x == 'Recent':
        return -1
    if x == 'Y':
        return 0
    try:
        return int(x.split('-')[1])
    except Exception:
        return 999


def norm_df(df):
    """DataFrame 컬럼의 timezone-aware datetime을 tz-naive로 표준화."""
    if df is not None and not df.empty:
        df.columns = [d.replace(tzinfo=None) if hasattr(d, 'tzinfo') else d for d in df.columns]
    return df


def _to_price_series(df, col='Close'):
    """yf.download / stock.history 결과에서 1-D 가격 Series 추출.

    yfinance >= 0.2.31 에서 yf.download 가 multi-level column DataFrame 을
    반환할 수 있으므로, 항상 1-D Series 를 보장한다.
    """
    if isinstance(df, pd.Series):
        return df
    if col in df.columns:
        s = df[col]
    else:
        s = df.iloc[:, 0]
    # multi-level column 이면 첫 번째 열만 선택
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0]
    return s


# ==========================================
# 1. Helper Functions (v17 Logic + Beta Calculation)
# ==========================================

def get_market_index(ticker):
    """
    티커 기반으로 거래소 및 시장지수 코드 반환
    Returns: (exchange_name, index_symbol)
    """
    ticker_upper = ticker.upper()

    # 아시아
    if ticker_upper.endswith('.KS'):
        return 'KOSPI', '^KS11'
    elif ticker_upper.endswith('.KQ'):
        return 'KOSDAQ', '^KQ11'
    elif ticker_upper.endswith('.T'):
        return 'TSE', '^N225'  # Nikkei 225
    elif ticker_upper.endswith('.SS'):
        return 'SSE', '000001.SS'  # Shanghai Composite
    elif ticker_upper.endswith('.SZ'):
        return 'SZSE', '399001.SZ'  # Shenzhen Component
    elif ticker_upper.endswith('.HK'):
        return 'HKEX', '^HSI'  # Hang Seng
    elif ticker_upper.endswith('.SI'):
        return 'SGX', '^STI'  # Straits Times Index
    elif ticker_upper.endswith('.TW'):
        return 'TWSE', '^TWII'  # Taiwan Weighted
    elif ticker_upper.endswith('.BO') or ticker_upper.endswith('.NS'):
        return 'NSE', '^NSEI'  # Nifty 50

    # 북미
    elif ticker_upper.endswith('.TO') or ticker_upper.endswith('.V'):
        return 'TSX', '^GSPTSE'  # S&P/TSX Composite
    elif ticker_upper.endswith('.MX'):
        return 'BMV', '^MXX'  # IPC Mexico

    # 유럽
    elif ticker_upper.endswith('.F') or ticker_upper.endswith('.DE') or ticker_upper.endswith('.BE'):
        return 'XETRA', '^GDAXI'  # DAX
    elif ticker_upper.endswith('.PA'):
        return 'Euronext', '^FCHI'  # CAC 40
    elif ticker_upper.endswith('.L'):
        return 'LSE', '^FTSE'  # FTSE 100
    elif ticker_upper.endswith('.MI'):
        return 'Borsa', 'FTSEMIB.MI'  # FTSE MIB
    elif ticker_upper.endswith('.MC'):
        return 'BME', '^IBEX'  # IBEX 35
    elif ticker_upper.endswith('.AS'):
        return 'Euronext', '^AEX'  # AEX Amsterdam
    elif ticker_upper.endswith('.SW') or ticker_upper.endswith('.VX'):
        return 'SIX', '^SSMI'  # Swiss Market Index
    elif ticker_upper.endswith('.VI'):
        return 'VSE', '^ATX'  # ATX Vienna
    elif ticker_upper.endswith('.BR'):
        return 'Euronext', '^BFX'  # BEL 20
    elif ticker_upper.endswith('.ST'):
        return 'OMX', '^OMX'  # OMX Stockholm 30
    elif ticker_upper.endswith('.OL'):
        return 'OSE', 'OSEBX.OL'  # Oslo Børs Index
    elif ticker_upper.endswith('.CO'):
        return 'OMX', '^OMXC20'  # OMX Copenhagen 20
    elif ticker_upper.endswith('.HE'):
        return 'OMX', '^OMXH25'  # OMX Helsinki 25
    elif ticker_upper.endswith('.IR'):
        return 'Euronext', '^ISEQ'  # ISEQ Overall

    # 오세아니아
    elif ticker_upper.endswith('.AX'):
        return 'ASX', '^AORD'  # All Ordinaries
    elif ticker_upper.endswith('.NZ'):
        return 'NZX', '^NZ50'  # NZX 50

    # 기타
    elif ticker_upper.endswith('.SA'):
        return 'B3', '^BVSP'  # Bovespa
    elif ticker_upper.endswith('.ME'):
        return 'MOEX', 'IMOEX.ME'  # MOEX Russia
    elif ticker_upper.endswith('.JO'):
        return 'JSE', 'J203.JO'  # FTSE/JSE Top 40

    # 기본값: 미국 S&P 500
    else:
        return 'US', '^GSPC'  # S&P 500

EMPTY_BETA = {'raw': None, 'adj': None, 'r2': None, 'stderr': None, 'n': 0}


def align_prices(stock_prices, market_prices):
    """종목/지수 가격 시계열을 공통 날짜로 맞춘 뒤 수익률을 계산.

    엑셀 Beta_Calculation 시트가 '공통 날짜로 맞춘 표에서 위아래 행 변동률'을
    계산하므로, 파이썬도 동일하게 (교집합 → 변동률) 순서로 계산해야
    두 결과가 일치한다.

    Returns: (stock_prices_aligned, market_prices_aligned, stock_returns, market_returns)
    """
    for s in (stock_prices, market_prices):
        if s is not None and getattr(s.index, 'tz', None) is not None:
            s.index = s.index.tz_localize(None)

    common_idx = stock_prices.index.intersection(market_prices.index)
    s_px = stock_prices.loc[common_idx].sort_index().dropna()
    m_px = market_prices.loc[common_idx].sort_index().dropna()

    common_idx2 = s_px.index.intersection(m_px.index)
    s_px = s_px.loc[common_idx2]
    m_px = m_px.loc[common_idx2]

    return s_px, m_px, s_px.pct_change().dropna(), m_px.pct_change().dropna()


def calculate_beta(stock_returns, market_returns, min_periods=20):
    """
    주식 수익률과 시장 수익률로부터 베타 및 회귀 통계량 계산

    Returns: dict(raw, adj, r2, stderr, n) — 산출 불가 시 값은 None
    - raw    : 회귀계수 (Raw Beta)
    - adj    : 조정 베타 = 2/3 × Raw + 1/3 × 1.0 (Bloomberg 방식)
    - r2     : 결정계수 (설명력). 통상 0.1 미만이면 베타 신뢰도 낮음
    - stderr : 회귀계수 표준오차. Raw Beta ± 2×stderr 가 약 95% 신뢰구간
    - n      : 회귀에 사용된 관측치 수
    """
    try:
        if stock_returns is None or market_returns is None:
            return dict(EMPTY_BETA)

        common_idx = stock_returns.index.intersection(market_returns.index)
        if len(common_idx) < min_periods:
            return dict(EMPTY_BETA)

        stock_ret = stock_returns.loc[common_idx]
        market_ret = market_returns.loc[common_idx]

        slope, intercept, r_value, p_value, std_err = stats.linregress(market_ret, stock_ret)
        raw_beta = slope
        adjusted_beta = (2/3) * raw_beta + (1/3) * 1.0

        # 값 검증: NaN, inf, 극단값(±10 초과) 필터링
        if not np.isfinite(raw_beta) or not np.isfinite(adjusted_beta):
            return dict(EMPTY_BETA)
        if abs(raw_beta) > 10 or abs(adjusted_beta) > 10:
            return dict(EMPTY_BETA)

        return {
            'raw': float(raw_beta),
            'adj': float(adjusted_beta),
            'r2': float(r_value ** 2) if np.isfinite(r_value) else None,
            'stderr': float(std_err) if np.isfinite(std_err) else None,
            'n': int(len(common_idx)),
        }

    except Exception:
        return dict(EMPTY_BETA)

@st.cache_data(ttl=86400)
def get_kpmg_tax_rates():
    """
    KPMG 사이트(OECD 데이터 기반) 법인세율 자동 크롤링 및 주요 국가 하드코딩 Fallback
    """
    # [Fallback] 주요 국가 최신 실효 법인세율 (Combined Rate, 2025 KPMG/OECD 기준)
    tax_rates = {
        'Korea': 0.264, 'United States': 0.26, 'Japan': 0.2974, 
        'France': 0.3613, 'Germany': 0.3006, 'Austria': 0.23, 
        'Canada': 0.265, 'United Kingdom': 0.25, 'China': 0.25, 
        'India': 0.30, 'Australia': 0.30, 'Netherlands': 0.258, 
        'Spain': 0.25, 'Italy': 0.24, 'Switzerland': 0.148, # 스위스 평균
        'Ireland': 0.125, 'Vietnam': 0.20, 'Sweden': 0.206, 'Norway': 0.22,
        'Belgium': 0.25, 'Mexico': 0.30
    }
    try:
        url = "https://kpmg.com/dk/en/services/tax/corporate-tax/corporate-tax-rates-table.html"
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=5)
        soup = BeautifulSoup(response.content, 'html.parser')
        table = soup.find('table')
        if table:
            rows = table.find_all('tr')
            for row in rows[1:]:
                cols = row.find_all('td')
                if len(cols) >= 2:
                    country = cols[0].get_text(strip=True)
                    rate_text = cols[1].get_text(strip=True).replace('%','')
                    try:
                        rate = float(rate_text) / 100
                        tax_rates[country] = rate
                    except Exception:
                        continue
    except Exception:
        pass
    return tax_rates

def get_corporate_tax_rates_from_wikipedia():
    # [KPMG 사이트 우선 사용하도록 랩핑]
    kpmg_rates = get_kpmg_tax_rates()
    # 기존 위키백과 로직은 백업용으로 유지하거나 KPMG 결과와 병합
    return kpmg_rates

# 한국 법인세 한계세율표 (사업연도별, 지방소득세 10% 포함)
# 각 구간: (과세표준 상한(백만원), 한계세율)  — 상한 None = 초과 구간
# · FY2018~2022 : 국세 10 / 20 / 22 / 25%
# · FY2023~2025 : 국세  9 / 19 / 21 / 24%  (2022년 세법개정, 1%p 인하)
# · FY2026~     : 국세 10 / 20 / 22 / 25%  (2025년 세법개정, 1%p 인상 환원)
KR_TAX_BRACKETS_PRE2023 = [(200, 0.110), (20000, 0.220), (300000, 0.242), (None, 0.275)]
KR_TAX_BRACKETS_2023 = [(200, 0.099), (20000, 0.209), (300000, 0.231), (None, 0.264)]
KR_TAX_BRACKETS_2026 = [(200, 0.110), (20000, 0.220), (300000, 0.242), (None, 0.275)]


def get_korean_tax_brackets(fiscal_year):
    """사업연도에 적용되는 한국 법인세 한계세율표 반환."""
    try:
        fy = int(fiscal_year)
    except (TypeError, ValueError):
        fy = datetime.now().year
    if fy >= 2026:
        return KR_TAX_BRACKETS_2026
    if fy >= 2023:
        return KR_TAX_BRACKETS_2023
    return KR_TAX_BRACKETS_PRE2023


def get_korean_marginal_tax_rate(pretax_income_millions, fiscal_year=None):
    """
    한국 법인세 한계세율 산출 (지방소득세 포함, 사업연도별 세율표 적용)

    Parameters:
    - pretax_income_millions: 세전이익 (백만원). 과세표준의 대용치로 사용.
    - fiscal_year: 해당 재무제표의 결산 연도. 미지정 시 현재 연도 기준.

    Note: 결손(음수) 기업은 한계세율 개념이 성립하지 않으므로 최고구간이 아닌
          '2억 초과 ~ 200억' 구간 세율(일반적인 중소·중견 법인 수준)을 적용하고
          Data_Quality 시트에 별도 표시한다.
    """
    brackets = get_korean_tax_brackets(fiscal_year)

    if pd.isna(pretax_income_millions) or pretax_income_millions <= 0:
        # 결손 또는 미확보: 2구간(2억~200억) 세율을 기본값으로 사용
        return brackets[1][1]

    for upper, rate in brackets:
        if upper is None or pretax_income_millions <= upper:
            return rate
    return brackets[-1][1]

def get_country_from_ticker(ticker, info=None):
    """
    티커 및 info로부터 국가 코드 추출
    """
    if info and 'country' in info:
        c_val = info.get('country')
        # yfinance info의 country는 'France', 'Austria', 'Germany' 등 풀네임인 경우가 많음
        c_map_inv = {
            'France':'FR', 'Austria':'AT', 'Germany':'DE', 'Japan':'JP', 'South Korea':'KR', 'Korea':'KR',
            'United States':'US', 'United Kingdom':'GB', 'Canada':'CA', 'Netherlands':'NL', 'Spain':'ES', 'Italy':'IT'
        }
        if c_val in c_map_inv: return c_map_inv[c_val]
        if len(str(c_val)) == 2: return c_val.upper()

    ticker_upper = ticker.upper()
    # [수송] 수동 매핑 추가 (사용자 특정 사례)
    if 'EZM.F' in ticker_upper: return 'FR' # OPmobility SE는 프랑스 본사지만 독일 상장
    if 'PYT.VI' in ticker_upper: return 'AT' # Polytec Holding AG는 오스트리아

    # 아시아
    if ticker_upper.endswith('.KS') or ticker_upper.endswith('.KQ'): return 'KR'
    elif ticker_upper.endswith('.T'): return 'JP'
    elif ticker_upper.endswith('.SS') or ticker_upper.endswith('.SZ'): return 'CN'
    elif ticker_upper.endswith('.HK'): return 'HK'
    elif ticker_upper.endswith('.SI'): return 'SG'
    elif ticker_upper.endswith('.TW'): return 'TW'
    elif ticker_upper.endswith('.BO') or ticker_upper.endswith('.NS'): return 'IN'
    # 북미
    elif ticker_upper.endswith('.TO') or ticker_upper.endswith('.V'): return 'CA'
    elif ticker_upper.endswith('.MX'): return 'MX'
    # 유럽 (Suffix 기반 Fallback)
    elif ticker_upper.endswith('.F') or ticker_upper.endswith('.DE') or ticker_upper.endswith('.BE'): return 'DE'
    elif ticker_upper.endswith('.PA'): return 'FR'
    elif ticker_upper.endswith('.L'): return 'GB'
    elif ticker_upper.endswith('.MI'): return 'IT'
    elif ticker_upper.endswith('.MC'): return 'ES'
    elif ticker_upper.endswith('.AS'): return 'NL'
    elif ticker_upper.endswith('.SW') or ticker_upper.endswith('.VX'): return 'CH'
    elif ticker_upper.endswith('.VI'): return 'AT'
    elif ticker_upper.endswith('.BR'): return 'BE'
    elif ticker_upper.endswith('.ST'): return 'SE'
    elif ticker_upper.endswith('.OL'): return 'NO'
    elif ticker_upper.endswith('.CO'): return 'DK'
    elif ticker_upper.endswith('.HE'): return 'FI'
    elif ticker_upper.endswith('.IR'): return 'IE'
    # 오세아니아
    elif ticker_upper.endswith('.AX'): return 'AU'
    elif ticker_upper.endswith('.NZ'): return 'NZ'
    
    return 'US'

def calculate_unlevered_beta(levered_beta, debt, equity, tax_rate):
    """
    하마다 모형으로 Unlevered Beta 계산
    Unlevered Beta = Levered Beta / (1 + (1 - Tax Rate) * (Debt / Equity))
    """
    if pd.isna(levered_beta) or levered_beta is None:
        return None
    if pd.isna(debt) or pd.isna(equity) or equity == 0:
        return levered_beta

    unlevered = levered_beta / (1 + (1 - tax_rate) * (debt / equity))
    return unlevered

def get_korea_10y_treasury_yield(base_date_str, user_rf_rate=0.033):
    """
    무위험이자율 반환 (사용자 입력값)

    Parameters:
    - user_rf_rate: 사용자가 직접 입력한 무위험이자율 (기본값 3.3%)
    """
    if user_rf_rate is None:
        user_rf_rate = 0.033
    return user_rf_rate


# ---------------------------------------------------------
# [설정] 계정 맵핑 (NOA Option, 투자부동산, 우선주 등)
# ---------------------------------------------------------
BS_HIGHLIGHT_MAP = {
    'Cash And Cash Equivalents':           'Cash',
    'Other Short Term Investments':        'Cash',
    'Current Debt And Capital Lease Obligation': 'IBD',
    'Long Term Debt And Capital Lease Obligation': 'IBD',
    'Minority Interest':                   'NCI',
    'Preferred Stock':                     'Preferred',
    'Preferred Securities Outside Stock Equity': 'Preferred',
    'Stockholders Equity':                 'Equity',
    'Long Term Equity Investment':                         'NOA(Option)',
    'Investments In Other Ventures Under Equity Method':   'NOA(Option)',
    'Investment In Financial Assets':                      'NOA(Option)',
    'Investmentin Financial Assets':                       'NOA(Option)',
    'Investment Properties':                               'NOA(Option)',
    'Non Current Note Receivables':                        'NOA(Option)',
    'Other Investments':                                   'NOA(Option)',
    'NOA':                                                 'NOA',
}
BS_SUBTOTAL_EXCLUDE = {
    'Cash Cash Equivalents And Short Term Investments', 'Cash And Short Term Investments',
    'Total Debt', 'Total Capitalization', 'Total Equity Gross Minority Interest',
}
PL_HIGHLIGHT_MAP = {
    'Total Revenue': 'Revenue', 'Operating Income': 'EBIT', 'EBIT': 'EBIT_yf',
    'EBITDA': 'EBITDA_yf', 'Normalized EBITDA': 'EBITDA_yf',
    'Net Income Common Stockholders': 'NI_Parent', 'Net Income': 'Net Income',
}
PL_CALC_KEY = {
    'Total Revenue': 'Revenue', 'Operating Income': 'OpIncome', 'EBIT': 'EBIT_yf',
    'EBITDA': 'EBITDA_yf', 'Normalized EBITDA': 'NormEBITDA_yf',
    'Net Income Common Stockholders': 'NI_Parent',
}
PL_SORT = {'Total Revenue': 10, 'Operating Revenue': 11, 'Cost Of Revenue': 20, 'Gross Profit': 30, 'Operating Expense': 35, 'Selling General And Administration': 36, 'Research And Development': 37, 'Operating Income': 50, 'EBIT': 55, 'EBITDA': 56, 'Normalized EBITDA': 57, 'Interest Expense': 60, 'Pretax Income': 70, 'Tax Provision': 75, 'Net Income': 90, 'Net Income Common Stockholders': 91, 'Basic EPS': 95, 'Diluted EPS': 96}

# LTM(최근 12개월) 검증 기준
# 정상적인 분기공시라면 최근 4개 분기의 '결산일 최초~최종 간격'은 약 9개월(270일)이다.
# 반기공시 기업(유럽 등)은 4개를 더하면 2년치가 되므로 이 범위를 크게 벗어난다.
LTM_SPAN_MIN_DAYS = 240
LTM_SPAN_MAX_DAYS = 320
LTM_KEY_ACCOUNTS = ('Total Revenue', 'Operating Income')

# Data_Quality 시트 심각도
SEV_ERROR = 'ERROR'
SEV_WARN = 'WARN'
SEV_INFO = 'INFO'


def _empty_gpcm(company_name, ticker, currency, exchange, market_idx, base_date_str, pl_source):
    """GPCM 계산용 딕셔너리 기본 구조."""
    return {
        'Company': company_name, 'Ticker': ticker, 'Currency': currency,
        'Base_Date': base_date_str,
        # 손익은 기준일과 다른 시점의 것이 쓰일 수 있다. 회사 결산월이 기준일과
        # 어긋나면 '기준일 이전 최신 회계연도'로 대체되고, LTM이면 최근 분기말이
        # 끝점이 된다. 라벨을 기준일로 두면 3월 데이터에 6월이 찍힌다.
        'PL_Date': base_date_str,
        'BS_Date': None,
        'Cash': 0, 'IBD': 0, 'NCI': 0, 'Preferred': 0, 'NOA(Option)': 0, 'NOA': 0, 'Equity': 0,
        'Revenue': 0, 'EBIT': 0, 'EBITDA': 0, 'DA': 0, 'NI_Parent': 0,
        'Close': 0, 'Shares': 0, 'Market_Cap_M': 0, 'Equity_Value_M': 0,
        'PL_Source': pl_source,
        'Exchange': exchange, 'Market_Index': market_idx,
        'Beta_5Y_Monthly_Raw': None, 'Beta_5Y_Monthly_Adj': None,
        'Beta_5Y_R2': None, 'Beta_5Y_StdErr': None, 'Beta_5Y_N': 0,
        'Beta_2Y_Weekly_Raw': None, 'Beta_2Y_Weekly_Adj': None,
        'Beta_2Y_R2': None, 'Beta_2Y_StdErr': None, 'Beta_2Y_N': 0,
        'Pretax_Income': 0, 'Tax_Rate': 0.209, 'Tax_Basis': '-',
        'DE_Ratio': 0, 'Debt_Ratio': 0,
        'Unlevered_Beta_5Y': None, 'Unlevered_Beta_2Y': None,
        'DA_Source': '-', 'DA_Imputed': False,
    }


def _price_on_or_before(price_series, target_dt, max_lookback_days=10):
    """target_dt 이전(포함) 가장 가까운 종가와 그 날짜를 반환."""
    if price_series is None or price_series.empty:
        return 0.0, '-'
    window = price_series.loc[:target_dt]
    if window.empty:
        return 0.0, '-'
    last_dt = window.index[-1]
    if (target_dt - last_dt).days > max_lookback_days:
        return 0.0, '-'
    return float(window.iloc[-1]), last_dt.strftime('%Y-%m-%d')


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_ticker_bundle(ticker, base_period_str, lookback, force_annual, include_recent):
    """
    단일 티커의 원천 데이터를 수집한다.

    [속도] 이 함수는 WACC 파라미터(Rf/MRP/Kd/세율/Size Premium)를 인자로 받지 않는다.
    따라서 사용자가 사이드바에서 WACC 가정만 바꾸면 캐시가 그대로 재사용되어
    야후 파이낸스를 다시 호출하지 않는다.

    [주의] 캐시 함수 안에서는 st.warning 등 화면 출력을 하지 않는다.
           문제 상황은 flags 리스트에 담아 반환하고, 호출부에서 화면에 표시한다.
    """
    base_dt = pd.to_datetime(base_period_str)
    flags = []

    def add_flag(level, label, item, message):
        flags.append({'Ticker': ticker, 'Label': label, 'Level': level,
                      'Item': item, 'Message': message})

    bundle = {
        'ticker': ticker, 'ok': False, 'error': None,
        'company': ticker, 'currency': 'USD', 'exchange': '', 'market_index': '',
        'periods': {}, 'bs_rows': [], 'pl_rows': [], 'market_rows': [],
        'price_series': None, 'flags': flags,
    }

    try:
        stock = yf.Ticker(ticker)
        info = stock.info or {}
        company_name = info.get('longName') or info.get('shortName') or ticker
        currency = info.get('currency', 'USD')
        exchange, market_idx = get_market_index(ticker)

        bundle.update({'company': company_name, 'currency': currency,
                       'exchange': exchange, 'market_index': market_idx})

        a_is = norm_df(stock.income_stmt)
        q_is = norm_df(stock.quarterly_income_stmt)
        a_bs = norm_df(stock.balance_sheet)
        q_bs = norm_df(stock.quarterly_balance_sheet)

        if a_is is None or a_is.empty:
            bundle['error'] = '연간 재무제표를 찾을 수 없습니다'
            add_flag(SEV_ERROR, '-', '재무제표', '연간 손익계산서 조회 실패 → 이 회사는 분석에서 제외됨')
            return bundle

        # -------------------------------------------------------------
        # [주가] 10년치를 한 번에 받아 두고 이후 슬라이싱해서 재사용한다.
        #  - hist_raw(auto_adjust=False): 시가총액 산출용 (실제 거래 주가)
        #  - hist_adj(auto_adjust=True) : 베타 산출용 (배당·분할 반영 총수익 기준)
        # -------------------------------------------------------------
        px_start = (base_dt - timedelta(days=365 * 10 + 20)).strftime('%Y-%m-%d')
        px_end = (base_dt + timedelta(days=1)).strftime('%Y-%m-%d')

        hist_raw = hist_adj = None
        try:
            hist_raw = _to_price_series(stock.history(start=px_start, end=px_end, auto_adjust=False))
            if getattr(hist_raw.index, 'tz', None) is not None:
                hist_raw.index = hist_raw.index.tz_localize(None)
        except Exception:
            hist_raw = None
        try:
            hist_adj = _to_price_series(stock.history(start=px_start, end=px_end, auto_adjust=True))
            if getattr(hist_adj.index, 'tz', None) is not None:
                hist_adj.index = hist_adj.index.tz_localize(None)
        except Exception:
            hist_adj = None

        if hist_raw is None or hist_raw.empty:
            add_flag(SEV_WARN, '-', '주가', '기준일 주가 조회 실패 → 시가총액이 0으로 계산됨')

        if hist_adj is not None and not hist_adj.empty:
            bundle['price_series'] = hist_adj

        # -------------------------------------------------------------
        # [라벨링] Y = 기준일(LTM), Y-1 = 직전 연간, Y-2 ...
        #  대원칙: 기준일 이후에 확정된 재무제표는 사용하지 않는다(Look-ahead 방지)
        # -------------------------------------------------------------
        all_fiscal_dates = sorted(list(a_is.columns), reverse=True)
        annuals_upto_base = [d for d in all_fiscal_dates if d <= base_dt + timedelta(days=7)]
        hist_annuals = [d for d in all_fiscal_dates if d < base_dt - timedelta(days=30)]

        labels_to_fetch = [('Y', base_dt)]
        for i in range(1, lookback):
            if i - 1 < len(hist_annuals):
                labels_to_fetch.append((f'Y-{i}', hist_annuals[i - 1]))

        for label, f_dt_obj in labels_to_fetch:
            if hasattr(f_dt_obj, 'tzinfo'):
                f_dt_obj = f_dt_obj.replace(tzinfo=None)
            f_dt_str = f_dt_obj.strftime('%Y-%m-%d')
            bs_shares = None

            gpcm = _empty_gpcm(company_name, ticker, currency, exchange, market_idx, f_dt_str, 'Annual')

            # ---------------------------------------------------------
            # [1] BS: '해당 라벨의 결산일' 이전 가장 최신 재무상태표
            #     (이전 버전은 모든 라벨에 기준일을 적용해 과거연도에도
            #      현재 BS가 들어가는 문제가 있었음)
            # ---------------------------------------------------------
            bs_cutoff = f_dt_obj + timedelta(days=7)
            all_possible_bs_dates = []
            if a_bs is not None and not a_bs.empty:
                all_possible_bs_dates.extend([d for d in a_bs.columns if d <= bs_cutoff])
            if q_bs is not None and not q_bs.empty:
                all_possible_bs_dates.extend([d for d in q_bs.columns if d <= bs_cutoff])
            all_possible_bs_dates = sorted(set(all_possible_bs_dates), reverse=True)

            target_bs = None
            actual_bs_date = None
            if all_possible_bs_dates:
                actual_bs_date = all_possible_bs_dates[0]
                if a_bs is not None and actual_bs_date in a_bs.columns:
                    target_bs = a_bs[actual_bs_date]
                elif q_bs is not None and actual_bs_date in q_bs.columns:
                    target_bs = q_bs[actual_bs_date]

            if target_bs is None:
                add_flag(SEV_WARN, label, '재무상태표', f'{f_dt_str} 이전 재무상태표 없음 → EV 구성항목이 0')
            else:
                gpcm['BS_Date'] = actual_bs_date.strftime('%Y-%m-%d')
                gap_days = (f_dt_obj - actual_bs_date).days
                if gap_days > 190:
                    add_flag(SEV_WARN, label, '재무상태표',
                             f'BS 기준일({gpcm["BS_Date"]})이 {f_dt_str}보다 {gap_days}일 이전 → 시점 불일치 확인 필요')

                for acct_name in target_bs.index:
                    val = target_bs.loc[acct_name]
                    if pd.isna(val):
                        continue
                    val_f = float(val)
                    if str(acct_name) == 'Ordinary Shares Number':
                        bs_shares = val_f
                    ev_tag = BS_HIGHLIGHT_MAP.get(str(acct_name), '')
                    if str(acct_name) in BS_SUBTOTAL_EXCLUDE:
                        ev_tag = ''

                    bundle['bs_rows'].append({
                        'Company': company_name, 'Ticker': ticker,
                        'Period': actual_bs_date.strftime('%Y-%m-%d'), 'Label': label,
                        'Currency': currency, 'Account': str(acct_name),
                        'EV_Tag': ev_tag, 'Amount_M': val_f / 1e6,
                    })
                    if ev_tag:
                        gpcm[ev_tag] += val_f / 1e6

            gpcm['Shares'] = bs_shares if bs_shares else float(info.get('sharesOutstanding', 0) or 0)

            # ---------------------------------------------------------
            # [2] 시가총액: 해당 결산일 시점의 실제 주가 사용
            # ---------------------------------------------------------
            close, p_date = _price_on_or_before(hist_raw, f_dt_obj)
            gpcm['Close'] = close
            gpcm['Market_Cap_M'] = (close * gpcm['Shares'] / 1e6) if gpcm['Shares'] else 0.0
            # 우선주 장부금액을 자기자본가치에 가산 (실무 관행)
            gpcm['Equity_Value_M'] = gpcm['Market_Cap_M'] + gpcm['Preferred']

            bundle['market_rows'].append({
                'Company': company_name, 'Ticker': ticker, 'Base_Date': f_dt_str,
                'Price_Date': p_date, 'Label': label, 'Currency': currency,
                'Close': close, 'Shares': gpcm['Shares'],
                'Preferred_M': round(gpcm['Preferred'], 1),
                'Market_Cap_M': round(gpcm['Market_Cap_M'], 1),
                'Equity_Value_M': round(gpcm['Equity_Value_M'], 1),
            })

            # ---------------------------------------------------------
            # [3] PL: 해당 결산일의 연간 데이터
            #     기준일에 딱 맞는 회계연도가 없으면 '기준일 이전' 최신 연간으로
            #     대체한다. (기준일 이후 회계연도를 쓰면 Look-ahead가 된다)
            # ---------------------------------------------------------
            calc_sums = {'Revenue': 0, 'OpIncome': 0, 'EBIT_yf': 0, 'EBITDA_yf': 0,
                         'NormEBITDA_yf': 0, 'NI_Parent': 0}

            f_pl_lookup = f_dt_obj
            if f_dt_obj not in a_is.columns:
                f_pl_lookup = annuals_upto_base[0] if (label == 'Y' and annuals_upto_base) else None
                if label == 'Y' and f_pl_lookup is not None and f_pl_lookup != f_dt_obj:
                    add_flag(SEV_INFO, label, '손익계산서',
                             f'기준일({f_dt_str}) 회계연도 없음 → 기준일 이전 최신 연간({f_pl_lookup.strftime("%Y-%m-%d")}) 사용')

            if f_pl_lookup is not None and f_pl_lookup in a_is.columns:
                gpcm['PL_Date'] = f_pl_lookup.strftime('%Y-%m-%d')
                for acct in a_is.index:
                    acct_str = str(acct)
                    hl_tag = PL_HIGHLIGHT_MAP.get(acct_str, '')
                    calc_key = PL_CALC_KEY.get(acct_str, '')
                    is_eps = 'EPS' in acct_str or 'Per Share' in acct_str
                    is_shares = 'Average Shares' in acct_str

                    val = a_is.loc[acct, f_pl_lookup]
                    if pd.isna(val):
                        continue
                    val_f = float(val)
                    if is_eps:
                        unit, amt = 'per share', val_f
                    elif is_shares:
                        unit, amt = 'M shares', val_f / 1e6
                    else:
                        unit, amt = 'M', val_f / 1e6

                    bundle['pl_rows'].append({
                        'Company': company_name, 'Ticker': ticker, 'Currency': currency,
                        'Account': acct_str, 'GPCM_Tag': hl_tag, 'PL_Source': 'Annual',
                        'Q_Label': 'Annual', 'Period': f_pl_lookup.strftime('%Y-%m-%d'),
                        'Label': label, 'Amount_M': amt, 'Unit': unit,
                        '_sort': PL_SORT.get(acct_str, 500),
                    })
                    if calc_key and not is_eps and not is_shares:
                        calc_sums[calc_key] += val_f / 1e6
            else:
                add_flag(SEV_WARN, label, '손익계산서', f'{f_dt_str} 시점 연간 손익 데이터 없음')

            gpcm['Revenue'] = calc_sums['Revenue']
            gpcm['EBIT'] = calc_sums['OpIncome']
            da_amount, da_source, da_imputed = _derive_da(calc_sums)
            gpcm['DA'] = da_amount
            gpcm['DA_Source'] = da_source
            gpcm['DA_Imputed'] = da_imputed
            gpcm['EBITDA'] = calc_sums['OpIncome'] + da_amount
            gpcm['NI_Parent'] = calc_sums['NI_Parent']

            # ---------------------------------------------------------
            # [4] 법인세 한계세율 (사업연도 기준)
            # ---------------------------------------------------------
            country_code = get_country_from_ticker(ticker, info)
            pl_year = f_pl_lookup.year if f_pl_lookup is not None else f_dt_obj.year
            if 'Pretax Income' in a_is.index and f_pl_lookup is not None and f_pl_lookup in a_is.columns:
                pt_val = a_is.loc['Pretax Income', f_pl_lookup]
                if pd.notna(pt_val):
                    gpcm['Pretax_Income'] = float(pt_val) / 1e6
            _apply_tax_rate(gpcm, country_code, pl_year)

            # ---------------------------------------------------------
            # [5] LTM 조정 (기준연도 Y 한정)
            # ---------------------------------------------------------
            if label == 'Y' and not force_annual:
                _apply_ltm(gpcm, bundle, q_is, base_dt, ticker, company_name, currency,
                           label, country_code, add_flag)

            # ---------------------------------------------------------
            # [6] 자본구조 (이전 버전에서 계산 자체가 누락되어 있던 항목)
            #     D/E   = IBD / (자기자본가치 + NCI)
            #     D/V   = IBD / (IBD + 자기자본가치 + NCI)
            # ---------------------------------------------------------
            equity_mv = gpcm['Equity_Value_M'] + gpcm['NCI']
            gpcm['DE_Ratio'] = (gpcm['IBD'] / equity_mv) if equity_mv > 0 else 0.0
            total_cap = gpcm['IBD'] + equity_mv
            gpcm['Debt_Ratio'] = (gpcm['IBD'] / total_cap) if total_cap > 0 else 0.0
            if gpcm['Market_Cap_M'] <= 0:
                add_flag(SEV_WARN, label, '자본구조', '시가총액 0 → 부채비율·멀티플 산출 불가')

            if gpcm['DA_Imputed'] and gpcm['EBIT'] != 0:
                add_flag(SEV_WARN, label, 'EBITDA',
                         'yfinance에 EBITDA/EBIT 항목이 없어 감가상각비를 0으로 처리함 '
                         '→ EBITDA = 영업이익. 사업보고서로 D&A 직접 확인 필요')

            bundle['periods'][label] = gpcm

        # -------------------------------------------------------------
        # [7] 최신 분기(Recent) — 과거재무정보 요약 모드용
        # -------------------------------------------------------------
        if include_recent:
            _collect_recent(bundle, stock, info, q_is, q_bs, company_name, ticker,
                            currency, exchange, market_idx, add_flag)

        # -------------------------------------------------------------
        # [8] 베타 — 기준연도(Y)에 대해서만 산출
        #     종목·지수 모두 배당 반영(총수익) 기준으로 통일한다.
        # -------------------------------------------------------------
        base_gpcm = bundle['periods'].get('Y')
        if base_gpcm is not None:
            _compute_betas(base_gpcm, hist_adj, market_idx, base_dt, add_flag)

        bundle['ok'] = True
        return bundle

    except Exception as e:
        bundle['error'] = str(e)
        add_flag(SEV_ERROR, '-', '수집', f'데이터 수집 중 오류: {e}')
        return bundle


def _derive_da(calc_sums):
    """
    감가상각비(D&A) 역산.

    야후 파이낸스는 현금흐름표 D&A를 안정적으로 제공하지 않으므로
    'yfinance EBITDA − yfinance EBIT'로 역산한다.
    둘 중 하나라도 없으면 역산이 불가능하며, 이 경우 D&A를 0으로 두고
    (= EBITDA가 영업이익과 같아짐) 사용자가 직접 확인하도록 플래그를 남긴다.

    Returns: (da_amount, da_source, da_imputed)
    """
    ebitda_yf = calc_sums['EBITDA_yf'] if calc_sums['EBITDA_yf'] != 0 else calc_sums['NormEBITDA_yf']
    ebit_yf = calc_sums['EBIT_yf']
    if ebitda_yf != 0 and ebit_yf != 0:
        return (ebitda_yf - ebit_yf), 'yfinance (EBITDA − EBIT)', False
    return 0.0, '⚠ 역산 불가 (D&A=0, EBITDA=영업이익)', True


def _apply_tax_rate(gpcm, country_code, fiscal_year):
    """국가별 법인세율 적용. 한국은 사업연도별 한계세율표를 사용한다."""
    if country_code == 'KR':
        gpcm['Tax_Rate'] = get_korean_marginal_tax_rate(gpcm['Pretax_Income'], fiscal_year)
        if pd.isna(gpcm['Pretax_Income']) or gpcm['Pretax_Income'] <= 0:
            gpcm['Tax_Basis'] = f'KR 한계세율 {fiscal_year} (결손 → 2구간 적용)'
        else:
            gpcm['Tax_Basis'] = f'KR 한계세율 {fiscal_year}'
    else:
        k_rates = get_corporate_tax_rates_from_wikipedia()
        c_full_name = COUNTRY_NAME_MAP.get(country_code, '')
        if c_full_name in k_rates:
            gpcm['Tax_Rate'] = k_rates[c_full_name]
            gpcm['Tax_Basis'] = f'KPMG {c_full_name}'
        else:
            gpcm['Tax_Rate'] = 0.25
            gpcm['Tax_Basis'] = f'기본값 25% ({country_code} 세율표 미보유)'


def _apply_ltm(gpcm, bundle, q_is, base_dt, ticker, company_name, currency,
               label, country_code, add_flag):
    """
    최근 4개 분기 합산(LTM)으로 손익을 대체한다.

    [검증] 단순히 '분기 4개'를 더하면 다음 두 가지를 놓친다.
      (1) 반기공시 기업 → 4개를 더하면 2년치가 된다
      (2) 특정 분기의 계정이 비어 있으면 3개 분기 합계가 조용히 LTM이 된다
    두 경우 모두 LTM을 포기하고 연간 데이터를 유지한다.
    """
    try:
        if q_is is None or q_is.empty:
            return

        q_cols = sorted(list(q_is.columns), reverse=True)
        valid_q_dates = [d for d in q_cols if d <= base_dt + timedelta(days=15)]
        if len(valid_q_dates) < 4:
            add_flag(SEV_INFO, label, 'LTM',
                     f'기준일 이전 분기 데이터 {len(valid_q_dates)}개뿐 → 연간 데이터 사용')
            return

        ltm_q_dates = valid_q_dates[:4]
        recent_q_dt = ltm_q_dates[0]
        oldest_q_dt = ltm_q_dates[-1]

        # (1) 기간 검증
        span_days = (recent_q_dt - oldest_q_dt).days
        if not (LTM_SPAN_MIN_DAYS <= span_days <= LTM_SPAN_MAX_DAYS):
            add_flag(SEV_WARN, label, 'LTM',
                     f'최근 4개 공시기간 간격이 {span_days}일 (정상 약 270일). '
                     f'반기공시 기업으로 추정되어 LTM 합산을 건너뛰고 연간 데이터를 사용함')
            return

        # (2) 핵심 계정 결측 검증
        missing = []
        for acct in LTM_KEY_ACCOUNTS:
            if acct not in q_is.index:
                continue
            n_null = int(q_is.loc[acct, ltm_q_dates].isna().sum())
            if n_null > 0:
                missing.append(f'{acct}({n_null}개 분기 누락)')
        if missing:
            add_flag(SEV_WARN, label, 'LTM',
                     f'분기 데이터 결측: {", ".join(missing)} → LTM 합산을 건너뛰고 연간 데이터를 사용함')
            return

        # 기존 연간 PL 행 제거 후 4개 분기 상세를 기록 (검증 가능하도록)
        bundle['pl_rows'][:] = [r for r in bundle['pl_rows']
                                if not (r['Ticker'] == ticker and r['Label'] == label)]

        for i, q_dt in enumerate(ltm_q_dates):
            q_col = q_is[q_dt]
            for acct in q_is.index:
                acct_str = str(acct)
                val_q = q_col.loc[acct]
                if pd.isna(val_q):
                    continue
                val_q = float(val_q)
                is_eps = 'EPS' in acct_str or 'Per Share' in acct_str
                bundle['pl_rows'].append({
                    'Company': company_name, 'Ticker': ticker, 'Currency': currency,
                    'Account': acct_str, 'GPCM_Tag': PL_HIGHLIGHT_MAP.get(acct_str, ''),
                    'PL_Source': 'Quarterly (4Q Sum)', 'Q_Label': f'D-{i}Q',
                    'Period': q_dt.strftime('%Y-%m-%d'), 'Label': label,
                    'Amount_M': val_q if is_eps else val_q / 1e6,
                    'Unit': 'per share' if is_eps else 'M',
                    '_sort': PL_SORT.get(acct_str, 500),
                })

        ltm_sum_vals = q_is[ltm_q_dates].sum(axis=1, min_count=1)
        calc_sums_ltm = {'Revenue': 0, 'OpIncome': 0, 'EBIT_yf': 0, 'EBITDA_yf': 0,
                         'NormEBITDA_yf': 0, 'NI_Parent': 0}
        for acct in ltm_sum_vals.index:
            acct_str = str(acct)
            calc_key = PL_CALC_KEY.get(acct_str, '')
            is_eps = 'EPS' in acct_str or 'Per Share' in acct_str
            is_shares = 'Average Shares' in acct_str
            val_ltm = ltm_sum_vals.loc[acct]
            if pd.isna(val_ltm):
                continue
            if calc_key and not is_eps and not is_shares:
                calc_sums_ltm[calc_key] += float(val_ltm) / 1e6

        gpcm['Revenue'] = calc_sums_ltm['Revenue']
        gpcm['EBIT'] = calc_sums_ltm['OpIncome']
        da_amount, da_source, da_imputed = _derive_da(calc_sums_ltm)
        gpcm['DA'] = da_amount
        gpcm['DA_Source'] = da_source
        gpcm['DA_Imputed'] = da_imputed
        gpcm['EBITDA'] = calc_sums_ltm['OpIncome'] + da_amount
        gpcm['NI_Parent'] = calc_sums_ltm['NI_Parent']
        gpcm['PL_Source'] = f'LTM (4Q Sum, {span_days}일)'
        gpcm['PL_Date'] = recent_q_dt.strftime('%Y-%m-%d')

        if 'Pretax Income' in ltm_sum_vals.index and pd.notna(ltm_sum_vals.loc['Pretax Income']):
            gpcm['Pretax_Income'] = float(ltm_sum_vals.loc['Pretax Income']) / 1e6
        _apply_tax_rate(gpcm, country_code, recent_q_dt.year)

    except Exception as e:
        add_flag(SEV_WARN, label, 'LTM', f'LTM 계산 중 오류 → 연간 데이터 사용 ({e})')


def _collect_recent(bundle, stock, info, q_is, q_bs, company_name, ticker,
                    currency, exchange, market_idx, add_flag):
    """최신 분기 재무정보 수집 (과거재무정보 요약 모드용)."""
    try:
        if q_is is None or q_is.empty or q_bs is None or q_bs.empty:
            add_flag(SEV_INFO, 'Recent', '분기재무', '분기 재무제표 없음 → 최신 분기 시트 생략')
            return

        recent_f_dt = sorted(list(q_is.columns), reverse=True)[0]
        recent_f_str = recent_f_dt.strftime('%Y-%m-%d')
        gpcm_r = _empty_gpcm(company_name, ticker, currency, exchange, market_idx,
                             recent_f_str, 'Quarterly (Recent)')

        recent_bs_dt = sorted(list(q_bs.columns), reverse=True)[0]
        recent_bs_data = q_bs[recent_bs_dt]
        gpcm_r['BS_Date'] = recent_bs_dt.strftime('%Y-%m-%d')
        bs_shares_r = None
        for acct_name in recent_bs_data.index:
            val = recent_bs_data.loc[acct_name]
            if pd.isna(val):
                continue
            val_f = float(val)
            if str(acct_name) == 'Ordinary Shares Number':
                bs_shares_r = val_f
            ev_tag = BS_HIGHLIGHT_MAP.get(str(acct_name), '')
            if str(acct_name) in BS_SUBTOTAL_EXCLUDE:
                ev_tag = ''
            bundle['bs_rows'].append({
                'Company': company_name, 'Ticker': ticker,
                'Period': gpcm_r['BS_Date'], 'Label': 'Recent', 'Currency': currency,
                'Account': str(acct_name), 'EV_Tag': ev_tag, 'Amount_M': val_f / 1e6,
            })
            if ev_tag:
                gpcm_r[ev_tag] += val_f / 1e6

        gpcm_r['Shares'] = bs_shares_r if bs_shares_r else float(info.get('sharesOutstanding', 0) or 0)

        try:
            hist_r = stock.history(period='5d', auto_adjust=False)
            close_r = float(hist_r['Close'].iloc[-1]) if not hist_r.empty else float(info.get('previousClose', 0) or 0)
            p_date_r = hist_r.index[-1].strftime('%Y-%m-%d') if not hist_r.empty else '-'
        except Exception:
            close_r, p_date_r = 0.0, '-'

        gpcm_r['Close'] = close_r
        gpcm_r['Market_Cap_M'] = (close_r * gpcm_r['Shares'] / 1e6) if gpcm_r['Shares'] else 0.0
        gpcm_r['Equity_Value_M'] = gpcm_r['Market_Cap_M'] + gpcm_r['Preferred']
        bundle['market_rows'].append({
            'Company': company_name, 'Ticker': ticker, 'Base_Date': recent_f_str,
            'Price_Date': p_date_r, 'Label': 'Recent', 'Currency': currency,
            'Close': close_r, 'Shares': gpcm_r['Shares'],
            'Preferred_M': round(gpcm_r['Preferred'], 1),
            'Market_Cap_M': round(gpcm_r['Market_Cap_M'], 1),
            'Equity_Value_M': round(gpcm_r['Equity_Value_M'], 1),
        })

        recent_pl_data = q_is[recent_f_dt]
        calc_sums_r = {'Revenue': 0, 'OpIncome': 0, 'EBIT_yf': 0, 'EBITDA_yf': 0,
                       'NormEBITDA_yf': 0, 'NI_Parent': 0}
        for acct in recent_pl_data.index:
            acct_str = str(acct)
            val = recent_pl_data.loc[acct]
            if pd.isna(val):
                continue
            val_f = float(val)
            is_eps = 'EPS' in acct_str or 'Per Share' in acct_str
            is_shares = 'Average Shares' in acct_str
            bundle['pl_rows'].append({
                'Company': company_name, 'Ticker': ticker, 'Currency': currency,
                'Account': acct_str, 'GPCM_Tag': PL_HIGHLIGHT_MAP.get(acct_str, ''),
                'PL_Source': 'Quarterly', 'Q_Label': 'Recent', 'Period': recent_f_str,
                'Label': 'Recent', 'Amount_M': val_f if is_eps else val_f / 1e6,
                'Unit': 'per share' if is_eps else ('M shares' if is_shares else 'M'),
                '_sort': PL_SORT.get(acct_str, 500),
            })
            calc_key = PL_CALC_KEY.get(acct_str, '')
            if calc_key and not is_eps and not is_shares:
                calc_sums_r[calc_key] += val_f / 1e6

        gpcm_r['Revenue'] = calc_sums_r['Revenue']
        gpcm_r['EBIT'] = calc_sums_r['OpIncome']
        da_amount, da_source, da_imputed = _derive_da(calc_sums_r)
        gpcm_r['DA'] = da_amount
        gpcm_r['DA_Source'] = da_source
        gpcm_r['DA_Imputed'] = da_imputed
        gpcm_r['EBITDA'] = calc_sums_r['OpIncome'] + da_amount
        gpcm_r['NI_Parent'] = calc_sums_r['NI_Parent']

        bundle['periods']['Recent'] = gpcm_r

    except Exception as e:
        add_flag(SEV_WARN, 'Recent', '분기재무', f'최신 분기 수집 오류: {e}')


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_market_index_prices(market_idx, start_str, end_str):
    """시장지수 종가(배당 반영 기준). 여러 티커가 같은 지수를 공유하므로 별도 캐시."""
    try:
        df = yf.download(market_idx, start=start_str, end=end_str,
                         progress=False, auto_adjust=True)
        if df is None or df.empty:
            return None
        s = _to_price_series(df)
        if not isinstance(s.index, pd.DatetimeIndex):
            s.index = pd.to_datetime(s.index)
        if s.index.tz is not None:
            s.index = s.index.tz_localize(None)
        return s
    except Exception:
        return None


def _compute_betas(base_gpcm, hist_adj, market_idx, base_dt, add_flag):
    """
    5년 월간 / 2년 주간 베타를 산출한다.

    [기준 통일] 종목·지수 모두 auto_adjust=True(배당·분할 반영 총수익) 기준을 쓴다.
                이전 버전은 종목만 배당 미반영이어서 분자·분모 정의가 달랐다.
    [엑셀 일치] 가격을 공통 날짜로 맞춘 뒤 변동률을 계산해 엑셀 SLOPE와 결과를 일치시킨다.
    """
    for key in ('Stock_Monthly_Prices_5Y', 'Market_Monthly_Prices_5Y',
                'Stock_Weekly_Prices_2Y', 'Market_Weekly_Prices_2Y'):
        base_gpcm[key] = None

    if hist_adj is None or hist_adj.empty:
        add_flag(SEV_WARN, 'Y', '베타', '종목 주가 데이터 없음 → 베타 산출 불가')
        return

    start_5y = base_dt - timedelta(days=365 * 5 + 20)
    start_2y = base_dt - timedelta(days=365 * 2 + 20)
    mkt_prices = fetch_market_index_prices(
        market_idx, start_5y.strftime('%Y-%m-%d'), base_dt.strftime('%Y-%m-%d'))
    if mkt_prices is None or mkt_prices.empty:
        add_flag(SEV_WARN, 'Y', '베타', f'시장지수({market_idx}) 데이터 없음 → 베타 산출 불가')
        return

    # ---- 5Y Monthly ----
    s_m = hist_adj.loc[start_5y:base_dt].resample('ME').last().dropna()
    m_m = mkt_prices.loc[start_5y:base_dt].resample('ME').last().dropna()
    if len(s_m) >= 12 and len(m_m) >= 12:
        s_px, m_px, s_ret, m_ret = align_prices(s_m, m_m)
        res = calculate_beta(s_ret, m_ret, min_periods=20)
        base_gpcm['Stock_Monthly_Prices_5Y'] = s_px
        base_gpcm['Market_Monthly_Prices_5Y'] = m_px
        base_gpcm['Beta_5Y_Monthly_Raw'] = res['raw']
        base_gpcm['Beta_5Y_Monthly_Adj'] = res['adj']
        base_gpcm['Beta_5Y_R2'] = res['r2']
        base_gpcm['Beta_5Y_StdErr'] = res['stderr']
        base_gpcm['Beta_5Y_N'] = res['n']
        if res['adj'] is not None:
            equity_m = base_gpcm['Equity_Value_M'] + base_gpcm['NCI']
            base_gpcm['Unlevered_Beta_5Y'] = calculate_unlevered_beta(
                res['adj'], base_gpcm['IBD'], equity_m, base_gpcm['Tax_Rate'])
        if res['r2'] is not None and res['r2'] < 0.10:
            add_flag(SEV_INFO, 'Y', '베타',
                     f'5Y 월간 베타 R²={res["r2"]:.2f} (설명력 낮음) → 베타 신뢰도 확인 필요')
    else:
        add_flag(SEV_WARN, 'Y', '베타', '5년 월간 관측치 12개 미만 → 5Y 베타 생략')

    # ---- 2Y Weekly ----
    s_w = hist_adj.loc[start_2y:base_dt].resample('W').last().dropna()
    m_w = mkt_prices.loc[start_2y:base_dt].resample('W').last().dropna()
    if len(s_w) >= 50 and len(m_w) >= 50:
        s_px, m_px, s_ret, m_ret = align_prices(s_w, m_w)
        res = calculate_beta(s_ret, m_ret, min_periods=40)
        base_gpcm['Stock_Weekly_Prices_2Y'] = s_px
        base_gpcm['Market_Weekly_Prices_2Y'] = m_px
        base_gpcm['Beta_2Y_Weekly_Raw'] = res['raw']
        base_gpcm['Beta_2Y_Weekly_Adj'] = res['adj']
        base_gpcm['Beta_2Y_R2'] = res['r2']
        base_gpcm['Beta_2Y_StdErr'] = res['stderr']
        base_gpcm['Beta_2Y_N'] = res['n']
        if res['adj'] is not None:
            equity_m = base_gpcm['Equity_Value_M'] + base_gpcm['NCI']
            base_gpcm['Unlevered_Beta_2Y'] = calculate_unlevered_beta(
                res['adj'], base_gpcm['IBD'], equity_m, base_gpcm['Tax_Rate'])
    else:
        add_flag(SEV_WARN, 'Y', '베타', '2년 주간 관측치 50개 미만 → 2Y 베타 생략')


def trimmed_stats(values, method='none', low_pct=25, high_pct=75, iqr_k=1.5):
    """
    멀티플 통계 산출 (이상치 처리 포함).

    method:
      - 'none' : 전체 사용
      - 'iqr'  : 1사분위 - k×IQR ~ 3사분위 + k×IQR 밖을 제외
      - 'pct'  : low_pct ~ high_pct 백분위 밖을 제외

    Returns: (stats_dict, kept_values, excluded_values)
    """
    vals = [float(v) for v in values if v is not None and pd.notna(v) and np.isfinite(v)]
    if not vals:
        return {'Mean': None, 'Median': None, 'Max': None, 'Min': None, 'N': 0}, [], []

    arr = np.array(vals, dtype=float)
    if method == 'iqr' and len(arr) >= 4:
        q1, q3 = np.percentile(arr, [25, 75])
        iqr = q3 - q1
        lo, hi = q1 - iqr_k * iqr, q3 + iqr_k * iqr
    elif method == 'pct' and len(arr) >= 4:
        lo, hi = np.percentile(arr, [low_pct, high_pct])
    else:
        lo, hi = -np.inf, np.inf

    kept = [v for v in vals if lo <= v <= hi]
    excluded = [v for v in vals if not (lo <= v <= hi)]
    if not kept:
        kept, excluded = vals, []

    k = np.array(kept, dtype=float)
    return ({'Mean': float(np.mean(k)), 'Median': float(np.median(k)),
             'Max': float(np.max(k)), 'Min': float(np.min(k)), 'N': len(kept)},
            kept, excluded)


def get_gpcm_data(tickers_list, target_periods, mrp=0.08, kd_pretax=0.035, size_premium=0.0402,
                  target_tax_rate=0.264, user_rf_rate=None, beta_type="5Y",
                  force_annual=False, include_recent=False, outlier_method='none'):
    """
    GPCM 데이터 수집 및 엑셀 생성을 위한 데이터 구조 반환

    Parameters:
    - mrp: Market Risk Premium
    - kd_pretax: 세전 타인자본비용
    - size_premium: Size Premium (한국공인회계사회 기준)
    - target_tax_rate: Target 기업 법인세율
    - user_rf_rate: 사용자 입력 무위험이자율
    - beta_type: WACC 계산에 사용할 베타 유형 ("5Y" 또는 "2Y")
    - outlier_method: 피어 평균 산출 시 이상치 처리 방식 ('none'|'iqr'|'pct')

    [속도] 원천 데이터 수집은 티커 단위로 캐시되는 fetch_ticker_bundle이 담당하고,
           이 함수는 WACC 계산만 수행한다. 따라서 WACC 가정을 바꿔도 재크롤링이 없다.
    """
    base_period_str = target_periods[-1]
    rf_rate = get_korea_10y_treasury_yield(base_period_str, user_rf_rate)
    lookback = len(target_periods)

    rel_labels = ["Y" if i == 0 else f"Y-{i}" for i in range(lookback)]
    all_period_data = {lbl: {} for lbl in rel_labels}
    if include_recent:
        all_period_data['Recent'] = {}

    raw_bs_rows, raw_pl_rows, market_rows = [], [], []
    price_abs_dfs, price_rel_dfs = [], []
    ticker_to_name, data_quality_rows = {}, []

    progress_bar = st.progress(0)
    status_text = st.empty()
    total_tickers = max(len(tickers_list), 1)
    failed_tickers = []

    for idx, ticker in enumerate(tickers_list):
        status_text.text(f"Processing: {ticker}... ({idx + 1}/{total_tickers})")
        progress_bar.progress((idx + 1) / total_tickers)

        bundle = fetch_ticker_bundle(ticker, base_period_str, lookback, force_annual, include_recent)
        data_quality_rows.extend(bundle['flags'])

        if not bundle['ok']:
            failed_tickers.append(f"{ticker} ({bundle['error']})")
            continue

        ticker_to_name[ticker] = bundle['company']
        raw_bs_rows.extend(bundle['bs_rows'])
        raw_pl_rows.extend(bundle['pl_rows'])
        market_rows.extend(bundle['market_rows'])
        for label, gpcm in bundle['periods'].items():
            if label in all_period_data:
                all_period_data[label][ticker] = gpcm

        if bundle['price_series'] is not None and not bundle['price_series'].empty:
            s = bundle['price_series'].copy()
            s.name = ticker
            price_abs_dfs.append(s)

    status_text.empty()
    progress_bar.empty()

    if failed_tickers:
        st.warning("⚠️ 데이터 수집 실패: " + ", ".join(failed_tickers))

    # 상대주가: 모든 종목이 공통으로 존재하는 첫 거래일을 100으로 재기준화
    if price_abs_dfs:
        df_abs = pd.concat(price_abs_dfs, axis=1).sort_index()
        df_common = df_abs.dropna()
        if not df_common.empty:
            base_row = df_common.iloc[0]
            df_rel = (df_abs / base_row) * 100
        else:
            df_rel = df_abs / df_abs.ffill().bfill().iloc[0] * 100
        price_rel_dfs = [df_rel[c].dropna().rename(c) for c in df_rel.columns]

    # ========================================
    # Target WACC 계산 (기준연도 'Y' 데이터 기준)
    # ========================================
    base_gpcm_list = list(all_period_data['Y'].values())

    debt_ratios = [g['Debt_Ratio'] for g in base_gpcm_list if g['Debt_Ratio'] > 0]
    dr_stats, dr_kept, dr_excluded = trimmed_stats(debt_ratios, outlier_method)
    if dr_stats['N'] > 0:
        avg_debt_ratio = dr_stats['Mean']
        debt_ratio_basis = f"피어 {dr_stats['N']}개 평균"
    else:
        avg_debt_ratio = 0.30
        debt_ratio_basis = "산출 불가 → 기본값 30% 적용"
        data_quality_rows.append({'Ticker': '(전체)', 'Label': 'Y', 'Level': SEV_ERROR,
                                  'Item': '목표자본구조',
                                  'Message': '유효한 피어 부채비율이 없어 기본값 30%를 사용함'})

    beta_field = 'Unlevered_Beta_5Y' if beta_type == '5Y' else 'Unlevered_Beta_2Y'
    beta_label = "5Y Monthly" if beta_type == '5Y' else "2Y Weekly"
    unlevered_betas = [g[beta_field] for g in base_gpcm_list
                       if g[beta_field] is not None and g[beta_field] > 0]
    ub_stats, ub_kept, ub_excluded = trimmed_stats(unlevered_betas, outlier_method)
    if ub_stats['N'] > 0:
        avg_unlevered_beta = ub_stats['Mean']
        beta_basis = f"피어 {ub_stats['N']}개 평균"
    else:
        avg_unlevered_beta = 1.0
        beta_basis = "산출 불가 → 기본값 1.0 적용"
        data_quality_rows.append({'Ticker': '(전체)', 'Label': 'Y', 'Level': SEV_ERROR,
                                  'Item': 'Unlevered Beta',
                                  'Message': '유효한 피어 베타가 없어 기본값 1.0을 사용함'})

    if dr_excluded:
        data_quality_rows.append({'Ticker': '(전체)', 'Label': 'Y', 'Level': SEV_INFO,
                                  'Item': '이상치 제외(부채비율)',
                                  'Message': f'{len(dr_excluded)}개 제외: ' +
                                             ', '.join(f'{v*100:.1f}%' for v in dr_excluded)})
    if ub_excluded:
        data_quality_rows.append({'Ticker': '(전체)', 'Label': 'Y', 'Level': SEV_INFO,
                                  'Item': '이상치 제외(Unlevered β)',
                                  'Message': f'{len(ub_excluded)}개 제외: ' +
                                             ', '.join(f'{v:.2f}' for v in ub_excluded)})

    if avg_debt_ratio < 1.0:
        target_de_ratio = avg_debt_ratio / (1 - avg_debt_ratio)
        target_relevered_beta = avg_unlevered_beta * (1 + (1 - target_tax_rate) * target_de_ratio)
    else:
        target_de_ratio = 0.0
        target_relevered_beta = avg_unlevered_beta

    target_ke = rf_rate + mrp * target_relevered_beta + size_premium
    target_kd_aftertax = kd_pretax * (1 - target_tax_rate)
    equity_weight = 1 - avg_debt_ratio
    debt_weight = avg_debt_ratio
    target_wacc = equity_weight * target_ke + debt_weight * target_kd_aftertax

    target_wacc_data = {
        'Rf': rf_rate,
        'MRP': mrp,
        'Size_Premium': size_premium,
        'Avg_Unlevered_Beta': avg_unlevered_beta,
        'Beta_Basis': beta_basis,
        'Beta_Label': beta_label,
        'Target_Tax_Rate': target_tax_rate,
        'Avg_Debt_Ratio': avg_debt_ratio,
        'Debt_Ratio_Basis': debt_ratio_basis,
        'Target_DE_Ratio': target_de_ratio,
        'Target_Relevered_Beta': target_relevered_beta,
        'Target_Ke': target_ke,
        'Kd_Pretax': kd_pretax,
        'Target_Kd_Aftertax': target_kd_aftertax,
        'Equity_Weight': equity_weight,
        'Debt_Weight': debt_weight,
        'Target_WACC': target_wacc,
        'Outlier_Method': outlier_method,
    }
    return (all_period_data, raw_bs_rows, raw_pl_rows, market_rows, price_abs_dfs,
            price_rel_dfs, ticker_to_name, avg_debt_ratio, target_wacc_data, data_quality_rows)



def create_excel(all_period_data, raw_bs_rows, raw_pl_rows, market_rows, price_abs_dfs, price_rel_dfs, base_period_str, target_periods, ticker_to_name, target_wacc_data, beta_type="5Y", data_quality_rows=None, target_inputs=None):
    output = io.BytesIO()
    wb = Workbook(); wb.remove(wb.active)

    data_quality_rows = data_quality_rows or []
    target_inputs = target_inputs or {}
    base_gpcm_data = all_period_data.get('Y', {})
    rel_labels = sorted([k for k in all_period_data.keys() if all_period_data[k]], key=label_sort_key)

    # Sheet Colors
    COLOR_DARK_BLUE = '00338D'
    COLOR_PURPLE = '6A1B9A'

    # Styles & Colors (v17)
    C_BL='00338D'; C_DB='1E2A5E'; C_MB='005EB8'; C_LB='C3D7EE'; C_PB='E8EFF8'
    C_DG='333333'; C_MG='666666'; C_LG='F5F5F5'; C_BG='B0B0B0'; C_W='FFFFFF'
    C_GR='E2EFDA'; C_YL='FFF8E1'; C_NOA='FCE4EC'

    S1=Side(style='thin',color=C_BG); BD=Border(left=S1,right=S1,top=S1,bottom=S1)
    fT=Font(name='Arial',bold=True,size=14,color=C_BL); fS=Font(name='Arial',size=9,color=C_MG,italic=True)
    fH=Font(name='Arial',bold=True,size=9,color=C_W); fA=Font(name='Arial',size=9,color=C_DG)
    fHL=Font(name='Arial',bold=True,size=9,color=C_DB); fSEC=Font(name='Arial',bold=True,size=10,color=C_W)
    fMUL=Font(name='Arial',bold=True,size=10,color=C_BL); fNOTE=Font(name='Arial',size=8,color=C_MG,italic=True)
    fSTAT=Font(name='Arial',bold=True,size=9,color=C_DB)
    fFRM=Font(name='Arial',size=9,color='000000'); fFRM_B=Font(name='Arial',bold=True,size=9,color='000000')
    fLINK=Font(name='Arial',size=9,color='008000'); fLINK_B=Font(name='Arial',bold=True,size=9,color='008000')

    pH=PatternFill('solid',fgColor=C_BL); pW=PatternFill('solid',fgColor=C_W)
    pST=PatternFill('solid',fgColor=C_LG); pSEC=PatternFill('solid',fgColor=C_DB)
    pSTAT=PatternFill('solid',fgColor=C_LB); pBETA=PatternFill('solid',fgColor='E8F5E9')
    pW2=PatternFill('solid',fgColor='FFF9C4')  # 2Y 주간 베타 계열

    ev_fills = {'Cash':PatternFill('solid',fgColor=C_GR), 'IBD':PatternFill('solid',fgColor=C_YL),
                'NCI':PatternFill('solid',fgColor=C_PB), 'NOA(Option)':PatternFill('solid',fgColor=C_NOA),
                'Preferred':PatternFill('solid',fgColor='EDE7F6'),
                'Equity':PatternFill('solid',fgColor=C_LB), 'PL_HL':PatternFill('solid',fgColor=C_YL),
                'NI_Parent':PatternFill('solid',fgColor=C_YL)}

    def sc(c,fo=None,fi=None,al=None,bd=None,nf=None):
        if fo:c.font=fo
        if fi:c.fill=fi
        if al:c.alignment=al
        if bd:c.border=bd
        if nf:c.number_format=nf

    aC=Alignment(horizontal='center',vertical='center',wrap_text=True)
    aL=Alignment(horizontal='left',vertical='center',indent=1)
    aR=Alignment(horizontal='right',vertical='center')
    NF_M='#,##0;(#,##0);"-"'; NF_M1='#,##0.0;(#,##0.0);"-"'; NF_PRC='#,##0.00;(#,##0.00);"-"'
    NF_INT='#,##0;(#,##0);"-"'; NF_EPS='#,##0.00;(#,##0.00);"-"'; NF_X='0.0"x";(0.0"x");"-"'

    # [Sheet 1] Multiples_Trend (Move to front as per user request)
    ws_trend = wb.create_sheet('Multiples_Trend', 0)
    ws_trend.sheet_properties.tabColor = COLOR_PURPLE
    
    ws_trend.merge_cells(start_row=1, start_column=1, end_row=1, end_column=27)
    sc(ws_trend.cell(1,1,'Multiples Trend Summary (Excel Formulas)'), fo=fT)
    ws_trend.merge_cells(start_row=2, start_column=1, end_row=2, end_column=27)
    sc(ws_trend.cell(2,1,'각 연도(Y, Y-1 …)는 해당 결산일 시점의 BS/PL을 각각 참조 | 단위: 백만(현지통화, Curr 열 확인) | EV = Equity Value + IBD − Cash + NCI − NOA'), fo=fS)

    tr_r = 4
    # 열: A Company, B Ticker, C Year Label, D Period, E Curr, F Close, G Shares,
    #     H Mkt Cap, I Preferred, J Equity Value, K Cash, L IBD, M NCI, N NOA, O Equity,
    #     P Revenue, Q EBIT, R D&A, S EBITDA, T NI_Parent, U EV,
    #     V EV/Sales, W EV/EBITDA, X EV/EBIT, Y PER, Z PBR, AA PSR
    trend_headers = ['Company', 'Ticker', 'Year Label', 'Period', 'Curr', 'Close', 'Shares',
                     'Mkt Cap', 'Preferred', 'Equity Value', 'Cash', 'IBD', 'NCI', 'NOA', 'Equity',
                     'Revenue', 'EBIT', 'D&A', 'EBITDA', 'NI_Parent', 'EV',
                     'EV/Sales', 'EV/EBITDA', 'EV/EBIT', 'PER', 'PBR', 'PSR']
    for i, h in enumerate(trend_headers, 1):
        sc(ws_trend.cell(tr_r, i, h), fo=fH, fi=pH, al=aC, bd=BD)

    ws_trend.column_dimensions['A'].width = 18
    ws_trend.column_dimensions['B'].width = 10
    ws_trend.column_dimensions['C'].width = 10
    ws_trend.column_dimensions['D'].width = 12
    ws_trend.column_dimensions['E'].width = 6
    for col_l in ['F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U']:
        ws_trend.column_dimensions[col_l].width = 14
    for col_l in ['V','W','X','Y','Z','AA']:
        ws_trend.column_dimensions[col_l].width = 11

    tr_r += 1
    mc_sht = "'Market_Cap'"
    for ticker_idx, (ticker, name) in enumerate(ticker_to_name.items()):
        for label in rel_labels:
            gpcm = all_period_data[label].get(ticker)
            if not gpcm: continue

            p_dt = gpcm.get('PL_Date') or gpcm.get('Base_Date', '-')
            row_fi = pW if ticker_idx % 2 == 0 else pST

            ws_trend.cell(tr_r, 1, name); sc(ws_trend.cell(tr_r, 1), fo=fA, fi=row_fi, al=aL, bd=BD)
            ws_trend.cell(tr_r, 2, ticker); sc(ws_trend.cell(tr_r, 2), fo=fA, fi=row_fi, al=aC, bd=BD)
            ws_trend.cell(tr_r, 3, label); sc(ws_trend.cell(tr_r, 3), fo=fA, fi=row_fi, al=aC, bd=BD)
            ws_trend.cell(tr_r, 4, p_dt); sc(ws_trend.cell(tr_r, 4), fo=fA, fi=row_fi, al=aC, bd=BD)
            # 통화 표시 (피어별 통화가 다르므로 절대금액 비교 시 반드시 확인할 것)
            ws_trend.cell(tr_r, 5, gpcm.get('Currency', '')); sc(ws_trend.cell(tr_r, 5), fo=fA, fi=row_fi, al=aC, bd=BD)

            def mc(col_letter):
                return (f'=SUMIFS({mc_sht}!{col_letter}:{col_letter}, {mc_sht}!B:B, $B{tr_r}, '
                        f'{mc_sht}!H:H, $C{tr_r})')

            ws_trend.cell(tr_r, 6).value = mc('F')   # Close
            sc(ws_trend.cell(tr_r, 6), fo=fLINK, fi=row_fi, al=aR, bd=BD, nf=NF_PRC)
            ws_trend.cell(tr_r, 7).value = mc('G')   # Shares
            sc(ws_trend.cell(tr_r, 7), fo=fLINK, fi=row_fi, al=aR, bd=BD, nf=NF_INT)
            ws_trend.cell(tr_r, 8).value = mc('I')   # Mkt Cap
            sc(ws_trend.cell(tr_r, 8), fo=fLINK, fi=row_fi, al=aR, bd=BD, nf=NF_M1)

            bs_sn = "BS_FULL (base)" if label == 'Y' else f"BS_Full_{label}"
            pl_sn = "PL_Data (base)" if label == 'Y' else f"PL_Data_{label}"

            def bst(tag):
                return (f"=SUMIFS('{bs_sn}'!G:G, '{bs_sn}'!B:B, $B{tr_r}, "
                        f"'{bs_sn}'!F:F, \"{tag}\")")

            ws_trend.cell(tr_r, 9).value = bst('Preferred')
            ws_trend.cell(tr_r, 10).value = f'=H{tr_r}+I{tr_r}'      # Equity Value
            ws_trend.cell(tr_r, 11).value = bst('Cash')
            ws_trend.cell(tr_r, 12).value = bst('IBD')
            ws_trend.cell(tr_r, 13).value = bst('NCI')
            # NOA: GPCM 시트와 동일하게 BS_Full의 EV Tag를 "NOA"로 바꾸면 EV에서 차감됨
            ws_trend.cell(tr_r, 14).value = bst('NOA')
            ws_trend.cell(tr_r, 15).value = bst('Equity')
            for c_idx in [9, 11, 12, 13, 14, 15]:
                sc(ws_trend.cell(tr_r, c_idx), fo=fLINK, fi=row_fi, al=aR, bd=BD, nf=NF_M)
            sc(ws_trend.cell(tr_r, 10), fo=fFRM_B, fi=row_fi, al=aR, bd=BD, nf=NF_M)

            def plt(tag):
                return (f"=SUMIFS('{pl_sn}'!$J:$J, '{pl_sn}'!$B:$B, $B{tr_r}, "
                        f"'{pl_sn}'!$E:$E, \"{tag}\")")

            def pla(acct):
                return (f"SUMIFS('{pl_sn}'!$J:$J, '{pl_sn}'!$B:$B, $B{tr_r}, "
                        f"'{pl_sn}'!$D:$D, \"{acct}\")")

            ws_trend.cell(tr_r, 16).value = plt('Revenue')
            ws_trend.cell(tr_r, 17).value = f'={pla("Operating Income")}'
            # D&A: EBITDA·EBIT 중 하나라도 없으면 0 (GPCM 시트와 동일한 안전장치)
            ws_trend.cell(tr_r, 18).value = (f'=IF(OR({pla("EBITDA")}=0,{pla("EBIT")}=0),0,'
                                             f'{pla("EBITDA")}-{pla("EBIT")})')
            ws_trend.cell(tr_r, 19).value = f'=Q{tr_r}+R{tr_r}'      # EBITDA
            ws_trend.cell(tr_r, 20).value = plt('NI_Parent')
            for c_idx in [16, 17, 20]:
                sc(ws_trend.cell(tr_r, c_idx), fo=fLINK, fi=row_fi, al=aR, bd=BD, nf=NF_M)
            for c_idx in [18, 19]:
                sc(ws_trend.cell(tr_r, c_idx), fo=fFRM_B, fi=row_fi, al=aR, bd=BD, nf=NF_M)

            # U: EV = Equity Value(J) + IBD(L) - Cash(K) + NCI(M) - NOA(N)
            ws_trend.cell(tr_r, 21).value = f'=J{tr_r}+L{tr_r}-K{tr_r}+M{tr_r}-N{tr_r}'
            sc(ws_trend.cell(tr_r, 21), fo=fFRM_B, fi=PatternFill('solid',fgColor='E3F2FD'), al=aR, bd=BD, nf=NF_M)

            # V~AA: 멀티플
            for c_idx, (den, num) in [(22, ('P', 'U')), (23, ('S', 'U')), (24, ('Q', 'U')),
                                      (25, ('T', 'J')), (26, ('O', 'J')), (27, ('P', 'J'))]:
                ws_trend.cell(tr_r, c_idx).value = (
                    f'=IFERROR(IF({den}{tr_r}>0, {num}{tr_r}/{den}{tr_r}, "N/M"), "N/M")')
                sc(ws_trend.cell(tr_r, c_idx), fo=fFRM_B,
                   fi=PatternFill('solid', fgColor='FFF9C4'), al=aR, bd=BD, nf=NF_X)

            tr_r += 1


    ws_trend.auto_filter.ref = f"A4:AA{tr_r-1}"
    ws_trend.freeze_panes = 'A5'

    # [Sheet 2] BS_Full
    for label in rel_labels:
        bs_rows_p = [r for r in raw_bs_rows if r.get('Label') == label]
        if bs_rows_p:
            sheet_name = "BS_FULL (base)" if label == 'Y' else f'BS_Full_{label}'
            ws_bs = wb.create_sheet(sheet_name)
            ws_bs.sheet_properties.tabColor = COLOR_DARK_BLUE
            df_bs = pd.DataFrame(bs_rows_p)
            cols = [('Company','Company',18),('Ticker','Ticker',10), ('Period','Period',12),('Currency','Curr',6), ('Account','Account',42),('EV_Tag','EV Tag',14), ('Amount_M','Amount (M)',18)]
            ws_bs.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols)); sc(ws_bs.cell(1,1,f'Balance Sheet ({label})'),fo=fT)
            r=3
            for i,(lb,key) in enumerate([('Cash','Cash'),('IBD','IBD'),('NCI','NCI'),('NOA(Option)','NOA(Option)'),('Equity','Equity')]):
                sc(ws_bs.cell(r,i+1,lb), fo=Font(name='Arial',size=8,bold=True),fi=ev_fills[key],al=aC,bd=BD)
            r=5
            for i,(col,disp,w) in enumerate(cols): ws_bs.column_dimensions[get_column_letter(i+1)].width=w; sc(ws_bs.cell(r,i+1,disp),fo=fH,fi=pH,al=aC,bd=BD)
            hdr=r; r+=1
            for _,rd in df_bs.iterrows():
                ev_tag=rd.get('EV_Tag', ''); is_hl=bool(ev_tag)
                row_fi=ev_fills.get(ev_tag, pST if r%2==0 else pW) if is_hl else (pST if r%2==0 else pW)
                row_font=fHL if is_hl else fA
                for i,(col,_,_) in enumerate(cols):
                    c=ws_bs.cell(r,i+1); v=rd.get(col, '')
                    if isinstance(v,(float,np.floating)): c.value=round(v,1) if pd.notna(v) else None
                    else: c.value=v
                    sc(c,fo=row_font,fi=row_fi,al=aR if col=='Amount_M' else aL,bd=BD,nf=NF_M if col=='Amount_M' else None)
                r+=1
            ws_bs.auto_filter.ref=f"A{hdr}:{get_column_letter(len(cols))}{r-1}"; ws_bs.freeze_panes=f'A{hdr+1}'

    # [Sheet 3] PL_Data
    for label in rel_labels:
        pl_rows_p = [r for r in raw_pl_rows if r.get('Label') == label]
        if pl_rows_p:
            is_base = (label == 'Y')
            sheet_name = "PL_Data (base)" if is_base else f'PL_Data_{label}'
            ws_pl = wb.create_sheet(sheet_name)
            ws_pl.sheet_properties.tabColor = COLOR_DARK_BLUE
            
            # 사용자 캡처 이미지 1의 컬럼 구성 및 순서 준수
            # Company, Ticker, Cur, Account, GPCM Tag, Unit, Source, Q Label, Period, Amount
            cols = [
                ('Company','Company',18),('Ticker','Ticker',10), ('Currency','Cur',6),
                ('Account','Account',42), ('GPCM_Tag','GPCM Tag',14), ('Unit','Unit',10),
                ('PL_Source','Source',16), ('Q_Label','Q Label',10), ('Period','Period',12),
                ('Amount_M','Amount',18)
            ]
            
            ws_pl.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols)); sc(ws_pl.cell(1,1,f'Income Statement ({label})'),fo=fT)
            r_idx=5
            for i,(col,disp,w) in enumerate(cols): ws_pl.column_dimensions[get_column_letter(i+1)].width=w; sc(ws_pl.cell(r_idx,i+1,disp),fo=fH,fi=pH,al=aC,bd=BD)
            hdr=r_idx; r_idx+=1
            
            df_pl = pd.DataFrame(pl_rows_p).sort_values(['Company','_sort','Q_Label'])
            for _,rd in df_pl.iterrows():
                is_hl=bool(rd.get('GPCM_Tag', '')); row_fi=ev_fills.get('PL_HL', pW) if is_hl else (pST if r_idx%2==0 else pW)
                row_font=fHL if is_hl else fA
                for i,(col,_,_) in enumerate(cols):
                    c=ws_pl.cell(r_idx,i+1); v=rd.get(col, '')
                    if isinstance(v,(float,np.floating)): c.value=round(v,1) if pd.notna(v) else None
                    else: c.value=v
                    is_eps=rd.get('Unit','')=='per share'; nf=NF_EPS if is_eps else NF_M
                    sc(c,fo=row_font,fi=row_fi,al=aR if col=='Amount_M' else aL,bd=BD,nf=nf if col=='Amount_M' else None)
                r_idx+=1
            ws_pl.auto_filter.ref=f"A{hdr}:{get_column_letter(len(cols))}{r_idx-1}"; ws_pl.freeze_panes=f'A{hdr+1}'

    # [Sheet 4] Market_Cap
    ws_mc = wb.create_sheet('Market_Cap')
    ws_mc.sheet_properties.tabColor = COLOR_DARK_BLUE
    if market_rows:
        df_mkt = pd.DataFrame(market_rows)
        # 열 순서 변경 시 GPCM 시트의 SUMIFS 참조(MC_COL)도 함께 수정할 것
        cols = [('Company','Company',18),('Ticker','Ticker',10), ('Base_Date','Base Date',12),('Price_Date','Price Date',12), ('Currency','Curr',6),('Close','Close Price',14), ('Shares','Shares (Ord.)',18),('Label','Label',10), ('Market_Cap_M','Mkt Cap (M)',20), ('Preferred_M','Preferred (M)',16), ('Equity_Value_M','Equity Value (M)',20)]
        ws_mc.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols)); sc(ws_mc.cell(1,1,'Market Capitalization'),fo=fT)
        ws_mc.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(cols)); sc(ws_mc.cell(2,1,'Mkt Cap = Ordinary Shares Number (자기주식 차감) × Close Price (auto_adjust=False) | Equity Value = Mkt Cap + 우선주 장부금액'), fo=fS)
        r=4
        for i,(col,disp,w) in enumerate(cols): ws_mc.column_dimensions[get_column_letter(i+1)].width=w; sc(ws_mc.cell(r,i+1,disp),fo=fH,fi=pH,al=aC,bd=BD)
        mc_hdr=r; r+=1
        for _,rd in df_mkt.iterrows():
            ev=(r%2==0)
            for i,(col,_,_) in enumerate(cols):
                c=ws_mc.cell(r,i+1); v=rd.get(col, '')
                if isinstance(v,(float,np.floating)): c.value=round(v,2) if pd.notna(v) else None
                else: c.value=v
                nf=NF_PRC if col=='Close' else (NF_INT if col=='Shares' else (NF_M1 if col in ('Market_Cap_M','Preferred_M','Equity_Value_M') else None))
                sc(c,fo=fA,fi=pST if ev else pW,al=aR if nf else aL,bd=BD,nf=nf)
            r+=1
        ws_mc.auto_filter.ref=f"A{mc_hdr}:{get_column_letter(len(cols))}{r-1}"; ws_mc.freeze_panes=f'A{mc_hdr+1}'

    # [Sheet 5] Beta_Calculation
    ws_beta = wb.create_sheet('Beta_Calculation')
    ws_beta.sheet_properties.tabColor = COLOR_DARK_BLUE

    # 제목
    ws_beta.merge_cells('A1:F1')
    sc(ws_beta['A1'], fo=Font(name='Arial', bold=True, size=14, color=C_BL))
    ws_beta['A1'] = 'Beta Calculation (Excel Formulas)'

    ws_beta.merge_cells('A2:F2')
    sc(ws_beta['A2'], fo=Font(name='Arial', size=9, color=C_MG, italic=True))
    ws_beta['A2'] = f'5-Year Monthly & 2-Year Weekly Returns | Base: {base_period_str}'

    r_beta = 4

    # ticker -> {raw_5y, adj_5y, r2_5y, se_5y, n_5y, raw_2y, adj_2y, r2_2y, se_2y, n_2y}
    beta_result_rows = {}

    def _beta_block(title, fill_hex, stock_px, market_px, date_fmt, tag):
        """한 종목의 한 기간(5Y/2Y) 베타 계산 블록을 기록하고 결과 행 번호를 반환."""
        nonlocal r_beta
        ws_beta.merge_cells(f'A{r_beta}:G{r_beta}')
        sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
           fi=PatternFill('solid', fgColor=fill_hex), al=Alignment(horizontal='center'))
        ws_beta.cell(r_beta, 1, title)
        r_beta += 1

        if stock_px is None or market_px is None or stock_px.empty or market_px.empty:
            ws_beta.cell(r_beta, 1, 'No price data available')
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', size=9, color='FF0000'))
            r_beta += 2
            return {}

        for ci, h in enumerate(['Date', 'Stock Price', 'Index Price', 'Stock Return', 'Index Return'], 1):
            ws_beta.cell(r_beta, ci, h)
            sc(ws_beta.cell(r_beta, ci), fo=Font(name='Arial', bold=True, size=9, color=C_W),
               fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
        r_beta += 1

        data_start_row = r_beta
        common_dates = stock_px.index.intersection(market_px.index)
        for d in common_dates:
            ws_beta.cell(r_beta, 1, d.strftime(date_fmt))
            ws_beta.cell(r_beta, 2, float(stock_px.loc[d]))
            ws_beta.cell(r_beta, 3, float(market_px.loc[d]))
            if r_beta > data_start_row:
                ws_beta.cell(r_beta, 4).value = f'=(B{r_beta}-B{r_beta-1})/B{r_beta-1}'
                ws_beta.cell(r_beta, 5).value = f'=(C{r_beta}-C{r_beta-1})/C{r_beta-1}'
            sc(ws_beta.cell(r_beta, 1), fo=fA, al=aC, bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=fA, al=aR, bd=BD, nf='#,##0.00')
            sc(ws_beta.cell(r_beta, 3), fo=fA, al=aR, bd=BD, nf='#,##0.00')
            sc(ws_beta.cell(r_beta, 4), fo=fA, al=aR, bd=BD, nf='0.00%')
            sc(ws_beta.cell(r_beta, 5), fo=fA, al=aR, bd=BD, nf='0.00%')
            r_beta += 1
        data_end_row = r_beta - 1

        if data_end_row <= data_start_row:
            ws_beta.cell(r_beta, 1, 'Not enough observations')
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', size=9, color='FF0000'))
            r_beta += 2
            return {}

        ys = f'D{data_start_row+1}:D{data_end_row}'
        xs = f'E{data_start_row+1}:E{data_end_row}'
        r_beta += 1
        pRES = PatternFill('solid', fgColor=fill_hex if fill_hex.startswith('E8') else 'FFF9C4')

        # (표시명, 수식, 숫자서식, 결과키, 설명)
        rows_spec = [
            (f'Raw Beta ({tag})', f'=SLOPE({ys},{xs})', '0.0000', f'raw_{tag}', '회귀 기울기'),
            (f'Adjusted Beta ({tag})', None, '0.0000', f'adj_{tag}', '2/3 x Raw + 1/3 x 1.0 (Bloomberg)'),
            (f'R² ({tag})', f'=RSQ({ys},{xs})', '0.0000', f'r2_{tag}', '설명력. 0.1 미만이면 베타 신뢰도 낮음'),
            (f'Std Error ({tag})', f'=STEYX({ys},{xs})/(STDEV({xs})*SQRT(COUNT({xs})-1))', '0.0000', f'se_{tag}', 'Raw Beta ± 2×SE ≈ 95% 신뢰구간'),
            (f'N obs ({tag})', f'=COUNT({ys})', '#,##0', f'n_{tag}', '회귀에 사용된 관측치 수'),
        ]
        out = {}
        raw_row = None
        for name, formula, nf, key, note in rows_spec:
            ws_beta.cell(r_beta, 1, name)
            if formula is None:  # Adjusted Beta는 바로 위 Raw Beta 참조
                ws_beta.cell(r_beta, 2).value = f'=2/3*B{raw_row}+1/3*1'
            else:
                ws_beta.cell(r_beta, 2).value = formula
            ws_beta.cell(r_beta, 3, note)
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', bold=True, size=9), fi=pRES,
               bd=BD, al=aR, nf=nf)
            sc(ws_beta.cell(r_beta, 3), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
            out[key] = r_beta
            if key.startswith('raw_'):
                raw_row = r_beta
            r_beta += 1

        r_beta += 1
        return out

    for idx, (ticker, gpcm) in enumerate(base_gpcm_data.items()):
        company_name = gpcm['Company']
        market_idx = gpcm['Market_Index']
        res = {}
        res.update(_beta_block(
            f'{company_name} ({ticker}) vs {market_idx} - 5Y Monthly (배당 반영 총수익 기준)',
            '607D8B', gpcm.get('Stock_Monthly_Prices_5Y'), gpcm.get('Market_Monthly_Prices_5Y'),
            '%Y-%m', '5y'))
        res.update(_beta_block(
            f'{company_name} ({ticker}) vs {market_idx} - 2Y Weekly (배당 반영 총수익 기준)',
            '455A64', gpcm.get('Stock_Weekly_Prices_2Y'), gpcm.get('Market_Weekly_Prices_2Y'),
            '%Y-%m-%d', '2y'))
        beta_result_rows[ticker] = res
        r_beta += 1

    for col_l, w in (('A', 22), ('B', 15), ('C', 42), ('D', 15), ('E', 15)):
        ws_beta.column_dimensions[col_l].width = w

    ws_beta.freeze_panes = 'A4'

    # [Sheet 7] GPCM (Valuation Summary)
    ws = wb.create_sheet('GPCM')
    ws.sheet_properties.tabColor = COLOR_PURPLE

    # ---------------------------------------------------------------------
    # 컬럼 정의 (key, 표시명, 너비, 섹션)
    # 열을 추가/삭제해도 아래 CI/CL 맵이 자동으로 따라가므로
    # 수식에 알파벳을 직접 쓰지 않는다.
    # ---------------------------------------------------------------------
    SEC_INFO = 'Company Info'; SEC_BS = 'BS → EV Components'
    SEC_PL = 'PL (LTM / Annual)'; SEC_MKT = 'Market Data'
    SEC_MUL = 'Valuation Multiples'; SEC_BETA = 'Beta & Risk Analysis'
    gpcm_col_defs = [
        ('Company', 'Company', 18, SEC_INFO),
        ('Ticker', 'Ticker', 10, SEC_INFO),
        ('PLDate', 'PL 기준일', 11, SEC_INFO),
        ('BSDate', 'BS 기준일', 11, SEC_INFO),
        ('Curr', 'Curr', 6, SEC_INFO),
        ('PLSource', 'PL Source', 18, SEC_INFO),
        ('Exchange', 'Exchange', 11, SEC_INFO),
        ('MktIndex', 'Mkt Index', 11, SEC_INFO),
        ('Cash', 'Cash', 14, SEC_BS),
        ('IBD', 'IBD', 14, SEC_BS),
        ('NetDebt', 'Net Debt', 14, SEC_BS),
        ('NCI', 'NCI', 12, SEC_BS),
        ('Preferred', 'Preferred', 12, SEC_BS),
        ('NOA', 'NOA', 12, SEC_BS),
        ('Equity', 'Equity', 14, SEC_BS),
        ('EV', 'EV', 16, SEC_BS),
        ('Revenue', 'Revenue', 14, SEC_PL),
        ('EBIT', 'EBIT', 14, SEC_PL),
        ('DA', 'D&A', 12, SEC_PL),
        ('EBITDA', 'EBITDA', 14, SEC_PL),
        ('NI', 'NI (Parent)', 14, SEC_PL),
        ('DASrc', 'D&A 산출근거', 26, SEC_PL),
        ('Price', 'Price', 12, SEC_MKT),
        ('Shares', 'Shares', 16, SEC_MKT),
        ('MktCap', 'Mkt Cap', 18, SEC_MKT),
        ('EqValue', 'Equity Value', 18, SEC_MKT),
        ('EV_Sales', 'EV/Sales', 11, SEC_MUL),
        ('EV_EBITDA', 'EV/EBITDA', 11, SEC_MUL),
        ('EV_EBIT', 'EV/EBIT', 11, SEC_MUL),
        ('PER', 'PER', 11, SEC_MUL),
        ('PBR', 'PBR', 11, SEC_MUL),
        ('PSR', 'PSR', 11, SEC_MUL),
        ('B5R', 'β 5Y Raw', 10, SEC_BETA),
        ('B5A', 'β 5Y Adj', 10, SEC_BETA),
        ('B5R2', 'R² 5Y', 9, SEC_BETA),
        ('B5SE', 'SE 5Y', 9, SEC_BETA),
        ('B5N', 'n 5Y', 7, SEC_BETA),
        ('B2R', 'β 2Y Raw', 10, SEC_BETA),
        ('B2A', 'β 2Y Adj', 10, SEC_BETA),
        ('B2R2', 'R² 2Y', 9, SEC_BETA),
        ('B2SE', 'SE 2Y', 9, SEC_BETA),
        ('B2N', 'n 2Y', 7, SEC_BETA),
        ('Pretax', 'Pretax Inc', 14, SEC_BETA),
        ('TaxRate', 'Tax Rate', 10, SEC_BETA),
        ('TaxBasis', '세율 근거', 22, SEC_BETA),
        ('DE', 'D/E Ratio', 10, SEC_BETA),
        ('DV', 'Debt Ratio (D/V)', 13, SEC_BETA),
        ('UB5', 'Unlevered β 5Y', 13, SEC_BETA),
        ('UB2', 'Unlevered β 2Y', 13, SEC_BETA),
    ]
    CI = {k: i + 1 for i, (k, _, _, _) in enumerate(gpcm_col_defs)}
    CL = {k: get_column_letter(v) for k, v in CI.items()}
    TOTAL_COLS = len(gpcm_col_defs)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS); sc(ws.cell(1,1,'GPCM Valuation Summary with Beta Analysis'), fo=fT)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS); sc(ws.cell(2,1,f'Base: {base_period_str} | Unit: Millions (local currency) | EV = Equity Value + IBD − Cash + NCI − NOA | Equity Value = Mkt Cap + 우선주 장부금액'), fo=fS)

    # 섹션 머리행
    r = 3
    sec_start = 1
    for i, (_, _, _, sec) in enumerate(gpcm_col_defs):
        is_last = (i == len(gpcm_col_defs) - 1)
        next_sec = None if is_last else gpcm_col_defs[i + 1][3]
        if is_last or next_sec != sec:
            ws.merge_cells(start_row=r, start_column=sec_start, end_row=r, end_column=i + 1)
            sc(ws.cell(r, sec_start, sec), fo=fSEC, fi=pSEC, al=aC, bd=BD)
            for c_idx in range(sec_start, i + 2):
                sc(ws.cell(r, c_idx), bd=BD)
            sec_start = i + 2

    r = 4
    for i, (_, disp, w, _) in enumerate(gpcm_col_defs):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
        sc(ws.cell(r, i + 1, disp), fo=fH, fi=pH, al=aC, bd=BD)

    DATA_START=5; n_companies=len(base_gpcm_data); DATA_END=DATA_START+n_companies-1
    NF_BETA='0.00;(0.00);"-"'; NF_PCT='0.0%;(0.0%);"-"'; NF_RATIO='0.00;(0.00);"-"'
    NF_R2='0.00;;"-"'

    bs_sn = "BS_FULL (base)"
    pl_sn = "PL_Data (base)"
    pMULT = PatternFill('solid', fgColor=C_PB)
    pFLAG = PatternFill('solid', fgColor='FFCDD2')   # 확인 필요(빨강)

    def bs_tag(row, tag):
        return (f"=SUMIFS('{bs_sn}'!$G:$G,'{bs_sn}'!$B:$B,$B{row},"
                f"'{bs_sn}'!$F:$F,\"{tag}\")")

    def pl_tag(row, tag):
        return (f"=SUMIFS('{pl_sn}'!$J:$J,'{pl_sn}'!$B:$B,$B{row},"
                f"'{pl_sn}'!$E:$E,\"{tag}\")")

    def pl_acct(row, acct):
        return (f"=SUMIFS('{pl_sn}'!$J:$J,'{pl_sn}'!$B:$B,$B{row},"
                f"'{pl_sn}'!$D:$D,\"{acct}\")")

    def mc_col(row, col_letter):
        return (f'=SUMIFS(Market_Cap!${col_letter}:${col_letter},'
                f'Market_Cap!$B:$B,$B{row},Market_Cap!$H:$H,"Y")')

    def put(row, key, value, fo=fA, fi=None, al=aR, nf=None):
        c = ws.cell(row, CI[key])
        c.value = value
        sc(c, fo=fo, fi=fi, al=al, bd=BD, nf=nf)

    for idx, (ticker, gpcm) in enumerate(base_gpcm_data.items()):
        r = DATA_START + idx
        base_fi = pST if (r % 2 == 0) else pW

        # --- Company Info ---
        info_vals = {
            'Company': gpcm['Company'], 'Ticker': ticker,
            'PLDate': gpcm.get('PL_Date') or gpcm['Base_Date'],
            'BSDate': gpcm.get('BS_Date') or '-',
            'Curr': gpcm['Currency'], 'PLSource': gpcm['PL_Source'],
            'Exchange': gpcm.get('Exchange', ''), 'MktIndex': gpcm.get('Market_Index', ''),
        }
        for k, v in info_vals.items():
            put(r, k, v, fo=fA, fi=base_fi, al=aL)

        # --- BS → EV Components ---
        put(r, 'Cash', bs_tag(r, 'Cash'), fo=fLINK_B, fi=ev_fills['Cash'], nf=NF_M)
        put(r, 'IBD', bs_tag(r, 'IBD'), fo=fLINK_B, fi=ev_fills['IBD'], nf=NF_M)
        put(r, 'NetDebt', f"={CL['IBD']}{r}-{CL['Cash']}{r}", fo=fFRM_B, fi=base_fi, nf=NF_M)
        put(r, 'NCI', bs_tag(r, 'NCI'), fo=fLINK_B, fi=ev_fills['NCI'], nf=NF_M)
        put(r, 'Preferred', bs_tag(r, 'Preferred'), fo=fLINK_B, fi=ev_fills['Preferred'], nf=NF_M)
        # NOA: BS_Full의 EV Tag를 "NOA(Option)" → "NOA"로 바꾸면 여기 잡혀 EV에서 차감됨
        put(r, 'NOA', bs_tag(r, 'NOA'), fo=fLINK_B, fi=ev_fills['NOA(Option)'], nf=NF_M)
        put(r, 'Equity', bs_tag(r, 'Equity'), fo=fLINK_B, fi=ev_fills['Equity'], nf=NF_M)
        put(r, 'EV', (f"={CL['EqValue']}{r}+{CL['IBD']}{r}-{CL['Cash']}{r}"
                      f"+{CL['NCI']}{r}-{CL['NOA']}{r}"),
            fo=fFRM_B, fi=PatternFill('solid', fgColor=C_PB), nf=NF_M)

        # --- PL ---
        put(r, 'Revenue', pl_tag(r, 'Revenue'), fo=fLINK_B, fi=ev_fills['PL_HL'], nf=NF_M)
        put(r, 'EBIT', pl_acct(r, 'Operating Income'), fo=fLINK_B, fi=ev_fills['PL_HL'], nf=NF_M)
        # D&A = yfinance EBITDA − yfinance EBIT. 둘 중 하나라도 없으면 0 (아래 산출근거 열에 표시)
        da_formula = (f"=IF(OR(SUMIFS('{pl_sn}'!$J:$J,'{pl_sn}'!$B:$B,$B{r},'{pl_sn}'!$D:$D,\"EBITDA\")=0,"
                      f"SUMIFS('{pl_sn}'!$J:$J,'{pl_sn}'!$B:$B,$B{r},'{pl_sn}'!$D:$D,\"EBIT\")=0),0,"
                      f"SUMIFS('{pl_sn}'!$J:$J,'{pl_sn}'!$B:$B,$B{r},'{pl_sn}'!$D:$D,\"EBITDA\")"
                      f"-SUMIFS('{pl_sn}'!$J:$J,'{pl_sn}'!$B:$B,$B{r},'{pl_sn}'!$D:$D,\"EBIT\"))")
        da_imputed = bool(gpcm.get('DA_Imputed'))
        put(r, 'DA', da_formula, fo=fFRM_B, fi=(pFLAG if da_imputed else ev_fills['PL_HL']), nf=NF_M)
        put(r, 'EBITDA', f"={CL['EBIT']}{r}+{CL['DA']}{r}", fo=fFRM_B, fi=ev_fills['PL_HL'], nf=NF_M)
        put(r, 'NI', pl_tag(r, 'NI_Parent'), fo=fLINK_B, fi=ev_fills['PL_HL'], nf=NF_M)
        put(r, 'DASrc', gpcm.get('DA_Source', '-'), fo=fA,
            fi=(pFLAG if da_imputed else base_fi), al=aL)

        # --- Market Data ---
        put(r, 'Price', mc_col(r, 'F'), fo=fLINK, fi=base_fi, nf=NF_PRC)
        put(r, 'Shares', mc_col(r, 'G'), fo=fLINK, fi=base_fi, nf=NF_INT)
        put(r, 'MktCap', mc_col(r, 'I'), fo=fLINK, fi=base_fi, nf=NF_M1)
        # 우선주 장부금액을 가산한 자기자본가치 (실무 관행)
        put(r, 'EqValue', f"={CL['MktCap']}{r}+{CL['Preferred']}{r}",
            fo=fFRM_B, fi=base_fi, nf=NF_M1)

        # --- Valuation Multiples ---
        ev_c, eq_c = CL['EV'], CL['EqValue']
        mult_specs = [
            ('EV_Sales', CL['Revenue'], ev_c),
            ('EV_EBITDA', CL['EBITDA'], ev_c),
            ('EV_EBIT', CL['EBIT'], ev_c),
            ('PER', CL['NI'], eq_c),
            ('PBR', CL['Equity'], eq_c),
            ('PSR', CL['Revenue'], eq_c),
        ]
        for key, den, num in mult_specs:
            put(r, key, f'=IF({den}{r}>0,{num}{r}/{den}{r},"N/M")',
                fo=fMUL, fi=pMULT, nf=NF_X)

        # --- Beta & Risk ---
        rows = beta_result_rows.get(ticker, {})
        beta_link_specs = [
            ('B5R', 'raw_5y', pBETA, NF_BETA), ('B5A', 'adj_5y', pBETA, NF_BETA),
            ('B5R2', 'r2_5y', pBETA, NF_R2), ('B5SE', 'se_5y', pBETA, NF_R2),
            ('B5N', 'n_5y', pBETA, NF_INT),
            ('B2R', 'raw_2y', pW2, NF_BETA), ('B2A', 'adj_2y', pW2, NF_BETA),
            ('B2R2', 'r2_2y', pW2, NF_R2), ('B2SE', 'se_2y', pW2, NF_R2),
            ('B2N', 'n_2y', pW2, NF_INT),
        ]
        for key, rk, fill, nf in beta_link_specs:
            src = rows.get(rk)
            put(r, key, (f'=Beta_Calculation!$B${src}' if src else None),
                fo=fLINK, fi=fill, nf=nf)

        put(r, 'Pretax', pl_acct(r, 'Pretax Income'), fo=fLINK, fi=base_fi, nf=NF_M)
        put(r, 'TaxRate', gpcm['Tax_Rate'], fo=fA, fi=base_fi, nf=NF_PCT)
        put(r, 'TaxBasis', gpcm.get('Tax_Basis', '-'), fo=fA, fi=base_fi, al=aL)

        # D/E = IBD / (자기자본가치 + NCI),  D/V = IBD / (IBD + 자기자본가치 + NCI)
        ibd_c, nci_c = CL['IBD'], CL['NCI']
        put(r, 'DE', f'=IF(({eq_c}{r}+{nci_c}{r})>0,{ibd_c}{r}/({eq_c}{r}+{nci_c}{r}),0)',
            fo=fFRM_B, fi=base_fi, nf=NF_RATIO)
        put(r, 'DV', (f'=IF(({eq_c}{r}+{ibd_c}{r}+{nci_c}{r})>0,'
                      f'{ibd_c}{r}/({eq_c}{r}+{ibd_c}{r}+{nci_c}{r}),0)'),
            fo=fFRM_B, fi=base_fi, nf=NF_RATIO)

        tax_c, de_c = CL['TaxRate'], CL['DE']
        put(r, 'UB5', f'=IF({CL["B5A"]}{r}<>0,{CL["B5A"]}{r}/(1+(1-{tax_c}{r})*{de_c}{r}),"")',
            fo=fFRM_B, fi=pBETA, nf=NF_BETA)
        put(r, 'UB2', f'=IF({CL["B2A"]}{r}<>0,{CL["B2A"]}{r}/(1+(1-{tax_c}{r})*{de_c}{r}),"")',
            fo=fFRM_B, fi=pBETA, nf=NF_BETA)

    # ---------------------------------------------------------------------
    # 통계 (이상치 제외 여부에 따라 대상 범위를 달리함)
    # ---------------------------------------------------------------------
    stat_rows = {}
    r = DATA_END + 2
    stat_labels = ['Mean', 'Median', 'Max', 'Min']
    func_map = {'Mean': 'AVERAGE', 'Median': 'MEDIAN', 'Max': 'MAX', 'Min': 'MIN'}
    mult_keys = ['EV_Sales', 'EV_EBITDA', 'EV_EBIT', 'PER', 'PBR', 'PSR']
    beta_keys = ['B5R', 'B5A', 'B5R2', 'B2R', 'B2A', 'B2R2', 'UB5', 'UB2']
    ratio_keys = ['DE', 'DV']

    for sn in stat_labels:
        stat_rows[sn] = r
        ws.cell(r, CI['MktCap'], sn); sc(ws.cell(r, CI['MktCap']), fo=fSTAT, fi=pSTAT, al=aC, bd=BD)
        fn = func_map[sn]
        for keys, nf in ((mult_keys, NF_X), (beta_keys, NF_BETA), (ratio_keys, NF_RATIO)):
            for k in keys:
                col = CL[k]
                ws.cell(r, CI[k]).value = f'=IFERROR({fn}({col}{DATA_START}:{col}{DATA_END}),"N/M")'
                sc(ws.cell(r, CI[k]), fo=fSTAT, fi=pSTAT, al=aR, bd=BD, nf=nf)
        r += 1

    # 사분위수 (Implied Valuation 밴드 산출용)
    for qname, qv in (('Q1 (25%)', 0.25), ('Q3 (75%)', 0.75)):
        stat_rows[qname] = r
        ws.cell(r, CI['MktCap'], qname); sc(ws.cell(r, CI['MktCap']), fo=fSTAT, fi=pSTAT, al=aC, bd=BD)
        for k in mult_keys:
            col = CL[k]
            ws.cell(r, CI[k]).value = f'=IFERROR(QUARTILE({col}{DATA_START}:{col}{DATA_END},{int(qv*4)}),"N/M")'
            sc(ws.cell(r, CI[k]), fo=fSTAT, fi=pSTAT, al=aR, bd=BD, nf=NF_X)
        r += 1

    # Notes
    r+=2
    notes = [
        '[ Valuation Methodology Notes ]',
        f'• Base Date: {base_period_str} | Unit: Millions (local currency)',
        '• Equity Value = Market Cap + 우선주 장부금액  (우선주는 자기자본에 가산하는 실무 관행 적용)',
        '• Market Cap = Ordinary Shares Number × Close Price (auto_adjust=False, 보통주 기준)',
        '• EV = Equity Value + IBD − Cash + NCI − NOA',
        '• Cash includes: Cash & Cash Equivalents + Other Short-Term Investments',
        '• NOA(Option): 기본적으로 EV에서 차감하지 않음 (EV Tag가 "NOA(Option)"이면 GPCM의 NOA 열에 집계되지 않음)',
        '    → BS_Full 시트의 EV Tag를 "NOA"로 바꾸면 즉시 EV에서 차감됨 (투자부동산·관계기업투자 등 선택 적용)',
        '    → 대상 계정: Long-Term Equity Investment, Investment In Financial Assets, Investment Properties 등',
        '• Net Debt = IBD − Cash',
        '• EBIT = Operating Income only',
        '• EBITDA = Operating Income + D&A',
        '    → D&A = yfinance EBITDA − yfinance EBIT (야후가 현금흐름표 D&A를 제공하지 않아 손익 역산)',
        '    → ⚠ 둘 중 하나라도 없으면 D&A = 0 으로 처리되어 EBITDA = 영업이익이 됨.',
        '      해당 회사는 "D&A 산출근거" 열이 붉게 표시되며 Data_Quality 시트에 별도 기재됨. 사업보고서로 직접 확인 필요.',
        '• EV/Sales = EV ÷ Revenue,  EV/EBITDA = EV ÷ EBITDA,  EV/EBIT = EV ÷ EBIT',
        '• PER = Equity Value ÷ Net Income Common Stockholders (NI Parent)',
        '• PBR = Equity Value ÷ Stockholders Equity',
        '• PSR = Equity Value ÷ Revenue',
        '• PL Source: LTM(최근 4개 분기 합산) 우선. 단, 아래 검증을 통과한 경우에만 적용',
        '    → 최근 4개 공시기간의 결산일 간격이 240~320일일 것 (반기공시 기업의 2년치 합산 방지)',
        '    → 핵심 계정(Total Revenue, Operating Income)에 결측 분기가 없을 것',
        '    → 미통과 시 연간 데이터를 사용하고 Data_Quality 시트에 사유를 기재함',
        '• 재무제표 선택 원칙: 각 라벨(Y, Y-1 …)의 결산일 기준 "그 시점 이전" 최신 재무제표만 사용 (Look-ahead 방지)',
        '',
        '[ Beta & Risk Analysis ]',
        '• Data Source: 개별종목 주가 및 시장지수 모두 Yahoo Finance (yfinance)',
        '• 수익률 기준: 종목·지수 모두 배당·분할 반영 총수익 기준(auto_adjust=True)으로 통일',
        '    → 시가총액 산출용 주가는 실제 거래가격(auto_adjust=False)을 별도로 사용함',
        '• Beta 계산 방법:',
        '  - 5Y Monthly Beta: 5년간 월말 종가 → 월간 수익률 → 시장지수 대비 선형회귀',
        '  - 2Y Weekly Beta: 2년간 주말 종가 → 주간 수익률 → 시장지수 대비 선형회귀',
        '  - Raw Beta = SLOPE(종목수익률, 지수수익률)',
        '  - Adjusted Beta = 2/3 × Raw Beta + 1/3 × 1.0 (Bloomberg 방법론)',
        '  - R² = RSQ(...) : 설명력. 0.1 미만이면 베타 신뢰도가 낮으므로 산업 베타 등 대체 검토 권장',
        '  - Std Error : Raw Beta ± 2 × SE 가 약 95% 신뢰구간',
        '  - n : 회귀에 사용된 관측치 수',
        '• Market Index: KOSPI (^KS11), KOSDAQ (^KQ11), Nikkei 225 (^N225), S&P/TSX (^GSPTSE), DAX (^GDAXI) 등',
        '• 값 검증: NaN, inf, 극단값(-10 ~ 10 범위 벗어남) 필터링 → 공란 처리',
        '',
        '[ Tax Rate ]',
        '• 한국: 사업연도별 법인세 한계세율 (지방소득세 10% 포함)',
        '   - FY2023~2025 : ≤2억 9.9% | 2억~200억 20.9% | 200억~3,000억 23.1% | >3,000억 26.4%',
        '   - FY2026~     : ≤2억 11.0% | 2억~200억 22.0% | 200억~3,000억 24.2% | >3,000억 27.5%  (2025년 세법개정, 1%p 인상)',
        '   - 결손(세전이익 ≤ 0) 기업은 한계세율 개념이 성립하지 않으므로 2구간 세율을 적용하고 "세율 근거" 열에 표시함',
        '• 해외: KPMG Corporate Tax Rates Table (Combined Rate). 세율표 미보유 국가는 25% 적용 후 "세율 근거" 열에 표시',
        '• D/E Ratio = IBD ÷ (Equity Value + NCI)',
        '• Debt Ratio (D/V) = IBD ÷ (Equity Value + IBD + NCI)',
        '• Unlevered Beta = Levered(Adj) Beta ÷ (1 + (1 - Tax Rate) × D/E Ratio) [Hamada Model]',
        '',
        '[ Target WACC Calculation ]',
        '• Target WACC은 "WACC_Calculation" 시트에서 계산되며, 모든 값이 이 GPCM 시트를 참조하는 수식임',
        '• Ke = Rf + MRP × Relevered Beta + Size Premium',
        '  - Relevered Beta = Avg Unlevered Beta × (1 + (1 - Tax) × Target D/E)',
        '  - Size Premium: 한국공인회계사회 기준 (Micro/Low/Mid, 3분위·5분위)',
        '• Kd (Aftertax) = Kd (Pretax) × (1 - Target Tax Rate)',
        '• Target D/V = 피어 평균 부채비율 (GPCM 시트 Mean 행 참조)',
        '• WACC = (E/V) × Ke + (D/V) × Kd (Aftertax)',
        '• 민감도 분석은 WACC_Calculation 시트 하단의 Sensitivity 표 참조',
        '',
        '• N/M = Not Meaningful (분모가 0 이하)',
        '• GPCM 시트의 모든 금액은 BS_Full / PL_Data / Market_Cap 시트를 참조하는 엑셀 수식임 (셀 클릭 시 추적 가능)',
        '', '⚠ 데이터 출처: Yahoo Finance (yfinance). 실제 공시자료와 대조 검증 필요.',
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
        sc(ws.cell(r,1,note), fo=fNOTE); r+=1

    # 틀고정: BS → EV Components 이전 (I5 = Company/Ticker/등 고정, BS부터 스크롤)
    ws.freeze_panes='I5'

    # [Sheet 6] WACC_Calculation (Target 기업의 WACC 계산)
    # 모든 계산값은 GPCM 시트를 참조하는 엑셀 수식이다.
    # (이전 버전의 'Format' 열은 파이썬이 따로 계산한 값이라 B열 수식 결과와
    #  달라질 수 있었으므로 삭제하고, 산출근거를 Note 열로 통합함)
    ws_wacc = wb.create_sheet('WACC_Calculation')
    ws_wacc.sheet_properties.tabColor = COLOR_DARK_BLUE

    ws_wacc.merge_cells('A1:C1')
    sc(ws_wacc['A1'], fo=fT)
    ws_wacc['A1'] = 'Target WACC Calculation'
    ws_wacc.merge_cells('A2:C2')
    sc(ws_wacc['A2'], fo=Font(name='Arial', size=9, color=C_MG, italic=True))
    ws_wacc['A2'] = (f'Base: {base_period_str} | Peer Average Method | '
                     f'파란 셀 = 입력값(수정 가능), 노란 셀 = 수식')

    pWACC_PARAM = PatternFill('solid', fgColor='E3F2FD')   # 입력값
    pWACC_CALC = PatternFill('solid', fgColor='FFF9C4')    # 계산값
    pWACC_RESULT = PatternFill('solid', fgColor='FFE082')  # 최종 WACC

    fWH = Font(name='Arial', bold=True, size=9, color=C_W)
    fWSEC = Font(name='Arial', bold=True, size=10, color=C_W)
    fWNOTE = Font(name='Arial', size=8, color=C_MG)
    aWR = Alignment(horizontal='right')
    aWC = Alignment(horizontal='center')
    aWL = Alignment(horizontal='left')

    mean_row = stat_rows['Mean']
    wacc_rows = {}

    def wacc_section(row, title):
        ws_wacc.merge_cells(f'A{row}:C{row}')
        sc(ws_wacc.cell(row, 1), fo=fWSEC, fi=PatternFill('solid', fgColor=C_MB), al=aWC)
        ws_wacc.cell(row, 1, title)
        return row + 1

    def wacc_header(row):
        for ci, h in enumerate(['Item', 'Value', 'Note / 산출근거'], 1):
            ws_wacc.cell(row, ci, h)
            sc(ws_wacc.cell(row, ci), fo=fWH, fi=PatternFill('solid', fgColor=C_BL), al=aWC, bd=BD)
        return row + 1

    def wacc_line(row, key, label, value, note, fill, nf='0.00%', bold=False):
        fo = Font(name='Arial', bold=True, size=10) if bold else fA
        ws_wacc.cell(row, 1, label); sc(ws_wacc.cell(row, 1), fo=fo, bd=BD, al=aWL)
        ws_wacc.cell(row, 2).value = value
        sc(ws_wacc.cell(row, 2), fo=fo, fi=fill, bd=BD, al=aWR, nf=nf)
        ws_wacc.cell(row, 3, note); sc(ws_wacc.cell(row, 3), fo=fWNOTE, bd=BD, al=aWL)
        wacc_rows[key] = row
        return row + 1

    r_wacc = 4
    r_wacc = wacc_section(r_wacc, '[ 1 ] Input Parameters  (사용자 입력 — 직접 수정 가능)')
    r_wacc = wacc_header(r_wacc)

    input_params = [
        ('rf', 'Risk-Free Rate (Rf)', target_wacc_data['Rf'], '한국 10년 국채수익률 (사용자 입력)', '0.00%'),
        ('mrp', 'Market Risk Premium (MRP)', target_wacc_data['MRP'], '한국공인회계사회 권장 7~9%', '0.00%'),
        ('sp', 'Size Premium', target_wacc_data['Size_Premium'], '한국공인회계사회 (시가총액 분위 기준)', '0.00%'),
        ('kdp', 'Kd (Pretax)', target_wacc_data['Kd_Pretax'], '세전 타인자본비용 (사용자 입력)', '0.00%'),
        ('tax', 'Target Tax Rate', target_wacc_data['Target_Tax_Rate'], 'Target 기업 법인세율 (지방소득세 포함)', '0.00%'),
    ]
    for key, label, val, note, nf in input_params:
        r_wacc = wacc_line(r_wacc, key, label, val, note, pWACC_PARAM, nf)

    r_wacc += 1
    r_wacc = wacc_section(r_wacc, '[ 2 ] Peer Analysis  (GPCM 시트 Mean 행 참조)')
    r_wacc = wacc_header(r_wacc)

    beta_label = target_wacc_data.get('Beta_Label', '5Y Monthly')
    beta_col = CL['UB5'] if beta_type == '5Y' else CL['UB2']
    r_wacc = wacc_line(r_wacc, 'ub', f'Avg Unlevered Beta ({beta_label})',
                       f'=GPCM!{beta_col}{mean_row}',
                       f"피어 평균 (GPCM!{beta_col}{mean_row}) · {target_wacc_data.get('Beta_Basis', '')}",
                       pWACC_CALC, '0.0000')
    r_wacc = wacc_line(r_wacc, 'dv', 'Avg Debt Ratio (D/V)',
                       f"=GPCM!{CL['DV']}{mean_row}",
                       f"피어 평균 자본구조 (GPCM!{CL['DV']}{mean_row}) · {target_wacc_data.get('Debt_Ratio_Basis', '')}",
                       pWACC_CALC, '0.00%')
    r_wacc = wacc_line(r_wacc, 'de', 'Target D/E Ratio',
                       f"=IF(B{wacc_rows['dv']}<1,B{wacc_rows['dv']}/(1-B{wacc_rows['dv']}),0)",
                       '= (D/V) ÷ (1 − D/V)', pWACC_CALC, '0.0000')

    r_wacc += 1
    r_wacc = wacc_section(r_wacc, '[ 3 ] Target WACC')
    r_wacc = wacc_header(r_wacc)

    r_wacc = wacc_line(r_wacc, 'rb', 'Relevered Beta',
                       f"=B{wacc_rows['ub']}*(1+(1-B{wacc_rows['tax']})*B{wacc_rows['de']})",
                       'Unlevered β × (1 + (1 − Tax) × Target D/E)', pWACC_CALC, '0.0000')
    r_wacc = wacc_line(r_wacc, 'ke', 'Ke (Cost of Equity)',
                       f"=B{wacc_rows['rf']}+B{wacc_rows['mrp']}*B{wacc_rows['rb']}+B{wacc_rows['sp']}",
                       'Rf + MRP × Relevered β + Size Premium', pWACC_CALC)
    r_wacc = wacc_line(r_wacc, 'kd', 'Kd (Aftertax)',
                       f"=B{wacc_rows['kdp']}*(1-B{wacc_rows['tax']})",
                       'Kd (Pretax) × (1 − Tax Rate)', pWACC_CALC)
    r_wacc = wacc_line(r_wacc, 'ew', 'Equity Weight (E/V)',
                       f"=1-B{wacc_rows['dv']}", '1 − Debt Ratio', pWACC_CALC)
    r_wacc = wacc_line(r_wacc, 'dw', 'Debt Weight (D/V)',
                       f"=B{wacc_rows['dv']}", 'Debt Ratio', pWACC_CALC)
    r_wacc = wacc_line(r_wacc, 'wacc', 'WACC',
                       f"=B{wacc_rows['ew']}*B{wacc_rows['ke']}+B{wacc_rows['dw']}*B{wacc_rows['kd']}",
                       '(E/V) × Ke + (D/V) × Kd (Aftertax)', pWACC_RESULT, bold=True)

    # -----------------------------------------------------------------
    # [ 4 ] 민감도 분석
    #   위쪽 입력 셀을 참조하는 수식이므로, 가정을 바꾸면 표 전체가 재계산된다.
    # -----------------------------------------------------------------
    r_wacc += 2
    r_wacc = wacc_section(r_wacc, '[ 4 ] Sensitivity Analysis')

    def sensitivity_table(row, title, mode, row_label, row_deltas, row_ref,
                          col_label, col_deltas, col_ref, is_pct_row, is_pct_col):
        """WACC 민감도 표 1개를 기록하고 다음 행 번호를 반환."""
        ws_wacc.merge_cells(f'A{row}:H{row}')
        sc(ws_wacc.cell(row, 1), fo=Font(name='Arial', bold=True, size=9, color=C_DB),
           fi=pSTAT, al=aWL)
        ws_wacc.cell(row, 1, title)
        row += 1

        ws_wacc.cell(row, 1, f'{row_label} \\ {col_label}')
        sc(ws_wacc.cell(row, 1), fo=fWH, fi=PatternFill('solid', fgColor=C_BL), al=aWC, bd=BD)
        for j, cd in enumerate(col_deltas):
            c = ws_wacc.cell(row, 2 + j)
            c.value = f'={col_ref}{cd:+}'
            sc(c, fo=fWH, fi=PatternFill('solid', fgColor=C_BL), al=aWC, bd=BD,
               nf='0.00%' if is_pct_col else '0.00')
        hdr_row = row
        row += 1

        for rd in row_deltas:
            c0 = ws_wacc.cell(row, 1)
            c0.value = f'={row_ref}{rd:+}'
            sc(c0, fo=fSTAT, fi=pSTAT, al=aWC, bd=BD,
               nf='0.00%' if is_pct_row else '0.00')
            for j in range(len(col_deltas)):
                col_letter = get_column_letter(2 + j)
                cell = ws_wacc.cell(row, 2 + j)
                # 행 = 첫 번째 변수, 열 = 두 번째 변수
                if mode == 'beta_mrp':                       # 행=Unlevered Beta, 열=MRP
                    reb = f"$A{row}*(1+(1-$B${wacc_rows['tax']})*$B${wacc_rows['de']})"
                    ke = f"$B${wacc_rows['rf']}+{col_letter}${hdr_row}*{reb}+$B${wacc_rows['sp']}"
                    cell.value = (f"=(1-$B${wacc_rows['dv']})*({ke})"
                                  f"+$B${wacc_rows['dv']}*$B${wacc_rows['kdp']}*(1-$B${wacc_rows['tax']})")
                else:                                        # 행=D/V, 열=Kd(Pretax)
                    de = f"IF($A{row}<1,$A{row}/(1-$A{row}),0)"
                    reb = f"$B${wacc_rows['ub']}*(1+(1-$B${wacc_rows['tax']})*{de})"
                    ke = f"$B${wacc_rows['rf']}+$B${wacc_rows['mrp']}*{reb}+$B${wacc_rows['sp']}"
                    cell.value = (f"=(1-$A{row})*({ke})"
                                  f"+$A{row}*{col_letter}${hdr_row}*(1-$B${wacc_rows['tax']})")
                sc(cell, fo=fA, fi=pWACC_CALC, al=aWR, bd=BD, nf='0.00%')
            row += 1
        return row + 1

    r_wacc = sensitivity_table(
        r_wacc, '① Unlevered Beta(행) × MRP(열) → WACC', 'beta_mrp',
        'Unlevered β', [-0.20, -0.10, 0, 0.10, 0.20], f"$B${wacc_rows['ub']}",
        'MRP', [-0.01, -0.005, 0, 0.005, 0.01], f"$B${wacc_rows['mrp']}",
        is_pct_row=False, is_pct_col=True)

    r_wacc = sensitivity_table(
        r_wacc, '② 목표 부채비율 D/V(행) × Kd 세전(열) → WACC', 'dv_kd',
        'D/V', [-0.10, -0.05, 0, 0.05, 0.10], f"$B${wacc_rows['dv']}",
        'Kd (Pretax)', [-0.01, -0.005, 0, 0.005, 0.01], f"$B${wacc_rows['kdp']}",
        is_pct_row=True, is_pct_col=True)

    ws_wacc.column_dimensions['A'].width = 26
    ws_wacc.column_dimensions['B'].width = 14
    ws_wacc.column_dimensions['C'].width = 52
    for cl in 'DEFGH':
        ws_wacc.column_dimensions[cl].width = 14
    ws_wacc.freeze_panes = 'A4'

    # Named Range 정의 (다른 시트에서 참조 가능)
    from openpyxl.workbook.defined_name import DefinedName

    wacc_named = {
        'Target_WACC': wacc_rows['wacc'], 'Target_Rf': wacc_rows['rf'],
        'Target_MRP': wacc_rows['mrp'], 'Target_Size_Premium': wacc_rows['sp'],
        'Target_Kd_Pretax': wacc_rows['kdp'], 'Target_Tax_Rate': wacc_rows['tax'],
        'Target_Ke': wacc_rows['ke'],
    }
    for nm, row_no in wacc_named.items():
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"'WACC_Calculation'!$B${row_no}")


    # =====================================================================
    # [Sheet] Implied_Valuation — 피어 멀티플로 Target 지분가치 밴드 산출
    #   회색 셀(Target 재무수치)에 값을 넣으면 아래가 전부 자동 계산된다.
    # =====================================================================
    ws_iv = wb.create_sheet('Implied_Valuation')
    ws_iv.sheet_properties.tabColor = 'C62828'

    ws_iv.merge_cells('A1:H1'); sc(ws_iv.cell(1, 1, 'Implied Valuation (피어 멀티플 적용 Target 가치)'), fo=fT)
    ws_iv.merge_cells('A2:H2')
    sc(ws_iv.cell(2, 1, '① 회색 셀에 Target 재무수치를 입력 → ② 멀티플별 EV/지분가치 자동 산출 → ③ 하단에서 프리미엄·할인 조정'), fo=fS)

    pIN = PatternFill('solid', fgColor='EEEEEE')       # 사용자 입력
    pOUT = PatternFill('solid', fgColor='FFF9C4')      # 계산값
    pRESULT = PatternFill('solid', fgColor='FFE082')   # 최종 결과

    def iv_section(row, title, span=8):
        ws_iv.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        sc(ws_iv.cell(row, 1), fo=fWSEC, fi=PatternFill('solid', fgColor=C_MB), al=aWC)
        ws_iv.cell(row, 1, title)
        return row + 1

    r_iv = 4
    r_iv = iv_section(r_iv, '[ 1 ] Target 재무수치 입력  (회색 셀 — 직접 입력)')
    for ci, h in enumerate(['항목', '금액 (백만, 현지통화)', '비고'], 1):
        sc(ws_iv.cell(r_iv, ci, h), fo=fWH, fi=PatternFill('solid', fgColor=C_BL), al=aWC, bd=BD)
    r_iv += 1

    iv_rows = {}
    target_metric_defs = [
        ('Revenue', 'Revenue (매출)', 'LTM 또는 최근 연간 기준'),
        ('EBITDA', 'EBITDA', '영업이익 + 감가상각비'),
        ('EBIT', 'EBIT (영업이익)', ''),
        ('NI', 'Net Income (지배주주순이익)', ''),
        ('Equity', 'Equity (자본총계)', 'PBR 적용 시 사용'),
        ('IBD', 'IBD (이자부부채)', 'EV → 지분가치 전환에 사용'),
        ('Cash', 'Cash (현금성자산)', 'EV → 지분가치 전환에 사용'),
        ('NCI', 'NCI (비지배지분)', 'EV → 지분가치 전환에 사용'),
        ('Shares', '발행주식수 (주)', '주당가치 산출용 (선택)'),
    ]
    target_metric_label = {k: lbl for k, lbl, _ in target_metric_defs}
    for key, label, note in target_metric_defs:
        ws_iv.cell(r_iv, 1, label); sc(ws_iv.cell(r_iv, 1), fo=fA, bd=BD, al=aWL)
        ws_iv.cell(r_iv, 2, target_inputs.get(key) or None)
        sc(ws_iv.cell(r_iv, 2), fo=fA, fi=pIN, bd=BD, al=aWR,
           nf=NF_INT if key == 'Shares' else NF_M)
        ws_iv.cell(r_iv, 3, note); sc(ws_iv.cell(r_iv, 3), fo=fWNOTE, bd=BD, al=aWL)
        iv_rows[key] = r_iv
        r_iv += 1

    r_iv += 1
    r_iv = iv_section(r_iv, '[ 2 ] 피어 멀티플 적용 → Implied EV / 지분가치')
    iv_headers = ['멀티플', '기준 지표', 'Min', 'Q1 (25%)', 'Median', 'Q3 (75%)', 'Max', '적용 결과']
    for ci, h in enumerate(iv_headers, 1):
        sc(ws_iv.cell(r_iv, ci, h), fo=fWH, fi=PatternFill('solid', fgColor=C_BL), al=aWC, bd=BD)
    r_iv += 1

    # (표시명, GPCM 컬럼키, Target 지표행, EV배수 여부)
    iv_mult_defs = [
        ('EV/Sales', 'EV_Sales', 'Revenue', True),
        ('EV/EBITDA', 'EV_EBITDA', 'EBITDA', True),
        ('EV/EBIT', 'EV_EBIT', 'EBIT', True),
        ('PER', 'PER', 'NI', False),
        ('PBR', 'PBR', 'Equity', False),
        ('PSR', 'PSR', 'Revenue', False),
    ]
    stat_order = [('Min', 'Min'), ('Q1 (25%)', 'Q1 (25%)'), ('Median', 'Median'),
                  ('Q3 (75%)', 'Q3 (75%)'), ('Max', 'Max')]
    iv_equity_rows = {}

    for disp, colkey, metric_key, is_ev in iv_mult_defs:
        # 멀티플 행: GPCM 통계 참조
        ws_iv.cell(r_iv, 1, disp); sc(ws_iv.cell(r_iv, 1), fo=fHL, bd=BD, al=aWL)
        ws_iv.cell(r_iv, 2, target_metric_label[metric_key])
        sc(ws_iv.cell(r_iv, 2), fo=fWNOTE, bd=BD, al=aWL)
        for j, (_, sname) in enumerate(stat_order):
            c = ws_iv.cell(r_iv, 3 + j)
            c.value = f'=IFERROR(GPCM!{CL[colkey]}{stat_rows[sname]},"N/M")'
            sc(c, fo=fA, fi=pOUT, bd=BD, al=aWR, nf=NF_X)
        ws_iv.cell(r_iv, 8, 'Multiple (x)'); sc(ws_iv.cell(r_iv, 8), fo=fWNOTE, bd=BD, al=aWL)
        mult_row = r_iv
        r_iv += 1

        # 가치 행
        val_label = 'Implied EV' if is_ev else 'Implied Equity Value'
        ws_iv.cell(r_iv, 1, f'  → {val_label}'); sc(ws_iv.cell(r_iv, 1), fo=fA, bd=BD, al=aWL)
        ws_iv.cell(r_iv, 2, f'× {target_metric_label[metric_key]}')
        sc(ws_iv.cell(r_iv, 2), fo=fWNOTE, bd=BD, al=aWL)
        for j in range(len(stat_order)):
            cl_ = get_column_letter(3 + j)
            c = ws_iv.cell(r_iv, 3 + j)
            c.value = f'=IFERROR({cl_}{mult_row}*$B${iv_rows[metric_key]},"")'
            sc(c, fo=fA, fi=pOUT, bd=BD, al=aWR, nf=NF_M)
        sc(ws_iv.cell(r_iv, 8, '백만'), fo=fWNOTE, bd=BD, al=aWL)
        ev_row = r_iv
        r_iv += 1

        # EV배수는 순부채·NCI를 차감해 지분가치로 환산
        if is_ev:
            ws_iv.cell(r_iv, 1, '  → Implied Equity Value')
            sc(ws_iv.cell(r_iv, 1), fo=fA, bd=BD, al=aWL)
            ws_iv.cell(r_iv, 2, '− IBD + Cash − NCI')
            sc(ws_iv.cell(r_iv, 2), fo=fWNOTE, bd=BD, al=aWL)
            for j in range(len(stat_order)):
                cl_ = get_column_letter(3 + j)
                c = ws_iv.cell(r_iv, 3 + j)
                c.value = (f'=IFERROR({cl_}{ev_row}-$B${iv_rows["IBD"]}'
                           f'+$B${iv_rows["Cash"]}-$B${iv_rows["NCI"]},"")')
                sc(c, fo=fA, fi=pOUT, bd=BD, al=aWR, nf=NF_M)
            sc(ws_iv.cell(r_iv, 8, '백만'), fo=fWNOTE, bd=BD, al=aWL)
            iv_equity_rows[disp] = r_iv
            r_iv += 1
        else:
            iv_equity_rows[disp] = ev_row

        r_iv += 1

    # ---- 요약 밴드 ----
    r_iv = iv_section(r_iv, '[ 3 ] 지분가치 밴드 요약 및 조정')
    for ci, h in enumerate(['구분', '적용 멀티플', 'Min', 'Q1 (25%)', 'Median', 'Q3 (75%)', 'Max', '비고'], 1):
        sc(ws_iv.cell(r_iv, ci, h), fo=fWH, fi=PatternFill('solid', fgColor=C_BL), al=aWC, bd=BD)
    r_iv += 1

    # 사용자가 어떤 멀티플을 대표로 쓸지 선택 (기본: EV/EBITDA)
    ws_iv.cell(r_iv, 1, '적용 멀티플 선택'); sc(ws_iv.cell(r_iv, 1), fo=fHL, bd=BD, al=aWL)
    ws_iv.cell(r_iv, 2, 'EV/EBITDA'); sc(ws_iv.cell(r_iv, 2), fo=fA, fi=pIN, bd=BD, al=aWC)
    ws_iv.cell(r_iv, 8, 'EV/Sales · EV/EBITDA · EV/EBIT · PER · PBR · PSR 중 입력')
    sc(ws_iv.cell(r_iv, 8), fo=fWNOTE, bd=BD, al=aWL)
    sel_row = r_iv
    r_iv += 1

    mult_names_arr = ','.join('"' + d + '"' for d, _, _, _ in iv_mult_defs)
    base_row = r_iv
    ws_iv.cell(r_iv, 1, 'Base 지분가치'); sc(ws_iv.cell(r_iv, 1), fo=fA, bd=BD, al=aWL)
    ws_iv.cell(r_iv, 2, f'=B{sel_row}'); sc(ws_iv.cell(r_iv, 2), fo=fA, bd=BD, al=aWC)
    for j in range(len(stat_order)):
        cl_ = get_column_letter(3 + j)
        c = ws_iv.cell(r_iv, 3 + j)
        # 전체 열 참조(INDEX(C:C,...))는 자기 셀을 포함해 순환참조 경고를 유발하므로
        # 해당 멀티플의 지분가치 셀을 직접 나열한다.
        cells = ','.join(f'{cl_}{iv_equity_rows[d]}' for d, _, _, _ in iv_mult_defs)
        c.value = f'=IFERROR(CHOOSE(MATCH($B${sel_row},{{{mult_names_arr}}},0),{cells}),"")'
        sc(c, fo=fA, fi=pOUT, bd=BD, al=aWR, nf=NF_M)
    sc(ws_iv.cell(r_iv, 8, '선택한 멀티플의 Implied Equity Value'), fo=fWNOTE, bd=BD, al=aWL)
    r_iv += 1

    adj_defs = [
        ('prem', '경영권 프리미엄 (%)', 0.0, '경영권 이전 거래 시 적용. 미적용이면 0'),
        ('dlom', '비유동성 할인 DLOM (%)', 0.0, '비상장 지분 평가 시 적용. 미적용이면 0'),
    ]
    adj_rows = {}
    for key, label, default, note in adj_defs:
        ws_iv.cell(r_iv, 1, label); sc(ws_iv.cell(r_iv, 1), fo=fA, bd=BD, al=aWL)
        ws_iv.cell(r_iv, 2, default); sc(ws_iv.cell(r_iv, 2), fo=fA, fi=pIN, bd=BD, al=aWR, nf='0.0%')
        ws_iv.cell(r_iv, 8, note); sc(ws_iv.cell(r_iv, 8), fo=fWNOTE, bd=BD, al=aWL)
        adj_rows[key] = r_iv
        r_iv += 1

    ws_iv.cell(r_iv, 1, '조정 후 지분가치'); sc(ws_iv.cell(r_iv, 1), fo=Font(name='Arial', bold=True, size=10), bd=BD, al=aWL)
    for j in range(len(stat_order)):
        cl_ = get_column_letter(3 + j)
        c = ws_iv.cell(r_iv, 3 + j)
        c.value = (f'=IFERROR({cl_}{base_row}*(1+$B${adj_rows["prem"]})'
                   f'*(1-$B${adj_rows["dlom"]}),"")')
        sc(c, fo=Font(name='Arial', bold=True, size=10), fi=pRESULT, bd=BD, al=aWR, nf=NF_M)
    sc(ws_iv.cell(r_iv, 8, '= Base × (1 + 프리미엄) × (1 − DLOM)'), fo=fWNOTE, bd=BD, al=aWL)
    adj_row = r_iv
    r_iv += 1

    ws_iv.cell(r_iv, 1, '주당가치'); sc(ws_iv.cell(r_iv, 1), fo=fA, bd=BD, al=aWL)
    for j in range(len(stat_order)):
        cl_ = get_column_letter(3 + j)
        c = ws_iv.cell(r_iv, 3 + j)
        c.value = f'=IFERROR({cl_}{adj_row}*1000000/$B${iv_rows["Shares"]},"")'
        sc(c, fo=fA, fi=pOUT, bd=BD, al=aWR, nf=NF_PRC)
    sc(ws_iv.cell(r_iv, 8, '조정 후 지분가치(백만) ÷ 발행주식수'), fo=fWNOTE, bd=BD, al=aWL)
    r_iv += 2

    iv_notes = [
        '[ 사용 방법 ]',
        '1. [1] 구역의 회색 셀에 Target 회사의 재무수치를 입력합니다 (단위: 백만, 피어와 동일 통화 기준).',
        '2. [2] 구역의 멀티플은 GPCM 시트 통계행을 참조하므로, 피어를 제외/추가하면 자동 반영됩니다.',
        '3. [3] 구역에서 대표 멀티플을 고르고 프리미엄·할인율을 입력하면 최종 밴드가 산출됩니다.',
        '',
        '[ 유의사항 ]',
        '• 피어와 Target의 통화가 다르면 반드시 동일 통화로 환산한 뒤 입력해야 합니다.',
        '• EV 배수는 순부채·비지배지분을 차감해 지분가치로 환산합니다 (= EV − IBD + Cash − NCI).',
        '• PER·PBR·PSR은 이미 지분가치 기준이므로 차감하지 않습니다.',
        '• 경영권 프리미엄과 DLOM은 상호 배타적이지 않으나, 중복 적용 시 근거를 문서화해야 합니다.',
        '• D&A 역산이 불가했던 피어가 있으면 EV/EBITDA 배수가 EV/EBIT과 동일해집니다 (Data_Quality 시트 확인).',
    ]
    for note in iv_notes:
        ws_iv.merge_cells(start_row=r_iv, start_column=1, end_row=r_iv, end_column=8)
        sc(ws_iv.cell(r_iv, 1, note), fo=fNOTE); r_iv += 1

    ws_iv.column_dimensions['A'].width = 28
    ws_iv.column_dimensions['B'].width = 24
    for cl in 'CDEFG':
        ws_iv.column_dimensions[cl].width = 16
    ws_iv.column_dimensions['H'].width = 46

    # =====================================================================
    # [Sheet] Data_Quality — 자동 수집 과정에서 확인이 필요한 항목
    # =====================================================================
    ws_dq = wb.create_sheet('Data_Quality')
    ws_dq.sheet_properties.tabColor = 'EF6C00'

    ws_dq.merge_cells('A1:E1'); sc(ws_dq.cell(1, 1, 'Data Quality Check (확인 필요 항목)'), fo=fT)
    ws_dq.merge_cells('A2:E2')
    sc(ws_dq.cell(2, 1, 'ERROR = 해당 값 사용 불가 / WARN = 수치가 왜곡될 수 있음 / INFO = 참고사항. '
                        '자동 수집의 한계를 표시한 것이므로 공시자료와 대조 검증하십시오.'), fo=fS)

    dq_cols = [('Level', 10), ('Ticker', 12), ('Label', 10), ('Item', 18), ('Message', 100)]
    r_dq = 4
    for ci, (h, w) in enumerate(dq_cols, 1):
        ws_dq.column_dimensions[get_column_letter(ci)].width = w
        sc(ws_dq.cell(r_dq, ci, h), fo=fWH, fi=PatternFill('solid', fgColor=C_BL), al=aWC, bd=BD)
    dq_hdr = r_dq
    r_dq += 1

    level_fill = {'ERROR': PatternFill('solid', fgColor='FFCDD2'),
                  'WARN': PatternFill('solid', fgColor='FFE0B2'),
                  'INFO': PatternFill('solid', fgColor='E3F2FD')}
    level_order = {'ERROR': 0, 'WARN': 1, 'INFO': 2}
    sorted_flags = sorted(data_quality_rows,
                          key=lambda x: (level_order.get(x.get('Level'), 9),
                                         str(x.get('Ticker', '')), str(x.get('Label', ''))))

    if sorted_flags:
        for fl in sorted_flags:
            lv = fl.get('Level', 'INFO')
            vals = [lv, fl.get('Ticker', ''), fl.get('Label', ''), fl.get('Item', ''), fl.get('Message', '')]
            for ci, v in enumerate(vals, 1):
                c = ws_dq.cell(r_dq, ci, v)
                sc(c, fo=fA, fi=level_fill.get(lv, pW), bd=BD,
                   al=aWC if ci <= 3 else aWL)
            r_dq += 1
    else:
        ws_dq.merge_cells(start_row=r_dq, start_column=1, end_row=r_dq, end_column=5)
        sc(ws_dq.cell(r_dq, 1, '✅ 자동 점검에서 특이사항이 발견되지 않았습니다.'), fo=fA)
        r_dq += 1

    ws_dq.auto_filter.ref = f"A{dq_hdr}:E{r_dq-1}"
    ws_dq.freeze_panes = f'A{dq_hdr+1}'

    # [Sheet 11] Price_History (Set tab color to black)
    if price_abs_dfs:
        ws_ph = wb.create_sheet('Price_History')
        ws_ph.sheet_properties.tabColor = "000000" # Black tab
        df_abs = pd.concat(price_abs_dfs, axis=1).sort_index().ffill()
        df_rel = pd.concat(price_rel_dfs, axis=1).sort_index().ffill()
        common_index = df_abs.index
        sc(ws_ph.cell(1,1,'Stock Price History (10 Years)'), fo=fT)
        ws_ph.merge_cells(start_row=1,start_column=1,end_row=1,end_column=10)
        r=3
        ws_ph.cell(r,1,'Date'); sc(ws_ph.cell(r,1), fo=fH, fi=pH, al=aC, bd=BD); ws_ph.column_dimensions['A'].width=12
        c_idx=2
        for col in df_abs.columns:
            sc(ws_ph.cell(r,c_idx,f"{ticker_to_name.get(col,col)} (Abs)"), fo=fH, fi=PatternFill('solid',fgColor='607D8B'), al=aC, bd=BD)
            ws_ph.column_dimensions[get_column_letter(c_idx)].width=16; c_idx+=1
        sc(ws_ph.cell(r,c_idx,""), fi=pW); ws_ph.column_dimensions[get_column_letter(c_idx)].width=2; c_idx+=1
        ws_ph.cell(r,c_idx,'Date'); sc(ws_ph.cell(r,c_idx), fo=fH, fi=pH, al=aC, bd=BD); rel_start=c_idx; c_idx+=1
        for col in df_rel.columns:
            sc(ws_ph.cell(r,c_idx,f"{ticker_to_name.get(col,col)} (Rel)"), fo=fH, fi=pH, al=aC, bd=BD)
            ws_ph.column_dimensions[get_column_letter(c_idx)].width=16; c_idx+=1
        r=4
        for date in common_index:
            dv=date.date(); ws_ph.cell(r,1,dv).number_format='yyyy-mm-dd'; sc(ws_ph.cell(r,1), fo=fA, al=aC, bd=BD); cc=2
            for v in df_abs.loc[date]: ws_ph.cell(r,cc,v).number_format='#,##0.00'; sc(ws_ph.cell(r,cc), fo=fA, al=aR, bd=BD); cc+=1
            cc+=1; ws_ph.cell(r,cc,dv).number_format='yyyy-mm-dd'; sc(ws_ph.cell(r,cc), fo=fA, al=aC, bd=BD); cc+=1
            for v in df_rel.loc[date]: ws_ph.cell(r,cc,v).number_format='#,##0'; sc(ws_ph.cell(r,cc), fo=fA, al=aR, bd=BD); cc+=1
            r+=1
        
        # Monthly Chart Data
        chart_start=r+2; df_m=df_rel.resample('ME').last().dropna(how='all')
        cr=chart_start; sc(ws_ph.cell(cr,1,'[ Chart Data - Monthly Sampled ]'), fo=fNOTE); cr+=1
        hdr_row=cr; sc(ws_ph.cell(cr,1,'Year-Month'), fo=Font(name='Arial',size=8,bold=True,color=C_MG), al=aC)
        for i,cn in enumerate(df_m.columns): sc(ws_ph.cell(cr,i+2,ticker_to_name.get(cn,cn)), fo=Font(name='Arial',size=8,bold=True,color=C_MG), al=aC)
        cr+=1; data_first=cr
        for date in df_m.index:
            ws_ph.cell(cr,1,date.strftime('%Y-%m')); sc(ws_ph.cell(cr,1), fo=Font(name='Arial',size=7,color=C_MG))
            for i,cn in enumerate(df_m.columns):
                v=df_m.loc[date,cn]; 
                if pd.notna(v): ws_ph.cell(cr,i+2,round(v,1))
                sc(ws_ph.cell(cr,i+2), fo=Font(name='Arial',size=7,color=C_MG))
            cr+=1
        data_last=cr-1
        
        # Chart
        chart=LineChart(); chart.title="10-Year Relative Performance (Base=100)"; chart.style=13; chart.height=18; chart.width=34
        chart.y_axis.title="Relative Index"; chart.y_axis.scaling.min=0; chart.y_axis.majorGridlines=ChartLines()
        chart.x_axis.title="Date"; chart.x_axis.tickLblSkip=max(1,(data_last-data_first)//10)
        cats=Reference(ws_ph, min_col=1, min_row=data_first, max_row=data_last); chart.set_categories(cats)
        colors=['1F77B4','FF7F0E','2CA02C','D62728','9467BD','8C564B','E377C2','7F7F7F','BCBD22','17BECF']
        for i in range(len(df_m.columns)):
            s=Series(Reference(ws_ph, min_col=i+2, min_row=hdr_row, max_row=data_last), title_from_data=True)
            s.graphicalProperties.line.solidFill=colors[i%len(colors)]; s.graphicalProperties.line.width=20000; s.smooth=False
            chart.series.append(s)
        ws_ph.add_chart(chart, f"{get_column_letter(rel_start+len(df_m.columns)+1)}3")

    # Temp 시트 삭제
    if 'Temp' in wb.sheetnames:
        del wb['Temp']

    # [Reorder Sheets] Strictly per user request:
    # 1. GPCM, 2. BS_FULL (base), 3. PL_Data (base), 4. WACC_Calculation, 5. Beta_Calculation, 6. Multiples_Trend, 7. Market_Cap, 8. Price_History, 9. Historicals...
    
    desired_order = ['GPCM', 'Implied_Valuation', 'WACC_Calculation', 'Data_Quality',
                     'BS_FULL (base)', 'PL_Data (base)', 'Beta_Calculation',
                     'Multiples_Trend', 'Market_Cap']
    
    # Get all actual sheet names
    current_sheets = wb.sheetnames
    
    # Sort history sheets (Y-1, Y-2...)
    hist_sheets = [s for s in current_sheets if any(x in s for x in ['BS_Full_Y-', 'PL_Data_Y-', 'BS_최신', 'PL_최신'])]
    def hist_sort(s):
        if '최신' in s: return -1
        try: return int(s.split('-')[1])
        except Exception: return 999
    hist_sheets.sort(key=hist_sort)
    
    final_order = []
    for d in desired_order:
        if d in current_sheets:
            final_order.append(d)
    
    # Add history sheets
    for s in hist_sheets:
        if s not in final_order:
            final_order.append(s)
            
    # Add any remaining sheets (excluding Price_History)
    for s in current_sheets:
        if s not in final_order and s != 'Price_History':
            final_order.append(s)
            
    # [CRITICAL] ALWAYS Price_History at the VERY END
    if 'Price_History' in current_sheets:
        final_order.append('Price_History')
            
    # Apply order
    new_sheets = []
    for name in final_order:
        new_sheets.append(wb[name])
    wb._sheets = new_sheets

    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# 5. Streamlit App Layout
# ==========================================
st.set_page_config(page_title="Global GPCM Generator", layout="wide", page_icon="📊")

# ---------------------------------------------------------
# [User Access Log] 접속자 로그 기록 (Console 출력)
# ---------------------------------------------------------
try:
    # Streamlit Cloud (Private) 환경에서 이메일 가져오기 (st.user: 최신, experimental_user: 구버전 fallback)
    _u = getattr(st, 'user', None) or getattr(st, 'experimental_user', None)
    user_email = (_u.email if _u and _u.email else "Anonymous")
except Exception:
    user_email = "Local_Dev"

# 현재 시간 (이미 import datetime이 되어 있으므로 바로 사용)
now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def export_historical_excel_global_v2(all_period_data, raw_bs_rows, raw_pl_rows, market_rows, target_periods, ticker_to_name):
    output = io.BytesIO()
    wb = Workbook(); wb.remove(wb.active)

    rel_labels = sorted([k for k in all_period_data.keys() if all_period_data[k]], key=label_sort_key)
    
    # Sheet Colors
    COLOR_DARK_BLUE = '00338D'
    COLOR_PURPLE = '6A1B9A'

    # Styles & Colors (NameError sc 방지용 로컬 정의)
    C_BL='00338D'; C_DB='1E2A5E'; C_MB='005EB8'; C_LB='C3D7EE'; C_PB='E8EFF8'
    C_DG='333333'; C_MG='666666'; C_LG='F5F5F5'; C_BG='B0B0B0'; C_W='FFFFFF'
    C_GR='E2EFDA'; C_YL='FFF8E1'; C_NOA='FCE4EC'

    S1=Side(style='thin',color=C_BG); BD=Border(left=S1,right=S1,top=S1,bottom=S1)
    fT=Font(name='Arial',bold=True,size=14,color=C_BL); fS=Font(name='Arial',size=9,color=C_MG,italic=True)
    fH=Font(name='Arial',bold=True,size=9,color=C_W); fA=Font(name='Arial',size=9,color=C_DG)
    fHL=Font(name='Arial',bold=True,size=9,color=C_DB)
    fFRM=Font(name='Arial',size=9,color='000000'); fFRM_B=Font(name='Arial',bold=True,size=9,color='000000')
    fLINK=Font(name='Arial',size=9,color='008000')

    pH=PatternFill('solid',fgColor=C_BL); pW=PatternFill('solid',fgColor=C_W)
    pST=PatternFill('solid',fgColor=C_LG); pSEC2=PatternFill('solid',fgColor=C_DB)
    pSEC1=PatternFill('solid',fgColor='1565C0') # 진한 파랑

    ev_fills = {'Cash':PatternFill('solid',fgColor=C_GR), 'IBD':PatternFill('solid',fgColor=C_YL),
                'NCI':PatternFill('solid',fgColor=C_PB), 'NOA(Option)':PatternFill('solid',fgColor=C_NOA),
                'Equity':PatternFill('solid',fgColor=C_LB), 'PL_HL':PatternFill('solid',fgColor=C_YL),
                'NI_Parent':PatternFill('solid',fgColor=C_YL)}
    
    aC=Alignment(horizontal='center',vertical='center',wrap_text=True)
    aL=Alignment(horizontal='left',vertical='center',indent=1)
    aR=Alignment(horizontal='right',vertical='center')
    NF_M='#,##0;(#,##0);"-"'; NF_M1='#,##0.0;(#,##0.0);"-"'; NF_PRC='#,##0.00;(#,##0.00);"-"'
    NF_INT='#,##0;(#,##0);"-"'; NF_EPS='#,##0.00;(#,##0.00);"-"'; NF_X='0.0"x";(0.0"x");"-"'; NF_PCT='0.0%;(0.0%);"-"'

    def sc(c,fo=None,fi=None,al=None,bd=None,nf=None):
        if fo:c.font=fo
        if fi:c.fill=fi
        if al:c.alignment=al
        if bd:c.border=bd
        if nf:c.number_format=nf

    # [1] Summary Sheet
    ws_summ = wb.create_sheet('Historical_Summary')
    ws_summ.sheet_properties.tabColor = COLOR_PURPLE
    
    ws_summ.merge_cells('A1:AA1')
    ws_summ['A1'] = f"Historical Financial Summary - yfinance Base (Label: Y, Y-1...)"
    sc(ws_summ['A1'], fo=fT)
    ws_summ.merge_cells('A2:AA2')
    ws_summ['A2'] = "Base on dynamic BS/PL Sheets per period | Unit: Millions (local currency)"
    sc(ws_summ['A2'], fo=fS)
    
    # ... (metrics definition remains same)
    
    mc_map = {} 
    
    ws_summ.cell(row=3, column=1, value="Company"); sc(ws_summ.cell(row=3, column=1), fo=fH, fi=pSEC1, al=aC, bd=BD)
    ws_summ.cell(row=3, column=2, value="Ticker"); sc(ws_summ.cell(row=3, column=2), fo=fH, fi=pSEC1, al=aC, bd=BD)
    ws_summ.merge_cells('A3:A4')
    ws_summ.merge_cells('B3:B4')

    col_idx = 3
    for m_key, m_name, _, _ in [
        ('Financial Date', 'Financial Date', 'Formula_Date', None),
        ('Revenue', 'Revenue(매출)', 'PL_Tag', NF_M), 
        ('Gross Profit', 'Gross Profit', 'PL_Acc', NF_M), 
        ('Operating Income', 'EBIT(영업이익)', 'PL_Acc', NF_M), 
        ('EBITDA_Calc', 'EBITDA', 'Formula', NF_M), 
        ('NI_Parent', 'Net Income(순이익)', 'PL_Tag', NF_M),
        ('Total Assets', 'Total Assets(자산)', 'BS_Acc', NF_M), 
        ('Total Liabilities Net Minority Interest', 'Total Liab(부채)', 'BS_Acc', NF_M), 
        ('Equity', 'Equity(자본)', 'BS_Tag', NF_M),
        ('Cash', 'Cash', 'BS_Tag', NF_M), 
        ('IBD', 'IBD(차입금)', 'BS_Tag', NF_M),
        ('NCI', 'NCI(비지배지분)', 'BS_Tag', NF_M),
        ('NetDebt', 'Net Debt(순부채)', 'Formula', NF_M),
        ('OPM', 'OPM(영업이익률)', 'Formula', NF_PCT), 
        ('DebtRatio', 'Debt Ratio(부채비율)', 'Formula', NF_PCT),
        ('Mkt_Cap', 'Equity Value(자기자본가치)', 'Mkt', NF_M1),
        ('EV', 'EV(기업가치)', 'Formula', NF_M),
    ]:
        start_col = col_idx
        end_col = col_idx + len(rel_labels) - 1
        ws_summ.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
        ws_summ.cell(row=3, column=start_col, value=m_name)
        sc(ws_summ.cell(row=3, column=start_col), fo=fH, fi=pSEC1, al=aC, bd=BD)
        
        for label in rel_labels:
            ws_summ.cell(row=4, column=col_idx, value=label)
            sc(ws_summ.cell(row=4, column=col_idx), fo=fH, fi=pSEC2, al=aC, bd=BD)
            mc_map[(m_key, label)] = get_column_letter(col_idx)
            col_idx += 1

    r = 5
    for ticker, comp_name in ticker_to_name.items():
        ws_summ.cell(row=r, column=1, value=comp_name); sc(ws_summ.cell(row=r, column=1), fo=fA, bd=BD)
        ws_summ.cell(row=r, column=2, value=ticker); sc(ws_summ.cell(row=r, column=2), fo=fA, al=aC, bd=BD)
        
        c = 3
        for m_key, m_name, m_type, fmt in [
            ('Financial Date', 'Financial Date', 'Formula_Date', None),
            ('Revenue', 'Revenue(매출)', 'PL_Tag', NF_M), 
            ('Gross Profit', 'Gross Profit', 'PL_Acc', NF_M), 
            ('Operating Income', 'EBIT(영업이익)', 'PL_Acc', NF_M), 
            ('EBITDA_Calc', 'EBITDA', 'Formula', NF_M), 
            ('NI_Parent', 'Net Income(순익)', 'PL_Tag', NF_M),
            ('Total Assets', 'Total Assets(자산)', 'BS_Acc', NF_M), 
            ('Total Liabilities Net Minority Interest', 'Total Liab(부채)', 'BS_Acc', NF_M), 
            ('Equity', 'Equity(자본)', 'BS_Tag', NF_M),
            ('Cash', 'Cash', 'BS_Tag', NF_M), 
            ('IBD', 'IBD(차입금)', 'BS_Tag', NF_M),
            ('NCI', 'NCI(비지배지분)', 'BS_Tag', NF_M),
            ('NetDebt', 'Net Debt(순부채)', 'Formula', NF_M),
            ('OPM', 'OPM(영업이익률)', 'Formula', NF_PCT), 
            ('DebtRatio', 'Debt Ratio(부채비율)', 'Formula', NF_PCT),
            ('Mkt_Cap', 'Equity Value(자기자본가치)', 'Mkt', NF_M1),
            ('EV', 'EV(기업가치)', 'Formula', NF_M),
        ]:
            for label in rel_labels:
                is_recent = (label == 'Recent')
                bs_sn = 'BS_최신' if is_recent else f"BS_Full_{label}"
                pl_sn = 'PL_최신' if is_recent else f"PL_Data_{label}"
                
                v = ""
                if m_type == 'Formula_Date':
                    # Get actual fiscal date for this peer and label
                    gpcm_p = all_period_data[label].get(ticker)
                    v = (gpcm_p.get('PL_Date') or gpcm_p.get('Base_Date', '-')) if gpcm_p else "-"
                elif m_type == 'BS_Tag':
                    v = f'=SUMIFS(\'{bs_sn}\'!G:G, \'{bs_sn}\'!B:B, $B{r}, \'{bs_sn}\'!F:F, "{m_key}")'
                elif m_type == 'BS_Acc':
                    v = f'=SUMIFS(\'{bs_sn}\'!G:G, \'{bs_sn}\'!B:B, $B{r}, \'{bs_sn}\'!E:E, "{m_key}")'
                elif m_type == 'PL_Tag':
                    v = f'=SUMIFS(\'{pl_sn}\'!J:J, \'{pl_sn}\'!B:B, $B{r}, \'{pl_sn}\'!E:E, "{m_key}")'
                elif m_type == 'PL_Acc':
                    v = f'=SUMIFS(\'{pl_sn}\'!J:J, \'{pl_sn}\'!B:B, $B{r}, \'{pl_sn}\'!D:D, "{m_key}")'
                elif m_type == 'Mkt':
                    # K열 = Equity Value (시가총액 + 우선주 장부금액)
                    v = f'=SUMIFS(Market_Cap!K:K, Market_Cap!B:B, $B{r}, Market_Cap!H:H, "{label}")'
                elif m_type == 'Formula':
                    if m_key == 'EBITDA_Calc':
                        ebit_addr = f"{mc_map[('Operating Income', label)]}{r}"
                        v = f'={ebit_addr} + SUMIFS(\'{pl_sn}\'!J:J, \'{pl_sn}\'!B:B, $B{r}, \'{pl_sn}\'!D:D, "EBITDA") - SUMIFS(\'{pl_sn}\'!J:J, \'{pl_sn}\'!B:B, $B{r}, \'{pl_sn}\'!D:D, "EBIT")'
                    elif m_key == 'NetDebt':
                        v = f"={mc_map[('IBD', label)]}{r} - {mc_map[('Cash', label)]}{r}"
                    elif m_key == 'OPM':
                        _num = mc_map[('Operating Income', label)]
                        _den = mc_map[('Revenue', label)]
                        v = f'=IFERROR({_num}{r}/{_den}{r}, "")'
                    elif m_key == 'DebtRatio':
                        _num = mc_map[('Total Liabilities Net Minority Interest', label)]
                        _den = mc_map[('Equity', label)]
                        v = f'=IFERROR({_num}{r}/{_den}{r}, "")'
                    elif m_key == 'EV':
                        v = f"={mc_map[('Mkt_Cap', label)]}{r} + {mc_map[('IBD', label)]}{r} - {mc_map[('Cash', label)]}{r} + {mc_map[('NCI', label)]}{r}"

                ws_summ.cell(row=r, column=c, value=v)
                font_style = fFRM_B if m_type == 'Formula' else (fA if m_type == 'Formula_Date' else fLINK)
                sc(ws_summ.cell(row=r, column=c), fo=font_style, nf=fmt, bd=BD)
                c += 1
        r += 1

    ws_summ.column_dimensions['A'].width = 18
    ws_summ.column_dimensions['B'].width = 10
    for i in range(3, c):
        ws_summ.column_dimensions[get_column_letter(i)].width = 14
    ws_summ.freeze_panes = "C5"

    # [2] BS_Full, PL_Data 시트들 생성
    # Recent 시트들을 먼저 생성하여 Summary 바로 뒤에 오도록 함 (index 1, 2)
    # bs_idx = 1 # Historical_Summary가 0번
    for label in rel_labels:
        bs_rows_p = [row for row in raw_bs_rows if row.get('Label') == label]
        if bs_rows_p:
            is_recent = (label == 'Recent')
            sheet_title = 'BS_최신' if is_recent else f'BS_Full_{label}'
            # Recent면 인덱스 1에 삽입, 아니면 맨 뒤에 추가
            ws_bs = wb.create_sheet(sheet_title, 1 if is_recent else None)
            ws_bs.sheet_properties.tabColor = COLOR_DARK_BLUE
            
            df_bs = pd.DataFrame(bs_rows_p)
            cols = [('Company','Company',18),('Ticker','Ticker',10), ('Period','Period',12),('Currency','Curr',6), ('Account','Account',42),('EV_Tag','EV Tag',14), ('Amount_M','Amount (M)',18)]
            ws_bs.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols)); sc(ws_bs.cell(1,1,f'Balance Sheet ({label})'),fo=fT)
            r_idx=3
            for i,(lb,key) in enumerate([('Cash','Cash'),('IBD','IBD'),('NCI','NCI'),('NOA(Option)','NOA(Option)'),('Equity','Equity')]):
                sc(ws_bs.cell(r_idx,i+1,lb), fo=Font(name='Arial',size=8,bold=True),fi=ev_fills[key],al=aC,bd=BD)
            r_idx=5
            for i,(col,disp,w) in enumerate(cols): ws_bs.column_dimensions[get_column_letter(i+1)].width=w; sc(ws_bs.cell(r_idx,i+1,disp),fo=fH,fi=pH,al=aC,bd=BD)
            hdr=r_idx; r_idx+=1
            for _,rd in df_bs.iterrows():
                ev_tag=rd.get('EV_Tag', ''); is_hl=bool(ev_tag)
                row_fi=ev_fills.get(ev_tag, pST if r_idx%2==0 else pW) if is_hl else (pST if r_idx%2==0 else pW)
                row_font=fHL if is_hl else fA
                for i,(col,_,_) in enumerate(cols):
                    c_cell=ws_bs.cell(r_idx,i+1); v=rd.get(col, '')
                    if isinstance(v,(float,np.floating)): c_cell.value=round(v,1) if pd.notna(v) else None
                    else: c_cell.value=v
                    sc(c_cell,fo=row_font,fi=row_fi,al=aR if col=='Amount_M' else aL,bd=BD,nf=NF_M if col=='Amount_M' else None)
                r_idx+=1
            ws_bs.auto_filter.ref=f"A{hdr}:{get_column_letter(len(cols))}{r_idx-1}"; ws_bs.freeze_panes=f'A{hdr+1}'

    for label in rel_labels:
        pl_rows_p = [row for row in raw_pl_rows if row.get('Label') == label]
        if pl_rows_p:
            is_recent = (label == 'Recent')
            sheet_title = 'PL_최신' if is_recent else f'PL_Data_{label}'
            # Recent면 인덱스 2에 삽입 (BS_최신이 인덱스 1이므로 그 뒤), 아니면 맨 뒤에 추가
            ws_pl = wb.create_sheet(sheet_title, 2 if is_recent else None)
            ws_pl.sheet_properties.tabColor = COLOR_DARK_BLUE
            
            df_pl = pd.DataFrame(pl_rows_p).sort_values(['Company','_sort','Q_Label'])
            cols = [('Company','Company',18),('Ticker','Ticker',10), ('Currency','Curr',6),('Account','Account',42), ('GPCM_Tag','GPCM Tag',14),('Unit','Unit',10), ('PL_Source','Source',14),('Q_Label','Q Label',9), ('Period','Period',12),('Amount_M','Amount',18)]
            ws_pl.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols)); sc(ws_pl.cell(1,1,f'Income Statement ({label})'),fo=fT)
            r_idx=5
            for i,(col,disp,w) in enumerate(cols): ws_pl.column_dimensions[get_column_letter(i+1)].width=w; sc(ws_pl.cell(r_idx,i+1,disp),fo=fH,fi=pH,al=aC,bd=BD)
            hdr=r_idx; r_idx+=1
            for _,rd in df_pl.iterrows():
                is_hl=bool(rd.get('GPCM_Tag', '')); row_fi=ev_fills.get('PL_HL', pW) if is_hl else (pST if r_idx%2==0 else pW)
                row_font=fHL if is_hl else fA
                for i,(col,_,_) in enumerate(cols):
                    if col=='_sort': continue
                    c_cell=ws_pl.cell(r_idx,i+1); v=rd.get(col, '')
                    if isinstance(v,(float,np.floating)): c_cell.value=round(v,1) if pd.notna(v) else None
                    else: c_cell.value=v
                    is_eps=rd.get('Unit','')=='per share'; nf=NF_EPS if is_eps else NF_M
                    sc(c_cell,fo=row_font,fi=row_fi,al=aR if col=='Amount_M' else aL,bd=BD,nf=nf if col=='Amount_M' else None)
                r_idx+=1
            ws_pl.auto_filter.ref=f"A{hdr}:{get_column_letter(len(cols))}{r_idx-1}"; ws_pl.freeze_panes=f'A{hdr+1}'

    # [3] Market_Cap Sheet
    ws_mc = wb.create_sheet('Market_Cap')
    ws_mc.sheet_properties.tabColor = COLOR_DARK_BLUE
    if market_rows:
        df_mkt = pd.DataFrame(market_rows)
        cols = [('Company','Company',18),('Ticker','Ticker',10), ('Base_Date','Base Date',12),('Price_Date','Price Date',12), ('Currency','Curr',6),('Close','Close Price',14), ('Shares','Shares (Ord.)',18),('Label','Label',10),('Market_Cap_M','Mkt Cap (M)',20),('Preferred_M','Preferred (M)',16),('Equity_Value_M','Equity Value (M)',20)]
        ws_mc.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols)); sc(ws_mc.cell(1,1,'Market Capitalization'),fo=fT)
        r_idx=4
        for i,(col,disp,w) in enumerate(cols): ws_mc.column_dimensions[get_column_letter(i+1)].width=w; sc(ws_mc.cell(r_idx,i+1,disp),fo=fH,fi=pH,al=aC,bd=BD)
        hdr=r_idx; r_idx+=1
        for _,rd in df_mkt.iterrows():
            ev=(r_idx%2==0)
            for i,(col,_,_) in enumerate(cols):
                c_cell=ws_mc.cell(r_idx,i+1); v=rd.get(col, '')
                if isinstance(v,(float,np.floating)): c_cell.value=round(v,2) if pd.notna(v) else None
                else: c_cell.value=v
                nf=NF_PRC if col=='Close' else (NF_INT if col=='Shares' else (NF_M1 if col in ('Market_Cap_M','Preferred_M','Equity_Value_M') else None))
                sc(c_cell,fo=fA,fi=pST if ev else pW,al=aR if nf else aL,bd=BD,nf=nf)
            r_idx+=1
        ws_mc.auto_filter.ref=f"A{hdr}:{get_column_letter(len(cols))}{r_idx-1}"; ws_mc.freeze_panes=f'A{hdr+1}'

    # [4] Price_History Sheet (Always at the end)
    ws_ph = wb.create_sheet('Price_History')
    ws_ph.sheet_properties.tabColor = '000000'
    # (Assuming market_rows or similar logic would populate this if requested, 
    # but for now ensuring empty sheet at least exists at the end as requested)

    wb.save(output)
    output.seek(0)
    return output

# 로그 출력 (Manage app > Logs 터미널에서 확인 가능)
print(f"👉 [접속 알림] {now_str} / 사용자: {user_email}")

st.title("📊 GPCM Calculator with yfinance")
st.write("yfinance 라이브러리를 통해 기준일 시점 선정된 Peer들의 재무제표, 주가, 시가총액 등을 크롤링하여 기준일 시점 Peer Group GPCM Multiple을 자동계산하는 어플리케이션입니다(Made by SGJ, 2026-02-10)")

# [Notes Section]
st.markdown("---")
with st.expander("📝 산출 방법론 (Methodology Notes)", expanded=False):
    st.markdown("""
**[ 가치 구성 ]**
- `Equity Value = Market Cap + 우선주 장부금액` — 우선주를 자기자본에 가산하는 실무 관행 적용
- `Market Cap = Ordinary Shares Number × Close Price` (보통주 기준, 배당 미조정 실제 주가)
- `EV = Equity Value + IBD − Cash + NCI − NOA`
- `Cash` = 현금및현금성자산 + 단기금융상품
- **NOA(Option)** — 투자부동산·관계기업투자 등은 **기본적으로 EV에서 차감하지 않습니다.**
  엑셀 `BS_Full` 시트의 **EV Tag를 `NOA`로 바꾸면 즉시 EV에서 차감**되며, GPCM·Multiples_Trend 시트에 함께 반영됩니다.

**[ 손익 ]**
- `EBIT = Operating Income`
- `EBITDA = Operating Income + D&A`, `D&A = yfinance EBITDA − yfinance EBIT`
  - ⚠️ 야후가 EBITDA 또는 EBIT을 제공하지 않으면 **D&A = 0** 으로 처리되어 **EBITDA = 영업이익**이 됩니다.
    해당 회사는 아래 경고와 엑셀 `Data_Quality` 시트에 표시되며, GPCM 시트의 `D&A 산출근거` 열이 붉게 칠해집니다.
- `PL Source`: **LTM(최근 4개 분기 합산) 우선**. 단 아래 검증을 모두 통과한 경우에만 적용
  - 최근 4개 공시기간의 결산일 간격이 **240~320일** (반기공시 기업의 2년치 합산 방지)
  - 핵심 계정(Total Revenue, Operating Income)에 결측 분기 없음
  - 미통과 시 연간 데이터를 사용하고 사유를 `Data_Quality` 시트에 기재

**[ 멀티플 ]**
- `EV/Sales`, `EV/EBITDA`, `EV/EBIT` — 분자는 EV
- `PER = Equity Value ÷ 지배주주순이익`, `PBR = Equity Value ÷ 자본총계`, `PSR = Equity Value ÷ 매출`

**[ 기간 선택 원칙 ]**
- 각 라벨(Y, Y-1, Y-2 …)은 **해당 결산일 시점 이전의 최신 재무제표만** 사용합니다 (Look-ahead 방지).
- 과거 연도의 재무상태표도 해당 연도 기준으로 각각 조회합니다.
""")

with st.expander("📊 베타 및 법인세율 산출 기준", expanded=False):
    st.markdown("""
**[ 베타 ]**
- 데이터 출처: 개별종목·시장지수 **모두 Yahoo Finance (yfinance)**
- 수익률 기준: 종목·지수 모두 **배당·분할 반영 총수익 기준(auto_adjust=True)** 으로 통일
  (시가총액용 주가는 실제 거래가격을 별도로 사용)
- `5Y Monthly Beta`: 5년 월말 종가 → 월간 수익률 → 시장지수 대비 선형회귀
- `2Y Weekly Beta`: 2년 주말 종가 → 주간 수익률 → 시장지수 대비 선형회귀
- `Adjusted Beta = 2/3 × Raw Beta + 1/3 × 1.0` (Bloomberg 방법론)
- **R²(설명력), 표준오차(SE), 관측치수(n)** 를 함께 산출합니다.
  - R² < 0.1 이면 베타 신뢰도가 낮으므로 산업 베타 등 대체 검토를 권장합니다.
  - `Raw Beta ± 2 × SE` 가 약 95% 신뢰구간입니다.
- 값 검증: NaN, inf, 극단값(±10 초과) 필터링 → 공란 처리

**[ 법인세율 — 한국 (지방소득세 10% 포함 한계세율) ]**

| 과세표준 | FY2023~2025 | FY2026~ |
|---|---|---|
| 2억 이하 | 9.9% | **11.0%** |
| 2억 ~ 200억 | 20.9% | **22.0%** |
| 200억 ~ 3,000억 | 23.1% | **24.2%** |
| 3,000억 초과 | 26.4% | **27.5%** |

- 2025년 세법개정으로 **2026.1.1 이후 개시 사업연도부터 각 구간 1%p 인상**되었습니다.
- 프로그램은 **각 피어의 결산연도에 맞는 세율표**를 자동 적용합니다.
- 결손(세전이익 ≤ 0) 기업은 한계세율 개념이 성립하지 않으므로 **2구간 세율**을 적용하고 `세율 근거` 열에 표시합니다.

**[ 법인세율 — 해외 ]**
- KPMG Corporate Tax Rates Table (Combined Rate) 기준
- 세율표 미보유 국가는 25%를 적용하고 `세율 근거` 열에 명시합니다.

**[ 자본구조 ]**
- `D/E Ratio = IBD ÷ (Equity Value + NCI)`
- `Debt Ratio (D/V) = IBD ÷ (Equity Value + IBD + NCI)`
- `Unlevered Beta = Levered(Adj) Beta ÷ (1 + (1 − Tax) × D/E)` [Hamada]
""")

with st.expander("💰 Target WACC 및 신규 시트 안내", expanded=False):
    st.markdown("""
**[ WACC ]**
- `Ke = Rf + MRP × Relevered Beta + Size Premium`
- `Relevered Beta = 피어 평균 Unlevered Beta × (1 + (1 − Tax) × Target D/E)`
- `Kd (Aftertax) = Kd (Pretax) × (1 − Target Tax Rate)`
- `Target D/V` = **피어 평균 부채비율** (GPCM 시트 Mean 행을 참조하는 엑셀 수식)
- `WACC = (E/V) × Ke + (D/V) × Kd (Aftertax)`
- WACC 시트 하단에 **민감도 분석표 2종**(Unlevered β × MRP, D/V × Kd)이 포함됩니다.

**[ 엑셀 신규 시트 ]**
- `Implied_Valuation` — Target 재무수치를 입력하면 피어 멀티플(Min/Q1/Median/Q3/Max)을 적용해
  **지분가치 밴드**를 산출합니다. 경영권 프리미엄·DLOM 조정란과 주당가치까지 포함됩니다.
- `Data_Quality` — LTM 검증 실패, D&A 역산 불가, 베타 신뢰도 등 **확인이 필요한 항목**을 모아둡니다.
""")

st.markdown("---")

# [Sidebar]
with st.sidebar:
    st.header("Settings")
    
    app_mode = st.selectbox("Analysis Mode", ["GPCM Valuation (Multi-Period)", "Historical FS Summary (과거재무정보 요약)"])
    
    st.subheader("1. 분석 기간 설정")
    lookback_years = st.slider("분석 년수 (N개년)", min_value=1, max_value=10, value=3 if app_mode == "Historical FS Summary (과거재무정보 요약)" else 1)
    
    col1, col2 = st.columns(2)
    with col1:
        base_year = st.number_input("Base Year (Y)", min_value=2010, max_value=2030, value=2024)
    with col2:
        base_quarter = st.selectbox("Base Quarter (Q)", ["1Q", "2Q", "3Q", "4Q"], index=3)
            
    if app_mode == "GPCM Valuation (Multi-Period)":
        cycle_choice = "Quarterly" # Always for LTM logic
    else:
        cycle_choice = "Annual" # Force Annual for summary mode as per user request
            
    # Annual: User might want actual fiscal date, but for selection we use Y-12-31 as proxy
    q_map = {"1Q": "-03-31", "2Q": "-06-30", "3Q": "-09-30", "4Q": "-12-31"}
    
    # Base period date proxy
    if cycle_choice == "Quarterly":
        base_period_str = f"{base_year}{q_map[base_quarter]}"
    else:
        base_period_str = f"{base_year}-12-31"

    target_periods = []
    # Base period is Label 'Y'
    target_periods.append(base_period_str)
    
    # 1) If base is quarter, we must include the latest annual (base_year-12-31 or base_year-1-12-31)
    # The logic: always add N annual periods starting from the current year OR previous year if current is not yet annual-closed.
    # User said: if base is 2025-09-30, 2024-12-31 must be in history.
    
    current_annual_proxy = f"{base_year}-12-31"
    if current_annual_proxy != base_period_str:
        # If base is a mid-year quarter, the latest "finished" annual is base_year - 1
        for i in range(1, lookback_years + 1):
            hist_year = base_year - i
            target_periods.append(f"{hist_year}-12-31")
    else:
        # If base is 12-31, then hist starts from base_year - 1
        for i in range(1, lookback_years):
            hist_year = base_year - i
            target_periods.append(f"{hist_year}-12-31")
    
    target_periods = sorted(list(set(target_periods))) 
    
    if app_mode == "Historical FS Summary (과거재무정보 요약)":
        st.info(f"분석 기준: Y={base_year}부터 과거 {lookback_years}개년 실제 연간재무제표")
    else:
        st.info(f"기준일: {base_period_str} | 과거 추이: {lookback_years-1}개년 연간 데이터")

    # 2. Ticker Input
    st.subheader("2. Ticker List")
    st.markdown("야후파이낸스 웹사이트 기준 Ticker를 한줄씩 입력하세요")
    default_tickers = """3116.T
7273.T
4246.T
7282.T
MG.TO
EZM.F
200880.KS
038110.KQ
012330.KS
PYT.VI"""
    txt_input = st.text_area("Tickers", value=default_tickers, height=250)

    # 3. WACC 파라미터 설정 (Target 기업용)
    st.subheader("3. Target WACC Parameters")

    st.markdown("**무위험이자율 (Rf)**")
    rf_input = st.number_input("Rf - 무위험이자율 (%)", min_value=0.0, max_value=10.0, value=3.3, step=0.1, format="%.2f",
                                help="한국 10년 국채수익률 (한국은행 경제통계시스템 참고)") / 100

    st.markdown("**자기자본비용 (Ke) 파라미터**")
    mrp_input = st.slider("MRP (시장위험프리미엄)", min_value=7.0, max_value=9.0, value=8.0, step=0.1, format="%.1f%%",
                         help="한국공인회계사회 권장: 7~9%") / 100

    st.markdown("**Size Premium (한국공인회계사회 기준, 2023)**")

    # Size Premium 표 보여주기
    with st.expander("📊 시가총액별 Size Premium 참고표"):
        st.markdown("**3분위수 기준**")
        st.markdown("""
        | 구분 | 시가총액 범위 (억원) | Size Premium |
        |------|---------------------|--------------|
        | **Micro** | < 2,000 | **4.02%** |
        | **Low** | 2,000 ~ 20,000 | 1.37% |
        | **Mid** | > 20,000 | -0.36% |
        """)

        st.markdown("**5분위수 기준**")
        st.markdown("""
        | 구분 | 시가총액 범위 (억원) | Size Premium |
        |------|---------------------|--------------|
        | **5분위 (최소)** | < 2,000 | **4.66%** |
        | **4분위** | 2,000 ~ 5,000 | 3.02% |
        | **3분위** | 5,000 ~ 20,000 | 1.21% |
        | **2분위** | 20,000 ~ 50,000 | 0.06% |
        | **1분위 (최대)** | > 50,000 | -0.58% |
        """)

        st.info("💡 Target 기업의 시가총액에 맞는 Size Premium을 선택하세요.")

    size_premium_options = {
        "3분위 - Micro (4.02%): < 2,000억": 0.0402,
        "3분위 - Low (1.37%): 2,000~20,000억": 0.0137,
        "3분위 - Mid (-0.36%): > 20,000억": -0.0036,
        "5분위 - 5분위/최소 (4.66%): < 2,000억": 0.0466,
        "5분위 - 4분위 (3.02%): 2,000~5,000억": 0.0302,
        "5분위 - 3분위 (1.21%): 5,000~20,000억": 0.0121,
        "5분위 - 2분위 (0.06%): 20,000~50,000억": 0.0006,
        "5분위 - 1분위/최대 (-0.58%): > 50,000억": -0.0058,
        "Size Premium 없음 (0%)": 0.0
    }
    size_premium_choice = st.selectbox("기업 규모 선택", list(size_premium_options.keys()), index=0,
                                       help="Target 기업의 시가총액에 맞는 Size Premium 선택")
    size_premium_input = size_premium_options[size_premium_choice]

    st.markdown("**Beta 계산 기준 선택**")
    beta_type_options = {
        "5년 월간 베타 (5Y Monthly)": "5Y",
        "2년 주간 베타 (2Y Weekly)": "2Y"
    }
    beta_type_choice = st.selectbox("WACC 계산에 사용할 Beta", list(beta_type_options.keys()), index=0,
                                    help="Target WACC 계산 시 사용할 베타 유형을 선택하세요. 두 베타 모두 엑셀에 표시됩니다.")
    beta_type_input = beta_type_options[beta_type_choice]

    st.markdown("**타인자본비용 (Kd) 파라미터**")
    kd_pretax_input = st.number_input("Kd (Pretax) - 세전 이자율 (%)", min_value=0.0, max_value=15.0, value=3.5, step=0.1, format="%.1f") / 100

    st.markdown("**Target 기업 법인세율**")
    target_tax_rate_input = st.number_input("Target 법인세율 (%)", min_value=0.0, max_value=50.0, value=26.4, step=0.1, format="%.1f",
                                            help="KPMG 2025 Combined Rate 기준 | 한국: 26.4% | 미국: 26.0% | 일본: 29.7% | 프랑스: 36.1% | 독일: 30.1%") / 100

    st.info("💡 목표 부채비율은 피어들의 평균 자본구조로 자동 계산됩니다.")

    # 4. 이상치 처리
    st.subheader("4. 피어 통계 이상치 처리")
    outlier_options = {
        "사용 안 함 (전체 피어 사용)": "none",
        "IQR 방식 (1.5×IQR 밖 제외) — 권장": "iqr",
        "백분위 방식 (25~75% 밖 제외)": "pct",
    }
    outlier_choice = st.selectbox(
        "이상치 처리 방식", list(outlier_options.keys()), index=1,
        help="WACC 산출용 피어 평균(부채비율·Unlevered Beta)에 적용됩니다. "
             "제외된 피어는 엑셀 Data_Quality 시트에 기재됩니다.")
    outlier_method_input = outlier_options[outlier_choice]

    # 5. Target 재무수치 (Implied Valuation 시트 사전 입력용 — 선택)
    st.subheader("5. Target 재무수치 (선택)")
    st.caption("입력하면 엑셀 Implied_Valuation 시트에 미리 채워집니다. 비워두면 엑셀에서 직접 입력하시면 됩니다.")
    with st.expander("Target 재무수치 입력", expanded=False):
        st.caption("단위: 백만 (피어와 동일 통화 기준)")
        ti_revenue = st.number_input("Revenue (매출)", value=0.0, step=1000.0, format="%.0f")
        ti_ebitda = st.number_input("EBITDA", value=0.0, step=1000.0, format="%.0f")
        ti_ebit = st.number_input("EBIT (영업이익)", value=0.0, step=1000.0, format="%.0f")
        ti_ni = st.number_input("Net Income (지배주주순이익)", value=0.0, step=1000.0, format="%.0f")
        ti_equity = st.number_input("Equity (자본총계)", value=0.0, step=1000.0, format="%.0f")
        ti_ibd = st.number_input("IBD (이자부부채)", value=0.0, step=1000.0, format="%.0f")
        ti_cash = st.number_input("Cash (현금성자산)", value=0.0, step=1000.0, format="%.0f")
        ti_nci = st.number_input("NCI (비지배지분)", value=0.0, step=1000.0, format="%.0f")
        ti_shares = st.number_input("발행주식수 (주)", value=0.0, step=1000.0, format="%.0f")

    target_inputs_dict = {
        'Revenue': ti_revenue or None, 'EBITDA': ti_ebitda or None, 'EBIT': ti_ebit or None,
        'NI': ti_ni or None, 'Equity': ti_equity or None, 'IBD': ti_ibd or None,
        'Cash': ti_cash or None, 'NCI': ti_nci or None, 'Shares': ti_shares or None,
    }

    # 6. Run Button
    btn_run = st.button("Go, Go, Go 🚀", type="primary")

# [Main Execution]
if btn_run:
    target_tickers = [t.strip() for t in txt_input.split('\n') if t.strip()]

    if not target_tickers:
        st.warning("분석할 Ticker를 입력하세요.")
        st.stop()

    with st.spinner(f"데이터 추출 및 분석 중 ({app_mode})... 잠시만 기다려주세요..."):
        if app_mode == "GPCM Valuation (Multi-Period)":
            (all_period_data, raw_bs, raw_pl, mkt_rows, p_abs, p_rel, t_map,
             avg_debt_ratio, target_wacc_data, dq_rows) = get_gpcm_data(
                target_tickers,
                target_periods,
                mrp=mrp_input,
                kd_pretax=kd_pretax_input,
                size_premium=size_premium_input,
                target_tax_rate=target_tax_rate_input,
                user_rf_rate=rf_input,
                beta_type=beta_type_input,
                outlier_method=outlier_method_input,
            )

            base_gpcm_data = all_period_data.get('Y', {})

            if not base_gpcm_data:
                st.error("데이터를 불러오지 못했습니다. 티커를 확인하시거나 잠시 후 다시 시도해주세요 (Yahoo Rate Limit).")
                st.stop()

            # ---------------------------------------------------------
            # 0. 데이터 품질 경고 (숫자를 보기 전에 먼저 확인해야 할 항목)
            # ---------------------------------------------------------
            errors = [f for f in dq_rows if f.get('Level') == 'ERROR']
            warns = [f for f in dq_rows if f.get('Level') == 'WARN']
            infos = [f for f in dq_rows if f.get('Level') == 'INFO']

            da_issues = sorted({f['Ticker'] for f in dq_rows if f.get('Item') == 'EBITDA'})
            ltm_issues = sorted({f['Ticker'] for f in dq_rows
                                 if f.get('Item') == 'LTM' and f.get('Level') == 'WARN'})

            if da_issues:
                st.error(
                    "⚠️ **EBITDA 확인 필요** — 아래 회사는 야후 파이낸스에 EBITDA/EBIT 항목이 없어 "
                    "**감가상각비를 0으로 처리**했습니다. 따라서 **EBITDA = 영업이익**이며 "
                    "EV/EBITDA가 EV/EBIT과 같게 나옵니다. 사업보고서에서 D&A를 직접 확인하십시오.\n\n"
                    "· " + ", ".join(da_issues))
            if ltm_issues:
                st.warning(
                    "⚠️ **LTM 합산 미적용** — 아래 회사는 최근 4개 공시기간이 12개월에 해당하지 않거나(반기공시 등) "
                    "결측 분기가 있어 **연간 데이터로 대체**했습니다.\n\n· " + ", ".join(ltm_issues))
            if errors:
                st.error("❌ **수집 실패 항목이 있습니다** — 엑셀 Data_Quality 시트를 확인하십시오. "
                         f"(ERROR {len(errors)}건)")

            if dq_rows:
                with st.expander(f"🔍 데이터 품질 점검 결과  (ERROR {len(errors)} · WARN {len(warns)} · INFO {len(infos)})"):
                    st.dataframe(pd.DataFrame(dq_rows)[['Level', 'Ticker', 'Label', 'Item', 'Message']],
                                 use_container_width=True, hide_index=True)

            # ---------------------------------------------------------
            # 1. Summary Table
            #    엑셀 GPCM 시트와 동일한 정의를 사용한다.
            #    (EV = Equity Value + IBD − Cash + NCI − NOA, 기본적으로 NOA는 0)
            # ---------------------------------------------------------
            st.subheader(f"📋 GPCM Multiples Summary (Base: {base_period_str})")
            summary_list = []
            for t, g in base_gpcm_data.items():
                eq_val = g['Equity_Value_M']
                ev = eq_val + g['IBD'] - g['Cash'] + g['NCI'] - g.get('NOA', 0)
                summary_list.append({
                    'Ticker': t, 'Company': g['Company'],
                    'PL 기준일': g.get('PL_Date') or g['Base_Date'],
                    'PL Source': g['PL_Source'],
                    'EV/Sales': ev / g['Revenue'] if g['Revenue'] > 0 else None,
                    'EV/EBITDA': ev / g['EBITDA'] if g['EBITDA'] > 0 else None,
                    'EV/EBIT': ev / g['EBIT'] if g['EBIT'] > 0 else None,
                    'PER': eq_val / g['NI_Parent'] if g['NI_Parent'] > 0 else None,
                    'PBR': eq_val / g['Equity'] if g['Equity'] > 0 else None,
                    'PSR': eq_val / g['Revenue'] if g['Revenue'] > 0 else None,
                    'D/V': g['Debt_Ratio'],
                    'Tax': g['Tax_Rate'],
                    'β 5Y Adj': g['Beta_5Y_Monthly_Adj'],
                    'R² 5Y': g['Beta_5Y_R2'],
                    'Unlev β 5Y': g['Unlevered_Beta_5Y'],
                })
            df_sum = pd.DataFrame(summary_list)
            mult_cols_ui = ['EV/Sales', 'EV/EBITDA', 'EV/EBIT', 'PER', 'PBR', 'PSR']
            st.dataframe(df_sum.style.format(
                {**{c: '{:.1f}x' for c in mult_cols_ui},
                 'D/V': '{:.1%}', 'Tax': '{:.1%}',
                 'β 5Y Adj': '{:.2f}', 'R² 5Y': '{:.2f}', 'Unlev β 5Y': '{:.2f}'},
                na_rep='N/M'), use_container_width=True, hide_index=True)

            low_r2 = [r['Ticker'] for r in summary_list
                      if r['R² 5Y'] is not None and r['R² 5Y'] < 0.10]
            if low_r2:
                st.info("ℹ️ 아래 회사는 5Y 베타의 R²가 0.1 미만으로 설명력이 낮습니다. "
                        "베타 채택 시 유의하십시오: " + ", ".join(low_r2))

            # ---------------------------------------------------------
            # 2. WACC 요약
            # ---------------------------------------------------------
            st.subheader("💰 Target WACC")
            c1, c2, c3 = st.columns(3)
            c1.metric("Target WACC", f"{target_wacc_data['Target_WACC']*100:.2f}%")
            c2.metric("Ke (자기자본비용)", f"{target_wacc_data['Target_Ke']*100:.2f}%")
            c3.metric("Kd (세후)", f"{target_wacc_data['Target_Kd_Aftertax']*100:.2f}%")
            c4, c5, c6 = st.columns(3)
            c4.metric("피어 평균 D/V", f"{target_wacc_data['Avg_Debt_Ratio']*100:.1f}%",
                      help=target_wacc_data.get('Debt_Ratio_Basis', ''))
            c5.metric(f"평균 Unlevered β ({target_wacc_data.get('Beta_Label','')})",
                      f"{target_wacc_data['Avg_Unlevered_Beta']:.4f}",
                      help=target_wacc_data.get('Beta_Basis', ''))
            c6.metric("Relevered β", f"{target_wacc_data['Target_Relevered_Beta']:.4f}")
            st.caption(f"목표 자본구조 근거: {target_wacc_data.get('Debt_Ratio_Basis','')} · "
                       f"베타 근거: {target_wacc_data.get('Beta_Basis','')} · "
                       f"이상치 처리: {outlier_choice}")

            # ---------------------------------------------------------
            # 3. Multiples Statistics (이상치 처리 반영)
            # ---------------------------------------------------------
            st.subheader("📊 Multiples Statistics")
            stat_list = []
            for col in mult_cols_ui:
                st_d, kept, excl = trimmed_stats(df_sum[col].tolist(), outlier_method_input)
                stat_list.append({
                    'Metric': col, 'N': st_d['N'],
                    'Min': st_d['Min'],
                    'Q1': float(np.percentile(kept, 25)) if kept else None,
                    'Median': st_d['Median'],
                    'Q3': float(np.percentile(kept, 75)) if kept else None,
                    'Max': st_d['Max'], 'Mean': st_d['Mean'],
                    '제외': len(excl),
                })
            df_stat = pd.DataFrame(stat_list).set_index('Metric')
            st.dataframe(df_stat.style.format(
                {c: '{:.1f}x' for c in ['Min', 'Q1', 'Median', 'Q3', 'Max', 'Mean']},
                na_rep='N/M'), use_container_width=True)
            st.caption("대표값은 이상치에 강한 **Median** 사용을 권장합니다. "
                       "'제외' 열은 선택한 이상치 처리 방식으로 통계에서 빠진 피어 수입니다.")

            # ---------------------------------------------------------
            # 4. Excel Download
            # ---------------------------------------------------------
            excel_data = create_excel(
                all_period_data, raw_bs, raw_pl, mkt_rows, p_abs, p_rel,
                base_period_str, target_periods, t_map, target_wacc_data,
                beta_type=beta_type_input, data_quality_rows=dq_rows,
                target_inputs=target_inputs_dict)

            st.download_button(
                label="📥 GPCM Report Download (Excel)",
                data=excel_data,
                file_name=f"Global_GPCM_{base_period_str.replace('-','')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        elif app_mode == "Historical FS Summary (과거재무정보 요약)":
            all_period_data, raw_bs, raw_pl, mkt_rows, p_abs, p_rel, t_map, _, _, dq_rows = get_gpcm_data(
                target_tickers,
                target_periods,
                mrp=mrp_input,
                kd_pretax=kd_pretax_input,
                size_premium=size_premium_input,
                target_tax_rate=target_tax_rate_input,
                user_rf_rate=rf_input,
                beta_type="5Y",
                force_annual=True,
                include_recent=True
            )
            
            if dq_rows:
                errs = [f for f in dq_rows if f.get('Level') in ('ERROR', 'WARN')]
                if errs:
                    st.warning(f"⚠️ 확인이 필요한 항목이 {len(errs)}건 있습니다. 아래 표를 확인하십시오.")
                with st.expander(f"🔍 데이터 품질 점검 결과 ({len(dq_rows)}건)"):
                    st.dataframe(pd.DataFrame(dq_rows)[['Level', 'Ticker', 'Label', 'Item', 'Message']],
                                 use_container_width=True, hide_index=True)

            excel_data = export_historical_excel_global_v2(all_period_data, raw_bs, raw_pl, mkt_rows, target_periods, t_map)
            
            if excel_data:
                st.success("✅ 기간별(년/분기별) 재무정보 및 요약 엑셀 생성이 완료되었습니다.")
                st.download_button(
                    label="📥 Historical FS Summary Download (Excel)",
                    data=excel_data,
                    file_name=f"Historical_FS_Summary_{target_tickers[0]}_etc.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("추출된 재무/주가 데이터가 없습니다.")
