import streamlit as st
import pandas as pd
import OpenDartReader
import FinanceDataReader as fdr
from datetime import datetime, timedelta
from pathlib import Path
import os
import warnings
import numpy as np
import re
import requests
import time
import io # 엑셀 메모리 저장을 위해 추가
import yfinance as yf # 지수 정보 조회를 위해 추가
from bs4 import BeautifulSoup # 주식수 크롤링을 위해 추가

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# 최신 수정: 2026-02-17 15:00 KST
# 주요 변경사항:
# - Beta 계산 기능 추가 (5Y Monthly, 2Y Weekly) - FinanceDataReader 사용
# - Beta_Calculation 시트 추가
# - WACC_Calculation 시트 완전 구현 (GPCM.py와 동일)
# - GPCM 시트에 Beta & Risk Analysis 컬럼 추가 (총 35개 컬럼)
# - D/E Ratio 컬럼 추가 (컬럼 32): IBD/(시총+NCI)
# - Debt Ratio 컬럼 이동 (컬럼 33): IBD/(시총+IBD+NCI)
# - Unlevered Beta 수식 수정: D/E 사용 (하마다 모형)
# - 한국 법인세 한계세율 적용 (지방세 포함)
# - Streamlit 사용자 입력 추가: Rf, MRP, Size Premium, Beta Type, Kd, Target Tax Rate
# - 모든 데이터 소스: FinanceDataReader (yfinance 미사용)

# Streamlit 페이지 설정 (가장 먼저 와야 함)
st.set_page_config(page_title="GPCM Calculator", layout="wide")

# 이 앱은 "설치"가 없다 — run_kr.bat 이 이 폴더의 파일을 그대로 실행하므로
# 내려받은 폴더가 곧 버전이다. 옛 폴더에서 돌리고 있는지 화면에서 보이게 한다.
# MCP 서버(gpcm_mcp.py)와 같은 출처(manifest.json)를 읽어 두 경로의 판을 대조할 수 있다.
APP_DIR = Path(__file__).resolve().parent


def app_version():
    try:
        import json
        return json.loads((APP_DIR / "manifest.json").read_text(encoding="utf-8"))["version"]
    except Exception:
        return "unknown"


warnings.filterwarnings('ignore')

# ==========================================
# 0. 전역 설정 및 상수
# ==========================================
RCODE_MAP = {'1Q': '11013', '2Q': '11012', '3Q': '11014', '4Q': '11011'}
QUARTER_INFO = {'1Q': '03-31', '2Q': '06-30', '3Q': '09-30', '4Q': '12-31'}

BETA_5Y_DAYS = 365 * 5 + 20
BETA_2Y_DAYS = 365 * 2 + 20
# 창(window)을 다 채웠을 때의 관측치 — 미달이면 경고한다(제외는 하지 않는다)
FULL_MONTHLY_OBS = 60   # 5년 × 12개월
FULL_WEEKLY_OBS = 104   # 2년 × 52주
BETA_SANITY_LIMIT = 3.0  # 통상 범위. 넘으면 경고만 하고 값은 그대로 둔다
MIN_MONTHLY_PTS = 12
MIN_WEEKLY_PTS = 50

# Data_Quality 시트 심각도
SEV_ERROR = 'ERROR'   # 그 값을 쓰면 안 된다
SEV_WARN = 'WARN'     # 값은 나왔지만 왜곡됐을 수 있다
SEV_INFO = 'INFO'     # 알아두면 되는 것


class QualityLog:
    """자동 수집이 실패하거나 값을 채우지 못한 지점을 모은다.

    DART 조회는 회사·기간별로 조용히 실패한다. 지금까지는 실패한 자리에 0이 남고
    나머지가 계속 돌아, 엑셀만 봐서는 그 0이 '정말 0'인지 '못 가져온 것'인지
    구분할 수 없었다. 특히 LTM은 (당기누계 + 전기연간 - 전년동기)라서 셋 중
    하나만 빠져도 숫자가 그럴듯하게 틀린다.
    """

    def __init__(self):
        self.rows = []

    def add(self, level, ticker, company, item, message):
        self.rows.append({
            'Level': level, 'Ticker': ticker, 'Company': company,
            'Item': item, 'Message': message,
        })

    def has(self, level):
        return any(r['Level'] == level for r in self.rows)

# ==========================================
# 1. Helper Functions
# ==========================================

def _to_price_series(df, col='Close'):
    """yf.download / fdr.DataReader 결과에서 1-D 가격 Series 추출."""
    if isinstance(df, pd.Series):
        return df
    if col in df.columns:
        s = df[col]
    else:
        s = df.iloc[:, 0]
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0]
    return s

def _slice_from(df, start_date_str):
    """이미 받아둔 넓은 구간 시계열에서 start_date 이후만 잘라낸다."""
    if df is None or len(df) == 0:
        return df
    try:
        idx = df.index
        if not isinstance(idx, pd.DatetimeIndex):
            idx = pd.to_datetime(idx)
        start = pd.to_datetime(start_date_str)
        if getattr(idx, 'tz', None) is not None:
            start = start.tz_localize(idx.tz)
        return df[idx >= start]
    except Exception:
        return df


def _get_market_index_data(market_idx, start, end, cache):
    """시장지수는 전 종목 공통이므로 (지수, 시작, 종료) 기준으로 1회만 조회한다."""
    key = (market_idx, start, end)
    if key in cache:
        return cache[key]
    if market_idx.startswith('^'):
        data = yf.download(market_idx, start=start, end=end, progress=False)
    else:
        data = fdr.DataReader(market_idx, start, end)
    cache[key] = data
    return data


BETA_BENCHMARK = '^KS11'  # KOSPI. fdr 의 KS11 이 자주 실패해 yfinance 심볼을 쓴다


def get_market_index(ticker=None):
    """베타 기준지수를 돌려준다. 종목과 무관하게 **KOSPI 단일 기준**이다.

    코스닥 종목도 KOSPI 대비로 잰다. 시장별로 다른 지수를 쓰지 않는 이유:

    1) 이 모델은 피어들의 무차입베타를 평균해 하나의 WACC 을 만든다. 평균이
       의미를 가지려면 모든 β 가 같은 지수 기준이어야 한다 — 코스닥 기준 β 와
       코스피 기준 β 를 섞어 평균하는 것은 단위가 다른 값을 더하는 것이다.
    2) Ke = rf + β × MRP 에서 β 와 MRP 의 기준이 같아야 한다. 여기서 쓰는 MRP 는
       시장 전체(KOSPI) 기준 추정치다. 코스닥 지수 기준 β 에 이 MRP 를 곱하면
       기준이 어긋난다. 코스닥 기준으로 재려면 MRP 도 코스닥 기준으로 다시
       추정해야 한다.

    코스닥 소형주는 이 기준에서 R² 가 낮게 나오는데, 그건 고유위험이 크다는
    뜻이지 기준을 잘못 골랐다는 뜻이 아니다 (고유위험은 분산 가능해 가격에
    반영되지 않는다). 낮은 R² 는 Beta 시트와 gpcm_review 에 드러난다.

    ticker 인자는 호출부 호환을 위해 남겨 두었고 쓰지 않는다.
    Returns: (exchange_name, index_symbol)
    """
    return 'KRX', BETA_BENCHMARK

# 한국 법인세 한계세율표 (사업연도별, 지방소득세 10% 포함)
# 각 구간: (과세표준 상한(억원), 한계세율)  — 상한 None = 초과 구간
# · FY2018~2022 : 국세 10 / 20 / 22 / 25%
# · FY2023~2025 : 국세  9 / 19 / 21 / 24%  (2022년 세법개정, 1%p 인하)
# · FY2026~     : 국세 10 / 20 / 22 / 25%  (2025년 세법개정, 1%p 인상 환원)
KR_TAX_BRACKETS_PRE2023 = [(2, 0.110), (200, 0.220), (3000, 0.242), (None, 0.275)]
KR_TAX_BRACKETS_2023 = [(2, 0.099), (200, 0.209), (3000, 0.231), (None, 0.264)]
KR_TAX_BRACKETS_2026 = [(2, 0.110), (200, 0.220), (3000, 0.242), (None, 0.275)]



# ==========================================
# 시장금리 (WACC 입력값의 근거)
# ==========================================
# rf·Kd 는 판단이 아니라 기준일의 시장 관측치다. 값만 손으로 넣으면 나중에
# "이 3.3%는 어디서 왔나"에 답할 근거가 산출물에 남지 않는다. 값과 함께
# 출처·금리기준일을 돌려주고, 그 문장을 그대로 엑셀 Notes 에 싣는다.
# 앱(사이드바 조회 버튼)과 MCP(get_wacc_inputs)가 같은 코드를 쓴다.
# MRP·size premium 은 조회 대상이 아니다 — 시장에서 관측되지 않는 판단 항목이다.

ECOS_MARKET_RATES = "817Y002"  # 한국은행 ECOS 통계표: 시장금리(일별)
BOND_GRADES = ("AA-", "BBB-")
RATE_BONDS = {"rf": "국고채", "kd": "회사채"}
# 만기는 고정하지 않는다. 평가에서 rf 만기는 현금흐름 기간에 맞춰 5년·10년을 오간다.
DEFAULT_MATURITY = {"rf": "5년", "kd": "3년"}
# 무키 경로 — FDR 에 있는 국채 만기만. 회사채는 FDR 에 없어 ECOS 로 간다.
FDR_TREASURY = {"1년": "KR1YT=RR", "3년": "KR3YT=RR", "5년": "KR5YT=RR", "10년": "KR10YT=RR"}

MATURITY_RE = re.compile(r"^\s*(\d+)\s*(?:년|y|Y)?\s*$")
_QTR_END = {"1Q": (3, 31), "2Q": (6, 30), "3Q": (9, 30), "4Q": (12, 31)}


def maturity(text, kind):
    """"10" · "10Y" · "10년" 을 모두 "10년" 으로 읽는다.

    어떤 만기가 실제로 있는지는 조회할 때 소스가 답한다 — 여기서 허용 목록을
    박으면 소스가 늘 때 막힌다.
    """
    raw = (text or "").strip()
    if not raw:
        return DEFAULT_MATURITY[kind]
    m = MATURITY_RE.match(raw)
    if not m:
        raise ValueError(f'만기는 "5년"·"10Y"·"10" 처럼 적습니다: {text!r}')
    return f"{int(m.group(1))}년"


def as_of_date(as_of):
    """기준일을 읽는다. "2026-06-30" 도 되고 GPCM 기간 표기 "2026.2Q" 도 된다.

    금리는 기준일 시점을 써야 한다 — 평가기준일이 작년 말인데 오늘 금리를 넣으면
    조서의 다른 숫자와 시점이 어긋난다.
    """
    text = (as_of or "").strip()
    if not text:
        return datetime.now()
    if re.match(r"^\d{4}\.[1-4]Q$", text):
        year, qtr = text.split(".")
        month, day = _QTR_END[qtr]
        return datetime(int(year), month, day)
    try:
        return datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        raise ValueError(f'기준일은 "2026-06-30" 또는 "2026.2Q" 형식입니다: {as_of!r}')


def ecos_key_from_env():
    """환경변수의 ECOS 키. 확장 설치에서 입력칸이 비면 `${...}` 리터럴이 들어온다."""
    value = (os.environ.get("ECOS_API_KEY") or "").strip()
    if value.startswith("${") and value.endswith("}"):
        return ""
    return value


def _rate_from_fdr(kind, asof, term):
    """무키 경로. 못 구하면 None — 여기서 조용히 기본값을 만들지 않는다."""
    symbol = FDR_TREASURY.get(term) if kind == "rf" else None
    if symbol is None:
        return None
    try:
        df = fdr.DataReader(symbol, (asof - timedelta(days=30)).strftime("%Y-%m-%d"),
                            asof.strftime("%Y-%m-%d"))
    except Exception:
        return None
    if df is None or len(df) == 0 or "Close" not in getattr(df, "columns", []):
        return None
    row = df.dropna(subset=["Close"]).tail(1)
    if len(row) == 0:
        return None
    return {"value": round(float(row["Close"].iloc[0]), 3),
            "rateDate": str(row.index[0])[:10],
            "source": f"FinanceDataReader {symbol}"}


def _ecos_get(path):
    response = requests.get(f"https://ecos.bok.or.kr/api/{path}", timeout=20)
    response.raise_for_status()
    payload = response.json()
    if "RESULT" in payload:  # ECOS 는 오류도 200 으로 준다
        raise RuntimeError(payload["RESULT"].get("MESSAGE", "ECOS 오류"))
    return payload


def _ecos_item_code(words, grade, key):
    """통계표 항목 목록에서 이름으로 찾는다 — 코드를 박아두지 않는다.

    ECOS 가 항목코드를 바꾸면 조용히 엉뚱한 금리를 물어오게 되기 때문이다.
    """
    payload = _ecos_get(f"StatisticItemList/{key}/json/kr/1/500/{ECOS_MARKET_RATES}")
    rows = payload.get("StatisticItemList", {}).get("row", [])
    wanted = list(words) + ([grade] if grade else [])
    hits = [r for r in rows if all(w in r.get("ITEM_NAME", "") for w in wanted)]
    if not hits:
        available = [r.get("ITEM_NAME", "") for r in rows
                     if words and words[0] in r.get("ITEM_NAME", "")]
        raise RuntimeError(
            f"ECOS 통계표 {ECOS_MARKET_RATES} 에서 '{' '.join(wanted)}' 항목을 못 찾았습니다."
            + (f" 이 통계표에 있는 {words[0]} 항목: {', '.join(available)}" if available else ""))
    hit = min(hits, key=lambda r: len(r.get("ITEM_NAME", "")))
    return hit["ITEM_CODE"], hit["ITEM_NAME"]


def _rate_from_ecos(kind, asof, grade, term, key):
    if not key:
        return None
    code, name = _ecos_item_code((RATE_BONDS[kind], term), grade, key)
    start = (asof - timedelta(days=30)).strftime("%Y%m%d")
    payload = _ecos_get(
        f"StatisticSearch/{key}/json/kr/1/100/{ECOS_MARKET_RATES}/D/"
        f"{start}/{asof.strftime('%Y%m%d')}/{code}")
    rows = [r for r in payload.get("StatisticSearch", {}).get("row", [])
            if (r.get("DATA_VALUE") or "").strip()]
    if not rows:
        raise RuntimeError(f"ECOS 에 {start}~{asof:%Y%m%d} 구간 '{name}' 값이 없습니다.")
    last = rows[-1]
    return {"value": round(float(last["DATA_VALUE"]), 3),
            "rateDate": f'{last["TIME"][:4]}-{last["TIME"][4:6]}-{last["TIME"][6:8]}',
            "source": f"한국은행 ECOS {ECOS_MARKET_RATES} {name}"}


def fetch_market_rate(kind, asof, grade=None, term=None, key=""):
    """무키(FDR) → ECOS 순으로 시도하고, 실패한 경로를 그대로 남긴다.

    Returns: (결과 dict 또는 None, 시도한 경로 설명 목록)
    """
    term = term or DEFAULT_MATURITY[kind]
    tried = []
    for label, getter in (("FinanceDataReader", lambda: _rate_from_fdr(kind, asof, term)),
                          ("한국은행 ECOS", lambda: _rate_from_ecos(kind, asof, grade, term, key))):
        try:
            got = getter()
        except Exception as exc:
            tried.append(f"{label}: {exc}")
            continue
        if got:
            return got, tried
        tried.append(f"{label}: 값 없음"
                     + ("" if label != "한국은행 ECOS" or key else " (인증키 미설정)"))
    return None, tried

def get_korean_tax_brackets(fiscal_year):
    """사업연도에 적용되는 한국 법인세 한계세율표 반환."""
    try:
        fy = int(fiscal_year)
    except (TypeError, ValueError):
        fy = datetime.now().year
    if fy >= 2026:
        return KR_TAX_BRACKETS_2026
    if fy >= 2023:
        return KR_TAX_BRACKETS_2023
    return KR_TAX_BRACKETS_PRE2023


def korean_tax_rate_formula(cell_ref, fiscal_year):
    """엑셀 Tax Rate 수식을 사업연도 세율표에서 만든다.

    종전에는 FY2023~25 세율이 수식에 박혀 있어, 2025년 세법개정(FY2026~)이
    반영되지 않았다 — 같은 워크북 안에서 엑셀 세율과 파이썬 세율이 갈렸다.
    get_korean_marginal_tax_rate 와 같은 표를 써서 두 값이 항상 맞물리게 한다.
    """
    brackets = get_korean_tax_brackets(fiscal_year)
    finite = [(u, r) for u, r in brackets if u is not None]
    last_rate = brackets[-1][1]
    formula = f'{last_rate:g}'
    for upper, rate in reversed(finite):
        formula = f'IF({cell_ref}<={upper:g}, {rate:g}, {formula})'
    return '=' + formula


def get_korean_marginal_tax_rate(pretax_income_100m, fiscal_year=None):
    """
    한국 법인세 한계세율 산출 (지방소득세 포함, 사업연도별 세율표 적용)

    Parameters:
    - pretax_income_100m: 세전이익 (억원). 과세표준의 대용치로 사용.
    - fiscal_year: 해당 재무제표의 결산 연도. 미지정 시 현재 연도 기준.

    Note: 결손(음수) 기업은 한계세율 개념이 성립하지 않으므로
          '2억 초과 ~ 200억' 구간 세율을 적용한다.
    """
    brackets = get_korean_tax_brackets(fiscal_year)

    if pd.isna(pretax_income_100m) or pretax_income_100m <= 0:
        return brackets[1][1]

    for upper, rate in brackets:
        if upper is None or pretax_income_100m <= upper:
            return rate
    return brackets[-1][1]

def calculate_unlevered_beta(levered_beta, debt, equity, tax_rate):
    """
    하마다 모형으로 Unlevered Beta 계산
    Unlevered Beta = Levered Beta / (1 + (1 - Tax Rate) * (Debt / Equity))
    """
    if pd.isna(levered_beta) or levered_beta is None:
        return None
    if pd.isna(debt) or pd.isna(equity) or equity == 0:
        return levered_beta

    unlevered = levered_beta / (1 + (1 - tax_rate) * (debt / equity))
    return unlevered

def parse_period(p: str):
    parts = p.strip().split('.')
    return int(parts[0]), parts[1]

def get_base_date_str(year: int, qtr: str):
    return f"{year}-{QUARTER_INFO[qtr]}"

# 정기보고서 법정 제출기한 (결산일로부터 경과일수)
FILING_DEADLINE_DAYS = {'1Q': 45, '2Q': 45, '3Q': 45, '4Q': 90}

def is_period_filed(year: int, qtr: str, asof: datetime = None):
    """해당 분기 정기보고서의 제출기한이 지났는지 여부 (= 조회 가능한지)"""
    asof = asof or datetime.now()
    qtr_end = pd.to_datetime(f"{year}-{QUARTER_INFO[qtr]}")
    return (qtr_end + timedelta(days=FILING_DEADLINE_DAYS[qtr])) <= asof

def get_latest_filed_period(asof: datetime = None):
    """오늘 기준으로 실제 공시가 끝난 가장 최근 분기를 (연도, 분기)로 반환"""
    asof = asof or datetime.now()
    y, q_idx = asof.year, 3
    for _ in range(12):
        qtr = ['1Q', '2Q', '3Q', '4Q'][q_idx]
        if is_period_filed(y, qtr, asof):
            return y, qtr
        q_idx -= 1
        if q_idx < 0:
            y -= 1
            q_idx = 3
    return asof.year - 1, '4Q'

def get_ltm_required_periods(year: int, qtr: str):
    if qtr == '4Q':
        return [(year, '4Q', 'annual')]
    return [
        (year, qtr, 'current_cum'),
        (year - 1, '4Q', 'prior_annual'),
        (year - 1, qtr, 'prior_same_q'),
    ]

@st.cache_resource(ttl=3600)
def get_krx_listing():
    """KRX 상장종목 목록 조회 - 재시도 및 fallback 포함 (Streamlit Cloud 에러 대응)"""
    # 1차 시도: KRX 전체 (최대 3번)
    for attempt in range(3):
        try:
            df = fdr.StockListing('KRX')
            if df is not None and not df.empty:
                return df
        except Exception:
            if attempt < 2:
                time.sleep(1.0)

    # 2차 시도: KOSPI + KOSDAQ 개별 조회 후 병합
    frames = []
    for mkt in ['KOSPI', 'KOSDAQ']:
        for attempt in range(2):
            try:
                df_mkt = fdr.StockListing(mkt)
                if df_mkt is not None and not df_mkt.empty:
                    frames.append(df_mkt)
                    break
            except Exception:
                if attempt < 1:
                    time.sleep(1.0)
    if frames:
        return pd.concat(frames, ignore_index=True).drop_duplicates(subset=['Code'])

    # 최후 fallback: 빈 DataFrame 반환 (코드만으로 진행 가능하도록)
    return pd.DataFrame(columns=['Code', 'Name', 'Stocks'])

@st.cache_resource(ttl=3600)
def get_krx_industry_listing():
    """업종·주요제품이 붙은 상장사 목록 (KRX 상장회사목록).

    get_krx_listing()이 쓰는 'KRX'는 가격·시총 목록이라 업종이 없다. 업종으로
    Peer 후보를 추리려면 이쪽이 필요하다. 인증키는 필요 없다.

    반환: Code, Name, Market, Sector(업종), Industry(주요제품), SettleMonth(결산월)
    실패하면 빈 DataFrame — 호출부는 종목코드 직접 입력으로 되돌아간다.
    """
    for attempt in range(2):
        try:
            df = fdr.StockListing('KRX-DESC')
            if df is not None and not df.empty and 'Sector' in df.columns:
                return df
        except Exception:
            if attempt < 1:
                time.sleep(1.0)
    return pd.DataFrame(columns=['Code', 'Name', 'Market', 'Sector', 'Industry', 'SettleMonth'])


def peer_candidate_rows(df_ind, sector: str):
    """업종 하나에 속한 상장사를 종목코드 순으로 정리한다.

    결산월이 12월이 아니면 비교기간이 어긋나므로 골라내기 전에 보이게 표시한다.
    """
    if df_ind is None or df_ind.empty or not sector:
        return []
    sub = df_ind[df_ind['Sector'] == sector].sort_values('Code')
    rows = []
    for _, r in sub.iterrows():
        settle = str(r.get('SettleMonth') or '').strip()
        rows.append({
            'Code': str(r.get('Code') or '').zfill(6),
            'Name': str(r.get('Name') or ''),
            'Market': str(r.get('Market') or ''),
            'Product': str(r.get('Industry') or '').strip(),
            'SettleMonth': settle,
            'FiscalNot12': bool(settle) and '12' not in settle,
        })
    return rows


def resolve_company_info(dart_instance, ticker: str):
    df_krx = get_krx_listing()
    rows = df_krx[df_krx['Code'] == ticker]
    krx_name = rows.iloc[0]['Name'] if not rows.empty else None

    # DART 내장 corp_codes 로 직접 이름 검색 (KRX 실패 대비)
    if krx_name is None:
        try:
            dart_rows = dart_instance.corp_codes[dart_instance.corp_codes['stock_code'] == ticker]
            if not dart_rows.empty:
                krx_name = dart_rows.iloc[0]['corp_name']
        except Exception:
            pass

    corp_code = None
    try:
        corp_code = dart_instance.find_corp_code(ticker)
    except Exception:
        corp_code = None

    if not corp_code and krx_name:
        try:
            corp_code = dart_instance.find_corp_code(krx_name)
        except Exception:
            corp_code = None

    return corp_code, krx_name

def get_stock_price(ticker: str, date_str: str):
    try:
        td = pd.to_datetime(date_str)
        if td > datetime.now():
            return None, None
        df = fdr.DataReader(ticker, td - timedelta(days=10), td)
        if df is not None and not df.empty:
            return float(df.iloc[-1]['Close']), df.index[-1].strftime('%Y-%m-%d')
        return None, None
    except Exception:
        return None, None

def _to_int(x):
    try:
        if x is None:
            return None
        s = str(x).strip().replace(',', '')
        if s == '' or s.lower() == 'nan':
            return None
        return int(float(s))
    except Exception:
        return None

# --- DART 유통주식수 ---
DART_STOCKTOTQY_URL = "https://opendart.fss.or.kr/api/stockTotqySttus.json"

# 매 호출마다 새 TLS 연결을 맺지 않도록 세션 재사용
_DART_SESSION = requests.Session()

def fetch_dart_distb_shares(api_key, corp_code: str, bsns_year: int, reprt_code: str, cache=None):
    """같은 (회사, 연도, 보고서)를 기간별로 반복 조회하므로 결과를 캐시한다.
    '데이터 없음' 응답도 캐시해야 fallback 탐색의 재조회가 사라진다.
    다만 네트워크 오류(ERR)는 캐시하지 않아 일시적 실패가 고정되지 않도록 한다."""
    ck = ('shares', corp_code, int(bsns_year), str(reprt_code))
    if cache is not None and ck in cache:
        return cache[ck]

    shares, meta = _fetch_dart_distb_shares(api_key, corp_code, bsns_year, reprt_code)

    if cache is not None and meta.get('status') != 'ERR':
        cache[ck] = (shares, meta)
    return shares, meta


def _fetch_dart_distb_shares(api_key, corp_code: str, bsns_year: int, reprt_code: str):
    meta = {'shares': None, 'rcept_no': None, 'stlm_dt': None, 'se': None, 'status': None, 'message': None}
    try:
        params = {
            'crtfc_key': api_key,
            'corp_code': corp_code,
            'bsns_year': str(bsns_year),
            'reprt_code': str(reprt_code),
        }
        resp = _DART_SESSION.get(DART_STOCKTOTQY_URL, params=params, timeout=10)
        resp.raise_for_status()
        js = resp.json()

        meta['status'] = js.get('status')
        meta['message'] = js.get('message')

        if js.get('status') != '000':
            return None, meta

        df = pd.DataFrame(js.get('list', []))
        if df.empty:
            return None, meta

        if 'se' in df.columns:
            c1 = df[df['se'].astype(str).str.contains('보통', na=False)]
            c2 = df[df['se'].astype(str).str.contains('합계', na=False)]
            pick = c1 if not c1.empty else (c2 if not c2.empty else df)
        else:
            pick = df

        row = pick.iloc[0].to_dict()
        meta['rcept_no'] = row.get('rcept_no')
        meta['stlm_dt'] = row.get('stlm_dt')
        meta['se'] = row.get('se')

        shares = _to_int(row.get('distb_stock_co'))
        if shares is None:
            istc = _to_int(row.get('istc_totqy'))
            tes = _to_int(row.get('tesstk_co'))
            if istc is not None and tes is not None:
                shares = istc - tes

        meta['shares'] = shares
        return shares, meta

    except Exception as e:
        meta['status'] = meta['status'] or 'ERR'
        meta['message'] = str(e)
        return None, meta

def get_outstanding_shares(api_key, corp_code: str, ticker: str, bsns_year: int, reprt_code: str, df_krx: pd.DataFrame, cache=None):
    # 1. DART API 조회 (요청한 기준년도/분기)
    shares, meta = fetch_dart_distb_shares(api_key, corp_code, bsns_year, reprt_code, cache=cache)
    if shares is not None and shares > 0:
        return shares, f"DART({reprt_code})", meta

    # 2. 직전 보고서들에서 주식수 조회 시도 (요청 분기에 주식수 누락 시)
    # 시간순: 11013 (1Q), 11012 (반기), 11014 (3Q), 11011 (사업보고서)
    order = ['11013', '11012', '11014', '11011']
    try:
        current_idx = order.index(reprt_code)
    except Exception:
        current_idx = 3 # 기본 사업보고서 매핑

    cy = bsns_year
    ci = current_idx - 1
    
    # 최근 8개 분기(약 2년치)를 역순으로 훑어 가장 최근 공시된 주식총수를 찾음
    for _ in range(8):
        if ci < 0:
            cy -= 1
            ci = 3
        
        fb_code = order[ci]
        fb_shares, fb_meta = fetch_dart_distb_shares(api_key, corp_code, cy, fb_code, cache=cache)
        
        if fb_shares is not None and fb_shares > 0:
            # 과거 분기 정보를 찾았을 경우, 해당 출처(년도와 보고서 코드) 명시하여 반환
            return fb_shares, f"DART(Fallback:{cy}-{fb_code})", fb_meta
            
        ci -= 1

    # 3. KRX 캐시 조회 (작동 안 할 확률 높음)
    try:
        row = df_krx[df_krx['Code'] == ticker]
        if not row.empty:
            shares_krx = _to_int(row.iloc[0].get('Stocks'))
            if shares_krx is not None and shares_krx > 0:
                meta_f = dict(meta)
                meta_f['shares'] = shares_krx
                return shares_krx, 'KRX', meta_f
    except Exception:
        pass

    return None, 'N/A', meta

# --- BS Matching Logic ---
IBD_AID_ALWAYS = {
    'ifrs-full_CurrentBorrowingsAndCurrentPortionOfNoncurrentBorrowings',
    'ifrs-full_LongtermBorrowings',
    'ifrs-full_CurrentLeaseLiabilities',
    'ifrs-full_CurrentPortionOfLongtermBorrowings',
    'ifrs-full_ShorttermBorrowings',
    'ifrs-full_NoncurrentLeaseLiabilities',
    'dart_CurrentPortionOfBonds',
    'ifrs-full_BondsIssued',
    'ifrs-full_Borrowings',
}
IBD_AID_PATTERN = re.compile(r'(Borrowings|Bonds|LeaseLiabilit)', re.IGNORECASE)
MEZZ_KW_KR = ['전환사채', '교환사채', '신주인수권부사채', 'BW', 'CB', 'EB', '전환', '상환', '신주인수', '교환']
MEZZ_KW_EN = ['convertible', 'exchangeable', 'bond with warrant', 'bonds with warrants', 'warrant']
IBD_KW_NAME = ['차입금', '사채', '리스부채', 'Borrowings', 'Bond', 'Bonds', 'LeaseLiabilit', 'Lease Liability']
IBD_EXCLUDE = [
    '매입채무', '미지급', '충당', '선수', '예수', '보증금',
    '자산', '대여금', '미수', '매출채권', '미수금', '미수수익',
    '선급', '선급금', '선급비용', '예치금', '보증금',
    '리스채권', '대여', '대출금(자산)',
]

def _norm(s):
    s = "" if s is None else str(s)
    return re.sub(r"\s+", "", s).strip()

def match_bs_ev_component(account_nm, account_id):
    acct = "" if account_nm is None else str(account_nm).strip()
    aid = "" if account_id is None else str(account_id).strip()
    acct_n = _norm(acct)
    acct_u = acct_n.upper()
    acct_l = acct_n.lower()

    if aid in ['ifrs-full_CashAndCashEquivalents', 'ifrs-full_ShorttermDepositsNotClassifiedAsCashEquivalents']:
        return 'Cash', '현금및단기예금'
    if aid == 'ifrs-full_Equity':
        return 'Equity_Total', '자본총계'
    if aid == 'ifrs-full_EquityAttributableToOwnersOfParent':
        return 'Equity_P', '지배기업지분'
    # 계정명 표기가 회사마다 달라(비지배지분/비지배주주지분/소수주주지분) 표준계정코드를 우선 사용
    if aid == 'ifrs-full_NoncontrollingInterests':
        return 'NCI', '비지배지분'
    # 우선주 자본금: 시가총액(보통주)에 잡히지 않으므로 자기자본가치에 별도 가산
    # '자본금'을 요구해 부채로 분류된 상환우선주(상환전환우선주부채 등)의 중복계상을 방지
    if '우선주' in acct_n and '자본금' in acct_n and '부채' not in acct_n:
        return 'Preferred', acct
    if aid == 'dart_ElementsOfOtherStockholdersEquity':
        return None, None

    if '우선주' not in acct_n:
        mezz_hit = False
        for kw in MEZZ_KW_KR:
            if kw.replace(" ", "") in acct_n: mezz_hit = True; break
        if (not mezz_hit) and any(kw in acct_l for kw in MEZZ_KW_EN): mezz_hit = True
        if (not mezz_hit) and re.search(r'(\bCB\b|\bEB\b|\bBW\b)', acct_u): mezz_hit = True
        if mezz_hit: return 'IBD(Option)', acct

    if not any(ex.replace(" ", "") in acct_n for ex in IBD_EXCLUDE):
        if aid in IBD_AID_ALWAYS: return 'IBD', acct
        if aid and IBD_AID_PATTERN.search(aid): return 'IBD', acct

    if any(k.replace(" ", "") in acct_n for k in IBD_KW_NAME):
        if not any(ex.replace(" ", "") in acct_n for ex in IBD_EXCLUDE):
            return 'IBD', acct

    if (('비지배' in acct_n and '지분' in acct_n) or '소수주주지분' in acct_n) and ('귀속' not in acct):
        return 'NCI', '비지배지분'

    noa_keywords = ['관계기업', '지분법', '공동기업', '종속기업', '금융자산', '금융상품']
    noa_exclude = ['단기', '현금', '매출', '보증금', '미수', '대여금', '예치금', '부채', '충당', '손실', '리스채권']
    if any(kw in acct for kw in noa_keywords) and not any(ex in acct for ex in noa_exclude):
        if aid not in ['ifrs-full_CashAndCashEquivalents', 'ifrs-full_ShorttermDepositsNotClassifiedAsCashEquivalents']:
            return 'NOA(Option)', acct
    return None, None

# --- PL Logic ---
PL_REVENUE = {
    '매출액', '수익(매출액)', '수익(매출)', '영업수익',
    '수익', '매출', '총매출액', '총수익', '영업수익',
    '매출액합계', '수익합계', '총영업수익'
}
PL_EBIT    = {'영업이익', '영업이익(손실)', '영업손실', '영업손익'}
PL_NI      = {
    '당기순이익', '당기순이익(손실)', '당기순손실', '당기순손익',
    '분기순이익', '분기순이익(손실)', '분기순손실', '분기순손익',
    '반기순이익', '반기순이익(손실)', '반기순손실', '반기순손익',
    '연결당기순이익', '연결당기순이익(손실)', '연결당기순손실', '연결당기순손익',
    'ProfitLoss', 'ifrs-full_ProfitLoss'
}
PL_PRETAX_INCOME = {
    '법인세비용차감전순이익', '법인세비용차감전순이익(손실)', '법인세차감전순이익',
    '법인세비용차감전계속사업이익', '법인세비용차감전이익', '세전순이익',
    '법인세비용차감전순손실', '세전이익', '법인세차감전이익'
}

_norm_pl = _norm

# DART 는 계정명과 함께 IFRS 표준태그(account_id)를 준다. 이름은 회사마다
# "Ⅰ. 영업수익"·"1. 매출액"처럼 제각각이지만 태그는 같다.
# 총액 태그만 넣는다 — 품목별 수익(RevenueFromSaleOfGoods 등)을 잡으면 매출이
# 실제보다 작게 나오면서 조용히 틀린다. 그게 제일 위험한 실패다.
PL_TAGS = {
    'ifrs-full_Revenue': 'Revenue',
    'ifrs-full_RevenueFromContractsWithCustomers': 'Revenue',
    'dart_OperatingIncomeLoss': 'EBIT',
    'ifrs-full_ProfitLossFromOperatingActivities': 'EBIT',
    'ifrs-full_ProfitLoss': 'NI',
    'ifrs-full_ProfitLossBeforeTax': 'Pretax_Income',
}

# 앞머리 번호("Ⅰ.", "1.", "(1)", "가.")와 꼬리 괄호("(주7)")를 떼기 위한 패턴
_PL_PREFIX = re.compile(r'^[(\[]?[0-9ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩIVXivx가-힣]{1,3}[.)\]]')
_PL_SUFFIX = re.compile(r'[(\[][^()\[\]]*[)\]]$')


def _pl_name_forms(account_nm):
    """계정명 표기 흔들림을 흡수한 후보들. 원문 형태가 항상 첫 번째다."""
    base = _norm_pl(account_nm)
    forms = [base]
    stripped = base
    for _ in range(2):  # "Ⅰ.1.매출액" 처럼 두 겹인 경우까지만
        nxt = _PL_PREFIX.sub('', stripped)
        if nxt == stripped:
            break
        stripped = nxt
    if stripped != base:
        forms.append(stripped)
    for f in list(forms):
        tail = _PL_SUFFIX.sub('', f)
        if tail and tail != f:
            forms.append(tail)
    return forms


def match_pl_lenient(account_nm, aid=None):
    """2단계 매칭 — 1단계(match_pl_core_only)가 못 찾은 항목에만 쓴다.

    표준태그를 먼저 보고, 없으면 표기 흔들림을 흡수한 이름으로 본다.
    제외 규칙은 1단계와 같게 원문 기준으로 먼저 적용한다.
    """
    a = _norm_pl(account_nm)
    if '지배' in a or '포괄' in a:
        return None
    tagged = PL_TAGS.get(str(aid or '').strip())
    if tagged:
        return tagged
    for form in _pl_name_forms(account_nm):
        if form in PL_REVENUE: return 'Revenue'
        if form in PL_EBIT:    return 'EBIT'
        if form in PL_NI:      return 'NI'
        if form in PL_PRETAX_INCOME: return 'Pretax_Income'
    return None


def match_pl_core_only(account_nm, aid=None):
    if aid == 'ifrs-full_ProfitLoss': return 'NI'
    a = _norm_pl(account_nm)
    if '지배' in a: return None # Exclude subset (지배기업, 비지배기업)
    if '포괄' in a: return None # Exclude Comprehensive Income
    if a in PL_REVENUE: return 'Revenue'
    if a in PL_EBIT:    return 'EBIT'
    if a in PL_NI:      return 'NI'
    if a in PL_PRETAX_INCOME: return 'Pretax_Income'
    return None

def _parse_amount(x):
    v = pd.to_numeric(str(x).replace(',', ''), errors='coerce')
    if pd.isna(v) or v == 0: return None
    return float(v)

def pick_pl_value(row: pd.Series, qtr: str):
    if qtr == '4Q':
        for col in ['thstrm_amount', 'thstrm_add_amount']:
            v = _parse_amount(row.get(col, ''))
            if v is not None: return v
    else:
        for col in ['thstrm_add_amount', 'thstrm_amount']:
            v = _parse_amount(row.get(col, ''))
            if v is not None: return v
    return None

# --- DART PL Fetch Functions (Need dart instance) ---
def _safe_dart_call(fn, *args, max_retry=2, **kwargs):
    last_err = None
    for _ in range(max_retry + 1):
        try:
            df = fn(*args, **kwargs)
            return df, None
        except Exception as e:
            last_err = e
            time.sleep(0.4)
    return None, last_err

def safe_finstate(dart_instance, corp_code, year, reprt_code, max_retry=2):
    # OpenDartReader.finstate 는 fs_div 인자를 받지 않는다 (finstate_all 에만 존재)
    return _safe_dart_call(dart_instance.finstate, corp_code, year, max_retry=max_retry, reprt_code=reprt_code)

def safe_finstate_all(dart_instance, corp_code, year, reprt_code, fs_div=None, max_retry=2):
    kwargs = {'reprt_code': reprt_code}
    if fs_div is not None:
        kwargs['fs_div'] = fs_div
    return _safe_dart_call(dart_instance.finstate_all, corp_code, year, max_retry=max_retry, **kwargs)

def fetch_pl_df(dart_instance, corp_code, year, reprt_code):
    df, err = safe_finstate(dart_instance, corp_code, year, reprt_code)
    if df is not None and not df.empty: return df, 'finstate', None

    for fs in ['CFS', 'OFS']:
        df, err = safe_finstate_all(dart_instance, corp_code, year, reprt_code, fs_div=fs)
        if df is not None and not df.empty: return df, f'finstate_all|{fs}', None

    df, err = safe_finstate_all(dart_instance, corp_code, year, reprt_code, fs_div=None)
    if df is not None and not df.empty: return df, 'finstate_all|no_fs_div', None
    
    return None, 'N/A', 'NO_DATA'

def filter_income_statement(df: pd.DataFrame):
    if df is None or df.empty: return df
    if 'sj_div' in df.columns:
        df2 = df[df['sj_div'].astype(str) == 'IS'].copy()
        if not df2.empty: return df2
    if 'sj_nm' in df.columns:
        df2 = df[df['sj_nm'].astype(str).str.contains('손익|포괄손익', na=False)].copy()
        return df2
    return df

# ==========================================
# 4. 스타일 및 엑셀 유틸
# ==========================================
C_BL='00338D'; C_DB='1E2A5E'; C_LB='C3D7EE'; C_PB='E8EFF8'
C_DG='333333'; C_MG='666666'; C_LG='F5F5F5'; C_BG='B0B0B0'; C_W='FFFFFF'
C_GR='E2EFDA'; C_YL='FFF8E1'; C_NOA='FCE4EC'

S1=Side(style='thin',color=C_BG); BD=Border(left=S1,right=S1,top=S1,bottom=S1)
fT=Font(name='KoPub돋움체 Medium',bold=True,size=14,color=C_BL)
fS=Font(name='KoPub돋움체 Medium',size=9,color=C_MG,italic=True)
fH=Font(name='KoPub돋움체 Medium',bold=True,size=9,color=C_W)
fA=Font(name='KoPub돋움체 Medium',size=9,color=C_DG)
fHL=Font(name='KoPub돋움체 Medium',bold=True,size=9,color=C_DB)
fMUL=Font(name='KoPub돋움체 Medium',bold=True,size=10,color=C_BL)
fNOTE=Font(name='KoPub돋움체 Medium',size=8,color=C_MG,italic=True)
fSTAT=Font(name='KoPub돋움체 Medium',bold=True,size=9,color=C_DB)
fFRM=Font(name='KoPub돋움체 Medium',size=9,color='000000')  # 검은색 (수식/계산값)
fHARD=Font(name='KoPub돋움체 Medium',size=9,color='FF0000') # 빨간색 (하드코딩/외부데이터)
fASSUM=Font(name='KoPub돋움체 Medium',size=9,color='FFC000') # 노란색 (주요 가정사항)
fLINK=Font(name='KoPub돋움체 Medium',size=9,color='008000') # 초록색 (내부시트 링크)
fSEC = Font(name='KoPub돋움체 Medium', bold=True, size=10, color=C_W)

pH=PatternFill('solid',fgColor=C_BL); pW=PatternFill('solid',fgColor=C_W)
pST=PatternFill('solid',fgColor=C_LG); pSTAT=PatternFill('solid',fgColor=C_LB)
pSEC1 = PatternFill('solid', fgColor=C_DB); pSEC2 = PatternFill('solid', fgColor=C_BL)

# ==========================================
# 4.5. [신규] 다기간 재무제표 요약 로직 추가
# ==========================================
def fetch_historical_financials(api_key, target_code_list, periods_to_fetch, dart, status_container, progress_bar, df_krx):
    total = len(target_code_list) * len(periods_to_fetch)
    cnt = 0
    hist_summary = []
    hist_details = []

    for ticker in target_code_list:
        corp_code, _ = resolve_company_info(dart, ticker)
        if not corp_code:
            cnt += len(periods_to_fetch); progress_bar.progress(cnt/total)
            continue
        
        comp_name = dart.company(corp_code).get('corp_name', ticker)

        for p in periods_to_fetch:
            year = p['year']
            hist_qtr = p['qtr']
            plabel = p['label']
            
            if not hist_qtr: 
                req_periods = [(year, '4Q', 'annual')]
            else:
                req_periods = [(year, hist_qtr, 'current_cum')]
            
            # BS & EV Components
            ast, liab, eq = np.nan, np.nan, np.nan
            cash, ibd, noa, nci = 0.0, 0.0, 0.0, 0.0
            
            # PL & CF LTM Aggregator
            pl_agg = {'Revenue': 0.0, 'GrossProfit': 0.0, 'EBIT': 0.0, 'NI': 0.0, 'CFO': 0.0, 'CFI': 0.0, 'CFF': 0.0}
            valid_pl_flags = {'Revenue': False, 'GrossProfit': False, 'EBIT': False, 'NI': False, 'CFO': False, 'CFI': False, 'CFF': False}
            
            used_code_current = 'N/A'
            df_fs_current = None
            
            for req_year, req_qtr, role in req_periods:
                primary = RCODE_MAP.get(req_qtr, '11013')
                fallbacks = [c for c in ['11013', '11014', '11012', '11011'] if c != primary]
                target_qtrs = [primary] + fallbacks
                
                df_fs = None
                used_code = None
                for rcode in target_qtrs:
                    df_fs, _ = safe_finstate_all(dart, corp_code, req_year, rcode, fs_div='CFS')
                    if df_fs is None or df_fs.empty:
                        df_fs, _ = safe_finstate_all(dart, corp_code, req_year, rcode, fs_div='OFS')
                    if df_fs is not None and not df_fs.empty:
                        used_code = rcode
                        break
                
                if df_fs is None or df_fs.empty:
                    continue # Skip if data is missing
                    
                if role in ('current_cum', 'annual'):
                    used_code_current = used_code
                    df_fs_current = df_fs
                    
                temp_pl = {'Revenue': np.nan, 'GrossProfit': np.nan, 'EBIT': np.nan, 'NI': np.nan, 'CFO': np.nan, 'CFI': np.nan, 'CFF': np.nan}
                
                for row_idx, row in df_fs.iterrows():
                    sj = str(row.get('sj_nm', ''))
                    acc = str(row.get('account_nm', '')).strip()
                    aid = str(row.get('account_id', '')).strip()
                    _raw = _parse_amount(str(row.get('thstrm_amount', '')))
                    val_1m = (_raw / 1000000) if _raw is not None else np.nan
                    
                    if pd.isna(val_1m): continue
                    
                    m_key = ""
                    if '상태' in sj and role in ('current_cum', 'annual'):
                        if acc == '자산총계': m_key = 'Assets'
                        elif acc == '부채총계': m_key = 'Liabilities'
                        elif acc == '자본총계': m_key = 'Equity_Total'
                        ev_comp, _ = match_bs_ev_component(acc, aid)
                        if ev_comp:
                            m_key = ev_comp # 'Cash', 'Cash(Option)', 'IBD', 'IBD(Option)', 'NOA', 'NOA(Option)', 'NCI'
                            
                    elif '손익' in sj and role in ('current_cum', 'annual'):
                        n_acc = _norm_pl(acc)
                        if '지배' not in n_acc and '포괄' not in n_acc:
                            if n_acc in PL_REVENUE: m_key = 'Revenue'
                            elif '매출총이익' in acc: m_key = 'GrossProfit'
                            elif '영업이익' in acc: m_key = 'EBIT'
                            elif '당기순이익' in acc or '분기순이익' in acc or '반기순이익' in acc or aid == 'ifrs-full_ProfitLoss': m_key = 'NI'
                            
                    elif '현금' in sj and role in ('current_cum', 'annual'):
                        if '영업활동' in acc and '흐름' in acc: m_key = 'CFO'
                        elif '투자활동' in acc and '흐름' in acc: m_key = 'CFI'
                        elif '재무활동' in acc and '흐름' in acc: m_key = 'CFF'

                    # Store Raw Data for Details Sheet (Only for current period)
                    if role in ('current_cum', 'annual') and val_1m != 0:
                        hist_details.append({
                            'Company': comp_name, 'Ticker': ticker, 'Period': plabel, 'Report': used_code_current,
                            'M_Key': m_key, 'Type': sj, 'Account_ID': aid, 'Account_NM': acc, 
                            'Amount': val_1m, 'Row_Idx': row_idx
                        })
                    
                    if '상태' in sj and role in ('current_cum', 'annual'):
                        if acc == '자산총계': ast = val_1m
                        elif acc == '부채총계': liab = val_1m
                        elif acc == '자본총계': eq = val_1m
                        
                        ev_comp, _ = match_bs_ev_component(acc, aid)
                        if ev_comp:
                            if ev_comp == 'Cash': cash += val_1m
                            elif ev_comp == 'IBD': ibd += val_1m
                            elif ev_comp == 'NCI': nci += val_1m
                            elif ev_comp == 'NOA': noa += val_1m
                            
                    elif '손익' in sj:
                        n_acc = _norm_pl(acc)
                        _raw_pl = pick_pl_value(row, req_qtr)
                        val_pl = (_raw_pl / 1000000) if _raw_pl is not None else np.nan
                        if not pd.isna(val_pl) and '지배' not in n_acc and '포괄' not in n_acc:
                            if pd.isna(temp_pl['Revenue']) and n_acc in PL_REVENUE: temp_pl['Revenue'] = val_pl
                            if pd.isna(temp_pl['GrossProfit']) and '매출총이익' in acc: temp_pl['GrossProfit'] = val_pl
                            if pd.isna(temp_pl['EBIT']) and '영업이익' in acc: temp_pl['EBIT'] = val_pl
                            if pd.isna(temp_pl['NI']) and '당기순이익' in acc: temp_pl['NI'] = val_pl
                            
                    elif '현금' in sj:
                        if pd.isna(temp_pl['CFO']) and '영업활동' in acc and '흐름' in acc: temp_pl['CFO'] = val_1m
                        if pd.isna(temp_pl['CFI']) and '투자활동' in acc and '흐름' in acc: temp_pl['CFI'] = val_1m
                        if pd.isna(temp_pl['CFF']) and '재무활동' in acc and '흐름' in acc: temp_pl['CFF'] = val_1m
                
                # Apply to aggregator
                for k in temp_pl:
                    v = temp_pl[k]
                    if pd.notna(v):
                        pl_agg[k] += v
                        valid_pl_flags[k] = True

            if df_fs_current is None or df_fs_current.empty:
                hist_summary.append({
                    'Company': comp_name, 'Ticker': ticker, 'Period': plabel, 'Report': 'N/A',
                    'Revenue': np.nan, 'GrossProfit': np.nan, 'EBIT': np.nan, 'NI': np.nan,
                    'Assets': np.nan, 'Liabilities': np.nan, 'Equity': np.nan,
                    'CFO': np.nan, 'CFI': np.nan, 'CFF': np.nan,
                    'Cash': np.nan, 'IBD': np.nan, 'NOA': np.nan, 'NCI': np.nan,
                    'Shares': np.nan, 'Price': np.nan, 'MarketCap': np.nan
                })
                cnt += 1; progress_bar.progress(cnt/total)
                continue

            hist_summary.append({
                'Company': comp_name, 'Ticker': ticker, 'Period': plabel, 'Report': used_code_current,
                'Revenue': pl_agg['Revenue'] if valid_pl_flags['Revenue'] else np.nan, 
                'GrossProfit': pl_agg['GrossProfit'] if valid_pl_flags['GrossProfit'] else np.nan, 
                'EBIT': pl_agg['EBIT'] if valid_pl_flags['EBIT'] else np.nan, 
                'NI': pl_agg['NI'] if valid_pl_flags['NI'] else np.nan,
                'Assets': ast, 'Liabilities': liab, 'Equity_Total': eq,
                'CFO': pl_agg['CFO'] if valid_pl_flags['CFO'] else np.nan, 
                'CFI': pl_agg['CFI'] if valid_pl_flags['CFI'] else np.nan, 
                'CFF': pl_agg['CFF'] if valid_pl_flags['CFF'] else np.nan,
                'Cash': cash, 'IBD': ibd, 'NOA': noa, 'NCI': nci
            })
            
            status_container.update(label=f"다기간 재무데이터 수집 중... {comp_name} ({plabel})")
            cnt += 1; progress_bar.progress(cnt/total)
            
    return pd.DataFrame(hist_summary), pd.DataFrame(hist_details)

def calculate_historical_metrics(df_summ):
    if df_summ.empty: return df_summ
    
    for col in ['OPM', 'GPM', 'ROE', 'DebtRatio', 'NetDebt']:
        df_summ[col] = np.nan
        
    for i, row in df_summ.iterrows():
        rev = row.get('Revenue'); ebit = row.get('EBIT'); gp = row.get('GrossProfit'); ni = row.get('NI')
        eq = row.get('Equity_Total'); liab = row.get('Liabilities')
        cash = row.get('Cash', 0.0); ibd = row.get('IBD', 0.0)
        noa = row.get('NOA', 0.0); nci = row.get('NCI', 0.0)
        
        if rev and rev > 0:
            df_summ.at[i, 'OPM'] = ebit / rev if pd.notna(ebit) else np.nan
            df_summ.at[i, 'GPM'] = gp / rev if pd.notna(gp) else np.nan
        if eq and eq > 0:
            df_summ.at[i, 'ROE'] = ni / eq if pd.notna(ni) else np.nan
            if pd.notna(liab): df_summ.at[i, 'DebtRatio'] = liab / eq
                
        nd = (ibd if pd.notna(ibd) else 0) - (cash if pd.notna(cash) else 0)
        df_summ.at[i, 'NetDebt'] = nd

    return df_summ

def export_historical_excel(df_summ, df_details, periods_to_fetch):
    output = io.BytesIO()
    wb = Workbook()
    
    # ---------------------------------------------------------
    # 1. Summary 시트 생성 (Layout A안: 회사 세로, 연도/지표 가로)
    # ---------------------------------------------------------
    ws_summ = wb.active
    ws_summ.title = "Summary"
    
    ws_summ.merge_cells('A1:Z1')
    p_start = periods_to_fetch[0]['label'] if periods_to_fetch else ""
    p_end = periods_to_fetch[-1]['label'] if periods_to_fetch else ""
    ws_summ['A1'] = f"Historical Financial Summary ({p_start} ~ {p_end})"
    sc(ws_summ['A1'], fo=fT)
    
    if df_summ.empty:
        ws_summ['A3'] = "No data available."
    else:
        metrics = [
            ('Revenue', '매출액', NB), ('GrossProfit', '매출총이익', NB), 
            ('EBIT', '영업이익', NB), ('NI', '당기순이익', NB),
            ('Assets', '자산총계', NB), ('Liabilities', '부채총계', NB), ('Equity_Total', '자본총계', NB),
            ('Cash', 'Cash', NB), ('IBD', 'IBD', NB),
            ('NOA', 'NOA', NB), ('NCI', 'NCI', NB),
            ('NetDebt', '순부채(Net Debt)', NB),
            ('CFO', '영업활동현금흐름', NB), ('CFI', '투자활동현금흐름', NB), ('CFF', '재무활동현금흐름', NB),
            ('OPM', '영업이익률', '0.0%'), ('GPM', '매출총이익률', '0.0%'), 
            ('ROE', 'ROE', '0.0%'), ('DebtRatio', '부채비율', '0.0%')
        ]
        
        labels = [p['label'] for p in periods_to_fetch]
        
        # 헤더 그리기 (Row 3: 지표명, Row 4: 기간)
        ws_summ.cell(row=3, column=1, value="Company"); sc(ws_summ.cell(row=3, column=1), fo=fH, fi=pH, al=aC, bd=BD)
        ws_summ.cell(row=3, column=2, value="Ticker"); sc(ws_summ.cell(row=3, column=2), fo=fH, fi=pH, al=aC, bd=BD)
        ws_summ.merge_cells('A3:A4')
        ws_summ.merge_cells('B3:B4')
        
        # 메트릭별 컬럼 알파벳 매핑 저장용
        mc_map = {} # (m_key, plabel) -> 'C'
        plabel_col_idx = {plabel: 5 + i for i, plabel in enumerate(labels)} # Detail 시트의 데이터 컬럼 E부터 시작
        
        col_idx = 3
        for m_key, m_name, _ in metrics:
            start_col = col_idx
            end_col = col_idx + len(labels) - 1
            ws_summ.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
            ws_summ.cell(row=3, column=start_col, value=m_name)
            sc(ws_summ.cell(row=3, column=start_col), fo=fH, fi=pSEC1, al=aC, bd=BD)
            
            for plabel in labels:
                ws_summ.cell(row=4, column=col_idx, value=plabel)
                sc(ws_summ.cell(row=4, column=col_idx), fo=fH, fi=pSEC2, al=aC, bd=BD)
                
                mc_map[(m_key, plabel)] = get_column_letter(col_idx)
                col_idx += 1
                
        # 데이터 쓰기 (Row 5부터 ~ 회사별)
        r = 5
        companies = df_summ['Company'].unique()
        for comp in companies:
            df_comp = df_summ[df_summ['Company'] == comp]
            ticker = df_comp['Ticker'].iloc[0] if not df_comp.empty else ""
            comp_sht = comp[:31] # 엑셀 시트 참조용 이름
            
            ws_summ.cell(row=r, column=1, value=comp); sc(ws_summ.cell(row=r, column=1), fo=fA, bd=BD)
            ws_summ.cell(row=r, column=2, value=ticker); sc(ws_summ.cell(row=r, column=2), fo=fA, al=aC, bd=BD)
            
            # 수식 적용 그룹 (비율/멀티플)
            ratio_keys = ['OPM', 'GPM', 'ROE', 'DebtRatio']
            # Raw Data SUMIFS 그룹
            sumifs_keys = ['Revenue', 'GrossProfit', 'EBIT', 'NI', 'Assets', 'Liabilities', 'Equity_Total', 'Cash', 'IBD', 'NOA', 'NCI', 'CFO', 'CFI', 'CFF']
            
            c = 3
            for m_key, m_name, fmt in metrics:
                for plabel in labels:
                    v = ""
                    dtl_col = get_column_letter(plabel_col_idx[plabel]) # Detail 시트의 타겟 Period 열
                    
                    if m_key in sumifs_keys:
                        # 엑셀 SUMIFS 수식 주입 (매핑 키 A열, 금액 Dtl_Col열)
                        v = f"=SUMIFS('{comp_sht}'!{dtl_col}:{dtl_col}, '{comp_sht}'!$A:$A, \"{m_key}\")"
                    elif m_key == 'NetDebt':
                        v = f"={mc_map[('IBD', plabel)]}{r} - {mc_map[('Cash', plabel)]}{r}"
                    elif m_key in ratio_keys:
                        if m_key == 'OPM': v = f"=IFERROR({mc_map[('EBIT', plabel)]}{r}/{mc_map[('Revenue', plabel)]}{r}, \"\")"
                        elif m_key == 'GPM': v = f"=IFERROR({mc_map[('GrossProfit', plabel)]}{r}/{mc_map[('Revenue', plabel)]}{r}, \"\")"
                        elif m_key == 'ROE': v = f"=IFERROR({mc_map[('NI', plabel)]}{r}/{mc_map[('Equity_Total', plabel)]}{r}, \"\")"
                        elif m_key == 'DebtRatio': v = f"=IFERROR({mc_map[('Liabilities', plabel)]}{r}/{mc_map[('Equity_Total', plabel)]}{r}, \"\")"
                        
                    ws_summ.cell(row=r, column=c, value=v)
                    font_style = fA if m_key in sumifs_keys else fFRM
                    sc(ws_summ.cell(row=r, column=c), fo=font_style, nf=fmt, bd=BD)
                    
                    c += 1
            r += 1
        
        ws_summ.column_dimensions['A'].width = 18
        ws_summ.column_dimensions['B'].width = 10
        for i in range(3, c):
            ws_summ.column_dimensions[get_column_letter(i)].width = 14
        
        ws_summ.freeze_panes = "C5"

    # ---------------------------------------------------------
    # 2. 개별 회사 상세 시트 생성 (세로: 계정, 가로: 연도 피벗 형태)
    # ---------------------------------------------------------
    if not df_details.empty:
        companies = df_details['Company'].unique()
        for comp in companies:
            ws_dtl = wb.create_sheet(title=comp[:31]) # 시트명 제한 31자
            df_c = df_details[df_details['Company'] == comp].copy()
            
            # 헤더 2줄 그리미 (Simplified)
            ws_dtl.merge_cells('A1:H1'); ws_dtl['A1'] = f"{comp} - 상세 재무제표 (Report 기반)"; sc(ws_dtl['A1'], fo=fT)
            ws_dtl.merge_cells('A2:H2'); ws_dtl['A2'] = "DART finstate_all 원본 계정 정보 (최다 추출)"; sc(ws_dtl['A2'], fo=fS)
            
            if df_c.empty:
                ws_dtl['A4'] = "No detailed data available."
                continue
                
            pivot_df = df_c.pivot_table(
                index=['M_Key', 'Type', 'Account_ID', 'Account_NM'], 
                columns='Period', 
                values='Amount', 
                aggfunc='sum'
            ).reset_index()
            
            order_df = df_c.groupby(['M_Key', 'Type', 'Account_ID', 'Account_NM'])['Row_Idx'].min().reset_index()
            pivot_df = pd.merge(pivot_df, order_df, on=['M_Key', 'Type', 'Account_ID', 'Account_NM'], how='left')
            
            # Type 내림차순 정렬 유도 및 DART 한계 극복용 계층형 정렬 맵핑 (KPMG Style)
            sort_map = {'재무상태표': 1, '손익계산서': 2, '포괄손익계산서': 3, '현금흐름표': 4}
            def get_heuristic_rank(row):
                t = str(row['Type']).split()[0].replace('연결', '')
                acc = str(row['Account_NM'])
                idx = row['Row_Idx']
                t_rank = sort_map.get(t, 99) * 1000000
                
                if t in ['손익계산서', '포괄손익계산서']:
                    if '매출액' in acc or '영업수익' in acc: return t_rank + 10000
                    if '원가' in acc: return t_rank + 20000
                    if '매출총이익' in acc: return t_rank + 30000
                    if '판매비' in acc or '관리비' in acc: return t_rank + 40000
                    if '영업이익' in acc or '영업손실' in acc: return t_rank + 50000
                    if '법인세비용' in acc: return t_rank + 80000
                    if '당기순이익' in acc or '당기순손실' in acc: return t_rank + 90000
                    return t_rank + 60000 + idx
                elif t == '현금흐름표':
                    if acc == '영업활동현금흐름' or ('영업활동' in acc and '흐름' in acc): return t_rank + 10000
                    if acc == '투자활동현금흐름' or ('투자활동' in acc and '흐름' in acc): return t_rank + 40000
                    if acc == '재무활동현금흐름' or ('재무활동' in acc and '흐름' in acc): return t_rank + 70000
                    return t_rank + 10000 + idx
                return t_rank + idx
                
            pivot_df['SortKey'] = pivot_df.apply(get_heuristic_rank, axis=1)
            pivot_df = pivot_df.sort_values('SortKey').drop(columns=['SortKey', 'Row_Idx'])
            
            # 헤더 그리기
            labels = [p['label'] for p in periods_to_fetch]
            ws_dtl.cell(row=3, column=1, value="M_Key"); sc(ws_dtl.cell(row=3, column=1), fo=fH, fi=pH, al=aC, bd=BD)
            ws_dtl.cell(row=3, column=2, value="Type"); sc(ws_dtl.cell(row=3, column=2), fo=fH, fi=pH, al=aC, bd=BD)
            ws_dtl.cell(row=3, column=3, value="Account ID"); sc(ws_dtl.cell(row=3, column=3), fo=fH, fi=pH, al=aC, bd=BD)
            ws_dtl.cell(row=3, column=4, value="Account Name"); sc(ws_dtl.cell(row=3, column=4), fo=fH, fi=pH, al=aC, bd=BD)
            
            col_idx = 5
            for plabel in labels:
                ws_dtl.cell(row=3, column=col_idx, value=plabel); sc(ws_dtl.cell(row=3, column=col_idx), fo=fH, fi=pSEC2, al=aC, bd=BD)
                col_idx += 1
                
            r = 4
            for _, row in pivot_df.iterrows():
                ws_dtl.cell(row=r, column=1, value=row.get('M_Key', '')); sc(ws_dtl.cell(row=r, column=1), fo=fA, al=aL, bd=BD)
                ws_dtl.cell(row=r, column=2, value=row.get('Type', '')); sc(ws_dtl.cell(row=r, column=2), fo=fA, al=aL, bd=BD)
                ws_dtl.cell(row=r, column=3, value=row.get('Account_ID', '')); sc(ws_dtl.cell(row=r, column=3), fo=fA, al=aL, bd=BD)
                ws_dtl.cell(row=r, column=4, value=row.get('Account_NM', '')); sc(ws_dtl.cell(row=r, column=4), fo=fA, al=aL, bd=BD)
                
                c = 5
                for plabel in labels:
                    val = row.get(plabel)
                    v = val if pd.notna(val) else ""
                    ws_dtl.cell(row=r, column=c, value=v); sc(ws_dtl.cell(row=r, column=c), fo=fHARD, nf=NB, bd=BD)
                    c += 1
                r += 1
                
            ws_dtl.column_dimensions['A'].width = 15
            ws_dtl.column_dimensions['B'].width = 15
            ws_dtl.column_dimensions['C'].width = 25
            ws_dtl.column_dimensions['D'].width = 35
            for i in range(5, c):
                ws_dtl.column_dimensions[get_column_letter(i)].width = 15
                
            ws_dtl.freeze_panes = "E5"


    wb.save(output)
    output.seek(0)
    return output

pSEC3 = PatternFill('solid', fgColor='2E7D32'); pSEC4 = PatternFill('solid', fgColor='6A1B9A')
pSEC5 = PatternFill('solid', fgColor='C62828'); pSEC6 = PatternFill('solid', fgColor='455A64')

ev_fills = {
    'Cash': PatternFill('solid',fgColor=C_GR), 'IBD': PatternFill('solid',fgColor=C_YL),
    'IBD(Option)': PatternFill('solid',fgColor=C_YL), 'NOA(Option)': PatternFill('solid',fgColor=C_NOA),
    'NOA': PatternFill('solid',fgColor=C_NOA), 'NCI': PatternFill('solid',fgColor=C_PB),
    'Equity': PatternFill('solid',fgColor=C_LB), 'PL_HL': PatternFill('solid',fgColor=C_YL),
    'Preferred': PatternFill('solid',fgColor='D1C4E9'),  # 우선주(장부) — 자기자본가치 가산 항목
}

aC=Alignment(horizontal='center',vertical='center',wrap_text=True)
aL=Alignment(horizontal='left',vertical='center',indent=1)
aR=Alignment(horizontal='right',vertical='center')

NB='#,##0;(#,##0);"-"'; NB1='#,##0.0;(#,##0.0);"-"'; NI_FMT='#,##0;(#,##0);"-"'
NP='₩#,##0;(₩#,##0);"-"'; NF_X='#,##0.0x;(#,##0.0x);"-"'

def sc(c,fo=None,fi=None,al=None,bd=None,nf=None):
    if fo: c.font=fo
    if fi: c.fill=fi
    if al: c.alignment=al
    if bd: c.border=bd
    if nf: c.number_format=nf

def style_range(ws, r1, c1, r2, c2, fo=None, fi=None, al=None, bd=None, nf=None):
    for rr in range(r1, r2+1):
        for cc in range(c1, c2+1):
            sc(ws.cell(rr, cc), fo=fo, fi=fi, al=al, bd=bd, nf=nf)



# GPCM 시트 열 배치 — 이 목록의 순서가 곧 열이다. 열을 옮기거나 끼울 때 여기만 고친다.
# 예전에는 열 번호·열 문자가 45곳에 흩어져 있어, 열 하나를 옮기면 수식 참조가
# 조용히 어긋났다. 섹션·헤더·수식·통계가 전부 이 목록에서 파생된다.
GPCM_COL_DEFS = [
    ('Company', 'Company', 18), ('Ticker', 'Ticker', 10), ('BaseDate', 'Base Date', 11),
    ('Curr', 'Curr', 6), ('PLSource', 'PL Source', 13),
    ('Cash', 'Cash', 13), ('IBD', 'IBD', 13), ('NOA', 'NOA', 13), ('NetDebt', 'Net Debt', 13),
    ('NCI', 'NCI', 12), ('Pref', '우선주(장부)', 13), ('Equity', 'Equity', 13), ('EV', 'EV', 15),
    ('Revenue', 'Revenue', 13), ('EBIT', 'EBIT', 13), ('DA', 'D&A', 10),
    ('EBITDA', 'EBITDA', 13), ('NI', 'NI', 13),
    ('Price', 'Price', 12), ('Shares', 'Shares', 15), ('MktCap', 'Mkt Cap', 15),
    ('EVEBITDA', 'EV/EBITDA', 12), ('EVEBIT', 'EV/EBIT', 12),
    ('PER', 'PER', 10), ('PBR', 'PBR', 10), ('PSR', 'PSR', 10),
    ('B5Raw', 'β 5Y Raw', 10), ('B5Adj', 'β 5Y Adj', 10),
    ('B2Raw', 'β 2Y Raw', 10), ('B2Adj', 'β 2Y Adj', 10),
    ('Pretax', 'Pretax Inc', 13), ('TaxRate', 'Tax Rate', 9), ('DERatio', 'D/E Ratio', 10),
    ('DVRatio', 'Debt Ratio (D/V)', 10), ('UB5', 'Unlevered β 5Y', 12), ('UB2', 'Unlevered β 2Y', 12),
]
GPCM_CI = {key: i for i, (key, _, _) in enumerate(GPCM_COL_DEFS, 1)}            # 키 → 열 번호
GPCM_CL = {key: get_column_letter(i) for key, i in GPCM_CI.items()}             # 키 → 열 문자


def add_gpcm_section_row(ws):
    sec_row = 4
    C = GPCM_CI
    sections = [
        (C['Company'], C['Ticker'],  "Company Info",        pSEC1),
        (C['BaseDate'], C['PLSource'], "Other Info",        pSEC2),
        (C['Cash'], C['EV'],         "BS & EV Components",  pSEC3),
        (C['Revenue'], C['NI'],      "PL(Annual & LTM)",    pSEC4),
        (C['Price'], C['MktCap'],    "Market Data",         pSEC5),
        (C['EVEBITDA'], C['PSR'],    "Valuation Multiples", pSEC6),
        (C['B5Raw'], C['UB2'],       "Beta & Risk Analysis", PatternFill('solid', fgColor='6A1B9A')),
    ]
    for c1, c2, label, fill in sections:
        ws.merge_cells(start_row=sec_row, start_column=c1, end_row=sec_row, end_column=c2)
        ws.cell(sec_row, c1).value = label
        style_range(ws, sec_row, c1, sec_row, c2, fo=fSEC, fi=fill, al=aC, bd=BD)


# ==========================================

def fetch_financial_data(api_key_input, target_code_list, target_periods, dart, status_container, progress_bar):
    df_krx = get_krx_listing()
    
    # 변수 초기화
    base_period_str = target_periods[-1]
    base_year, base_qtr = parse_period(base_period_str)
    base_date_str = get_base_date_str(base_year, base_qtr)

    raw_bs_rows = []
    raw_pl_rows = []
    all_mkt = []
    ticker_to_name = {}

    screen_summary_data = []
    all_multiples = []
    quality = QualityLog()

    total_tickers = len(target_code_list)
    dart_fs_cache = {}  # DART API Call 최소화를 위한 캐시 (ticker 포함 키로 충돌 방지)
    market_idx_cache = {}  # 시장지수 시계열 캐시 (전 종목 공통)

    for idx, ticker in enumerate(target_code_list):
        status_container.write(f"Processing [{ticker}] ({idx+1}/{total_tickers})...")
        progress_bar.progress((idx) / total_tickers)

        corp_code, krx_name = resolve_company_info(dart, ticker)
        if not corp_code:
            status_container.write(f"❌ [{ticker}] DART 고유번호 조회 실패")
            quality.add(SEV_ERROR, ticker, '', '회사 조회',
                        'DART 고유번호를 찾지 못해 이 회사는 분석에서 통째로 빠졌습니다. '
                        '종목코드 6자리가 맞는지, 상장사가 맞는지 확인하세요.')
            continue

        display_name = krx_name if krx_name else f"Company_{ticker}"
        ticker_to_name[ticker] = display_name

        # 임시 저장소 (화면 출력용) - 최신 기준일 데이터
        temp_metrics = {
            'Company': display_name, 'Ticker': ticker,
            'Market_Cap': 0, 'Cash': 0, 'IBD': 0, 'NCI': 0, 'NOA': 0, 'Equity': 0, 'Preferred': 0,
            'Revenue': 0, 'EBIT': 0, 'NI': 0, 'Pretax_Income': 0,
            'Stock_Monthly_Prices_5Y': None, 'Market_Monthly_Prices_5Y': None,
            'Stock_Weekly_Prices_2Y': None, 'Market_Weekly_Prices_2Y': None,
            'Exchange': 'KRX', 'Market_Index': 'KS11',
        }

        for tp in target_periods:
            tyear, tqtr = parse_period(tp)
            required_periods = get_ltm_required_periods(tyear, tqtr)
            
            period_metrics = {
                'Market_Cap': 0, 'Cash': 0, 'IBD': 0, 'NCI': 0, 'NOA': 0, 'Equity': 0, 'Preferred': 0,
                'Revenue': 0, 'EBIT': 0, 'NI': 0, 'Pretax_Income': 0
            }

            for year, qtr, role in required_periods:
                r_code = RCODE_MAP[qtr]
                bds = get_base_date_str(year, qtr)

                # 1) Market Cap (기준시점만)
                if role in ('current_cum', 'annual'):
                    price, price_date = get_stock_price(ticker, bds)
                    shares, shares_src, sh_meta = get_outstanding_shares(api_key_input, corp_code, ticker, year, r_code, df_krx, cache=dart_fs_cache)

                    mkt_100m = 0
                    if price is not None and shares is not None and shares > 0:
                        mkt_100m = round((price * shares) / 1e8, 1)
                    else:
                        # 시가총액이 0이면 EV·자본구조·베타가 한꺼번에 무너진다
                        missing = []
                        if price is None:
                            missing.append(f'{bds} 종가')
                        if not shares:
                            missing.append('발행주식수')
                        quality.add(SEV_ERROR, ticker, display_name, f'시가총액 {tp}',
                                    f"{' · '.join(missing)}를 못 가져와 시가총액이 0입니다. "
                                    f"EV와 자본구조가 왜곡되니 Market_Cap 시트에서 사유를 확인하세요."
                                    + (f" (DART: {sh_meta.get('message')})" if sh_meta.get('message') else ''))

                    period_metrics['Market_Cap'] = mkt_100m

                    all_mkt.append({
                        'Company': display_name, 'Ticker': ticker, 'Period': tp,
                        'Price_Date': price_date or bds, 'Close': price,
                        'Outstanding_Shares': shares, 'Market_Cap_100M': mkt_100m,
                        'Shares_Source': shares_src, 'Shares_RceptNo': sh_meta.get('rcept_no'),
                        'Shares_StlmDt': sh_meta.get('stlm_dt'), 'Shares_Se': sh_meta.get('se'),
                        'DART_Status': sh_meta.get('status'), 'DART_Message': sh_meta.get('message'),
                    })

                # 2) BS Fetch (finstate_all: 상세) - CFS 우선 → OFS
                if role in ('current_cum', 'annual'):
                    df_all = None
                    cache_key = f"all_{ticker}_{year}_{r_code}"
                    if cache_key in dart_fs_cache:
                        df_all = dart_fs_cache[cache_key]
                    else:
                        for fs in ['CFS', 'OFS']:
                            try:
                                _df = dart.finstate_all(corp_code, year, reprt_code=r_code, fs_div=fs)
                                if _df is not None and not _df.empty:
                                    df_all = _df
                                    dart_fs_cache[cache_key] = _df
                                    break
                            except Exception:
                                continue

                    if df_all is not None and not df_all.empty:
                        df_bs = df_all[df_all['sj_nm'].astype(str).str.contains('상태표|재정상태', na=False)].copy()
                        for _, row in df_bs.iterrows():
                            amt = pd.to_numeric(str(row.get('thstrm_amount', '')).replace(',', ''), errors='coerce')
                            if pd.isna(amt) or amt == 0: continue

                            acct = str(row.get('account_nm', '')).strip()
                            aid = str(row.get('account_id', '')).strip()
                            ev_comp, _ = match_bs_ev_component(acct, aid)

                            if ev_comp:
                                # 화면 출력용 집계
                                val_100m = amt / 1e8
                                if ev_comp == 'Cash': period_metrics['Cash'] += val_100m
                                elif ev_comp == 'IBD': period_metrics['IBD'] += val_100m
                                elif ev_comp == 'NCI': period_metrics['NCI'] += val_100m
                                elif ev_comp == 'NOA': period_metrics['NOA'] += val_100m
                                elif ev_comp == 'Preferred': period_metrics['Preferred'] += val_100m
                                elif ev_comp in ['Equity_Total', 'Equity_P']: period_metrics['Equity'] += val_100m

                            raw_bs_rows.append({
                                'Company': display_name, 'Ticker': ticker, 'Period': tp,
                                'sj_nm': row.get('sj_nm', ''), 'account_nm': acct, 'account_id': aid,
                                'EV_Component': ev_comp or '', 'Amount_100M': amt / 1e8,
                            })
                    else:
                        # 재무상태표를 못 받으면 현금·차입금·자본이 전부 0으로 남는다
                        quality.add(SEV_ERROR, ticker, display_name, f'재무상태표 {tp}',
                                    f'{year}년 {qtr} 재무상태표를 연결·별도 모두 가져오지 못했습니다. '
                                    f'현금·이자부부채·비지배지분·자본이 0으로 처리됩니다.')

                # 3) PL Fetch
                df_is = None
                cache_key_pl = f"pl_{ticker}_{year}_{r_code}"
                if cache_key_pl in dart_fs_cache:
                    df_is, pl_src = dart_fs_cache[cache_key_pl]
                else:
                    df_pl_raw, pl_src, _ = fetch_pl_df(dart, corp_code, year, r_code)
                    if df_pl_raw is not None and not df_pl_raw.empty:
                        df_is = filter_income_statement(df_pl_raw)
                        dart_fs_cache[cache_key_pl] = (df_is, pl_src)
                    
                # LTM = 당기누계 + 전기연간 - 전년동기. 셋 중 하나만 빠져도 합계가
                # 그럴듯하게 틀리므로, 어느 조각이 빠졌는지 남긴다.
                ltm_part = {'current_cum': '당기누계', 'prior_annual': '전기연간',
                            'prior_same_q': '전년동기'}.get(role)
                if df_is is None or df_is.empty:
                    if ltm_part:
                        quality.add(SEV_ERROR, ticker, display_name, f'LTM {tp}',
                                    f'{ltm_part}({year} {qtr}) 손익계산서를 못 가져왔습니다. '
                                    f'이 조각을 뺀 채로 합산되어 매출·영업이익이 실제와 다릅니다.')
                    else:
                        quality.add(SEV_ERROR, ticker, display_name, f'손익계산서 {tp}',
                                    f'{year}년 {qtr} 손익계산서를 가져오지 못해 매출·영업이익이 0입니다.')
                    continue

                wanted = {'Revenue', 'EBIT', 'NI', 'Pretax_Income'}
                picked = set()
                seen_names = []

                def take(row, acct, calc_key):
                    val = pick_pl_value(row, qtr)
                    if val is None: return False

                    amt_100m = val / 1e8
                    raw_pl_rows.append({
                        'Company': display_name, 'Ticker': ticker, 'Period': tp,
                        'Role': role, 'PL_Source': pl_src, 'account_nm': acct,
                        'calc_key': calc_key, 'Amount_100M': amt_100m,
                    })

                    if role in ('current_cum', 'annual'):
                        period_metrics[calc_key] += amt_100m
                    elif role == 'prior_annual':
                        period_metrics[calc_key] += amt_100m
                    elif role == 'prior_same_q':
                        period_metrics[calc_key] -= amt_100m

                    picked.add(calc_key)
                    return True

                # 1단계 — 완전 일치. 여기서 잡히면 2단계는 돌지 않는다(기존 동작 보존).
                for _, row in df_is.iterrows():
                    acct = str(row.get('account_nm', '')).strip()
                    aid = str(row.get('account_id', '')).strip()
                    if acct: seen_names.append(acct)
                    calc_key = match_pl_core_only(acct, aid)
                    if not calc_key or calc_key not in wanted: continue
                    if calc_key in picked: continue
                    take(row, acct, calc_key)
                    if picked == wanted: break

                # 2단계 — 못 찾은 것만. 표기가 "Ⅰ. 영업수익"·"매출액(주7)" 처럼
                # 흔들리거나 표준태그로만 알 수 있는 경우를 여기서 줍는다.
                if picked != wanted:
                    for _, row in df_is.iterrows():
                        acct = str(row.get('account_nm', '')).strip()
                        aid = str(row.get('account_id', '')).strip()
                        calc_key = match_pl_lenient(acct, aid)
                        if not calc_key or calc_key not in wanted: continue
                        if calc_key in picked: continue
                        take(row, acct, calc_key)
                        if picked == wanted: break

                # 그래도 못 찾으면 그 계정만 0이 된다. 무엇을 봤는지 같이 남긴다 —
                # 안 그러면 어느 표기 때문에 빠졌는지 확인할 방법이 없다.
                unmatched = wanted - picked
                if unmatched:
                    label = {'Revenue': '매출액', 'EBIT': '영업이익',
                             'NI': '당기순이익', 'Pretax_Income': '법인세비용차감전순이익'}
                    shown = ', '.join(seen_names[:10]) or '(계정과목명이 비어 있습니다)'
                    more = f' 외 {len(seen_names) - 10}개' if len(seen_names) > 10 else ''
                    quality.add(SEV_WARN, ticker, display_name, f'계정 매칭 {tp}',
                                f"{year} {qtr} 손익계산서에서 "
                                f"{', '.join(label[k] for k in sorted(unmatched))}을(를) "
                                f"찾지 못했습니다. 해당 계정은 0으로 집계됩니다. "
                                f"읽은 계정과목: {shown}{more}")

            # Period loop ends, append to all_multiples
            all_multiples.append({
                'Company': display_name, 'Ticker': ticker, 'Period': tp,
                **period_metrics
            })
            
            # If this is the main base period, update temp_metrics
            if tp == base_period_str:
                temp_metrics.update(period_metrics)

        # 4) Beta Calculation (5Y Monthly, 2Y Weekly)
        exchange, market_idx = get_market_index(ticker)
        temp_metrics['Exchange'] = exchange
        temp_metrics['Market_Index'] = market_idx

        try:
            end_date = base_date_str
            start_5y = (pd.to_datetime(base_date_str) - timedelta(days=BETA_5Y_DAYS)).strftime('%Y-%m-%d')
            start_2y = (pd.to_datetime(base_date_str) - timedelta(days=BETA_2Y_DAYS)).strftime('%Y-%m-%d')

            # 5년 월간 베타 데이터
            # 시장지수는 모든 종목이 동일하므로 종목마다 다시 받지 않고 캐시에서 재사용
            stock_data_5y = fdr.DataReader(ticker, start_5y, end_date)
            market_data_5y = _get_market_index_data(market_idx, start_5y, end_date, market_idx_cache)

            if stock_data_5y is not None and not stock_data_5y.empty and market_data_5y is not None and not market_data_5y.empty:
                stock_prices_5y = _to_price_series(stock_data_5y)
                market_prices_5y = _to_price_series(market_data_5y)

                if not isinstance(stock_prices_5y.index, pd.DatetimeIndex):
                    stock_prices_5y.index = pd.to_datetime(stock_prices_5y.index)
                if stock_prices_5y.index.tz is not None:
                    stock_prices_5y.index = stock_prices_5y.index.tz_localize(None)
                if not isinstance(market_prices_5y.index, pd.DatetimeIndex):
                    market_prices_5y.index = pd.to_datetime(market_prices_5y.index)
                if market_prices_5y.index.tz is not None:
                    market_prices_5y.index = market_prices_5y.index.tz_localize(None)

                stock_monthly_prices = stock_prices_5y.resample('ME').last().dropna()
                market_monthly_prices = market_prices_5y.resample('ME').last().dropna()

                if len(stock_monthly_prices) >= MIN_MONTHLY_PTS and len(market_monthly_prices) >= MIN_MONTHLY_PTS:
                    temp_metrics['Stock_Monthly_Prices_5Y'] = stock_monthly_prices
                    temp_metrics['Market_Monthly_Prices_5Y'] = market_monthly_prices

            # 2년 주간 베타 데이터
            # 2년 구간은 위에서 받은 5년 구간의 부분집합이므로 잘라 쓴다 (재조회 불필요)
            stock_data_2y = _slice_from(stock_data_5y, start_2y)
            market_data_2y = _slice_from(market_data_5y, start_2y)

            if stock_data_2y is not None and not stock_data_2y.empty and market_data_2y is not None and not market_data_2y.empty:
                stock_prices_2y = _to_price_series(stock_data_2y)
                market_prices_2y = _to_price_series(market_data_2y)

                if not isinstance(stock_prices_2y.index, pd.DatetimeIndex):
                    stock_prices_2y.index = pd.to_datetime(stock_prices_2y.index)
                if stock_prices_2y.index.tz is not None:
                    stock_prices_2y.index = stock_prices_2y.index.tz_localize(None)
                if not isinstance(market_prices_2y.index, pd.DatetimeIndex):
                    market_prices_2y.index = pd.to_datetime(market_prices_2y.index)
                if market_prices_2y.index.tz is not None:
                    market_prices_2y.index = market_prices_2y.index.tz_localize(None)

                stock_weekly_prices = stock_prices_2y.resample('W-FRI').last().dropna()
                market_weekly_prices = market_prices_2y.resample('W-FRI').last().dropna()

                if len(stock_weekly_prices) >= MIN_WEEKLY_PTS and len(market_weekly_prices) >= MIN_WEEKLY_PTS:
                    temp_metrics['Stock_Weekly_Prices_2Y'] = stock_weekly_prices
                    temp_metrics['Market_Weekly_Prices_2Y'] = market_weekly_prices

        except Exception as e:
            beta_failed = True
            quality.add(SEV_WARN, ticker, display_name, '베타',
                        f'주가 시계열을 받지 못해 베타를 계산할 수 없습니다 ({type(e).__name__}). '
                        f'이 회사는 WACC 평균에서 빠집니다.')
        else:
            beta_failed = False

        # 조회는 됐지만 관측치가 모자라 시계열을 담지 못한 경우도 조용히 넘어간다
        if not beta_failed and temp_metrics['Stock_Monthly_Prices_5Y'] is None:
            quality.add(SEV_WARN, ticker, display_name, '베타 5Y',
                        f'월간 관측치가 {MIN_MONTHLY_PTS}개에 못 미쳐 5년 월간 베타를 산출하지 않았습니다. '
                        f'상장한 지 얼마 안 된 회사에서 주로 발생합니다.')
        if not beta_failed and temp_metrics['Stock_Weekly_Prices_2Y'] is None:
            quality.add(SEV_WARN, ticker, display_name, '베타 2Y',
                        f'주간 관측치가 {MIN_WEEKLY_PTS}개에 못 미쳐 2년 주간 베타를 산출하지 않았습니다.')

        screen_summary_data.append(temp_metrics)
        time.sleep(0.5) # API 호출 간격 조절

    progress_bar.progress(1.0)
    status_container.update(label="분석 완료!", state="complete", expanded=False)

    # --- 결과 처리 및 엑셀 생성 ---

    return raw_bs_rows, raw_pl_rows, all_mkt, ticker_to_name, screen_summary_data, base_year, base_qtr, base_date_str, all_multiples, quality

def calculate_wacc_and_beta(target_code_list, screen_summary_data, target_tax_rate_input, rf_input, mrp_input, size_premium_input, kd_pretax_input, beta_type_input, fiscal_year=None, quality=None):
    # 1.5. WACC Calculation (Target 기업용)
    # Beta 시트에서 계산될 Unlevered Beta를 엑셀에서 참조할 것이므로,
    # 여기서는 대략적인 값만 계산 (정확한 값은 엑셀 수식 기반)

    # 피어들의 평균 계산을 위한 준비
    avg_debt_ratios = []
    avg_unlevered_betas_5y = []
    avg_unlevered_betas_2y = []

    for ticker in target_code_list:
        comp_data = next((item for item in screen_summary_data if item["Ticker"] == ticker), None)
        if not comp_data:
            continue

        mkt_cap = comp_data.get('Market_Cap', 0)
        ibd = comp_data.get('IBD', 0)
        nci = comp_data.get('NCI', 0)
        pref = comp_data.get('Preferred', 0)  # 우선주 자본금 (시가총액은 보통주만 반영)
        equity = comp_data.get('Equity', 0)
        pretax_income = comp_data.get('Pretax_Income', 0)

        # Debt Ratio (D/V) = IBD / (Mkt Cap + 우선주 + IBD + NCI)
        total_value = mkt_cap + pref + ibd + nci
        if total_value > 0:
            debt_ratio = ibd / total_value
            avg_debt_ratios.append(debt_ratio)

        # D/E Ratio = IBD / (Mkt Cap + 우선주 + NCI)
        equity_value = mkt_cap + pref + nci
        de_ratio = ibd / equity_value if equity_value > 0 else 0

        # 한계세율 계산 (사업연도별 한국 법인세율표, 지방소득세 포함)
        tax_rate = get_korean_marginal_tax_rate(pretax_income, fiscal_year)
        comp_data['Tax_Rate'] = tax_rate  # 저장 (나중에 Excel 출력용)

        # Beta 계산 (간단히 수익률 기반)
        stock_monthly_5y = comp_data.get('Stock_Monthly_Prices_5Y')
        market_monthly_5y = comp_data.get('Market_Monthly_Prices_5Y')
        stock_weekly_2y = comp_data.get('Stock_Weekly_Prices_2Y')
        market_weekly_2y = comp_data.get('Market_Weekly_Prices_2Y')

        # 5Y Monthly Beta
        if stock_monthly_5y is not None and market_monthly_5y is not None and not stock_monthly_5y.empty and not market_monthly_5y.empty:
            try:
                common_dates = stock_monthly_5y.index.intersection(market_monthly_5y.index)
                stock_ret = stock_monthly_5y.loc[common_dates].pct_change().dropna()
                market_ret = market_monthly_5y.loc[common_dates].pct_change().dropna()
                common_idx = stock_ret.index.intersection(market_ret.index)
                # 관측치가 적다고 여기서 빼지 않는다 — 엑셀 SLOPE 는 그대로 계산하므로
                # 파이썬만 빼면 두 평균의 대상이 달라진다. 대신 아래에서 경고한다.
                if len(common_idx) >= 2:
                    stock_ret_aligned = stock_ret.loc[common_idx]
                    market_ret_aligned = market_ret.loc[common_idx]
                    cov_matrix = np.cov(stock_ret_aligned, market_ret_aligned)
                    beta_raw = cov_matrix[0, 1] / cov_matrix[1, 1] if cov_matrix[1, 1] != 0 else np.nan
                    beta_adj = (2/3) * beta_raw + (1/3) * 1

                    # 신뢰도 — 값 계산은 그대로 두고 판단 근거만 덧붙인다.
                    # R²: 주가 변동 중 시장으로 설명되는 비중 (낮으면 기울기가
                    # 관계를 요약한 것이 아니다). n: 회귀에 실제로 들어간 관측치.
                    n_obs = int(len(common_idx))
                    comp_data['Beta_5Y_Raw'] = None if np.isnan(beta_raw) else float(beta_raw)
                    comp_data['Beta_5Y_N'] = n_obs
                    with np.errstate(invalid='ignore'):
                        corr = np.corrcoef(stock_ret_aligned, market_ret_aligned)[0, 1]
                    comp_data['Beta_5Y_R2'] = None if np.isnan(corr) else float(corr ** 2)

                    if quality is not None and n_obs < FULL_MONTHLY_OBS:
                        quality.add(SEV_WARN, ticker, comp_data.get('Company', ''), 'Beta',
                                    f'5년 월간 관측치가 {n_obs}/{FULL_MONTHLY_OBS}개입니다. 상장이 늦었거나 '
                                    f'거래정지 구간이 있는지 확인하세요 — 값은 그대로 씁니다.')
                    if quality is not None and not np.isnan(beta_raw) and abs(beta_raw) > BETA_SANITY_LIMIT:
                        quality.add(SEV_WARN, ticker, comp_data.get('Company', ''), 'Beta',
                                    f'5년 월간 Raw 베타가 {beta_raw:.2f} 로 통상 범위'
                                    f'(±{BETA_SANITY_LIMIT})를 벗어납니다. 버리지 않고 그대로 두니 '
                                    f'쓸지 여부를 판단하세요.')

                    # 엑셀과 같은 조건으로 평균에 넣는다:
                    #   =IF(조정베타>0, 조정베타/(1+(1-세율)*D/E), "")  → 빈칸은 AVERAGE 에서 빠짐
                    # 자기자본(장부) 조건은 두지 않는다 — 엑셀에 없는 조건이라
                    # 파이썬만 빼면 두 평균이 달라진다.
                    if not np.isnan(beta_adj) and beta_adj > 0:
                        unlevered_beta_5y = beta_adj / (1 + (1 - tax_rate) * de_ratio)
                        avg_unlevered_betas_5y.append(unlevered_beta_5y)
                    elif quality is not None:
                        quality.add(SEV_WARN, ticker, comp_data.get('Company', ''), 'Beta',
                                    f'5년 월간 조정베타가 0 이하이거나 산출되지 않아 평균에서 '
                                    f'빠집니다(엑셀도 동일). 피어 평균이 이 회사 없이 계산됩니다.')
                elif quality is not None:
                    quality.add(SEV_WARN, ticker, comp_data.get('Company', ''), 'Beta',
                                f'5년 월간 수익률이 {len(common_idx)}개뿐이라 회귀가 불가능합니다.')
            except Exception as e:
                if quality is not None:
                    quality.add(SEV_WARN, ticker, comp_data.get('Company', ''), 'Beta',
                                f'5년 월간 베타 계산이 실패해 평균에서 빠졌습니다: {e}')

        # 2Y Weekly Beta
        if stock_weekly_2y is not None and market_weekly_2y is not None and not stock_weekly_2y.empty and not market_weekly_2y.empty:
            try:
                common_dates = stock_weekly_2y.index.intersection(market_weekly_2y.index)
                stock_ret = stock_weekly_2y.loc[common_dates].pct_change().dropna()
                market_ret = market_weekly_2y.loc[common_dates].pct_change().dropna()
                common_idx = stock_ret.index.intersection(market_ret.index)
                # 관측치가 적다고 여기서 빼지 않는다 — 엑셀 SLOPE 는 그대로 계산하므로
                # 파이썬만 빼면 두 평균의 대상이 달라진다. 대신 아래에서 경고한다.
                if len(common_idx) >= 2:
                    stock_ret_aligned = stock_ret.loc[common_idx]
                    market_ret_aligned = market_ret.loc[common_idx]
                    cov_matrix = np.cov(stock_ret_aligned, market_ret_aligned)
                    beta_raw = cov_matrix[0, 1] / cov_matrix[1, 1] if cov_matrix[1, 1] != 0 else np.nan
                    beta_adj = (2/3) * beta_raw + (1/3) * 1

                    # 신뢰도 — 값 계산은 그대로 두고 판단 근거만 덧붙인다.
                    # R²: 주가 변동 중 시장으로 설명되는 비중 (낮으면 기울기가
                    # 관계를 요약한 것이 아니다). n: 회귀에 실제로 들어간 관측치.
                    n_obs = int(len(common_idx))
                    comp_data['Beta_2Y_Raw'] = None if np.isnan(beta_raw) else float(beta_raw)
                    comp_data['Beta_2Y_N'] = n_obs
                    with np.errstate(invalid='ignore'):
                        corr = np.corrcoef(stock_ret_aligned, market_ret_aligned)[0, 1]
                    comp_data['Beta_2Y_R2'] = None if np.isnan(corr) else float(corr ** 2)

                    if quality is not None and n_obs < FULL_WEEKLY_OBS:
                        quality.add(SEV_WARN, ticker, comp_data.get('Company', ''), 'Beta',
                                    f'2년 주간 관측치가 {n_obs}/{FULL_WEEKLY_OBS}개입니다. 상장이 늦었거나 '
                                    f'거래정지 구간이 있는지 확인하세요 — 값은 그대로 씁니다.')
                    if quality is not None and not np.isnan(beta_raw) and abs(beta_raw) > BETA_SANITY_LIMIT:
                        quality.add(SEV_WARN, ticker, comp_data.get('Company', ''), 'Beta',
                                    f'2년 주간 Raw 베타가 {beta_raw:.2f} 로 통상 범위'
                                    f'(±{BETA_SANITY_LIMIT})를 벗어납니다. 버리지 않고 그대로 두니 '
                                    f'쓸지 여부를 판단하세요.')

                    # 엑셀과 같은 조건으로 평균에 넣는다:
                    #   =IF(조정베타>0, 조정베타/(1+(1-세율)*D/E), "")  → 빈칸은 AVERAGE 에서 빠짐
                    # 자기자본(장부) 조건은 두지 않는다 — 엑셀에 없는 조건이라
                    # 파이썬만 빼면 두 평균이 달라진다.
                    if not np.isnan(beta_adj) and beta_adj > 0:
                        unlevered_beta_2y = beta_adj / (1 + (1 - tax_rate) * de_ratio)
                        avg_unlevered_betas_2y.append(unlevered_beta_2y)
                    elif quality is not None:
                        quality.add(SEV_WARN, ticker, comp_data.get('Company', ''), 'Beta',
                                    f'2년 주간 조정베타가 0 이하이거나 산출되지 않아 평균에서 '
                                    f'빠집니다(엑셀도 동일). 피어 평균이 이 회사 없이 계산됩니다.')
                elif quality is not None:
                    quality.add(SEV_WARN, ticker, comp_data.get('Company', ''), 'Beta',
                                f'2년 주간 수익률이 {len(common_idx)}개뿐이라 회귀가 불가능합니다.')
            except Exception as e:
                if quality is not None:
                    quality.add(SEV_WARN, ticker, comp_data.get('Company', ''), 'Beta',
                                f'2년 주간 베타 계산이 실패해 평균에서 빠졌습니다: {e}')

    # 평균값 계산
    avg_debt_ratio = np.mean(avg_debt_ratios) if avg_debt_ratios else 0.3

    # Beta Type에 따라 선택
    used_betas = avg_unlevered_betas_5y if beta_type_input == "5Y" else avg_unlevered_betas_2y
    avg_unlevered_beta = np.mean(used_betas) if used_betas else 0.8
    if not used_betas and quality is not None:
        # 값(0.8)은 종전 그대로 둔다 — 다만 말없이 쓰이면 안 된다
        quality.add(SEV_ERROR, '', '', 'Beta',
                    f'피어 중 어느 곳에서도 {beta_type_input} 베타를 못 구해 '
                    f'기본값 0.8 을 사용했습니다. WACC 이 이 가정 위에 서 있습니다.')
    elif quality is not None and len(used_betas) < len(target_code_list):
        quality.add(SEV_WARN, '', '', 'Beta',
                    f'{beta_type_input} 무차입베타 평균에 {len(used_betas)}곳만 기여했습니다 '
                    f'(대상 {len(target_code_list)}곳). 빠진 회사는 위 Beta 경고를 보세요.')

    # Target D/E Ratio 계산
    target_de_ratio = avg_debt_ratio / (1 - avg_debt_ratio) if avg_debt_ratio < 1 else 0

    # Relevered Beta 계산
    target_relevered_beta = avg_unlevered_beta * (1 + (1 - target_tax_rate_input) * target_de_ratio)

    # Ke (자기자본비용) 계산
    target_ke = rf_input + mrp_input * target_relevered_beta + size_premium_input

    # Kd (타인자본비용, 세후)
    kd_aftertax = kd_pretax_input * (1 - target_tax_rate_input)

    # E/V, D/V
    equity_weight = 1 - avg_debt_ratio
    debt_weight = avg_debt_ratio

    # Target WACC
    target_wacc = equity_weight * target_ke + debt_weight * kd_aftertax

    # WACC 데이터 저장
    target_wacc_data = {
        'Rf': rf_input,
        'MRP': mrp_input,
        'Size_Premium': size_premium_input,
        'Avg_Unlevered_Beta': avg_unlevered_beta,
        'Target_Tax_Rate': target_tax_rate_input,
        'Avg_Debt_Ratio': avg_debt_ratio,
        'Target_DE_Ratio': target_de_ratio,
        'Target_Relevered_Beta': target_relevered_beta,
        'Target_Ke': target_ke,
        'Kd_Pretax': kd_pretax_input,
        'Kd_Aftertax': kd_aftertax,
        'Equity_Weight': equity_weight,
        'Debt_Weight': debt_weight,
        'Target_WACC': target_wacc
    }
    return target_wacc_data, avg_debt_ratio

def export_gpcm_excel(base_period_str, base_qtr, target_code_list, screen_summary_data, raw_bs_rows, raw_pl_rows, all_mkt, ticker_to_name, target_wacc_data, beta_type_input, notes_list, avg_debt_ratio, base_date_str, df_screen, target_periods, quality, peer_selection=None):
    # 2. 엑셀 생성 (메모리)
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    # 세율 수식을 만들 사업연도 — 파이썬 WACC 계산에 쓴 것과 같은 기준이어야 한다
    base_fiscal_year, _ = parse_period(base_period_str)

    # (기존 엑셀 생성 로직 그대로 활용 - 함수화 하지 않고 바로 실행)
    # Sheet 1: BS_Full
    ws_bs = wb.create_sheet('BS_Full')
    ws_bs.merge_cells('A1:H1'); ws_bs['A1'] = "BS_Full (Balance Sheet Detail)"; sc(ws_bs['A1'], fo=fT)
    ws_bs.merge_cells('A2:H2'); ws_bs['A2'] = "Logic: finstate_all(CFS→OFS) 재무상태표 라인아이템 수집 후 EV_Component 태깅 | Unit: 억원"; sc(ws_bs['A2'], fo=fS)
    cols = [('Company',15), ('Ticker',10), ('Period',10), ('sj_nm',15),('account_nm',35), ('account_id',40), ('EV_Component',12), ('Amount_100M',15)]
    header_row = 4
    ws_bs.append([]); ws_bs.append([c[0] for c in cols])
    for i, (_, w) in enumerate(cols): ws_bs.column_dimensions[get_column_letter(i+1)].width = w; sc(ws_bs.cell(header_row, i+1), fo=fH, fi=pH, al=aC, bd=BD)
    r = header_row + 1
    if raw_bs_rows:
        for rd in raw_bs_rows:
            ev_comp = rd['EV_Component']; is_hl = bool(ev_comp)
            fill_key = 'Equity' if ev_comp in ['Equity_P', 'Equity_Total'] else ev_comp
            row_fi = ev_fills.get(fill_key, pW) if is_hl else pW; row_fo = fHL if is_hl else fA
            vals = [rd['Company'], rd['Ticker'], rd['Period'], rd['sj_nm'],rd['account_nm'], rd['account_id'], rd['EV_Component'], rd['Amount_100M']]
            for i, v in enumerate(vals): sc(ws_bs.cell(r, i+1), fo=row_fo, fi=row_fi, al=aR if i==7 else aL, nf=NB if i==7 else None, bd=BD); ws_bs.cell(r, i+1).value = v
            r += 1
    ws_bs.auto_filter.ref = f"A{header_row}:H{r-1}"; ws_bs.freeze_panes = f"A{header_row+1}"

    # Sheet 2: PL_Data
    ws_pl = wb.create_sheet('PL_Data')
    ws_pl.merge_cells('A1:H1'); ws_pl['A1'] = "PL_Data (Income Statement Core Only)"; sc(ws_pl['A1'], fo=fT)
    ws_pl.merge_cells('A2:H2'); ws_pl['A2'] = "Logic: IS 추출 후 매출/영업이익/순이익 3개 계정만 엄격 추출 | Unit: 억원"; sc(ws_pl['A2'], fo=fS)
    cols = [('Company',15), ('Ticker',10), ('Period',10), ('Role',15),('PL_Source',16), ('account_nm',35), ('calc_key',12), ('Amount_100M',15)]
    header_row = 4
    ws_pl.append([]); ws_pl.append([c[0] for c in cols])
    for i, (_, w) in enumerate(cols): ws_pl.column_dimensions[get_column_letter(i+1)].width = w; sc(ws_pl.cell(header_row, i+1), fo=fH, fi=pH, al=aC, bd=BD)
    r = header_row + 1
    if raw_pl_rows:
        for rd in raw_pl_rows:
            vals = [rd['Company'], rd['Ticker'], rd['Period'], rd['Role'],rd['PL_Source'], rd['account_nm'], rd['calc_key'], rd['Amount_100M']]
            for i, v in enumerate(vals): sc(ws_pl.cell(r, i+1), fo=fHL, fi=ev_fills['PL_HL'], al=aR if i==7 else aL, nf=NB if i==7 else None, bd=BD); ws_pl.cell(r, i+1).value = v
            r += 1
    ws_pl.auto_filter.ref = f"A{header_row}:H{r-1}"; ws_pl.freeze_panes = f"A{header_row+1}"

    # Sheet 3: Market_Cap
    ws_mc = wb.create_sheet('Market_Cap')
    ws_mc.merge_cells('A1:M1'); ws_mc['A1'] = "Market_Cap (Price & Shares)"; sc(ws_mc['A1'], fo=fT)
    ws_mc.merge_cells('A2:M2'); ws_mc['A2'] = "Logic: 종가(FDR) × 유통주식수(DART) | Unit: 억원"; sc(ws_mc['A2'], fo=fS)
    cols = [('Company',15), ('Ticker',10), ('Period',10), ('Price_Date',12), ('Close',12),('Shares',16), ('Market_Cap_100M',18),('Shares_Source',12), ('Shares_RcpNo',16), ('Shares_StlmDt',12), ('Shares_Se',10),('DART_Status',10), ('DART_Message',40)]
    header_row = 4
    ws_mc.append([]); ws_mc.append([c[0] for c in cols])
    for i, (_, w) in enumerate(cols): ws_mc.column_dimensions[get_column_letter(i+1)].width = w; sc(ws_mc.cell(header_row, i+1), fo=fH, fi=pH, al=aC, bd=BD)
    r = header_row + 1
    if all_mkt:
        for rd in all_mkt:
            vals = [rd.get('Company'), rd.get('Ticker'), rd.get('Period'), rd.get('Price_Date'), rd.get('Close'),rd.get('Outstanding_Shares'), rd.get('Market_Cap_100M'),rd.get('Shares_Source'), rd.get('Shares_RceptNo'), rd.get('Shares_StlmDt'), rd.get('Shares_Se'),rd.get('DART_Status'), rd.get('DART_Message')]
            for i, v in enumerate(vals):
                c = ws_mc.cell(r, i+1); c.value = v
                nf = NP if i==4 else (NI_FMT if i==5 else (NB1 if i==6 else None)); al = aR if i in [4,5,6] else aL
                sc(c, fo=fA, fi=pW, al=al, nf=nf, bd=BD)
            r += 1
    ws_mc.auto_filter.ref = f"A{header_row}:M{r-1}"; ws_mc.freeze_panes = f"A{header_row+1}"

    # Sheet 4: LTM_Calc
    ws_ltm = wb.create_sheet('LTM_Calc')
    ws_ltm.merge_cells('A1:I1'); ws_ltm['A1'] = "LTM_Calc (Revenue/EBIT/NI/Pretax Inc)"; sc(ws_ltm['A1'], fo=fT)
    ws_ltm.merge_cells('A2:I2'); ws_ltm['A2'] = "모든 선택 기간별 LTM 계산 내역 | Unit: 억원"; sc(ws_ltm['A2'], fo=fS)
    cols = [('Company',15), ('Ticker',10), ('Period',10), ('Calc_Key',12),('Current_Cum(A)',15), ('Prior_Annual(B)',15), ('Prior_SameQ(C)',15), ('LTM_Value',15), ('Note',10)]
    header_row = 4
    ws_ltm.append([]); ws_ltm.append([c[0] for c in cols])
    for i, (_, w) in enumerate(cols): ws_ltm.column_dimensions[get_column_letter(i+1)].width = w; sc(ws_ltm.cell(header_row, i+1), fo=fH, fi=pH, al=aC, bd=BD)
    r = header_row + 1
    ltm_keys = ['Revenue', 'EBIT', 'NI', 'Pretax_Income']
    for ticker in target_code_list:
        comp_name = ticker_to_name.get(ticker, ticker)
        for tp in target_periods:
            qtr_suffix = tp.split('.')[-1] if '.' in tp else '4Q'
            for k in ltm_keys:
                ws_ltm.cell(r, 1, comp_name); sc(ws_ltm.cell(r, 1), fo=fA, fi=pW, al=aL, bd=BD)
                ws_ltm.cell(r, 2, ticker);    sc(ws_ltm.cell(r, 2), fo=fA, fi=pW, al=aL, bd=BD)
                ws_ltm.cell(r, 3, tp);        sc(ws_ltm.cell(r, 3), fo=fA, fi=pW, al=aL, bd=BD)
                ws_ltm.cell(r, 4, k);         sc(ws_ltm.cell(r, 4), fo=fA, fi=pW, al=aL, bd=BD)
                # Formula: SUMIFS sum_range, r1, criteria1, r2, criteria2...
                ws_ltm.cell(r, 5).value = f'=SUMIFS(PL_Data!H:H, PL_Data!B:B, B{r}, PL_Data!C:C, C{r}, PL_Data!G:G, D{r}, PL_Data!D:D, "current_cum")'; sc(ws_ltm.cell(r,5), fo=fLINK, fi=pW, nf=NB, bd=BD)
                ws_ltm.cell(r, 6).value = f'=SUMIFS(PL_Data!H:H, PL_Data!B:B, B{r}, PL_Data!C:C, C{r}, PL_Data!G:G, D{r}, PL_Data!D:D, "prior_annual")'; sc(ws_ltm.cell(r,6), fo=fLINK, fi=pW, nf=NB, bd=BD)
                ws_ltm.cell(r, 7).value = f'=SUMIFS(PL_Data!H:H, PL_Data!B:B, B{r}, PL_Data!C:C, C{r}, PL_Data!G:G, D{r}, PL_Data!D:D, "prior_same_q")'; sc(ws_ltm.cell(r,7), fo=fLINK, fi=pW, nf=NB, bd=BD)
                if qtr_suffix == '4Q':
                    ws_ltm.cell(r, 8).value = f'=SUMIFS(PL_Data!H:H, PL_Data!B:B, B{r}, PL_Data!C:C, C{r}, PL_Data!G:G, D{r}, PL_Data!D:D, "annual")'; note = 'Annual'
                else:
                    ws_ltm.cell(r, 8).value = f'=E{r}+F{r}-G{r}'; note = 'A+B-C'
                sc(ws_ltm.cell(r,8), fo=fFRM, fi=pW, nf=NB, bd=BD); ws_ltm.cell(r, 9).value = note; sc(ws_ltm.cell(r,9), fo=fA, fi=pW, al=aC, bd=BD)
                r += 1
    ws_ltm.auto_filter.ref = f"A{header_row}:I{r-1}"; ws_ltm.freeze_panes = f"A{header_row+1}"

    # Sheet 3.5: Beta_Calculation
    ws_beta = wb.create_sheet('Beta_Calculation')
    ws_beta.merge_cells('A1:F1')
    sc(ws_beta['A1'], fo=Font(name='Arial', bold=True, size=14, color=C_BL))
    ws_beta['A1'] = 'Beta Calculation (Excel Formulas)'

    ws_beta.merge_cells('A2:F2')
    sc(ws_beta['A2'], fo=Font(name='Arial', size=9, color=C_MG, italic=True))
    ws_beta['A2'] = f'5-Year Monthly & 2-Year Weekly Returns | Base: {base_period_str}'

    r_beta = 4
    beta_result_rows = {}  # ticker: (raw_5y, adj_5y, raw_2y, adj_2y) 매핑

    for idx, ticker in enumerate(target_code_list):
        comp_data = next((item for item in screen_summary_data if item["Ticker"] == ticker), None)
        if not comp_data:
            continue

        company_name = comp_data['Company']
        market_idx = comp_data['Market_Index']

        # ========== 5Y Monthly Beta Section ==========
        ws_beta.merge_cells(f'A{r_beta}:F{r_beta}')
        sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
           fi=PatternFill('solid', fgColor='607D8B'), al=Alignment(horizontal='center'))
        ws_beta.cell(r_beta, 1, f'{company_name} ({ticker}) vs {market_idx} - 5Y Monthly')
        r_beta += 1

        stock_prices_5y = comp_data.get('Stock_Monthly_Prices_5Y')
        market_prices_5y = comp_data.get('Market_Monthly_Prices_5Y')
        raw_5y_row = None
        adj_5y_row = None

        if stock_prices_5y is not None and market_prices_5y is not None and not stock_prices_5y.empty and not market_prices_5y.empty:
            # 헤더
            ws_beta.cell(r_beta, 1, 'Date')
            ws_beta.cell(r_beta, 2, f'{ticker} Price')
            ws_beta.cell(r_beta, 3, f'{market_idx} Price')
            ws_beta.cell(r_beta, 4, f'{ticker} Return')
            ws_beta.cell(r_beta, 5, f'{market_idx} Return')
            for col in range(1, 6):
                sc(ws_beta.cell(r_beta, col), fo=Font(name='Arial', bold=True, size=9, color=C_W),
                   fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
            r_beta += 1

            data_start_row = r_beta

            # 공통 날짜 인덱스
            common_dates = stock_prices_5y.index.intersection(market_prices_5y.index)

            # 데이터 행 작성
            for date in common_dates:
                ws_beta.cell(r_beta, 1, date.strftime('%Y-%m'))
                ws_beta.cell(r_beta, 2, float(stock_prices_5y.loc[date]))
                ws_beta.cell(r_beta, 3, float(market_prices_5y.loc[date]))

                # 수익률 계산 (엑셀 수식)
                if r_beta > data_start_row:
                    ws_beta.cell(r_beta, 4).value = f'=(B{r_beta}-B{r_beta-1})/B{r_beta-1}'
                    ws_beta.cell(r_beta, 5).value = f'=(C{r_beta}-C{r_beta-1})/C{r_beta-1}'
                else:
                    ws_beta.cell(r_beta, 4, None)
                    ws_beta.cell(r_beta, 5, None)

                # 스타일
                sc(ws_beta.cell(r_beta, 1), fo=fA, al=aC, bd=BD)
                sc(ws_beta.cell(r_beta, 2), fo=fA, al=aR, bd=BD, nf='#,##0.00')
                sc(ws_beta.cell(r_beta, 3), fo=fA, al=aR, bd=BD, nf='#,##0.00')
                sc(ws_beta.cell(r_beta, 4), fo=fA, al=aR, bd=BD, nf='0.00%')
                sc(ws_beta.cell(r_beta, 5), fo=fA, al=aR, bd=BD, nf='0.00%')

                r_beta += 1

            data_end_row = r_beta - 1

            # 베타 계산 (SLOPE 함수)
            r_beta += 1
            ws_beta.cell(r_beta, 1, 'Raw Beta (5Y Monthly)')
            ws_beta.cell(r_beta, 2).value = f'=SLOPE(D{data_start_row+1}:D{data_end_row},E{data_start_row+1}:E{data_end_row})'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', bold=True, size=9), fi=PatternFill('solid', fgColor='E8F5E9'),
               bd=BD, al=aR, nf='0.0000')
            raw_5y_row = r_beta
            r_beta += 1

            # Adjusted Beta
            ws_beta.cell(r_beta, 1, 'Adjusted Beta (5Y)')
            ws_beta.cell(r_beta, 2).value = f'=2/3*B{r_beta-1}+1/3*1'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', bold=True, size=9), fi=PatternFill('solid', fgColor='E8F5E9'),
               bd=BD, al=aR, nf='0.0000')
            adj_5y_row = r_beta

            # 신뢰도 — 베타를 그대로 써도 되는지 판단할 근거.
            # R²: 주가 변동 중 시장으로 설명되는 비중. 낮으면 그 기울기는 관계가
            #     아니라 흩어진 점에 그은 선이다. n: 회귀에 실제로 들어간 관측치 수.
            r_beta += 1
            ws_beta.cell(r_beta, 1, 'R² (시장 설명력)')
            ws_beta.cell(r_beta, 2).value = f'=RSQ(D{data_start_row+1}:D{data_end_row},E{data_start_row+1}:E{data_end_row})'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', size=9), bd=BD, al=aR, nf='0.000')
            r_beta += 1
            ws_beta.cell(r_beta, 1, '관측치 수 n')
            ws_beta.cell(r_beta, 2).value = f'=COUNT(D{data_start_row+1}:D{data_end_row})'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', size=9), bd=BD, al=aR, nf='#,##0')

        else:
            ws_beta.cell(r_beta, 1, 'No 5Y price data available')
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', size=9, color='FF0000'))

        r_beta += 2  # 간격

        # ========== 2Y Weekly Beta Section ==========
        ws_beta.merge_cells(f'A{r_beta}:F{r_beta}')
        sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
           fi=PatternFill('solid', fgColor='455A64'), al=Alignment(horizontal='center'))
        ws_beta.cell(r_beta, 1, f'{company_name} ({ticker}) vs {market_idx} - 2Y Weekly')
        r_beta += 1

        stock_prices_2y = comp_data.get('Stock_Weekly_Prices_2Y')
        market_prices_2y = comp_data.get('Market_Weekly_Prices_2Y')
        raw_2y_row = None
        adj_2y_row = None

        if stock_prices_2y is not None and market_prices_2y is not None and not stock_prices_2y.empty and not market_prices_2y.empty:
            # 헤더
            ws_beta.cell(r_beta, 1, 'Date')
            ws_beta.cell(r_beta, 2, f'{ticker} Price')
            ws_beta.cell(r_beta, 3, f'{market_idx} Price')
            ws_beta.cell(r_beta, 4, f'{ticker} Return')
            ws_beta.cell(r_beta, 5, f'{market_idx} Return')
            for col in range(1, 6):
                sc(ws_beta.cell(r_beta, col), fo=Font(name='Arial', bold=True, size=9, color=C_W),
                   fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
            r_beta += 1

            data_start_row = r_beta

            # 공통 날짜 인덱스
            common_dates = stock_prices_2y.index.intersection(market_prices_2y.index)

            # 데이터 행 작성
            for date in common_dates:
                ws_beta.cell(r_beta, 1, date.strftime('%Y-%m-%d'))
                ws_beta.cell(r_beta, 2, float(stock_prices_2y.loc[date]))
                ws_beta.cell(r_beta, 3, float(market_prices_2y.loc[date]))

                # 수익률 계산 (엑셀 수식)
                if r_beta > data_start_row:
                    ws_beta.cell(r_beta, 4).value = f'=(B{r_beta}-B{r_beta-1})/B{r_beta-1}'
                    ws_beta.cell(r_beta, 5).value = f'=(C{r_beta}-C{r_beta-1})/C{r_beta-1}'
                else:
                    ws_beta.cell(r_beta, 4, None)
                    ws_beta.cell(r_beta, 5, None)

                # 스타일
                sc(ws_beta.cell(r_beta, 1), fo=fA, al=aC, bd=BD)
                sc(ws_beta.cell(r_beta, 2), fo=fA, al=aR, bd=BD, nf='#,##0.00')
                sc(ws_beta.cell(r_beta, 3), fo=fA, al=aR, bd=BD, nf='#,##0.00')
                sc(ws_beta.cell(r_beta, 4), fo=fA, al=aR, bd=BD, nf='0.00%')
                sc(ws_beta.cell(r_beta, 5), fo=fA, al=aR, bd=BD, nf='0.00%')

                r_beta += 1

            data_end_row = r_beta - 1

            # 베타 계산 (SLOPE 함수)
            r_beta += 1
            ws_beta.cell(r_beta, 1, 'Raw Beta (2Y Weekly)')
            ws_beta.cell(r_beta, 2).value = f'=SLOPE(D{data_start_row+1}:D{data_end_row},E{data_start_row+1}:E{data_end_row})'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', bold=True, size=9), fi=PatternFill('solid', fgColor='FFF9C4'),
               bd=BD, al=aR, nf='0.0000')
            raw_2y_row = r_beta
            r_beta += 1

            # Adjusted Beta
            ws_beta.cell(r_beta, 1, 'Adjusted Beta (2Y)')
            ws_beta.cell(r_beta, 2).value = f'=2/3*B{r_beta-1}+1/3*1'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', bold=True, size=9), fi=PatternFill('solid', fgColor='FFF9C4'),
               bd=BD, al=aR, nf='0.0000')
            adj_2y_row = r_beta

            # 신뢰도 — 베타를 그대로 써도 되는지 판단할 근거.
            # R²: 주가 변동 중 시장으로 설명되는 비중. 낮으면 그 기울기는 관계가
            #     아니라 흩어진 점에 그은 선이다. n: 회귀에 실제로 들어간 관측치 수.
            r_beta += 1
            ws_beta.cell(r_beta, 1, 'R² (시장 설명력)')
            ws_beta.cell(r_beta, 2).value = f'=RSQ(D{data_start_row+1}:D{data_end_row},E{data_start_row+1}:E{data_end_row})'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', size=9), bd=BD, al=aR, nf='0.000')
            r_beta += 1
            ws_beta.cell(r_beta, 1, '관측치 수 n')
            ws_beta.cell(r_beta, 2).value = f'=COUNT(D{data_start_row+1}:D{data_end_row})'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', size=9), bd=BD, al=aR, nf='#,##0')

        else:
            ws_beta.cell(r_beta, 1, 'No 2Y price data available')
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', size=9, color='FF0000'))

        # 결과 저장
        beta_result_rows[ticker] = (raw_5y_row, adj_5y_row, raw_2y_row, adj_2y_row)

        r_beta += 2  # 다음 회사와 간격

    ws_beta.column_dimensions['A'].width = 15
    ws_beta.column_dimensions['B'].width = 15
    ws_beta.column_dimensions['C'].width = 15
    ws_beta.column_dimensions['D'].width = 15
    ws_beta.column_dimensions['E'].width = 15

    ws_beta.freeze_panes = 'A4'

    # Sheet 4: WACC_Calculation (완전 구현 - GPCM.py와 동일)
    ws_wacc = wb.create_sheet('WACC_Calculation')
    ws_wacc.merge_cells('A1:D1')
    sc(ws_wacc['A1'], fo=Font(name='Arial', bold=True, size=14, color=C_BL))
    ws_wacc['A1'] = 'Target WACC Calculation'

    ws_wacc.merge_cells('A2:D2')
    sc(ws_wacc['A2'], fo=Font(name='Arial', size=9, color=C_MG, italic=True))
    ws_wacc['A2'] = f'Base: {base_period_str} | Peer Average Method'

    # 스타일 정의
    C_MB = '005EB8'
    pWACC_PARAM = PatternFill('solid', fgColor='E3F2FD')
    pWACC_CALC = PatternFill('solid', fgColor='FFF9C4')
    pWACC_RESULT = PatternFill('solid', fgColor='FFE082')

    r_wacc = 4

    # Section 1: Input Parameters
    ws_wacc.merge_cells(f'A{r_wacc}:D{r_wacc}')
    sc(ws_wacc.cell(r_wacc, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
       fi=PatternFill('solid', fgColor=C_MB), al=Alignment(horizontal='center'))
    ws_wacc.cell(r_wacc, 1, '[ 1 ] Input Parameters')
    r_wacc += 1

    # 헤더
    ws_wacc['A' + str(r_wacc)] = 'Parameter'
    ws_wacc['B' + str(r_wacc)] = 'Value'
    ws_wacc['C' + str(r_wacc)] = 'Format'
    ws_wacc['D' + str(r_wacc)] = 'Note'
    for col in ['A', 'B', 'C', 'D']:
        sc(ws_wacc[col + str(r_wacc)], fo=Font(name='Arial', bold=True, size=9, color=C_W),
           fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
    r_wacc += 1

    # Calculate GPCM stats row position for formulas
    # DATA_START = 6 (header_row + 1), DATA_END depends on number of companies
    # Mean row = DATA_END + 2
    n_companies = len(target_code_list)
    DATA_START = 6
    DATA_END = 6 + n_companies - 1
    mean_row = DATA_END + 2

    # 데이터 행 - Input Parameters
    wacc_params = [
        ('Risk-Free Rate (Rf)', target_wacc_data['Rf'], f"{target_wacc_data['Rf']*100:.2f}%", '10-year Korea Treasury Yield'),
        ('Market Risk Premium (MRP)', target_wacc_data['MRP'], f"{target_wacc_data['MRP']*100:.1f}%", '한국공인회계사회 기준'),
        ('Size Premium', target_wacc_data['Size_Premium'], f"{target_wacc_data['Size_Premium']*100:.2f}%", '한국공인회계사회 (시가총액 기준)'),
        ('Kd (Pretax)', target_wacc_data['Kd_Pretax'], f"{target_wacc_data['Kd_Pretax']*100:.1f}%", '세전 타인자본비용 (사용자 입력)'),
        ('Target Tax Rate', target_wacc_data['Target_Tax_Rate'], f"{target_wacc_data['Target_Tax_Rate']*100:.1f}%", '한국 법인세 한계세율 (지방소득세 포함)'),
    ]

    for param, value, formatted, note in wacc_params:
        ws_wacc.cell(r_wacc, 1, param)
        ws_wacc.cell(r_wacc, 2, value)
        ws_wacc.cell(r_wacc, 3, formatted)
        ws_wacc.cell(r_wacc, 4, note)
        sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD, al=Alignment(horizontal='left'))
        sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_PARAM, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
        sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
        sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
        r_wacc += 1

    r_wacc += 1

    # Section 2: Peer Analysis
    ws_wacc.merge_cells(f'A{r_wacc}:D{r_wacc}')
    sc(ws_wacc.cell(r_wacc, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
       fi=PatternFill('solid', fgColor=C_MB), al=Alignment(horizontal='center'))
    ws_wacc.cell(r_wacc, 1, '[ 2 ] Peer Analysis')
    r_wacc += 1

    # 헤더
    ws_wacc['A' + str(r_wacc)] = 'Metric'
    ws_wacc['B' + str(r_wacc)] = 'Value'
    ws_wacc['C' + str(r_wacc)] = 'Format'
    ws_wacc['D' + str(r_wacc)] = 'Note'
    for col in ['A', 'B', 'C', 'D']:
        sc(ws_wacc[col + str(r_wacc)], fo=Font(name='Arial', bold=True, size=9, color=C_W),
           fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
    r_wacc += 1

    # Avg Unlevered Beta - 엑셀 수식으로 GPCM 시트 참조
    row_unlevered_beta = r_wacc
    beta_label = "5Y Monthly" if beta_type_input == "5Y" else "2Y Weekly"
    beta_col = GPCM_CL['UB5'] if beta_type_input == "5Y" else GPCM_CL['UB2']
    ws_wacc.cell(r_wacc, 1, f'Avg Unlevered Beta ({beta_label})')
    ws_wacc.cell(r_wacc, 2).value = f'=GPCM!{beta_col}{mean_row}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Avg_Unlevered_Beta']:.4f}")
    ws_wacc.cell(r_wacc, 4, '피어 평균 (GPCM Mean)')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.0000')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Avg Debt Ratio - 엑셀 수식으로 GPCM 시트 참조
    row_debt_ratio = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Avg Debt Ratio (D/V)')
    ws_wacc.cell(r_wacc, 2).value = f'=GPCM!{GPCM_CL["DVRatio"]}{mean_row}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Avg_Debt_Ratio']*100:.1f}%")
    ws_wacc.cell(r_wacc, 4, '피어 평균 자본구조 (GPCM Mean)')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Target D/E Ratio - 엑셀 수식으로 계산
    row_de_ratio = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Target D/E Ratio')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_debt_ratio}/(1-B{row_debt_ratio})'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Target_DE_Ratio']:.4f}")
    ws_wacc.cell(r_wacc, 4, '= D/V ÷ (1 - D/V)')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.0000')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    r_wacc += 1

    # Section 3: Target WACC Calculation
    ws_wacc.merge_cells(f'A{r_wacc}:D{r_wacc}')
    sc(ws_wacc.cell(r_wacc, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
       fi=PatternFill('solid', fgColor=C_MB), al=Alignment(horizontal='center'))
    ws_wacc.cell(r_wacc, 1, '[ 3 ] Target WACC Calculation')
    r_wacc += 1

    # 헤더
    ws_wacc['A' + str(r_wacc)] = 'Component'
    ws_wacc['B' + str(r_wacc)] = 'Value'
    ws_wacc['C' + str(r_wacc)] = 'Format'
    ws_wacc['D' + str(r_wacc)] = 'Formula'
    for col in ['A', 'B', 'C', 'D']:
        sc(ws_wacc[col + str(r_wacc)], fo=Font(name='Arial', bold=True, size=9, color=C_W),
           fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
    r_wacc += 1

    # Row references for formulas
    row_rf = 6
    row_mrp = 7
    row_size_premium = 8
    row_kd_pretax = 9
    row_tax = 10

    # Relevered Beta
    row_relevered_beta = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Relevered Beta')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_unlevered_beta}*(1+(1-B{row_tax})*B{row_de_ratio})'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Target_Relevered_Beta']:.4f}")
    ws_wacc.cell(r_wacc, 4, 'Unlevered β × (1 + (1 - Tax) × D/E)')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.0000')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Ke (Cost of Equity)
    row_ke = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Ke (Cost of Equity)')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_rf}+B{row_mrp}*B{row_relevered_beta}+B{row_size_premium}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Target_Ke']*100:.2f}%")
    ws_wacc.cell(r_wacc, 4, 'Rf + MRP × Relevered β + Size Premium')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Kd (Aftertax)
    row_kd_aftertax = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Kd (Aftertax)')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_kd_pretax}*(1-B{row_tax})'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Kd_Aftertax']*100:.2f}%")
    ws_wacc.cell(r_wacc, 4, 'Kd (Pretax) × (1 - Tax Rate)')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Equity Weight (E/V)
    row_equity_weight = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Equity Weight (E/V)')
    ws_wacc.cell(r_wacc, 2).value = f'=1-B{row_debt_ratio}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Equity_Weight']*100:.1f}%")
    ws_wacc.cell(r_wacc, 4, '1 - Debt Ratio')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Debt Weight (D/V)
    row_debt_weight = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Debt Weight (D/V)')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_debt_ratio}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Debt_Weight']*100:.1f}%")
    ws_wacc.cell(r_wacc, 4, 'Debt Ratio')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # 구분선
    ws_wacc.cell(r_wacc, 1, '━━━━━━━━━━━━')
    ws_wacc.cell(r_wacc, 2, None)
    ws_wacc.cell(r_wacc, 3, '━━━━━━━━━━━━')
    ws_wacc.cell(r_wacc, 4, '━━━━━━━━━━━━━━━━━━━━━━━━━━━')
    for col_idx in range(1, 5):
        sc(ws_wacc.cell(r_wacc, col_idx), bd=BD)
    r_wacc += 1

    # WACC (최종 결과)
    row_wacc_final = r_wacc
    ws_wacc.cell(r_wacc, 1, 'WACC')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_equity_weight}*B{row_ke}+B{row_debt_weight}*B{row_kd_aftertax}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Target_WACC']*100:.2f}%")
    ws_wacc.cell(r_wacc, 4, '(E/V) × Ke + (D/V) × Kd (Aftertax)')
    sc(ws_wacc.cell(r_wacc, 1), fo=Font(name='Arial', bold=True, size=10), bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=Font(name='Arial', bold=True, size=10), fi=pWACC_RESULT,
       bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=Font(name='Arial', bold=True, size=10), bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG, italic=True), bd=BD)
    r_wacc += 1

    # 열 너비 조정
    ws_wacc.column_dimensions['A'].width = 25
    ws_wacc.column_dimensions['B'].width = 12
    ws_wacc.column_dimensions['C'].width = 15
    ws_wacc.column_dimensions['D'].width = 40

    ws_wacc.freeze_panes = 'A4'

    # Named Range 정의 (다른 시트에서 참조 가능)
    wb.defined_names['Target_WACC'] = DefinedName('Target_WACC', attr_text=f"'WACC_Calculation'!$B${row_wacc_final}")
    wb.defined_names['Target_Rf'] = DefinedName('Target_Rf', attr_text="'WACC_Calculation'!$B$6")
    wb.defined_names['Target_MRP'] = DefinedName('Target_MRP', attr_text="'WACC_Calculation'!$B$7")
    wb.defined_names['Target_Size_Premium'] = DefinedName('Target_Size_Premium', attr_text="'WACC_Calculation'!$B$8")
    wb.defined_names['Target_Kd_Pretax'] = DefinedName('Target_Kd_Pretax', attr_text="'WACC_Calculation'!$B$9")
    wb.defined_names['Target_Tax_Rate'] = DefinedName('Target_Tax_Rate', attr_text="'WACC_Calculation'!$B$10")

    # 참고용 셀 주소 표시
    ws_wacc['A' + str(r_wacc + 2)] = '[ Named Ranges for Reference ]'
    sc(ws_wacc.cell(r_wacc + 2, 1), fo=Font(name='Arial', bold=True, size=9, color=C_MG, italic=True))
    ws_wacc['A' + str(r_wacc + 3)] = '다른 시트에서 참조: =Target_WACC, =Target_Rf 등'
    sc(ws_wacc.cell(r_wacc + 3, 1), fo=Font(name='Arial', size=8, color=C_MG))

    # Sheet 1: GPCM (맨 앞)
    ws = wb.create_sheet('GPCM')
    wb.move_sheet('GPCM', offset=-6)  # 맨 앞으로 이동 (index 0)
    # 시트 순서: GPCM, WACC_Calculation, Beta_Calculation, BS_Full, PL_Data, Market_Cap, LTM_Calc
    wb.move_sheet('WACC_Calculation', offset=-4)  # GPCM 다음 (index 1)
    wb.move_sheet('Beta_Calculation', offset=-3)  # WACC 다음 (index 2)
    # 열 배치는 모듈 상단 GPCM_COL_DEFS 가 정한다 — 여기서는 이름으로만 참조한다.
    C, L = GPCM_CI, GPCM_CL
    TOTAL_COLS = len(GPCM_COL_DEFS)
    ws.merge_cells(f'A1:{get_column_letter(TOTAL_COLS)}1'); ws['A1'] = "GPCM Valuation Summary with Beta Analysis"; sc(ws['A1'], fo=fT)
    ws.merge_cells(f'A2:{get_column_letter(TOTAL_COLS)}2'); ws['A2'] = f"Base: {base_period_str} | Unit: 억원 | EV = MCap + 우선주 + IBD − Cash + NCI − NOA"; sc(ws['A2'], fo=fS)
    add_gpcm_section_row(ws)
    header_row = 5
    for i, (key, header, width) in enumerate(GPCM_COL_DEFS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        sc(ws.cell(header_row, i, header), fo=fH, fi=pH, al=aC, bd=BD)

    pMULT = PatternFill('solid', fgColor=C_PB)
    pBETA = PatternFill('solid', fgColor='E8F5E9')
    pBETA2 = PatternFill('solid', fgColor='FFF9C4')
    NF_BETA = '0.00;(0.00);"-"'
    NF_PCT = '0.0%;(0.0%);"-"'

    r = header_row + 1
    for ticker in target_code_list:
        comp_name = ticker_to_name.get(ticker, ticker); bg = pST if (r % 2 == 0) else pW
        ws.cell(r, C['Company'], comp_name); ws.cell(r, C['Ticker'], ticker); ws.cell(r, C['BaseDate'], base_period_str); ws.cell(r, C['Curr'], 'KRW'); ws.cell(r, C['PLSource'], 'LTM')
        for c in range(C['Company'], C['PLSource'] + 1): sc(ws.cell(r,c), fo=fA, fi=bg, al=aL, bd=BD)
        ws.cell(r, C['Cash']).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "Cash")'; sc(ws.cell(r, C['Cash']), fo=fLINK, fi=ev_fills['Cash'], nf=NB, bd=BD)
        ws.cell(r, C['IBD']).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "IBD")'; sc(ws.cell(r, C['IBD']), fo=fLINK, fi=ev_fills['IBD'], nf=NB, bd=BD)
        ws.cell(r, C['NOA']).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "NOA")'; sc(ws.cell(r, C['NOA']), fo=fLINK, fi=ev_fills['NOA'], nf=NB, bd=BD)
        ws.cell(r, C['NetDebt']).value = f'={L["IBD"]}{r}-{L["Cash"]}{r}-{L["NOA"]}{r}'; sc(ws.cell(r, C['NetDebt']), fo=fFRM, fi=bg, nf=NB, bd=BD)
        ws.cell(r, C['NCI']).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "NCI")'; sc(ws.cell(r, C['NCI']), fo=fLINK, fi=ev_fills['NCI'], nf=NB, bd=BD)
        # 우선주 자본금: 시가총액은 보통주만 반영하므로 자기자본가치에 별도 가산 (EV 구성요소 옆에 둔다)
        ws.cell(r, C['Pref']).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "Preferred")'; sc(ws.cell(r, C['Pref']), fo=fLINK, fi=ev_fills['Preferred'], nf=NB, bd=BD)
        ws.cell(r, C['Equity']).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "Equity_Total")'; sc(ws.cell(r, C['Equity']), fo=fLINK, fi=ev_fills['Equity'], nf=NB, bd=BD)
        ws.cell(r, C['EV']).value = f'={L["MktCap"]}{r}+{L["Pref"]}{r}+{L["IBD"]}{r}-{L["Cash"]}{r}+{L["NCI"]}{r}-{L["NOA"]}{r}'; sc(ws.cell(r, C['EV']), fo=fFRM, fi=bg, nf=NB, bd=BD)
        ws.cell(r, C['Revenue']).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{r}, LTM_Calc!C:C, C{r}, LTM_Calc!D:D, "Revenue")'; sc(ws.cell(r, C['Revenue']), fo=fLINK, fi=ev_fills['PL_HL'], nf=NB, bd=BD)
        ws.cell(r, C['EBIT']).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{r}, LTM_Calc!C:C, C{r}, LTM_Calc!D:D, "EBIT")'; sc(ws.cell(r, C['EBIT']), fo=fLINK, fi=ev_fills['PL_HL'], nf=NB, bd=BD)
        sc(ws.cell(r, C['DA']), fi=PatternFill('solid', fgColor='FFFF00'), nf=NB, bd=BD) # D&A 수기
        ws.cell(r, C['EBITDA']).value = f'={L["EBIT"]}{r}+{L["DA"]}{r}'; sc(ws.cell(r, C['EBITDA']), fo=fFRM, fi=bg, nf=NB, bd=BD)
        ws.cell(r, C['NI']).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{r}, LTM_Calc!C:C, C{r}, LTM_Calc!D:D, "NI")'; sc(ws.cell(r, C['NI']), fo=fLINK, fi=ev_fills['PL_HL'], nf=NB, bd=BD)
        ws.cell(r, C['Price']).value = f'=SUMIFS(Market_Cap!E:E, Market_Cap!B:B, B{r}, Market_Cap!C:C, C{r})'; sc(ws.cell(r, C['Price']), fo=fLINK, nf=NP, bd=BD)
        ws.cell(r, C['Shares']).value = f'=SUMIFS(Market_Cap!F:F, Market_Cap!B:B, B{r}, Market_Cap!C:C, C{r})'; sc(ws.cell(r, C['Shares']), fo=fLINK, nf=NI_FMT, bd=BD)
        ws.cell(r, C['MktCap']).value = f'=SUMIFS(Market_Cap!G:G, Market_Cap!B:B, B{r}, Market_Cap!C:C, C{r})'; sc(ws.cell(r, C['MktCap']), fo=fLINK, nf=NB1, bd=BD)
        ws.cell(r, C['EVEBITDA']).value = f'=IF({L["EBITDA"]}{r}>0, {L["EV"]}{r}/{L["EBITDA"]}{r}, "N/M")'; sc(ws.cell(r, C['EVEBITDA']), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
        ws.cell(r, C['EVEBIT']).value = f'=IF({L["EBIT"]}{r}>0, {L["EV"]}{r}/{L["EBIT"]}{r}, "N/M")'; sc(ws.cell(r, C['EVEBIT']), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
        ws.cell(r, C['PER']).value = f'=IF({L["NI"]}{r}>0, {L["MktCap"]}{r}/{L["NI"]}{r}, "N/M")'; sc(ws.cell(r, C['PER']), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
        ws.cell(r, C['PBR']).value = f'=IF({L["Equity"]}{r}>0, {L["MktCap"]}{r}/{L["Equity"]}{r}, "N/M")'; sc(ws.cell(r, C['PBR']), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
        ws.cell(r, C['PSR']).value = f'=IF({L["Revenue"]}{r}>0, {L["MktCap"]}{r}/{L["Revenue"]}{r}, "N/M")'; sc(ws.cell(r, C['PSR']), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)

        # Beta 값은 Beta_Calculation 시트에서 참조
        beta_rows = beta_result_rows.get(ticker, (None, None, None, None))
        for key, row_idx, fill in (('B5Raw', beta_rows[0], pBETA), ('B5Adj', beta_rows[1], pBETA),
                                   ('B2Raw', beta_rows[2], pBETA2), ('B2Adj', beta_rows[3], pBETA2)):
            if row_idx:
                ws.cell(r, C[key]).value = f'=Beta_Calculation!B{row_idx}'
                sc(ws.cell(r, C[key]), fo=fLINK, fi=fill, al=aR, nf=NF_BETA, bd=BD)
            else:
                ws.cell(r, C[key], ''); sc(ws.cell(r, C[key]), fo=fA, fi=fill, al=aR, nf=NF_BETA, bd=BD)

        ws.cell(r, C['Pretax']).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{r}, LTM_Calc!C:C, C{r}, LTM_Calc!D:D, "Pretax_Income")'; sc(ws.cell(r, C['Pretax']), fo=fLINK, fi=bg, al=aR, nf=NB, bd=BD)

        # Tax Rate (한국 법인세 한계세율, 사업연도별 세율표, 지방소득세 포함)
        # 세율표는 파이썬과 같은 출처(get_korean_tax_brackets)에서 만든다 —
        # 수식에 박아두면 세법이 바뀔 때 엑셀만 옛 세율로 남는다.
        ws.cell(r, C['TaxRate']).value = korean_tax_rate_formula(f'{L["Pretax"]}{r}', base_fiscal_year)
        sc(ws.cell(r, C['TaxRate']), fo=fFRM, fi=bg, al=aR, nf=NF_PCT, bd=BD)

        # D/E Ratio = IBD / (Mkt Cap + 우선주 + NCI)
        ws.cell(r, C['DERatio']).value = f'=IF({L["MktCap"]}{r}+{L["Pref"]}{r}+{L["NCI"]}{r}>0, {L["IBD"]}{r}/({L["MktCap"]}{r}+{L["Pref"]}{r}+{L["NCI"]}{r}), 0)'; sc(ws.cell(r, C['DERatio']), fo=fFRM, fi=bg, al=aR, nf=NF_PCT, bd=BD)

        # Debt Ratio (D/V) = IBD / (Mkt Cap + 우선주 + IBD + NCI)
        ws.cell(r, C['DVRatio']).value = f'=IF({L["MktCap"]}{r}+{L["Pref"]}{r}+{L["IBD"]}{r}+{L["NCI"]}{r}>0, {L["IBD"]}{r}/({L["MktCap"]}{r}+{L["Pref"]}{r}+{L["IBD"]}{r}+{L["NCI"]}{r}), 0)'; sc(ws.cell(r, C['DVRatio']), fo=fFRM, fi=bg, al=aR, nf=NF_PCT, bd=BD)

        # Unlevered Beta = Adj Beta / (1 + (1 - Tax Rate) × D/E Ratio)
        ws.cell(r, C['UB5']).value = f'=IF({L["B5Adj"]}{r}>0, {L["B5Adj"]}{r}/(1+(1-{L["TaxRate"]}{r})*{L["DERatio"]}{r}), "")'; sc(ws.cell(r, C['UB5']), fo=fFRM, fi=pBETA, al=aR, nf=NF_BETA, bd=BD)
        ws.cell(r, C['UB2']).value = f'=IF({L["B2Adj"]}{r}>0, {L["B2Adj"]}{r}/(1+(1-{L["TaxRate"]}{r})*{L["DERatio"]}{r}), "")'; sc(ws.cell(r, C['UB2']), fo=fFRM, fi=pBETA2, al=aR, nf=NF_BETA, bd=BD)
        r += 1
    r_end = r - 1; r += 1
    beta_stat_cols = {C[k] for k in ('B5Raw', 'B5Adj', 'B2Raw', 'B2Adj', 'UB5', 'UB2')}
    pct_stat_cols = {C[k] for k in ('TaxRate', 'DERatio', 'DVRatio')}
    for stat, fn in [('Mean','AVERAGE'), ('Median','MEDIAN'), ('Max','MAX'), ('Min','MIN')]:
        ws.cell(r, C['MktCap'], stat); sc(ws.cell(r, C['MktCap']), fo=fSTAT, fi=pSTAT, al=aC, bd=BD)
        for c in range(C['EVEBITDA'], C['PSR'] + 1):        # Valuation Multiples
            col = get_column_letter(c)
            ws.cell(r, c).value = f'=IFERROR({fn}({col}{header_row+1}:{col}{r_end}), "N/M")'
            sc(ws.cell(r,c), fo=fSTAT, fi=pSTAT, nf=NF_X, bd=BD)
        for c in range(C['B5Raw'], C['UB2'] + 1):           # Beta & Risk
            col = get_column_letter(c)
            nf = NF_BETA if c in beta_stat_cols else (NF_PCT if c in pct_stat_cols else NB)
            ws.cell(r, c).value = f'=IFERROR({fn}({col}{header_row+1}:{col}{r_end}), "")'
            sc(ws.cell(r,c), fo=fSTAT, fi=pSTAT, nf=nf, bd=BD)
        r += 1
    r += 2
    for note in notes_list: ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS); sc(ws.cell(r, 1, note), fo=fNOTE); r += 1
    ws.freeze_panes = f"{GPCM_CL['Cash']}{header_row+1}"  # Cash 컬럼부터 스크롤

    # === Multiples_Trend Sheet generation ===
    ws_trend = wb.create_sheet('Multiples_Trend')
    ws_trend.merge_cells('A1:M1'); ws_trend['A1'] = "Multiples Trend (PER, PBR, PSR, EV/EBIT)"; sc(ws_trend['A1'], fo=fT)
    ws_trend.merge_cells('A2:M2'); ws_trend['A2'] = "모든 타겟 기간별 Valuation Multiples 흐름 요약 (Formula 기반)"; sc(ws_trend['A2'], fo=fS)
    
    # 0:Comp, 1:Tick, 2:Per, 3:MC, 4:EV, 5:Rev, 6:EBIT, 7:NI, 8:Eq, 9:EV/EB, 10:PER, 11:PSR, 12:PBR
    cols_t = [('Company',15), ('Ticker',10), ('Period',10), ('Market_Cap',15), ('EV',15), ('Revenue(LTM)',15), ('EBIT(LTM)',15), ('NI(LTM)',15), ('Equity', 15), ('EV/EBIT',12), ('PER',10), ('PSR',10), ('PBR', 10)]
    header_row_t = 4
    ws_trend.append([]); ws_trend.append([c[0] for c in cols_t])
    for i, (_, w) in enumerate(cols_t): ws_trend.column_dimensions[get_column_letter(i+1)].width = w; sc(ws_trend.cell(header_row_t, i+1), fo=fH, fi=pH, al=aC, bd=BD)
    
    rt = header_row_t + 1
    if df_screen is not None and not df_screen.empty:
        # Ticker x Period 순회 (df_screen 기반)
        for _, row_data in df_screen.iterrows():
            ticker = row_data.get('Ticker')
            period = row_data.get('Period')
            comp_name = row_data.get('Company')
            
            # Static basic info
            ws_trend.cell(rt, 1, comp_name); sc(ws_trend.cell(rt, 1), fo=fA, fi=pW, al=aL, bd=BD)
            ws_trend.cell(rt, 2, ticker);    sc(ws_trend.cell(rt, 2), fo=fA, fi=pW, al=aL, bd=BD)
            ws_trend.cell(rt, 3, period);    sc(ws_trend.cell(rt, 3), fo=fA, fi=pW, al=aL, bd=BD)
            
            # MC & EV from Market_Cap sheet
            ws_trend.cell(rt, 4).value = f'=SUMIFS(Market_Cap!G:G, Market_Cap!B:B, B{rt}, Market_Cap!C:C, C{rt})'; sc(ws_trend.cell(rt, 4), fo=fLINK, nf=NB, bd=BD)
            ws_trend.cell(rt, 5).value = f'=D{rt}+SUMIFS(BS_Full!H:H, BS_Full!B:B, B{rt}, BS_Full!C:C, C{rt}, BS_Full!G:G, "IBD")-SUMIFS(BS_Full!H:H, BS_Full!B:B, B{rt}, BS_Full!C:C, C{rt}, BS_Full!G:G, "Cash")+SUMIFS(BS_Full!H:H, BS_Full!B:B, B{rt}, BS_Full!C:C, C{rt}, BS_Full!G:G, "NCI")-SUMIFS(BS_Full!H:H, BS_Full!B:B, B{rt}, BS_Full!C:C, C{rt}, BS_Full!G:G, "NOA")'
            sc(ws_trend.cell(rt, 5), fo=fFRM, nf=NB, bd=BD)
            
            # LTM Figures from LTM_Calc (Sumifs calc_key)
            ws_trend.cell(rt, 6).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{rt}, LTM_Calc!C:C, C{rt}, LTM_Calc!D:D, "Revenue")'; sc(ws_trend.cell(rt, 6), fo=fLINK, nf=NB, bd=BD)
            ws_trend.cell(rt, 7).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{rt}, LTM_Calc!C:C, C{rt}, LTM_Calc!D:D, "EBIT")'; sc(ws_trend.cell(rt, 7), fo=fLINK, nf=NB, bd=BD)
            ws_trend.cell(rt, 8).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{rt}, LTM_Calc!C:C, C{rt}, LTM_Calc!D:D, "NI")'; sc(ws_trend.cell(rt, 8), fo=fLINK, nf=NB, bd=BD)
            
            # Equity from BS_Full
            ws_trend.cell(rt, 9).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{rt}, BS_Full!C:C, C{rt}, BS_Full!G:G, "Equity_Total")'; sc(ws_trend.cell(rt, 9), fo=fLINK, nf=NB, bd=BD)
            
            # Multiples by Formula
            pMULT = PatternFill('solid', fgColor=C_PB)
            ws_trend.cell(rt, 10).value = f'=IF(G{rt}>0, E{rt}/G{rt}, "N/M")'; sc(ws_trend.cell(rt, 10), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
            ws_trend.cell(rt, 11).value = f'=IF(H{rt}>0, D{rt}/H{rt}, "N/M")'; sc(ws_trend.cell(rt, 11), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
            ws_trend.cell(rt, 12).value = f'=IF(F{rt}>0, D{rt}/F{rt}, "N/M")'; sc(ws_trend.cell(rt, 12), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
            ws_trend.cell(rt, 13).value = f'=IF(I{rt}>0, D{rt}/I{rt}, "N/M")'; sc(ws_trend.cell(rt, 13), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
            
            rt += 1
            
    ws_trend.auto_filter.ref = f"A{header_row_t}:M{rt-1}"
    ws_trend.freeze_panes = f"D{header_row_t+1}" # Scroll from Market_Cap

    # === Data_Quality Sheet — 자동 수집이 채우지 못한 자리 ===
    # 못 가져온 값은 0으로 남고 나머지 계산은 계속 돌아간다. 파일만 받아 본 사람이
    # 그 0을 '정말 0'으로 읽지 않도록, 어디가 비었는지 여기에 모아 둔다.
    ws_dq = wb.create_sheet('Data_Quality')
    ws_dq.sheet_properties.tabColor = 'EF6C00'

    ws_dq.merge_cells('A1:E1'); sc(ws_dq.cell(1, 1, 'Data Quality Check (확인 필요 항목)'), fo=fT)
    ws_dq.merge_cells('A2:E2')
    sc(ws_dq.cell(2, 1,
                  'ERROR = 그 값을 쓰면 안 됨 / WARN = 값은 나왔으나 왜곡 가능 / INFO = 참고. '
                  '자동 수집의 한계를 표시한 것이므로 공시자료와 대조 검증하십시오.'), fo=fS)

    dq_cols = [('Level', 10), ('Ticker', 12), ('Company', 20), ('Item', 20), ('Message', 100)]
    dq_hdr = 4
    for ci, (h, w) in enumerate(dq_cols, 1):
        ws_dq.column_dimensions[get_column_letter(ci)].width = w
        sc(ws_dq.cell(dq_hdr, ci, h), fo=fH, fi=pH, al=aC, bd=BD)

    level_fill = {SEV_ERROR: PatternFill('solid', fgColor='FFCDD2'),
                  SEV_WARN: PatternFill('solid', fgColor='FFE0B2'),
                  SEV_INFO: PatternFill('solid', fgColor='E3F2FD')}
    level_order = {SEV_ERROR: 0, SEV_WARN: 1, SEV_INFO: 2}
    dq_rows = sorted(quality.rows,
                     key=lambda x: (level_order.get(x.get('Level'), 9),
                                    str(x.get('Ticker', '')), str(x.get('Item', ''))))

    r_dq = dq_hdr + 1
    if dq_rows:
        for fl in dq_rows:
            lv = fl.get('Level', SEV_INFO)
            vals = [lv, fl.get('Ticker', ''), fl.get('Company', ''),
                    fl.get('Item', ''), fl.get('Message', '')]
            for ci, v in enumerate(vals, 1):
                sc(ws_dq.cell(r_dq, ci, v), fo=fA, fi=level_fill.get(lv, pW), bd=BD,
                   al=aC if ci == 1 else aL)
            r_dq += 1
    else:
        ws_dq.merge_cells(start_row=r_dq, start_column=1, end_row=r_dq, end_column=5)
        sc(ws_dq.cell(r_dq, 1, '✅ 자동 점검에서 특이사항이 발견되지 않았습니다. '
                               '점검 대상이 아닌 항목은 여전히 직접 확인이 필요합니다.'), fo=fA)
        r_dq += 1

    ws_dq.auto_filter.ref = f"A{dq_hdr}:E{r_dq-1}"
    ws_dq.freeze_panes = f'A{dq_hdr+1}'
    # GPCM 바로 뒤 — 숫자를 보기 전에 눈에 걸리도록 맨 앞쪽에 둔다
    wb.move_sheet('Data_Quality', offset=1 - wb.sheetnames.index('Data_Quality'))

    # === Peer_Selection Sheet — 업종에서 고른 경우, 무엇을 보고 골랐는지 ===
    # 후보 목록은 실행 시점의 상장 현황이라 나중에 다시 만들 수 없다. "Peer 를 왜
    # 이 회사들로 했나"에 답하려면 그때 무엇이 떠 있었는지가 파일에 남아 있어야 한다.
    if peer_selection:
        ws_ps = wb.create_sheet('Peer_Selection')
        ws_ps.sheet_properties.tabColor = '00695C'
        picked_set = set(peer_selection.get('picked') or [])

        ws_ps.merge_cells('A1:F1'); sc(ws_ps.cell(1, 1, 'Peer Selection (모집단과 선택 근거)'), fo=fT)
        ws_ps.merge_cells('A2:F2')
        sc(ws_ps.cell(2, 1,
                      f"업종: {peer_selection.get('sector', '')} | "
                      f"후보 {len(peer_selection.get('candidates') or [])}개 중 {len(picked_set)}개 선택 | "
                      f"조회 시점: {datetime.now().strftime('%Y-%m-%d %H:%M')} — "
                      f"상장·폐지에 따라 후보 목록은 시점마다 달라집니다."), fo=fS)

        ps_cols = [('선택', 8), ('Code', 12), ('Name', 24), ('Market', 10),
                   ('주요제품', 60), ('결산월', 10)]
        ps_hdr = 4
        for ci, (h, w) in enumerate(ps_cols, 1):
            ws_ps.column_dimensions[get_column_letter(ci)].width = w
            sc(ws_ps.cell(ps_hdr, ci, h), fo=fH, fi=pH, al=aC, bd=BD)

        pPICK = PatternFill('solid', fgColor='C8E6C9')
        r_ps = ps_hdr + 1
        for c in (peer_selection.get('candidates') or []):
            chosen = c['Code'] in picked_set
            vals = ['●' if chosen else '', c['Code'], c['Name'], c['Market'],
                    c['Product'], c['SettleMonth']]
            for ci, v in enumerate(vals, 1):
                sc(ws_ps.cell(r_ps, ci, v), fo=fA,
                   fi=pPICK if chosen else (PatternFill('solid', fgColor='FFE0B2') if c['FiscalNot12'] else pW),
                   al=aC if ci in (1, 2, 4, 6) else aL, bd=BD)
            r_ps += 1

        ws_ps.auto_filter.ref = f"A{ps_hdr}:F{r_ps-1}"
        ws_ps.freeze_panes = f'A{ps_hdr+1}'
        wb.move_sheet('Peer_Selection', offset=2 - wb.sheetnames.index('Peer_Selection'))

    wb.save(output)
    output.seek(0)

    return output
# 5. Streamlit App Layout & Logic
# ==========================================

# 사이드바 UI
with st.sidebar:
    st.header("Settings")
    st.caption(f"gpcm {app_version()}  ·  {APP_DIR}")
    if app_version() == "unknown":
        st.caption("⚠️ manifest.json 이 없어 버전을 알 수 없습니다 — 폴더가 온전한지 확인하세요.")
    
    # 좌측 1 : 기능 모드 선택 (신규)
    ui_mode = st.radio(
        "분석 모드 선택",
        ["GPCM Valuation (기존)", "다기간 재무제표 요약 (신규)"],
        index=0,
        help="GPCM 기반 가치평가 모드와 여러 회사의 과거 N년치 재무제표 요약 모드 중 하나를 선택하세요."
    )
    
    st.markdown("---")
    
    # 공통 입력 1: OpenDart API Key
    api_key_input = st.text_input("OpenDart API Key", type="password", help="OpenDart API 키를 입력하세요.")
    
    # 모드별 입력 파라미터 분기
    if ui_mode == "GPCM Valuation (기존)":
        # 다기간 GPCM 파라미터
        # 기본값은 '아직 공시되지 않은 미래 분기'가 잡히지 않도록 실제 공시 완료된 최신 분기로 설정
        latest_year, latest_qtr = get_latest_filed_period()

        st.write("**GPCM 분석 대상 기간**")
        st.caption(f"📅 현재 조회 가능한 최신 공시: **{latest_year}년 {latest_qtr}**")
        g_cycle = st.radio("분석 주기", ["분기별 (Quarterly)", "연간별 (Annual)"], index=0, horizontal=True, help="연간별 선택 시 각 연도의 4Q(사업보고서) 데이터만 추출하여 트렌드를 구성합니다.")

        col1, col2 = st.columns(2)
        with col1:
            st.write("**시작 기간**")
            g_start_year = st.number_input("시작 연도", min_value=2015, max_value=2030, value=latest_year - 1, step=1, key="gsy")
            g_start_qtr = "1Q"
            if g_cycle == "분기별 (Quarterly)":
                g_start_qtr = st.selectbox("시작 분기", ["1Q", "2Q", "3Q", "4Q"], index=0, key="gsq")
        with col2:
            st.write("**종료 기간 (기본 Base Date)**")
            g_end_year = st.number_input("종료 연도", min_value=2015, max_value=2030, value=latest_year, step=1, key="gey")
            g_end_qtr = "4Q"
            if g_cycle == "분기별 (Quarterly)":
                g_end_qtr = st.selectbox("종료 분기", ["1Q", "2Q", "3Q", "4Q"],
                                         index=["1Q", "2Q", "3Q", "4Q"].index(latest_qtr), key="geq")

        target_periods = []
        qtrs = ["1Q", "2Q", "3Q", "4Q"]
        for y in range(g_start_year, g_end_year + 1):
            if g_cycle == "연간별 (Annual)":
                # 연간 모드에서는 해당 연도 4Q 데이터를 추가
                # 단, 종료 연도(gey)의 경우 사용자가 의도한 최신 데이터가 반영되도록 함.
                target_periods.append(f"{y}.4Q")
            else:
                s_idx = qtrs.index(g_start_qtr) if y == g_start_year else 0
                e_idx = qtrs.index(g_end_qtr) if y == g_end_year else 3
                for q_idx in range(s_idx, e_idx + 1):
                    target_periods.append(f"{y}.{qtrs[q_idx]}")
                
        if not target_periods:
            st.error("종료 기간이 시작 기간보다 빠릅니다.")
            st.stop()
            
        base_period_str = target_periods[-1]
        base_year, base_qtr = parse_period(base_period_str)
        base_date_display = get_base_date_str(base_year, base_qtr)

        unfiled = [p for p in target_periods if not is_period_filed(*parse_period(p))]
        if unfiled:
            st.warning(
                f"⚠️ 아직 공시되지 않은 기간이 포함되어 있습니다: {', '.join(unfiled)}\n\n"
                f"해당 기간은 재무·주가 데이터가 비어 있어 결과가 0으로 나옵니다. "
                f"종료 기간을 **{latest_year}년 {latest_qtr}** 이하로 맞춰주세요."
            )

        st.info(f"Target WACC 기준일 (최신기간 적용): {base_date_display} (말일)")

        st.subheader("Target Companies")
        # 종목코드를 손으로 치면 오타 난 회사가 조용히 빠지고, 어느 회사를 왜 골랐는지도
        # 남지 않는다. 업종으로 후보를 띄우고 고르게 해, 그 선택을 엑셀에 기록한다.
        df_ind = get_krx_industry_listing()
        peer_selection = None

        if df_ind.empty:
            st.caption("업종 목록을 받지 못했습니다. 종목코드를 직접 입력하세요.")
            pick_mode = "종목코드 직접 입력"
        else:
            pick_mode = st.radio("종목 선택 방식", ["업종에서 고르기", "종목코드 직접 입력"],
                                 horizontal=True, label_visibility="collapsed")

        if pick_mode == "업종에서 고르기":
            sectors = sorted(s for s in df_ind['Sector'].dropna().unique() if str(s).strip())
            sector = st.selectbox("업종", sectors,
                                  help="KRX 상장회사목록의 업종 구분입니다. 같은 업종 안에서도 "
                                       "사업이 다를 수 있으니 주요제품을 보고 판단하세요.")
            candidates = peer_candidate_rows(df_ind, sector)
            label_of = {
                c['Code']: f"{c['Code']} {c['Name']}"
                          + (f" — {c['Product']}" if c['Product'] else "")
                          + (f"  ⚠️{c['SettleMonth']} 결산" if c['FiscalNot12'] else "")
                for c in candidates
            }
            picked = st.multiselect(f"비교 대상 ({len(candidates)}개 후보)",
                                    options=[c['Code'] for c in candidates],
                                    format_func=lambda c: label_of.get(c, c))
            tickers_input = "\n".join(picked)
            peer_selection = {'sector': sector, 'candidates': candidates, 'picked': picked}
            if not picked:
                st.caption("후보 중에서 비교할 회사를 고르세요.")
        else:
            tickers_input = st.text_area("대상회사의 종목코드를 한줄씩 입력하세요", value="000250\n039030\n005290", height=150)

        st.subheader("Target WACC Parameters")

        # 기준일 시장금리 조회 — rf·Kd 는 판단이 아니라 그날의 관측치다.
        # 조회값을 그대로 쓰지 않고 아래 입력칸의 기본값으로만 넣는다(확정은 사용자).
        with st.expander("📉 기준일 시장금리 조회 (선택)"):
            st.caption("국고채는 인증키 없이 조회됩니다. 회사채는 한국은행 ECOS 인증키가 필요합니다.")
            ecos_key_input = st.text_input(
                "한국은행 ECOS 인증키 (선택)", type="password",
                value=ecos_key_from_env(),
                help="https://ecos.bok.or.kr 에서 무료 발급. 비워도 국고채는 조회됩니다.")
            c1, c2 = st.columns(2)
            rf_term_input = c1.selectbox("국고채 만기", ["1년", "3년", "5년", "10년", "20년", "30년"], index=2)
            kd_grade_input = c2.selectbox("회사채 등급", list(BOND_GRADES), index=0)
            if st.button("기준일 금리 조회", key="btn_rates"):
                # 기준일은 화면에서 고른 종료 기간의 분기말이다 — 오늘 날짜가 아니다
                asof = as_of_date(f"{g_end_year}.{g_end_qtr}")
                lines, got_rf, got_kd = [], None, None
                got_rf, tried_rf = fetch_market_rate('rf', asof, None, rf_term_input, ecos_key_input)
                got_kd, tried_kd = fetch_market_rate('kd', asof, kd_grade_input, '3년', ecos_key_input)
                if got_rf:
                    st.session_state['rf_fetched'] = got_rf['value']
                    lines.append(f"국고채 {rf_term_input} {got_rf['value']}% ({got_rf['rateDate']}, {got_rf['source']})")
                else:
                    st.warning("국고채 금리를 못 받았습니다 — " + " / ".join(tried_rf))
                if got_kd:
                    st.session_state['kd_fetched'] = got_kd['value']
                    lines.append(f"회사채 3년 {kd_grade_input} {got_kd['value']}% ({got_kd['rateDate']}, {got_kd['source']})")
                else:
                    st.warning("회사채 금리를 못 받았습니다 — " + " / ".join(tried_kd))
                if lines:
                    st.session_state['rate_source'] = " | ".join(lines)
                    st.success("아래 입력칸에 반영했습니다. 값과 출처를 확인하고 확정하세요.")
            if st.session_state.get('rate_source'):
                st.caption("출처: " + st.session_state['rate_source'])

        rf_input = st.number_input("Rf - 무위험이자율 (%)", min_value=0.0, max_value=10.0,
                                   value=float(st.session_state.get('rf_fetched', 3.3)),
                                   step=0.1, format="%.2f") / 100
        mrp_input = st.slider("MRP (시장위험프리미엄)", min_value=7.0, max_value=9.0, value=8.0, step=0.1, format="%.1f%%") / 100

        with st.expander("📊 시가총액별 Size Premium 참고표"):
            st.markdown("**3분위수 기준**")
            st.markdown("""
            | 구분 | 시가총액 범위 (억원) | Size Premium |
            |------|---------------------|--------------|
            | **Micro** | < 2,000 | **4.02%** |
            | **Low** | 2,000 ~ 20,000 | 1.37% |
            | **Mid** | > 20,000 | -0.36% |
            """)
            st.markdown("**5분위수 기준**")
            st.markdown("""
            | 구분 | 시가총액 범위 (억원) | Size Premium |
            |------|---------------------|--------------|
            | **5분위 (최소)** | < 2,000 | **4.66%** |
            | **4분위** | 2,000 ~ 5,000 | 3.02% |
            | **3분위** | 5,000 ~ 20,000 | 1.21% |
            | **2분위** | 20,000 ~ 50,000 | 0.06% |
            | **1분위 (최대)** | > 50,000 | -0.58% |
            """)
        
        size_premium_options = {
            "3분위 - Micro (4.02%): < 2,000억": 0.0402,
            "3분위 - Low (1.37%): 2,000~20,000억": 0.0137,
            "3분위 - Mid (-0.36%): > 20,000억": -0.0036,
            "5분위 - 5분위/최소 (4.66%): < 2,000억": 0.0466,
            "5분위 - 4분위 (3.02%): 2,000~5,000억": 0.0302,
            "5분위 - 3분위 (1.21%): 5,000~20,000억": 0.0121,
            "5분위 - 2분위 (0.06%): 20,000~50,000억": 0.0006,
            "5분위 - 1분위/최대 (-0.58%): > 50,000억": -0.0058,
            "Size Premium 없음 (0%)": 0.0
        }
        size_premium_choice = st.selectbox("기업 규모 선택", list(size_premium_options.keys()), index=0)
        size_premium_input = size_premium_options[size_premium_choice]

        beta_type_options = {"5년 월간 베타 (5Y Monthly)": "5Y", "2년 주간 베타 (2Y Weekly)": "2Y"}
        beta_type_choice = st.selectbox("WACC 계산에 사용할 Beta", list(beta_type_options.keys()), index=0)
        beta_type_input = beta_type_options[beta_type_choice]

        kd_pretax_input = st.number_input("Kd (Pretax) - 세전 이자율 (%)", min_value=0.0, max_value=15.0,
                                          value=float(st.session_state.get('kd_fetched', 3.5)),
                                          step=0.1, format="%.1f") / 100

        # 세율은 회사 규모에 따라 다르다. 기본값 26.4%는 3,000억 초과 구간이라
        # 중소형 피평가회사에서는 대개 틀린다 — 세전이익에서 자동으로 정하게 한다.
        tax_auto_input = st.checkbox(
            "법인세율 자동 (피평가회사 세전이익 기준)", value=True,
            help="목록 첫 번째 회사의 LTM 세전이익을 해당 사업연도 한계세율표에 넣어 정합니다. "
                 "한국 법인 전제이며, 적용된 세율은 실행 후 화면과 Data_Quality에 남습니다.")
        target_tax_rate_input = st.number_input(
            "Target 법인세율 (%)", min_value=0.0, max_value=50.0, value=26.4, step=0.1, format="%.1f",
            disabled=tax_auto_input,
            help="지방소득세 포함 한계세율. 자동 체크를 끄면 이 값을 씁니다.") / 100

        run_btn = st.button("Go,Go,Go 🚀", type="primary", key="btn_gpcm")

    else:
        # 신규 다기간 재무제표 요약 파라미터
        current_year = datetime.now().year
        
        # 연간 vs 분기 조회 선택
        hist_period_type = st.radio(
            "조회 기준",
            ["연간 (사업보고서)", "분기 선택"],
            index=0,
            help="'연간'은 매년 연간 재무제표를 조회\n'분기 선택'은 특정 분기의 재무제표를 순차적으로 조회"
        )
        
        periods_to_fetch = []
        if hist_period_type == "연간 (사업보고서)":
            col1, col2 = st.columns(2)
            with col1:
                start_year = st.number_input("시작 연도", min_value=2015, max_value=2030, value=current_year - 3, step=1)
            with col2:
                end_year = st.number_input("종료 연도", min_value=2015, max_value=2030, value=current_year - 1, step=1)
            
            for y in range(start_year, end_year + 1):
                periods_to_fetch.append({'year': y, 'qtr': None, 'label': f"{y}년"})
        else:
            col1, col2 = st.columns(2)
            with col1:
                st.write("**시작 기간**")
                start_year = st.number_input("시작 연도", min_value=2015, max_value=2030, value=current_year - 1, step=1, key="sy_qtr")
                start_qtr = st.selectbox("시작 분기", ["1Q", "2Q", "3Q", "4Q"], index=0, key="sq_qtr")
            with col2:
                st.write("**종료 기간**")
                end_year = st.number_input("종료 연도", min_value=2015, max_value=2030, value=current_year, step=1, key="ey_qtr")
                end_qtr = st.selectbox("종료 분기", ["1Q", "2Q", "3Q", "4Q"], index=3, key="eq_qtr")
            
            qtrs = ["1Q", "2Q", "3Q", "4Q"]
            for y in range(start_year, end_year + 1):
                s_idx = qtrs.index(start_qtr) if y == start_year else 0
                e_idx = qtrs.index(end_qtr) if y == end_year else 3
                for q_idx in range(s_idx, e_idx + 1):
                    periods_to_fetch.append({'year': y, 'qtr': qtrs[q_idx], 'label': f"{y}년 {qtrs[q_idx]}"})
        
        st.subheader("Target Companies")
        tickers_input = st.text_area("대상회사의 종목코드를 한줄씩 입력하세요", value="000250\n039030\n005290", height=150)
        
        run_btn = st.button("재무제표 일괄 조회 🚀", type="primary", key="btn_hist")


# 메인 UI
if ui_mode == "GPCM Valuation (기존)":
    st.title("GPCM Calculator with Dart/KRX")
    st.markdown("""
    Opendartreader, Financedatareadr 라이브러리를 활용하여 기준일 시점 선정된 Peer의 재무제표, 주가, 유통주식수 등을 크롤링하여 GPCM Multiple을 계산하는 App 입니다. 
    해당 App 사용을 위해서는 **OpenDart API 인증키**를 개별적으로 발급받으셔야 합니다. 
    **감가상각비는 Dart에서 자동으로 불러올 수 없으니 EBITDA 계산 시 엑셀에서 수기로 입력하셔야 합니다.**
    (Made by SGJ _260211)
    """)

    st.markdown("---")
    st.subheader("📝 Valuation Methodology Notes")
    notes_list = [
        f'• Base Date: {base_period_str} ({base_date_display}) | Unit: 억원 (KRW 100M)',
    ]
    if st.session_state.get('rate_source'):
        # 조회한 금리를 썼으면 출처를 조서에 남긴다 ("이 rf 는 어디서 왔나")
        notes_list.append('• Rf/Kd 출처: ' + st.session_state['rate_source'])
    notes_list += [
        '• 공통: 연결재무제표 작성 시 CFS 우선, 미존재 시 OFS 기준으로 수집',
        '• PL: 요약 손익계산서에서 매출액/영업이익/당기순이익 3개 계정만 엄격 추출',
        '• PL Fetch: finstate(요약) → finstate_all(CFS/OFS) fallback',
        '• Shares: DART(stockTotqySttus) 유통주식수(distb_stock_co) 우선, 미공시 시 DART 과거보고서 fallback',
        '• EV = Market Cap + 우선주(장부) + IBD − Cash + NCI − NOA',
        '• Net Debt = IBD − Cash − NOA',
        '• IBD(Option): CB/EB/BW 등 메자닌은 기본적으로 IBD(Option)으로 태깅되어 EV/NetDebt에서 제외됨',
        '• NOA(Option): 투자자산/관계기업 등은 기본적으로 NOA(Option)으로 태깅되어 EV/NetDebt에서 제외됨',
        '• LTM = Current Cumulative + Prior Annual − Prior Same Quarter Cumulative (단, 4Q는 Annual)',
        '• Beta: 5년 월간 & 2년 주간 수익률 기준 (FinanceDataReader 사용)',
        '• Adjusted Beta = 2/3 × Raw Beta + 1/3 × 1 (조정을 Unlevered 계산 前에 적용)',
        '• Beta 벤치마크: 전 종목 KOSPI(^KS11) 단일 기준. 코스닥 종목도 KOSPI 대비로 산출한다 —',
        '  피어 무차입베타를 평균해 하나의 WACC 을 만들므로 모든 β 가 같은 지수 기준이어야 하고,',
        '  MRP 도 시장 전체(KOSPI) 기준 추정치라 β 와 기준을 맞춘 것이다',
        '• Beta 수익률: 배당 미반영 가격수익률 기준 (종목·지수 모두 동일 기준)',
        '• Beta 평균 대상: 조정베타가 0 초과인 회사 (GPCM 시트 Mean 행과 동일 모집단)',
        '• D/E Ratio = IBD / (Market Cap + 우선주 + NCI)',
        '• Debt Ratio (D/V) = IBD / (Market Cap + 우선주 + IBD + NCI)',
        '• 우선주: BS의 우선주자본금(액면) 기준. 시가총액은 보통주만 반영하므로 자기자본가치에 가산',
        '• Unlevered Beta = Levered Beta / (1 + (1 - Tax Rate) × D/E Ratio)',
        '• Tax Rate: 한국 법인세 한계세율 (지방소득세 포함, 세전순이익 기준, 사업연도별 세율표 적용)',
        '   - FY2023~2025: 2억 이하 9.9% | 2~200억 20.9% | 200~3,000억 23.1% | 3,000억 초과 26.4%',
        '   - FY2026~    : 2억 이하 11.0% | 2~200억 22.0% | 200~3,000억 24.2% | 3,000억 초과 27.5% (2025년 세법개정)',
    ]
    for note in notes_list:
        st.text(note)
    st.markdown("---")
else:
    st.title("📚 다기간 과거 재무제표 및 지표 요약 조회")
    st.markdown("""
    선택한 회사들의 과거 **연간 및 분기 재무제표(재무상태표, 손익계산서, 현금흐름표)**를 DART API를 통해 일괄 조회하고 엑셀로 추출합니다.
    - **조회 대상 기간**: 지정하신 시작/종료 연도 및 분기에 해당하는 공시 자료를 순차적으로 모두 수집합니다.
    - **결과물**: 여러 회사를 가로축으로 한눈에 비교할 수 있는 Summary 시트 + 각 회사별 상세 과거 재무제표 시트
    - **현금흐름표 처리**: 영업활동/투자활동/재무활동 등 대분류 현금흐름만 Summary에 표시되며 상세정보는 개별 시트에 기록됩니다.
    """)
    st.markdown("---")

# ▼▼▼▼▼ [추가 1] DART 접속 객체를 캐싱하는 함수 (if run_btn 바로 위에 넣으세요) ▼▼▼▼▼
def check_dart_reachable(timeout=10):
    """OpenDartReader 는 timeout 을 지정하지 않아 접속 불가 시 수 분간 멈춘다.
    실제 조회 전에 짧은 timeout 으로 도달 가능 여부만 먼저 확인한다."""
    try:
        requests.get('https://opendart.fss.or.kr/api/corpCode.xml',
                     params={'crtfc_key': 'preflight'}, timeout=timeout)
        return True, None
    except requests.exceptions.Timeout:
        return False, 'timeout'
    except requests.exceptions.ConnectionError:
        return False, 'unreachable'
    except Exception as e:
        return False, str(e)

@st.cache_resource
def get_dart_reader(api_key):
    return OpenDartReader(api_key)
# ▲▲▲▲▲ [여기까지 추가] ▲▲▲▲▲


# 실행 로직
if run_btn:
    if not api_key_input:
        st.error("OpenDart API Key를 입력해주세요.")
    else:
        target_code_list = [t.strip() for t in tickers_input.split('\n') if t.strip()]
        if not target_code_list:
            st.error("종목코드를 입력해주세요.")
        else:
            if ui_mode == "GPCM Valuation (기존)":
                # ==========================================
                # [모드 1] 기존 GPCM Valuation 로직
                # ==========================================
                status_container = st.status("GPCM 데이터 분석 중...", expanded=True)
                progress_bar = st.progress(0)

                ok, reason = check_dart_reachable()
                if not ok:
                    st.error(
                        "**DART 서버에 접속할 수 없습니다.**\n\n"
                        "API 키 문제가 아니라 네트워크에서 막힌 상태입니다.\n\n"
                        "- **국내에서 실행해주세요.** DART 는 해외 접속을 제한합니다.\n"
                        "- 사내망이라면 방화벽에서 `opendart.fss.or.kr` 을 허용해야 합니다.\n\n"
                        f"(원인: {reason})"
                    )
                    st.stop()

                try:
                    dart = get_dart_reader(api_key_input)
                except Exception as e:
                    st.error(f"DART 인증 실패 또는 조회 오류: {e}\n\nAPI 키가 올바른지 확인해주세요.")
                    st.stop()

                # 변수 초기화 및 데이터 수집
                raw_bs_rows, raw_pl_rows, all_mkt, ticker_to_name, screen_summary_data, base_year, base_qtr, base_date_str, all_multiples, quality = fetch_financial_data(
                    api_key_input, target_code_list, target_periods, dart, status_container, progress_bar)

                # 1. 화면 출력용 DataFrame 구성
                df_screen = pd.DataFrame(all_multiples)

                # 1-0. 수집 실패 진단 (조용히 0으로 넘어가는 것을 방지)
                problems = []
                base_rows = [m for m in all_multiples if m['Period'] == base_period_str]
                if not base_rows:
                    problems.append(f"기준 기간({base_period_str}) 데이터를 한 건도 수집하지 못했습니다.")
                for m in base_rows:
                    empty_fields = [k for k in ('Market_Cap', 'Revenue', 'Equity') if not m.get(k)]
                    if empty_fields:
                        problems.append(f"[{m['Company']}] {base_period_str} — {', '.join(empty_fields)} 값이 없습니다.")
                if problems:
                    st.warning(
                        "⚠️ **아래 항목을 수집하지 못했습니다. 배수·WACC 결과가 왜곡됩니다.**\n\n"
                        + "\n".join(f"- {p}" for p in problems)
                        + "\n\n대부분 아직 공시되지 않은 기간을 조회했을 때 발생합니다."
                    )
                    for p in problems:
                        quality.add(SEV_ERROR, '', '', f'기준기간 {base_period_str}', p)

                # 화면 경고는 창을 닫으면 사라진다. 파일만 남았을 때도 알 수 있게 개수를 알린다.
                n_err = sum(1 for r in quality.rows if r['Level'] == SEV_ERROR)
                n_warn = sum(1 for r in quality.rows if r['Level'] == SEV_WARN)
                if n_err or n_warn:
                    st.info(
                        f"📋 자동 점검 결과 **ERROR {n_err}건 · WARN {n_warn}건**을 "
                        f"엑셀 **Data_Quality** 시트에 기록했습니다. 숫자를 쓰기 전에 먼저 보십시오."
                    )

                if not df_screen.empty:
                    df_screen['EV'] = df_screen['Market_Cap'] + df_screen['Preferred'] + df_screen['IBD'] - df_screen['Cash'] + df_screen['NCI'] - df_screen['NOA']
                    df_screen['EV/EBIT'] = np.where(df_screen['EBIT'] > 0, df_screen['EV'] / df_screen['EBIT'], np.nan)
                    df_screen['PER'] = np.where(df_screen['NI'] > 0, df_screen['Market_Cap'] / df_screen['NI'], np.nan)
                    df_screen['PSR'] = np.where(df_screen['Revenue'] > 0, df_screen['Market_Cap'] / df_screen['Revenue'], np.nan)

                    st.subheader("📊 Multiples Table (Preview)")
                    st.dataframe(df_screen[['Company', 'Period', 'Market_Cap', 'EV', 'Revenue', 'EBIT', 'NI', 'EV/EBIT', 'PER', 'PSR']]
                                 .style.format("{:.1f}", subset=['Market_Cap','EV','Revenue','EBIT','NI'])
                                 .format("{:.1f}x", subset=['EV/EBIT','PER','PSR'], na_rep="N/M"))
                    
                    st.subheader("📈 Statistics (Mean/Median - Latest Period)")
                    latest_df = df_screen[df_screen['Period'] == base_period_str]
                    if not latest_df.empty:
                        stats = latest_df[['EV/EBIT', 'PER', 'PSR']].agg(['mean', 'median', 'max', 'min'])
                        st.dataframe(stats.style.format("{:.1f}x"))

                # 1.5. WACC Calculation (Target 기업용)
                # 세율 자동: 목록 첫 회사(피평가회사)의 세전이익을 사업연도 세율표에 넣는다.
                # 기본값 26.4%는 3,000억 초과 구간이라 중소형 회사에서는 대개 틀린다.
                tax_rate_used = target_tax_rate_input
                if tax_auto_input:
                    t_row = next((s for s in screen_summary_data
                                  if s.get('Ticker') == target_code_list[0]), None)
                    t_pretax = (t_row or {}).get('Pretax_Income')
                    tax_rate_used = get_korean_marginal_tax_rate(t_pretax, base_year)
                    st.info(
                        f"법인세율 **{tax_rate_used*100:.2f}%** 적용 — "
                        f"{(t_row or {}).get('Company', target_code_list[0])}의 세전이익 "
                        f"{'미상' if t_pretax is None else format(t_pretax, ',.0f') + '억원'} · "
                        f"FY{base_year} 한계세율표 (지방소득세 포함)")
                    quality.add(SEV_INFO, target_code_list[0],
                                (t_row or {}).get('Company', ''), 'Tax Rate',
                                f'타겟 법인세율을 세전이익 '
                                f'{"미상" if t_pretax is None else round(float(t_pretax), 1)}억원과 '
                                f'FY{base_year} 한계세율표로 {tax_rate_used*100:.2f}% 로 정했습니다 '
                                f'(한국 법인 전제, 지방소득세 포함).')

                target_wacc_data, avg_debt_ratio = calculate_wacc_and_beta(
                    target_code_list, screen_summary_data, tax_rate_used, rf_input, mrp_input, size_premium_input, kd_pretax_input, beta_type_input,
                    fiscal_year=base_year, quality=quality)

                if target_wacc_data['Target_WACC'] <= rf_input:
                    st.error(
                        f"❌ **계산된 WACC({target_wacc_data['Target_WACC']*100:.2f}%)이 "
                        f"무위험이자율({rf_input*100:.2f}%)보다 낮습니다.** 정상적인 결과가 아닙니다.\n\n"
                        "시가총액이 0으로 수집되어 자본구조·베타가 붕괴한 경우입니다. "
                        "위의 수집 실패 경고를 먼저 확인해주세요."
                    )

                # 2. 엑셀 생성 (메모리)
                output = export_gpcm_excel(
                    base_period_str, base_qtr, target_code_list, screen_summary_data, raw_bs_rows, raw_pl_rows, all_mkt, ticker_to_name,
                    target_wacc_data, beta_type_input, notes_list, avg_debt_ratio, base_date_str, df_screen, target_periods, quality, peer_selection)
                st.success("분석 완료! 아래 버튼을 눌러 리포트를 다운로드하세요.")
                st.download_button(
                    label="📥 Report Download (Excel)",
                    data=output,
                    file_name=f"KR_GPCM_Fixed_{base_period_str.replace('.','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            elif ui_mode == "다기간 재무제표 요약 (신규)":
                # ==========================================
                # [모드 2] 다기간 재무제표 요약 로직
                # ==========================================
                status_container = st.status(f"다기간 재무제표 데이터 수집 중... (대상: {len(target_code_list)}개 기업 / 기간: {len(periods_to_fetch)}개)", expanded=True)
                progress_bar = st.progress(0)

                ok, reason = check_dart_reachable()
                if not ok:
                    st.error(
                        "**DART 서버에 접속할 수 없습니다.**\n\n"
                        "API 키 문제가 아니라 네트워크에서 막힌 상태입니다.\n\n"
                        "- **국내에서 실행해주세요.** DART 는 해외 접속을 제한합니다.\n"
                        "- 사내망이라면 방화벽에서 `opendart.fss.or.kr` 을 허용해야 합니다.\n\n"
                        f"(원인: {reason})"
                    )
                    st.stop()

                try:
                    dart = get_dart_reader(api_key_input)
                except Exception as e:
                    st.error(f"DART 인증 실패 또는 조회 오류: {e}\n\nAPI 키가 올바른지 확인해주세요.")
                    st.stop()
                
                # 1. 데이터 수집
                df_krx = get_krx_listing()
                df_summ, df_details = fetch_historical_financials(
                    api_key_input, target_code_list, periods_to_fetch,
                    dart, status_container, progress_bar, df_krx
                )
                
                # 2. 지표 계산
                status_container.update(label="지표 계산 및 엑셀 리포트 생성 중...")
                df_summ = calculate_historical_metrics(df_summ)
                
                # 3. 엑셀 생성
                if not df_summ.empty:
                    output = export_historical_excel(df_summ, df_details, periods_to_fetch)
                    
                    status_container.update(label="분석 완료!", state="complete")
                    st.success("데이터 추출이 완료되었습니다. 아래 버튼을 눌러 리포트를 다운로드하세요.")
                    
                    st.subheader("📊 Summary Preview")
                    # 화면 표시용: 일부 핵심 지표만 나열
                    preview_cols = ['Company', 'Period', 'Revenue', 'EBIT', 'NI', 'OPM', 'ROE']
                    avail_cols = [c for c in preview_cols if c in df_summ.columns]
                    
                    # 포맷 적용 (문자열/정수 컬럼 제외)
                    num_cols = [c for c in avail_cols if c in ('Revenue', 'EBIT', 'NI')]
                    pct_cols = [c for c in avail_cols if c in ('OPM', 'ROE')]
                    
                    styler = df_summ[avail_cols].style
                    if num_cols: styler = styler.format("{:,.0f}", subset=num_cols, na_rep="")
                    if pct_cols: styler = styler.format("{:.1%}", subset=pct_cols, na_rep="")
                    st.dataframe(styler)
                    
                    st.download_button(
                        label="📥 Report Download (Excel)",
                        data=output,
                        file_name=f"KR_Historical_Financials_{start_year}_to_{end_year}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    status_container.update(label="데이터 수집 실패", state="error")
                    st.warning("수집된 데이터가 없습니다. 종목코드나 연도를 다시 한번 확인해주세요.")
