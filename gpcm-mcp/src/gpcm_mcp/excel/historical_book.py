"""모드 2 워크북 생성. gpcm_kr.py 에서 그대로 옮겼다."""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .styles import *  # noqa: F401,F403  (원본이 모듈 전역 스타일에 의존한다)
from .styles import NB, aC, aL, aR, fA, fFRM, fHARD, fS, fT, pSEC1, pSEC2, sc


def export_historical_excel(df_summ, df_details, periods_to_fetch):
    output = io.BytesIO()
    wb = Workbook()
    
    # ---------------------------------------------------------
    # 1. Summary 시트 생성 (Layout A안: 회사 세로, 연도/지표 가로)
    # ---------------------------------------------------------
    ws_summ = wb.active
    ws_summ.title = "Summary"
    
    ws_summ.merge_cells('A1:Z1')
    p_start = periods_to_fetch[0]['label'] if periods_to_fetch else ""
    p_end = periods_to_fetch[-1]['label'] if periods_to_fetch else ""
    ws_summ['A1'] = f"Historical Financial Summary ({p_start} ~ {p_end})"
    sc(ws_summ['A1'], fo=fT)
    
    if df_summ.empty:
        ws_summ['A3'] = "No data available."
    else:
        metrics = [
            ('Revenue', '매출액', NB), ('GrossProfit', '매출총이익', NB), 
            ('EBIT', '영업이익', NB), ('NI', '당기순이익', NB),
            ('Assets', '자산총계', NB), ('Liabilities', '부채총계', NB), ('Equity_Total', '자본총계', NB),
            ('Cash', 'Cash', NB), ('IBD', 'IBD', NB),
            ('NOA', 'NOA', NB), ('NCI', 'NCI', NB),
            ('NetDebt', '순부채(Net Debt)', NB),
            ('CFO', '영업활동현금흐름', NB), ('CFI', '투자활동현금흐름', NB), ('CFF', '재무활동현금흐름', NB),
            ('OPM', '영업이익률', '0.0%'), ('GPM', '매출총이익률', '0.0%'), 
            ('ROE', 'ROE', '0.0%'), ('DebtRatio', '부채비율', '0.0%')
        ]
        
        labels = [p['label'] for p in periods_to_fetch]
        
        # 헤더 그리기 (Row 3: 지표명, Row 4: 기간)
        ws_summ.cell(row=3, column=1, value="Company"); sc(ws_summ.cell(row=3, column=1), fo=fH, fi=pH, al=aC, bd=BD)
        ws_summ.cell(row=3, column=2, value="Ticker"); sc(ws_summ.cell(row=3, column=2), fo=fH, fi=pH, al=aC, bd=BD)
        ws_summ.merge_cells('A3:A4')
        ws_summ.merge_cells('B3:B4')
        
        # 메트릭별 컬럼 알파벳 매핑 저장용
        mc_map = {} # (m_key, plabel) -> 'C'
        plabel_col_idx = {plabel: 5 + i for i, plabel in enumerate(labels)} # Detail 시트의 데이터 컬럼 E부터 시작
        
        col_idx = 3
        for m_key, m_name, _ in metrics:
            start_col = col_idx
            end_col = col_idx + len(labels) - 1
            ws_summ.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
            ws_summ.cell(row=3, column=start_col, value=m_name)
            sc(ws_summ.cell(row=3, column=start_col), fo=fH, fi=pSEC1, al=aC, bd=BD)
            
            for plabel in labels:
                ws_summ.cell(row=4, column=col_idx, value=plabel)
                sc(ws_summ.cell(row=4, column=col_idx), fo=fH, fi=pSEC2, al=aC, bd=BD)
                
                mc_map[(m_key, plabel)] = get_column_letter(col_idx)
                col_idx += 1
                
        # 데이터 쓰기 (Row 5부터 ~ 회사별)
        r = 5
        companies = df_summ['Company'].unique()
        for comp in companies:
            df_comp = df_summ[df_summ['Company'] == comp]
            ticker = df_comp['Ticker'].iloc[0] if not df_comp.empty else ""
            comp_sht = comp[:31] # 엑셀 시트 참조용 이름
            
            ws_summ.cell(row=r, column=1, value=comp); sc(ws_summ.cell(row=r, column=1), fo=fA, bd=BD)
            ws_summ.cell(row=r, column=2, value=ticker); sc(ws_summ.cell(row=r, column=2), fo=fA, al=aC, bd=BD)
            
            # 수식 적용 그룹 (비율/멀티플)
            ratio_keys = ['OPM', 'GPM', 'ROE', 'DebtRatio']
            # Raw Data SUMIFS 그룹
            sumifs_keys = ['Revenue', 'GrossProfit', 'EBIT', 'NI', 'Assets', 'Liabilities', 'Equity_Total', 'Cash', 'IBD', 'NOA', 'NCI', 'CFO', 'CFI', 'CFF']
            
            c = 3
            for m_key, m_name, fmt in metrics:
                for plabel in labels:
                    v = ""
                    dtl_col = get_column_letter(plabel_col_idx[plabel]) # Detail 시트의 타겟 Period 열
                    
                    if m_key in sumifs_keys:
                        # 엑셀 SUMIFS 수식 주입 (매핑 키 A열, 금액 Dtl_Col열)
                        v = f"=SUMIFS('{comp_sht}'!{dtl_col}:{dtl_col}, '{comp_sht}'!$A:$A, \"{m_key}\")"
                    elif m_key == 'NetDebt':
                        v = f"={mc_map[('IBD', plabel)]}{r} - {mc_map[('Cash', plabel)]}{r}"
                    elif m_key in ratio_keys:
                        if m_key == 'OPM': v = f"=IFERROR({mc_map[('EBIT', plabel)]}{r}/{mc_map[('Revenue', plabel)]}{r}, \"\")"
                        elif m_key == 'GPM': v = f"=IFERROR({mc_map[('GrossProfit', plabel)]}{r}/{mc_map[('Revenue', plabel)]}{r}, \"\")"
                        elif m_key == 'ROE': v = f"=IFERROR({mc_map[('NI', plabel)]}{r}/{mc_map[('Equity_Total', plabel)]}{r}, \"\")"
                        elif m_key == 'DebtRatio': v = f"=IFERROR({mc_map[('Liabilities', plabel)]}{r}/{mc_map[('Equity_Total', plabel)]}{r}, \"\")"
                        
                    ws_summ.cell(row=r, column=c, value=v)
                    font_style = fA if m_key in sumifs_keys else fFRM
                    sc(ws_summ.cell(row=r, column=c), fo=font_style, nf=fmt, bd=BD)
                    
                    c += 1
            r += 1
        
        ws_summ.column_dimensions['A'].width = 18
        ws_summ.column_dimensions['B'].width = 10
        for i in range(3, c):
            ws_summ.column_dimensions[get_column_letter(i)].width = 14
        
        ws_summ.freeze_panes = "C5"

    # ---------------------------------------------------------
    # 2. 개별 회사 상세 시트 생성 (세로: 계정, 가로: 연도 피벗 형태)
    # ---------------------------------------------------------
    if not df_details.empty:
        companies = df_details['Company'].unique()
        for comp in companies:
            ws_dtl = wb.create_sheet(title=comp[:31]) # 시트명 제한 31자
            df_c = df_details[df_details['Company'] == comp].copy()
            
            # 헤더 2줄 그리미 (Simplified)
            ws_dtl.merge_cells('A1:H1'); ws_dtl['A1'] = f"{comp} - 상세 재무제표 (Report 기반)"; sc(ws_dtl['A1'], fo=fT)
            ws_dtl.merge_cells('A2:H2'); ws_dtl['A2'] = "DART finstate_all 원본 계정 정보 (최다 추출)"; sc(ws_dtl['A2'], fo=fS)
            
            if df_c.empty:
                ws_dtl['A4'] = "No detailed data available."
                continue
                
            pivot_df = df_c.pivot_table(
                index=['M_Key', 'Type', 'Account_ID', 'Account_NM'], 
                columns='Period', 
                values='Amount', 
                aggfunc='sum'
            ).reset_index()
            
            order_df = df_c.groupby(['M_Key', 'Type', 'Account_ID', 'Account_NM'])['Row_Idx'].min().reset_index()
            pivot_df = pd.merge(pivot_df, order_df, on=['M_Key', 'Type', 'Account_ID', 'Account_NM'], how='left')
            
            # Type 내림차순 정렬 유도 및 DART 한계 극복용 계층형 정렬 맵핑 (KPMG Style)
            sort_map = {'재무상태표': 1, '손익계산서': 2, '포괄손익계산서': 3, '현금흐름표': 4}
            def get_heuristic_rank(row):
                t = str(row['Type']).split()[0].replace('연결', '')
                acc = str(row['Account_NM'])
                idx = row['Row_Idx']
                t_rank = sort_map.get(t, 99) * 1000000
                
                if t in ['손익계산서', '포괄손익계산서']:
                    if '매출액' in acc or '영업수익' in acc: return t_rank + 10000
                    if '원가' in acc: return t_rank + 20000
                    if '매출총이익' in acc: return t_rank + 30000
                    if '판매비' in acc or '관리비' in acc: return t_rank + 40000
                    if '영업이익' in acc or '영업손실' in acc: return t_rank + 50000
                    if '법인세비용' in acc: return t_rank + 80000
                    if '당기순이익' in acc or '당기순손실' in acc: return t_rank + 90000
                    return t_rank + 60000 + idx
                elif t == '현금흐름표':
                    if acc == '영업활동현금흐름' or ('영업활동' in acc and '흐름' in acc): return t_rank + 10000
                    if acc == '투자활동현금흐름' or ('투자활동' in acc and '흐름' in acc): return t_rank + 40000
                    if acc == '재무활동현금흐름' or ('재무활동' in acc and '흐름' in acc): return t_rank + 70000
                    return t_rank + 10000 + idx
                return t_rank + idx
                
            pivot_df['SortKey'] = pivot_df.apply(get_heuristic_rank, axis=1)
            pivot_df = pivot_df.sort_values('SortKey').drop(columns=['SortKey', 'Row_Idx'])
            
            # 헤더 그리기
            labels = [p['label'] for p in periods_to_fetch]
            ws_dtl.cell(row=3, column=1, value="M_Key"); sc(ws_dtl.cell(row=3, column=1), fo=fH, fi=pH, al=aC, bd=BD)
            ws_dtl.cell(row=3, column=2, value="Type"); sc(ws_dtl.cell(row=3, column=2), fo=fH, fi=pH, al=aC, bd=BD)
            ws_dtl.cell(row=3, column=3, value="Account ID"); sc(ws_dtl.cell(row=3, column=3), fo=fH, fi=pH, al=aC, bd=BD)
            ws_dtl.cell(row=3, column=4, value="Account Name"); sc(ws_dtl.cell(row=3, column=4), fo=fH, fi=pH, al=aC, bd=BD)
            
            col_idx = 5
            for plabel in labels:
                ws_dtl.cell(row=3, column=col_idx, value=plabel); sc(ws_dtl.cell(row=3, column=col_idx), fo=fH, fi=pSEC2, al=aC, bd=BD)
                col_idx += 1
                
            r = 4
            for _, row in pivot_df.iterrows():
                ws_dtl.cell(row=r, column=1, value=row.get('M_Key', '')); sc(ws_dtl.cell(row=r, column=1), fo=fA, al=aL, bd=BD)
                ws_dtl.cell(row=r, column=2, value=row.get('Type', '')); sc(ws_dtl.cell(row=r, column=2), fo=fA, al=aL, bd=BD)
                ws_dtl.cell(row=r, column=3, value=row.get('Account_ID', '')); sc(ws_dtl.cell(row=r, column=3), fo=fA, al=aL, bd=BD)
                ws_dtl.cell(row=r, column=4, value=row.get('Account_NM', '')); sc(ws_dtl.cell(row=r, column=4), fo=fA, al=aL, bd=BD)
                
                c = 5
                for plabel in labels:
                    val = row.get(plabel)
                    v = val if pd.notna(val) else ""
                    ws_dtl.cell(row=r, column=c, value=v); sc(ws_dtl.cell(row=r, column=c), fo=fHARD, nf=NB, bd=BD)
                    c += 1
                r += 1
                
            ws_dtl.column_dimensions['A'].width = 15
            ws_dtl.column_dimensions['B'].width = 15
            ws_dtl.column_dimensions['C'].width = 25
            ws_dtl.column_dimensions['D'].width = 35
            for i in range(5, c):
                ws_dtl.column_dimensions[get_column_letter(i)].width = 15
                
            ws_dtl.freeze_panes = "E5"


    wb.save(output)
    output.seek(0)
    return output
