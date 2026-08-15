"""runner / summarize / output / 입력 검증."""

import json
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parents[1] / 'src'))

import fakes  # noqa: E402

import gpcm_mcp.gpcm as port_gpcm  # noqa: E402
import gpcm_mcp.historical as port_hist  # noqa: E402
import gpcm_mcp.listings as port_listings  # noqa: E402
from gpcm_mcp import output, runner  # noqa: E402

TICKERS = ['005930', '000660']
NAMES = {'005930': '테스트전자', '000660': '테스트반도체'}
PERIODS = ['2024.4Q', '2025.1Q']


@pytest.fixture
def patched(tmp_path, monkeypatch):
    monkeypatch.setenv(output.ENV_DIR, str(tmp_path))
    restore = fakes.install(port_gpcm, port_hist, port_listings,
                            names=NAMES, end_date='2025-03-31')
    yield
    restore()


def test_gpcm_result_is_json_safe(patched):
    """주가 시계열이 응답에 새어 나가면 전송 단계에서 죽는다."""
    result = runner.run_gpcm(TICKERS, PERIODS, dart=fakes.FakeDart(names=NAMES),
                             preflight=False)
    text = json.dumps(result, ensure_ascii=False)
    assert 'Stock_Monthly_Prices' not in text
    assert 'NaN' not in text  # JSON 에 NaN 은 없다


def test_gpcm_result_has_what_claude_needs(patched):
    result = runner.run_gpcm(TICKERS, PERIODS, dart=fakes.FakeDart(names=NAMES),
                             preflight=False)
    assert result['status'] == 'done'
    assert result['base_period'] == '2025.1Q'
    assert result['unit'].startswith('억원')

    # 배수는 파이썬이 계산한 값이어야 한다. 워크북 쪽은 수식이라 값이 없다.
    assert len(result['multiples']) == len(TICKERS)
    first = result['multiples'][0]
    assert first['Market_Cap'] == pytest.approx(700.0)
    assert first['EV'] == pytest.approx(700.0 + 500.0 - 1000.0 + 50.0)

    assert set(result['statistics']) == {'EV/EBIT', 'PER', 'PSR'}
    assert result['wacc']['Target_WACC'] > 0
    assert 'counts' in result['quality']
    assert len(result['notes']) > 10

    written = Path(result['file']['path'])
    assert written.exists() and written.suffix == '.xlsx'
    assert result['file']['bytes'] == written.stat().st_size


def test_historical_result_is_json_safe(patched):
    periods = [{'year': 2023, 'qtr': None, 'label': '2023년'},
               {'year': 2024, 'qtr': None, 'label': '2024년'}]
    result = runner.run_historical(TICKERS, periods,
                                   dart=fakes.FakeDart(names=NAMES),
                                   preflight=False)
    json.dumps(result, ensure_ascii=False)
    assert result['status'] == 'done'
    assert result['unit'].startswith('백만원')
    assert len(result['rows']) == len(TICKERS) * len(periods)
    assert Path(result['file']['path']).exists()


# --- 입력 검증: 조회를 시작하기 전에 걸러야 한다 ----------------------------

def test_empty_tickers_rejected():
    with pytest.raises(runner.InputError, match='종목코드'):
        runner.run_gpcm([], PERIODS, preflight=False)


def test_empty_periods_rejected():
    """원본은 여기서 ZeroDivisionError 로 죽는다."""
    with pytest.raises(runner.InputError, match='기간'):
        runner.run_gpcm(TICKERS, [], preflight=False)
    with pytest.raises(runner.InputError, match='기간'):
        runner.run_historical(TICKERS, [], preflight=False)


def test_malformed_ticker_rejected():
    with pytest.raises(runner.InputError, match='6자리'):
        runner.run_gpcm(['5930', 'AAPL'], PERIODS, preflight=False)


def test_bad_beta_type_rejected():
    """원본은 '5Y' 가 아니면 조용히 2Y 베타를 쓴다. 오타가 값을 바꾼다."""
    with pytest.raises(runner.InputError, match='beta_type'):
        runner.run_gpcm(TICKERS, PERIODS, beta_type='5y', preflight=False)


def test_output_name_does_not_clobber(tmp_path, monkeypatch):
    monkeypatch.setenv(output.ENV_DIR, str(tmp_path))
    first = output.build_name('KR_GPCM', '2025_1Q')
    assert first.startswith('KR_GPCM_2025_1Q_') and first.endswith('.xlsx')


def test_save_is_atomic(tmp_path, monkeypatch):
    import io
    monkeypatch.setenv(output.ENV_DIR, str(tmp_path))
    path = output.save(io.BytesIO(b'hello'), 'x.xlsx')
    assert path.read_bytes() == b'hello'
    assert not list(tmp_path.glob('*.tmp')), '임시 파일이 남았다'


# --- 시트명 정제가 실제로 워크북에서 통하는가 -------------------------------

def test_illegal_company_names_produce_valid_workbook(tmp_path, monkeypatch):
    """엑셀 금지문자가 든 회사명과, 앞 31자가 같은 두 회사를 한꺼번에 넣어 본다.

    원본은 전자에서 openpyxl 예외로 죽고, 후자에서는 시트명이 겹친다.
    """
    import re
    from openpyxl import load_workbook

    monkeypatch.setenv(output.ENV_DIR, str(tmp_path))
    names = {
        '005930': '한국ABC/DEF[주]',
        '000660': '가' * 31 + 'AAA',
        '005490': '가' * 31 + 'BBB',
    }
    tickers = sorted(names)
    restore = fakes.install(port_hist, port_listings, names=names,
                            end_date='2025-03-31')
    try:
        periods = [{'year': 2024, 'qtr': None, 'label': '2024년'}]
        result = runner.run_historical(tickers, periods,
                                       dart=fakes.FakeDart(names=names),
                                       preflight=False)
    finally:
        restore()

    assert result['status'] == 'done'
    wb = load_workbook(result['file']['path'])

    # 엑셀이 허용하지 않는 글자가 시트명에 남아 있으면 안 된다.
    for sheet in wb.sheetnames:
        assert not re.search(r'[\[\]:*?/\\]', sheet), sheet
        assert len(sheet) <= 31

    # 회사 셋이면 상세 시트도 셋이어야 한다 (겹치면 둘로 줄어든다).
    assert len(wb.sheetnames) == 1 + len(names), wb.sheetnames

    # Summary 의 SUMIFS 가 가리키는 시트가 실제로 있어야 한다.
    # 여기가 어긋나면 #REF! 가 되어 Summary 가 조용히 틀린다.
    referenced = set()
    ws = wb['Summary']
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('=SUMIFS'):
                referenced.update(re.findall(r"'([^']+)'!", cell.value))
    assert referenced, 'SUMIFS 수식을 하나도 찾지 못했다'
    missing = referenced - set(wb.sheetnames)
    assert not missing, f'없는 시트를 가리키는 수식이 있다: {missing}'
