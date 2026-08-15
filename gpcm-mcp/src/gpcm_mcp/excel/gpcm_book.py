"""모드 1 워크북 생성. gpcm_kr.py 에서 그대로 옮겼다."""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from ..constants import SEV_ERROR, SEV_INFO, SEV_WARN
from .styles import *  # noqa: F401,F403  (원본이 모듈 전역 스타일에 의존한다)
from .styles import (BD, NB, NB1, NF_X, NI_FMT, NP, aC, aL, aR,
                     add_gpcm_section_row, ev_fills, fA, fFRM, fH, fHL, fLINK,
                     fMUL, fNOTE, fS, fSEC, fSTAT, fT, pH, pST, pSTAT, pSEC1,
                     pSEC2, pSEC3, pSEC4, pSEC5, pSEC6, pW, sc, style_range)


def export_gpcm_excel(base_period_str, base_qtr, target_code_list, screen_summary_data, raw_bs_rows, raw_pl_rows, all_mkt, ticker_to_name, target_wacc_data, beta_type_input, notes_list, avg_debt_ratio, base_date_str, df_screen, target_periods, quality, peer_selection=None):
    # 2. 엑셀 생성 (메모리)
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    # (기존 엑셀 생성 로직 그대로 활용 - 함수화 하지 않고 바로 실행)
    # Sheet 1: BS_Full
    ws_bs = wb.create_sheet('BS_Full')
    ws_bs.merge_cells('A1:H1'); ws_bs['A1'] = "BS_Full (Balance Sheet Detail)"; sc(ws_bs['A1'], fo=fT)
    ws_bs.merge_cells('A2:H2'); ws_bs['A2'] = "Logic: finstate_all(CFS→OFS) 재무상태표 라인아이템 수집 후 EV_Component 태깅 | Unit: 억원"; sc(ws_bs['A2'], fo=fS)
    cols = [('Company',15), ('Ticker',10), ('Period',10), ('sj_nm',15),('account_nm',35), ('account_id',40), ('EV_Component',12), ('Amount_100M',15)]
    header_row = 4
    ws_bs.append([]); ws_bs.append([c[0] for c in cols])
    for i, (_, w) in enumerate(cols): ws_bs.column_dimensions[get_column_letter(i+1)].width = w; sc(ws_bs.cell(header_row, i+1), fo=fH, fi=pH, al=aC, bd=BD)
    r = header_row + 1
    if raw_bs_rows:
        for rd in raw_bs_rows:
            ev_comp = rd['EV_Component']; is_hl = bool(ev_comp)
            fill_key = 'Equity' if ev_comp in ['Equity_P', 'Equity_Total'] else ev_comp
            row_fi = ev_fills.get(fill_key, pW) if is_hl else pW; row_fo = fHL if is_hl else fA
            vals = [rd['Company'], rd['Ticker'], rd['Period'], rd['sj_nm'],rd['account_nm'], rd['account_id'], rd['EV_Component'], rd['Amount_100M']]
            for i, v in enumerate(vals): sc(ws_bs.cell(r, i+1), fo=row_fo, fi=row_fi, al=aR if i==7 else aL, nf=NB if i==7 else None, bd=BD); ws_bs.cell(r, i+1).value = v
            r += 1
    ws_bs.auto_filter.ref = f"A{header_row}:H{r-1}"; ws_bs.freeze_panes = f"A{header_row+1}"

    # Sheet 2: PL_Data
    ws_pl = wb.create_sheet('PL_Data')
    ws_pl.merge_cells('A1:H1'); ws_pl['A1'] = "PL_Data (Income Statement Core Only)"; sc(ws_pl['A1'], fo=fT)
    ws_pl.merge_cells('A2:H2'); ws_pl['A2'] = "Logic: IS 추출 후 매출/영업이익/순이익 3개 계정만 엄격 추출 | Unit: 억원"; sc(ws_pl['A2'], fo=fS)
    cols = [('Company',15), ('Ticker',10), ('Period',10), ('Role',15),('PL_Source',16), ('account_nm',35), ('calc_key',12), ('Amount_100M',15)]
    header_row = 4
    ws_pl.append([]); ws_pl.append([c[0] for c in cols])
    for i, (_, w) in enumerate(cols): ws_pl.column_dimensions[get_column_letter(i+1)].width = w; sc(ws_pl.cell(header_row, i+1), fo=fH, fi=pH, al=aC, bd=BD)
    r = header_row + 1
    if raw_pl_rows:
        for rd in raw_pl_rows:
            vals = [rd['Company'], rd['Ticker'], rd['Period'], rd['Role'],rd['PL_Source'], rd['account_nm'], rd['calc_key'], rd['Amount_100M']]
            for i, v in enumerate(vals): sc(ws_pl.cell(r, i+1), fo=fHL, fi=ev_fills['PL_HL'], al=aR if i==7 else aL, nf=NB if i==7 else None, bd=BD); ws_pl.cell(r, i+1).value = v
            r += 1
    ws_pl.auto_filter.ref = f"A{header_row}:H{r-1}"; ws_pl.freeze_panes = f"A{header_row+1}"

    # Sheet 3: Market_Cap
    ws_mc = wb.create_sheet('Market_Cap')
    ws_mc.merge_cells('A1:M1'); ws_mc['A1'] = "Market_Cap (Price & Shares)"; sc(ws_mc['A1'], fo=fT)
    ws_mc.merge_cells('A2:M2'); ws_mc['A2'] = "Logic: 종가(FDR) × 유통주식수(DART) | Unit: 억원"; sc(ws_mc['A2'], fo=fS)
    cols = [('Company',15), ('Ticker',10), ('Period',10), ('Price_Date',12), ('Close',12),('Shares',16), ('Market_Cap_100M',18),('Shares_Source',12), ('Shares_RcpNo',16), ('Shares_StlmDt',12), ('Shares_Se',10),('DART_Status',10), ('DART_Message',40)]
    header_row = 4
    ws_mc.append([]); ws_mc.append([c[0] for c in cols])
    for i, (_, w) in enumerate(cols): ws_mc.column_dimensions[get_column_letter(i+1)].width = w; sc(ws_mc.cell(header_row, i+1), fo=fH, fi=pH, al=aC, bd=BD)
    r = header_row + 1
    if all_mkt:
        for rd in all_mkt:
            vals = [rd.get('Company'), rd.get('Ticker'), rd.get('Period'), rd.get('Price_Date'), rd.get('Close'),rd.get('Outstanding_Shares'), rd.get('Market_Cap_100M'),rd.get('Shares_Source'), rd.get('Shares_RceptNo'), rd.get('Shares_StlmDt'), rd.get('Shares_Se'),rd.get('DART_Status'), rd.get('DART_Message')]
            for i, v in enumerate(vals):
                c = ws_mc.cell(r, i+1); c.value = v
                nf = NP if i==4 else (NI_FMT if i==5 else (NB1 if i==6 else None)); al = aR if i in [4,5,6] else aL
                sc(c, fo=fA, fi=pW, al=al, nf=nf, bd=BD)
            r += 1
    ws_mc.auto_filter.ref = f"A{header_row}:M{r-1}"; ws_mc.freeze_panes = f"A{header_row+1}"

    # Sheet 4: LTM_Calc
    ws_ltm = wb.create_sheet('LTM_Calc')
    ws_ltm.merge_cells('A1:I1'); ws_ltm['A1'] = "LTM_Calc (Revenue/EBIT/NI/Pretax Inc)"; sc(ws_ltm['A1'], fo=fT)
    ws_ltm.merge_cells('A2:I2'); ws_ltm['A2'] = "모든 선택 기간별 LTM 계산 내역 | Unit: 억원"; sc(ws_ltm['A2'], fo=fS)
    cols = [('Company',15), ('Ticker',10), ('Period',10), ('Calc_Key',12),('Current_Cum(A)',15), ('Prior_Annual(B)',15), ('Prior_SameQ(C)',15), ('LTM_Value',15), ('Note',10)]
    header_row = 4
    ws_ltm.append([]); ws_ltm.append([c[0] for c in cols])
    for i, (_, w) in enumerate(cols): ws_ltm.column_dimensions[get_column_letter(i+1)].width = w; sc(ws_ltm.cell(header_row, i+1), fo=fH, fi=pH, al=aC, bd=BD)
    r = header_row + 1
    ltm_keys = ['Revenue', 'EBIT', 'NI', 'Pretax_Income']
    for ticker in target_code_list:
        comp_name = ticker_to_name.get(ticker, ticker)
        for tp in target_periods:
            qtr_suffix = tp.split('.')[-1] if '.' in tp else '4Q'
            for k in ltm_keys:
                ws_ltm.cell(r, 1, comp_name); sc(ws_ltm.cell(r, 1), fo=fA, fi=pW, al=aL, bd=BD)
                ws_ltm.cell(r, 2, ticker);    sc(ws_ltm.cell(r, 2), fo=fA, fi=pW, al=aL, bd=BD)
                ws_ltm.cell(r, 3, tp);        sc(ws_ltm.cell(r, 3), fo=fA, fi=pW, al=aL, bd=BD)
                ws_ltm.cell(r, 4, k);         sc(ws_ltm.cell(r, 4), fo=fA, fi=pW, al=aL, bd=BD)
                # Formula: SUMIFS sum_range, r1, criteria1, r2, criteria2...
                ws_ltm.cell(r, 5).value = f'=SUMIFS(PL_Data!H:H, PL_Data!B:B, B{r}, PL_Data!C:C, C{r}, PL_Data!G:G, D{r}, PL_Data!D:D, "current_cum")'; sc(ws_ltm.cell(r,5), fo=fLINK, fi=pW, nf=NB, bd=BD)
                ws_ltm.cell(r, 6).value = f'=SUMIFS(PL_Data!H:H, PL_Data!B:B, B{r}, PL_Data!C:C, C{r}, PL_Data!G:G, D{r}, PL_Data!D:D, "prior_annual")'; sc(ws_ltm.cell(r,6), fo=fLINK, fi=pW, nf=NB, bd=BD)
                ws_ltm.cell(r, 7).value = f'=SUMIFS(PL_Data!H:H, PL_Data!B:B, B{r}, PL_Data!C:C, C{r}, PL_Data!G:G, D{r}, PL_Data!D:D, "prior_same_q")'; sc(ws_ltm.cell(r,7), fo=fLINK, fi=pW, nf=NB, bd=BD)
                if qtr_suffix == '4Q':
                    ws_ltm.cell(r, 8).value = f'=SUMIFS(PL_Data!H:H, PL_Data!B:B, B{r}, PL_Data!C:C, C{r}, PL_Data!G:G, D{r}, PL_Data!D:D, "annual")'; note = 'Annual'
                else:
                    ws_ltm.cell(r, 8).value = f'=E{r}+F{r}-G{r}'; note = 'A+B-C'
                sc(ws_ltm.cell(r,8), fo=fFRM, fi=pW, nf=NB, bd=BD); ws_ltm.cell(r, 9).value = note; sc(ws_ltm.cell(r,9), fo=fA, fi=pW, al=aC, bd=BD)
                r += 1
    ws_ltm.auto_filter.ref = f"A{header_row}:I{r-1}"; ws_ltm.freeze_panes = f"A{header_row+1}"

    # Sheet 3.5: Beta_Calculation
    ws_beta = wb.create_sheet('Beta_Calculation')
    ws_beta.merge_cells('A1:F1')
    sc(ws_beta['A1'], fo=Font(name='Arial', bold=True, size=14, color=C_BL))
    ws_beta['A1'] = 'Beta Calculation (Excel Formulas)'

    ws_beta.merge_cells('A2:F2')
    sc(ws_beta['A2'], fo=Font(name='Arial', size=9, color=C_MG, italic=True))
    ws_beta['A2'] = f'5-Year Monthly & 2-Year Weekly Returns | Base: {base_period_str}'

    r_beta = 4
    beta_result_rows = {}  # ticker: (raw_5y, adj_5y, raw_2y, adj_2y) 매핑

    for idx, ticker in enumerate(target_code_list):
        comp_data = next((item for item in screen_summary_data if item["Ticker"] == ticker), None)
        if not comp_data:
            continue

        company_name = comp_data['Company']
        market_idx = comp_data['Market_Index']

        # ========== 5Y Monthly Beta Section ==========
        ws_beta.merge_cells(f'A{r_beta}:F{r_beta}')
        sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
           fi=PatternFill('solid', fgColor='607D8B'), al=Alignment(horizontal='center'))
        ws_beta.cell(r_beta, 1, f'{company_name} ({ticker}) vs {market_idx} - 5Y Monthly')
        r_beta += 1

        stock_prices_5y = comp_data.get('Stock_Monthly_Prices_5Y')
        market_prices_5y = comp_data.get('Market_Monthly_Prices_5Y')
        raw_5y_row = None
        adj_5y_row = None

        if stock_prices_5y is not None and market_prices_5y is not None and not stock_prices_5y.empty and not market_prices_5y.empty:
            # 헤더
            ws_beta.cell(r_beta, 1, 'Date')
            ws_beta.cell(r_beta, 2, f'{ticker} Price')
            ws_beta.cell(r_beta, 3, f'{market_idx} Price')
            ws_beta.cell(r_beta, 4, f'{ticker} Return')
            ws_beta.cell(r_beta, 5, f'{market_idx} Return')
            for col in range(1, 6):
                sc(ws_beta.cell(r_beta, col), fo=Font(name='Arial', bold=True, size=9, color=C_W),
                   fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
            r_beta += 1

            data_start_row = r_beta

            # 공통 날짜 인덱스
            common_dates = stock_prices_5y.index.intersection(market_prices_5y.index)

            # 데이터 행 작성
            for date in common_dates:
                ws_beta.cell(r_beta, 1, date.strftime('%Y-%m'))
                ws_beta.cell(r_beta, 2, float(stock_prices_5y.loc[date]))
                ws_beta.cell(r_beta, 3, float(market_prices_5y.loc[date]))

                # 수익률 계산 (엑셀 수식)
                if r_beta > data_start_row:
                    ws_beta.cell(r_beta, 4).value = f'=(B{r_beta}-B{r_beta-1})/B{r_beta-1}'
                    ws_beta.cell(r_beta, 5).value = f'=(C{r_beta}-C{r_beta-1})/C{r_beta-1}'
                else:
                    ws_beta.cell(r_beta, 4, None)
                    ws_beta.cell(r_beta, 5, None)

                # 스타일
                sc(ws_beta.cell(r_beta, 1), fo=fA, al=aC, bd=BD)
                sc(ws_beta.cell(r_beta, 2), fo=fA, al=aR, bd=BD, nf='#,##0.00')
                sc(ws_beta.cell(r_beta, 3), fo=fA, al=aR, bd=BD, nf='#,##0.00')
                sc(ws_beta.cell(r_beta, 4), fo=fA, al=aR, bd=BD, nf='0.00%')
                sc(ws_beta.cell(r_beta, 5), fo=fA, al=aR, bd=BD, nf='0.00%')

                r_beta += 1

            data_end_row = r_beta - 1

            # 베타 계산 (SLOPE 함수)
            r_beta += 1
            ws_beta.cell(r_beta, 1, 'Raw Beta (5Y Monthly)')
            ws_beta.cell(r_beta, 2).value = f'=SLOPE(D{data_start_row+1}:D{data_end_row},E{data_start_row+1}:E{data_end_row})'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', bold=True, size=9), fi=PatternFill('solid', fgColor='E8F5E9'),
               bd=BD, al=aR, nf='0.0000')
            raw_5y_row = r_beta
            r_beta += 1

            # Adjusted Beta
            ws_beta.cell(r_beta, 1, 'Adjusted Beta (5Y)')
            ws_beta.cell(r_beta, 2).value = f'=2/3*B{r_beta-1}+1/3*1'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', bold=True, size=9), fi=PatternFill('solid', fgColor='E8F5E9'),
               bd=BD, al=aR, nf='0.0000')
            adj_5y_row = r_beta

        else:
            ws_beta.cell(r_beta, 1, 'No 5Y price data available')
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', size=9, color='FF0000'))

        r_beta += 2  # 간격

        # ========== 2Y Weekly Beta Section ==========
        ws_beta.merge_cells(f'A{r_beta}:F{r_beta}')
        sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
           fi=PatternFill('solid', fgColor='455A64'), al=Alignment(horizontal='center'))
        ws_beta.cell(r_beta, 1, f'{company_name} ({ticker}) vs {market_idx} - 2Y Weekly')
        r_beta += 1

        stock_prices_2y = comp_data.get('Stock_Weekly_Prices_2Y')
        market_prices_2y = comp_data.get('Market_Weekly_Prices_2Y')
        raw_2y_row = None
        adj_2y_row = None

        if stock_prices_2y is not None and market_prices_2y is not None and not stock_prices_2y.empty and not market_prices_2y.empty:
            # 헤더
            ws_beta.cell(r_beta, 1, 'Date')
            ws_beta.cell(r_beta, 2, f'{ticker} Price')
            ws_beta.cell(r_beta, 3, f'{market_idx} Price')
            ws_beta.cell(r_beta, 4, f'{ticker} Return')
            ws_beta.cell(r_beta, 5, f'{market_idx} Return')
            for col in range(1, 6):
                sc(ws_beta.cell(r_beta, col), fo=Font(name='Arial', bold=True, size=9, color=C_W),
                   fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
            r_beta += 1

            data_start_row = r_beta

            # 공통 날짜 인덱스
            common_dates = stock_prices_2y.index.intersection(market_prices_2y.index)

            # 데이터 행 작성
            for date in common_dates:
                ws_beta.cell(r_beta, 1, date.strftime('%Y-%m-%d'))
                ws_beta.cell(r_beta, 2, float(stock_prices_2y.loc[date]))
                ws_beta.cell(r_beta, 3, float(market_prices_2y.loc[date]))

                # 수익률 계산 (엑셀 수식)
                if r_beta > data_start_row:
                    ws_beta.cell(r_beta, 4).value = f'=(B{r_beta}-B{r_beta-1})/B{r_beta-1}'
                    ws_beta.cell(r_beta, 5).value = f'=(C{r_beta}-C{r_beta-1})/C{r_beta-1}'
                else:
                    ws_beta.cell(r_beta, 4, None)
                    ws_beta.cell(r_beta, 5, None)

                # 스타일
                sc(ws_beta.cell(r_beta, 1), fo=fA, al=aC, bd=BD)
                sc(ws_beta.cell(r_beta, 2), fo=fA, al=aR, bd=BD, nf='#,##0.00')
                sc(ws_beta.cell(r_beta, 3), fo=fA, al=aR, bd=BD, nf='#,##0.00')
                sc(ws_beta.cell(r_beta, 4), fo=fA, al=aR, bd=BD, nf='0.00%')
                sc(ws_beta.cell(r_beta, 5), fo=fA, al=aR, bd=BD, nf='0.00%')

                r_beta += 1

            data_end_row = r_beta - 1

            # 베타 계산 (SLOPE 함수)
            r_beta += 1
            ws_beta.cell(r_beta, 1, 'Raw Beta (2Y Weekly)')
            ws_beta.cell(r_beta, 2).value = f'=SLOPE(D{data_start_row+1}:D{data_end_row},E{data_start_row+1}:E{data_end_row})'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', bold=True, size=9), fi=PatternFill('solid', fgColor='FFF9C4'),
               bd=BD, al=aR, nf='0.0000')
            raw_2y_row = r_beta
            r_beta += 1

            # Adjusted Beta
            ws_beta.cell(r_beta, 1, 'Adjusted Beta (2Y)')
            ws_beta.cell(r_beta, 2).value = f'=2/3*B{r_beta-1}+1/3*1'
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', bold=True, size=9), bd=BD)
            sc(ws_beta.cell(r_beta, 2), fo=Font(name='Arial', bold=True, size=9), fi=PatternFill('solid', fgColor='FFF9C4'),
               bd=BD, al=aR, nf='0.0000')
            adj_2y_row = r_beta

        else:
            ws_beta.cell(r_beta, 1, 'No 2Y price data available')
            sc(ws_beta.cell(r_beta, 1), fo=Font(name='Arial', size=9, color='FF0000'))

        # 결과 저장
        beta_result_rows[ticker] = (raw_5y_row, adj_5y_row, raw_2y_row, adj_2y_row)

        r_beta += 2  # 다음 회사와 간격

    ws_beta.column_dimensions['A'].width = 15
    ws_beta.column_dimensions['B'].width = 15
    ws_beta.column_dimensions['C'].width = 15
    ws_beta.column_dimensions['D'].width = 15
    ws_beta.column_dimensions['E'].width = 15

    ws_beta.freeze_panes = 'A4'

    # Sheet 4: WACC_Calculation (완전 구현 - GPCM.py와 동일)
    ws_wacc = wb.create_sheet('WACC_Calculation')
    ws_wacc.merge_cells('A1:D1')
    sc(ws_wacc['A1'], fo=Font(name='Arial', bold=True, size=14, color=C_BL))
    ws_wacc['A1'] = 'Target WACC Calculation'

    ws_wacc.merge_cells('A2:D2')
    sc(ws_wacc['A2'], fo=Font(name='Arial', size=9, color=C_MG, italic=True))
    ws_wacc['A2'] = f'Base: {base_period_str} | Peer Average Method'

    # 스타일 정의
    C_MB = '005EB8'
    pWACC_PARAM = PatternFill('solid', fgColor='E3F2FD')
    pWACC_CALC = PatternFill('solid', fgColor='FFF9C4')
    pWACC_RESULT = PatternFill('solid', fgColor='FFE082')

    r_wacc = 4

    # Section 1: Input Parameters
    ws_wacc.merge_cells(f'A{r_wacc}:D{r_wacc}')
    sc(ws_wacc.cell(r_wacc, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
       fi=PatternFill('solid', fgColor=C_MB), al=Alignment(horizontal='center'))
    ws_wacc.cell(r_wacc, 1, '[ 1 ] Input Parameters')
    r_wacc += 1

    # 헤더
    ws_wacc['A' + str(r_wacc)] = 'Parameter'
    ws_wacc['B' + str(r_wacc)] = 'Value'
    ws_wacc['C' + str(r_wacc)] = 'Format'
    ws_wacc['D' + str(r_wacc)] = 'Note'
    for col in ['A', 'B', 'C', 'D']:
        sc(ws_wacc[col + str(r_wacc)], fo=Font(name='Arial', bold=True, size=9, color=C_W),
           fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
    r_wacc += 1

    # Calculate GPCM stats row position for formulas
    # DATA_START = 6 (header_row + 1), DATA_END depends on number of companies
    # Mean row = DATA_END + 2
    n_companies = len(target_code_list)
    DATA_START = 6
    DATA_END = 6 + n_companies - 1
    mean_row = DATA_END + 2

    # 데이터 행 - Input Parameters
    wacc_params = [
        ('Risk-Free Rate (Rf)', target_wacc_data['Rf'], f"{target_wacc_data['Rf']*100:.2f}%", '10-year Korea Treasury Yield'),
        ('Market Risk Premium (MRP)', target_wacc_data['MRP'], f"{target_wacc_data['MRP']*100:.1f}%", '한국공인회계사회 기준'),
        ('Size Premium', target_wacc_data['Size_Premium'], f"{target_wacc_data['Size_Premium']*100:.2f}%", '한국공인회계사회 (시가총액 기준)'),
        ('Kd (Pretax)', target_wacc_data['Kd_Pretax'], f"{target_wacc_data['Kd_Pretax']*100:.1f}%", '세전 타인자본비용 (사용자 입력)'),
        ('Target Tax Rate', target_wacc_data['Target_Tax_Rate'], f"{target_wacc_data['Target_Tax_Rate']*100:.1f}%", '한국 대기업 기준 (지방세 포함)'),
    ]

    for param, value, formatted, note in wacc_params:
        ws_wacc.cell(r_wacc, 1, param)
        ws_wacc.cell(r_wacc, 2, value)
        ws_wacc.cell(r_wacc, 3, formatted)
        ws_wacc.cell(r_wacc, 4, note)
        sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD, al=Alignment(horizontal='left'))
        sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_PARAM, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
        sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
        sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
        r_wacc += 1

    r_wacc += 1

    # Section 2: Peer Analysis
    ws_wacc.merge_cells(f'A{r_wacc}:D{r_wacc}')
    sc(ws_wacc.cell(r_wacc, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
       fi=PatternFill('solid', fgColor=C_MB), al=Alignment(horizontal='center'))
    ws_wacc.cell(r_wacc, 1, '[ 2 ] Peer Analysis')
    r_wacc += 1

    # 헤더
    ws_wacc['A' + str(r_wacc)] = 'Metric'
    ws_wacc['B' + str(r_wacc)] = 'Value'
    ws_wacc['C' + str(r_wacc)] = 'Format'
    ws_wacc['D' + str(r_wacc)] = 'Note'
    for col in ['A', 'B', 'C', 'D']:
        sc(ws_wacc[col + str(r_wacc)], fo=Font(name='Arial', bold=True, size=9, color=C_W),
           fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
    r_wacc += 1

    # Avg Unlevered Beta - 엑셀 수식으로 GPCM 시트 참조
    row_unlevered_beta = r_wacc
    beta_label = "5Y Monthly" if beta_type_input == "5Y" else "2Y Weekly"
    beta_col = 'AH' if beta_type_input == "5Y" else 'AI'  # AH = 컬럼 34 (Unlevered Beta 5Y), AI = 컬럼 35 (Unlevered Beta 2Y)
    ws_wacc.cell(r_wacc, 1, f'Avg Unlevered Beta ({beta_label})')
    ws_wacc.cell(r_wacc, 2).value = f'=GPCM!{beta_col}{mean_row}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Avg_Unlevered_Beta']:.4f}")
    ws_wacc.cell(r_wacc, 4, '피어 평균 (GPCM Mean)')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.0000')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Avg Debt Ratio - 엑셀 수식으로 GPCM 시트 참조
    row_debt_ratio = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Avg Debt Ratio (D/V)')
    ws_wacc.cell(r_wacc, 2).value = f'=GPCM!AG{mean_row}'  # 컬럼 33 (AG) = Debt Ratio (D/V)
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Avg_Debt_Ratio']*100:.1f}%")
    ws_wacc.cell(r_wacc, 4, '피어 평균 자본구조 (GPCM Mean)')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Target D/E Ratio - 엑셀 수식으로 계산
    row_de_ratio = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Target D/E Ratio')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_debt_ratio}/(1-B{row_debt_ratio})'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Target_DE_Ratio']:.4f}")
    ws_wacc.cell(r_wacc, 4, '= D/V ÷ (1 - D/V)')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.0000')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    r_wacc += 1

    # Section 3: Target WACC Calculation
    ws_wacc.merge_cells(f'A{r_wacc}:D{r_wacc}')
    sc(ws_wacc.cell(r_wacc, 1), fo=Font(name='Arial', bold=True, size=10, color=C_W),
       fi=PatternFill('solid', fgColor=C_MB), al=Alignment(horizontal='center'))
    ws_wacc.cell(r_wacc, 1, '[ 3 ] Target WACC Calculation')
    r_wacc += 1

    # 헤더
    ws_wacc['A' + str(r_wacc)] = 'Component'
    ws_wacc['B' + str(r_wacc)] = 'Value'
    ws_wacc['C' + str(r_wacc)] = 'Format'
    ws_wacc['D' + str(r_wacc)] = 'Formula'
    for col in ['A', 'B', 'C', 'D']:
        sc(ws_wacc[col + str(r_wacc)], fo=Font(name='Arial', bold=True, size=9, color=C_W),
           fi=PatternFill('solid', fgColor=C_BL), al=Alignment(horizontal='center'), bd=BD)
    r_wacc += 1

    # Row references for formulas
    row_rf = 6
    row_mrp = 7
    row_size_premium = 8
    row_kd_pretax = 9
    row_tax = 10

    # Relevered Beta
    row_relevered_beta = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Relevered Beta')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_unlevered_beta}*(1+(1-B{row_tax})*B{row_de_ratio})'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Target_Relevered_Beta']:.4f}")
    ws_wacc.cell(r_wacc, 4, 'Unlevered β × (1 + (1 - Tax) × D/E)')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.0000')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Ke (Cost of Equity)
    row_ke = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Ke (Cost of Equity)')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_rf}+B{row_mrp}*B{row_relevered_beta}+B{row_size_premium}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Target_Ke']*100:.2f}%")
    ws_wacc.cell(r_wacc, 4, 'Rf + MRP × Relevered β + Size Premium')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Kd (Aftertax)
    row_kd_aftertax = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Kd (Aftertax)')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_kd_pretax}*(1-B{row_tax})'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Kd_Aftertax']*100:.2f}%")
    ws_wacc.cell(r_wacc, 4, 'Kd (Pretax) × (1 - Tax Rate)')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Equity Weight (E/V)
    row_equity_weight = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Equity Weight (E/V)')
    ws_wacc.cell(r_wacc, 2).value = f'=1-B{row_debt_ratio}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Equity_Weight']*100:.1f}%")
    ws_wacc.cell(r_wacc, 4, '1 - Debt Ratio')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # Debt Weight (D/V)
    row_debt_weight = r_wacc
    ws_wacc.cell(r_wacc, 1, 'Debt Weight (D/V)')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_debt_ratio}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Debt_Weight']*100:.1f}%")
    ws_wacc.cell(r_wacc, 4, 'Debt Ratio')
    sc(ws_wacc.cell(r_wacc, 1), fo=fA, bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=fA, fi=pWACC_CALC, bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=fA, bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG), bd=BD)
    r_wacc += 1

    # 구분선
    ws_wacc.cell(r_wacc, 1, '━━━━━━━━━━━━')
    ws_wacc.cell(r_wacc, 2, None)
    ws_wacc.cell(r_wacc, 3, '━━━━━━━━━━━━')
    ws_wacc.cell(r_wacc, 4, '━━━━━━━━━━━━━━━━━━━━━━━━━━━')
    for col_idx in range(1, 5):
        sc(ws_wacc.cell(r_wacc, col_idx), bd=BD)
    r_wacc += 1

    # WACC (최종 결과)
    row_wacc_final = r_wacc
    ws_wacc.cell(r_wacc, 1, 'WACC')
    ws_wacc.cell(r_wacc, 2).value = f'=B{row_equity_weight}*B{row_ke}+B{row_debt_weight}*B{row_kd_aftertax}'
    ws_wacc.cell(r_wacc, 3, f"{target_wacc_data['Target_WACC']*100:.2f}%")
    ws_wacc.cell(r_wacc, 4, '(E/V) × Ke + (D/V) × Kd (Aftertax)')
    sc(ws_wacc.cell(r_wacc, 1), fo=Font(name='Arial', bold=True, size=10), bd=BD)
    sc(ws_wacc.cell(r_wacc, 2), fo=Font(name='Arial', bold=True, size=10), fi=pWACC_RESULT,
       bd=BD, al=Alignment(horizontal='right'), nf='0.00%')
    sc(ws_wacc.cell(r_wacc, 3), fo=Font(name='Arial', bold=True, size=10), bd=BD, al=Alignment(horizontal='center'))
    sc(ws_wacc.cell(r_wacc, 4), fo=Font(name='Arial', size=8, color=C_MG, italic=True), bd=BD)
    r_wacc += 1

    # 열 너비 조정
    ws_wacc.column_dimensions['A'].width = 25
    ws_wacc.column_dimensions['B'].width = 12
    ws_wacc.column_dimensions['C'].width = 15
    ws_wacc.column_dimensions['D'].width = 40

    ws_wacc.freeze_panes = 'A4'

    # Named Range 정의 (다른 시트에서 참조 가능)
    wb.defined_names['Target_WACC'] = DefinedName('Target_WACC', attr_text=f"'WACC_Calculation'!$B${row_wacc_final}")
    wb.defined_names['Target_Rf'] = DefinedName('Target_Rf', attr_text="'WACC_Calculation'!$B$6")
    wb.defined_names['Target_MRP'] = DefinedName('Target_MRP', attr_text="'WACC_Calculation'!$B$7")
    wb.defined_names['Target_Size_Premium'] = DefinedName('Target_Size_Premium', attr_text="'WACC_Calculation'!$B$8")
    wb.defined_names['Target_Kd_Pretax'] = DefinedName('Target_Kd_Pretax', attr_text="'WACC_Calculation'!$B$9")
    wb.defined_names['Target_Tax_Rate'] = DefinedName('Target_Tax_Rate', attr_text="'WACC_Calculation'!$B$10")

    # 참고용 셀 주소 표시
    ws_wacc['A' + str(r_wacc + 2)] = '[ Named Ranges for Reference ]'
    sc(ws_wacc.cell(r_wacc + 2, 1), fo=Font(name='Arial', bold=True, size=9, color=C_MG, italic=True))
    ws_wacc['A' + str(r_wacc + 3)] = '다른 시트에서 참조: =Target_WACC, =Target_Rf 등'
    sc(ws_wacc.cell(r_wacc + 3, 1), fo=Font(name='Arial', size=8, color=C_MG))

    # Sheet 1: GPCM (맨 앞)
    ws = wb.create_sheet('GPCM')
    wb.move_sheet('GPCM', offset=-6)  # 맨 앞으로 이동 (index 0)
    # 시트 순서: GPCM, WACC_Calculation, Beta_Calculation, BS_Full, PL_Data, Market_Cap, LTM_Calc
    wb.move_sheet('WACC_Calculation', offset=-4)  # GPCM 다음 (index 1)
    wb.move_sheet('Beta_Calculation', offset=-3)  # WACC 다음 (index 2)
    TOTAL_COLS = 36
    ws.merge_cells(f'A1:{get_column_letter(TOTAL_COLS)}1'); ws['A1'] = "GPCM Valuation Summary with Beta Analysis"; sc(ws['A1'], fo=fT)
    ws.merge_cells(f'A2:{get_column_letter(TOTAL_COLS)}2'); ws['A2'] = f"Base: {base_period_str} | Unit: 억원 | EV = MCap + 우선주 + IBD − Cash + NCI − NOA"; sc(ws['A2'], fo=fS)
    add_gpcm_section_row(ws)
    headers = ['Company','Ticker','Base Date','Curr','PL Source','Cash','IBD','NOA','Net Debt','NCI','Equity','EV','Revenue','EBIT','D&A','EBITDA','NI','Price','Shares','Mkt Cap','EV/EBITDA','EV/EBIT','PER','PBR','PSR','β 5Y Raw','β 5Y Adj','β 2Y Raw','β 2Y Adj','Pretax Inc','Tax Rate','D/E Ratio','Debt Ratio (D/V)','Unlevered β 5Y','Unlevered β 2Y','우선주(장부)']
    widths = [18, 10, 11, 6, 13, 13, 13, 13, 13, 12, 13, 15, 13, 13, 10, 13, 13, 12, 15, 15, 12, 12, 10, 10, 10, 10, 10, 10, 10, 13, 9, 10, 10, 12, 12, 13]
    header_row = 5
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        sc(ws.cell(header_row, i, h), fo=fH, fi=pH, al=aC, bd=BD)
    r = header_row + 1
    for ticker in target_code_list:
        comp_name = ticker_to_name.get(ticker, ticker); bg = pST if (r % 2 == 0) else pW
        ws.cell(r,1, comp_name); ws.cell(r,2, ticker); ws.cell(r,3, base_period_str); ws.cell(r,4, 'KRW'); ws.cell(r,5, 'LTM')
        for c in range(1, 6): sc(ws.cell(r,c), fo=fA, fi=bg, al=aL, bd=BD)
        ws.cell(r,6).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "Cash")'; sc(ws.cell(r,6), fo=fLINK, fi=ev_fills['Cash'], nf=NB, bd=BD)
        ws.cell(r,7).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "IBD")'; sc(ws.cell(r,7), fo=fLINK, fi=ev_fills['IBD'], nf=NB, bd=BD)
        ws.cell(r,8).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "NOA")'; sc(ws.cell(r,8), fo=fLINK, fi=ev_fills['NOA'], nf=NB, bd=BD)
        ws.cell(r,9).value = f'=G{r}-F{r}-H{r}'; sc(ws.cell(r,9), fo=fFRM, fi=bg, nf=NB, bd=BD)
        ws.cell(r,10).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "NCI")'; sc(ws.cell(r,10), fo=fLINK, fi=ev_fills['NCI'], nf=NB, bd=BD)
        ws.cell(r,11).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "Equity_Total")'; sc(ws.cell(r,11), fo=fLINK, fi=ev_fills['Equity'], nf=NB, bd=BD)
        ws.cell(r,12).value = f'=T{r}+AJ{r}+G{r}-F{r}+J{r}-H{r}'; sc(ws.cell(r,12), fo=fFRM, fi=bg, nf=NB, bd=BD)
        ws.cell(r,13).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{r}, LTM_Calc!C:C, C{r}, LTM_Calc!D:D, "Revenue")'; sc(ws.cell(r,13), fo=fLINK, fi=ev_fills['PL_HL'], nf=NB, bd=BD)
        ws.cell(r,14).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{r}, LTM_Calc!C:C, C{r}, LTM_Calc!D:D, "EBIT")'; sc(ws.cell(r,14), fo=fLINK, fi=ev_fills['PL_HL'], nf=NB, bd=BD)
        sc(ws.cell(r,15), fi=PatternFill('solid', fgColor='FFFF00'), nf=NB, bd=BD) # D&A 수기
        ws.cell(r,16).value = f'=N{r}+O{r}'; sc(ws.cell(r,16), fo=fFRM, fi=bg, nf=NB, bd=BD)
        ws.cell(r,17).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{r}, LTM_Calc!C:C, C{r}, LTM_Calc!D:D, "NI")'; sc(ws.cell(r,17), fo=fLINK, fi=ev_fills['PL_HL'], nf=NB, bd=BD)
        ws.cell(r,18).value = f'=SUMIFS(Market_Cap!E:E, Market_Cap!B:B, B{r}, Market_Cap!C:C, C{r})'; sc(ws.cell(r,18), fo=fLINK, nf=NP, bd=BD)
        ws.cell(r,19).value = f'=SUMIFS(Market_Cap!F:F, Market_Cap!B:B, B{r}, Market_Cap!C:C, C{r})'; sc(ws.cell(r,19), fo=fLINK, nf=NI_FMT, bd=BD)
        ws.cell(r,20).value = f'=SUMIFS(Market_Cap!G:G, Market_Cap!B:B, B{r}, Market_Cap!C:C, C{r})'; sc(ws.cell(r,20), fo=fLINK, nf=NB1, bd=BD)
        pMULT = PatternFill('solid', fgColor=C_PB)
        ws.cell(r,21).value = f'=IF(P{r}>0, L{r}/P{r}, "N/M")'; sc(ws.cell(r,21), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
        ws.cell(r,22).value = f'=IF(N{r}>0, L{r}/N{r}, "N/M")'; sc(ws.cell(r,22), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
        ws.cell(r,23).value = f'=IF(Q{r}>0, T{r}/Q{r}, "N/M")'; sc(ws.cell(r,23), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
        ws.cell(r,24).value = f'=IF(K{r}>0, T{r}/K{r}, "N/M")'; sc(ws.cell(r,24), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
        ws.cell(r,25).value = f'=IF(M{r}>0, T{r}/M{r}, "N/M")'; sc(ws.cell(r,25), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)

        # Z-AH: Beta & Risk Analysis (26-34번 컬럼)
        pBETA = PatternFill('solid', fgColor='E8F5E9')
        pBETA2 = PatternFill('solid', fgColor='FFF9C4')
        NF_BETA = '0.00;(0.00);"-"'
        NF_PCT = '0.0%;(0.0%);"-"'

        # Beta 값은 Beta_Calculation 시트에서 참조
        beta_rows = beta_result_rows.get(ticker, (None, None, None, None))
        if beta_rows[0]:  # Raw 5Y
            ws.cell(r,26).value = f'=Beta_Calculation!B{beta_rows[0]}'
            sc(ws.cell(r,26), fo=fLINK, fi=pBETA, al=aR, nf=NF_BETA, bd=BD)
        else:
            ws.cell(r,26, ''); sc(ws.cell(r,26), fo=fA, fi=pBETA, al=aR, nf=NF_BETA, bd=BD)

        if beta_rows[1]:  # Adj 5Y
            ws.cell(r,27).value = f'=Beta_Calculation!B{beta_rows[1]}'
            sc(ws.cell(r,27), fo=fLINK, fi=pBETA, al=aR, nf=NF_BETA, bd=BD)
        else:
            ws.cell(r,27, ''); sc(ws.cell(r,27), fo=fA, fi=pBETA, al=aR, nf=NF_BETA, bd=BD)

        if beta_rows[2]:  # Raw 2Y
            ws.cell(r,28).value = f'=Beta_Calculation!B{beta_rows[2]}'
            sc(ws.cell(r,28), fo=fLINK, fi=pBETA2, al=aR, nf=NF_BETA, bd=BD)
        else:
            ws.cell(r,28, ''); sc(ws.cell(r,28), fo=fA, fi=pBETA2, al=aR, nf=NF_BETA, bd=BD)

        if beta_rows[3]:  # Adj 2Y
            ws.cell(r,29).value = f'=Beta_Calculation!B{beta_rows[3]}'
            sc(ws.cell(r,29), fo=fLINK, fi=pBETA2, al=aR, nf=NF_BETA, bd=BD)
        else:
            ws.cell(r,29, ''); sc(ws.cell(r,29), fo=fA, fi=pBETA2, al=aR, nf=NF_BETA, bd=BD)

        # 컬럼 30: Pretax Inc (LTM_Calc에서 참조)
        ws.cell(r,30).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{r}, LTM_Calc!C:C, C{r}, LTM_Calc!D:D, "Pretax_Income")'; sc(ws.cell(r,30), fo=fLINK, fi=bg, al=aR, nf=NB, bd=BD)

        # 컬럼 31: Tax Rate (한국 법인세 한계세율, 사업연도별 세율표, 지방소득세 포함)
        # FY2023~2025: 2억 이하 9.9% / 2~200억 20.9% / 200~3000억 23.1% / 3000억 초과 26.4%
        # FY2026~    : 2억 이하 11.0% / 2~200억 22.0% / 200~3000억 24.2% / 3000억 초과 27.5%
        ws.cell(r,31).value = f'=IF(AD{r}<=2, 0.099, IF(AD{r}<=200, 0.209, IF(AD{r}<=3000, 0.231, 0.264)))'
        sc(ws.cell(r,31), fo=fFRM, fi=bg, al=aR, nf=NF_PCT, bd=BD)

        # 컬럼 32: D/E Ratio = IBD / (Mkt Cap + 우선주 + NCI)
        ws.cell(r,32).value = f'=IF(T{r}+AJ{r}+J{r}>0, G{r}/(T{r}+AJ{r}+J{r}), 0)'; sc(ws.cell(r,32), fo=fFRM, fi=bg, al=aR, nf=NF_PCT, bd=BD)

        # 컬럼 33: Debt Ratio (D/V) = IBD / (Mkt Cap + 우선주 + IBD + NCI)
        ws.cell(r,33).value = f'=IF(T{r}+AJ{r}+G{r}+J{r}>0, G{r}/(T{r}+AJ{r}+G{r}+J{r}), 0)'; sc(ws.cell(r,33), fo=fFRM, fi=bg, al=aR, nf=NF_PCT, bd=BD)

        # 컬럼 34: Unlevered Beta 5Y = β 5Y Adj / (1 + (1 - Tax Rate) × D/E Ratio)
        # 컬럼 27 (AA) = β 5Y Adj, 컬럼 31 (AE) = Tax Rate, 컬럼 32 (AF) = D/E Ratio
        ws.cell(r,34).value = f'=IF(AA{r}>0, AA{r}/(1+(1-AE{r})*AF{r}), "")'; sc(ws.cell(r,34), fo=fFRM, fi=pBETA, al=aR, nf=NF_BETA, bd=BD)

        # 컬럼 35: Unlevered Beta 2Y = β 2Y Adj / (1 + (1 - Tax Rate) × D/E Ratio)
        # 컬럼 29 (AC) = β 2Y Adj, 컬럼 31 (AE) = Tax Rate, 컬럼 32 (AF) = D/E Ratio
        ws.cell(r,35).value = f'=IF(AC{r}>0, AC{r}/(1+(1-AE{r})*AF{r}), "")'; sc(ws.cell(r,35), fo=fFRM, fi=pBETA2, al=aR, nf=NF_BETA, bd=BD)

        # 컬럼 36: 우선주 자본금 (시가총액은 보통주만 반영하므로 자기자본가치에 별도 가산)
        ws.cell(r,36).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{r}, BS_Full!C:C, C{r}, BS_Full!G:G, "Preferred")'
        sc(ws.cell(r,36), fo=fLINK, fi=ev_fills['Equity'], nf=NB, bd=BD)
        r += 1
    r_end = r - 1; r += 1
    for stat, fn in [('Mean','AVERAGE'), ('Median','MEDIAN'), ('Max','MAX'), ('Min','MIN')]:
        ws.cell(r, 20, stat); sc(ws.cell(r,20), fo=fSTAT, fi=pSTAT, al=aC, bd=BD)
        # Valuation Multiples (21-25)
        for c in range(21, 26):
            col = get_column_letter(c)
            ws.cell(r, c).value = f'=IFERROR({fn}({col}{header_row+1}:{col}{r_end}), "N/M")'
            sc(ws.cell(r,c), fo=fSTAT, fi=pSTAT, nf=NF_X, bd=BD)
        # Beta & Risk (26-35)
        for c in range(26, 36):
            col = get_column_letter(c)
            if c in [26, 27, 28, 29, 34, 35]:  # Beta 컬럼 (34=Unlevered β 5Y, 35=Unlevered β 2Y)
                ws.cell(r, c).value = f'=IFERROR({fn}({col}{header_row+1}:{col}{r_end}), "")'
                sc(ws.cell(r,c), fo=fSTAT, fi=pSTAT, nf=NF_BETA, bd=BD)
            elif c == 31:  # Tax Rate
                ws.cell(r, c).value = f'=IFERROR({fn}({col}{header_row+1}:{col}{r_end}), "")'
                sc(ws.cell(r,c), fo=fSTAT, fi=pSTAT, nf=NF_PCT, bd=BD)
            elif c in [32, 33]:  # D/E Ratio, Debt Ratio (D/V)
                ws.cell(r, c).value = f'=IFERROR({fn}({col}{header_row+1}:{col}{r_end}), "")'
                sc(ws.cell(r,c), fo=fSTAT, fi=pSTAT, nf=NF_PCT, bd=BD)
            else:  # Pretax Inc (30)
                ws.cell(r, c).value = f'=IFERROR({fn}({col}{header_row+1}:{col}{r_end}), "")'
                sc(ws.cell(r,c), fo=fSTAT, fi=pSTAT, nf=NB, bd=BD)
        r += 1
    r += 2
    for note in notes_list: ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS); sc(ws.cell(r, 1, note), fo=fNOTE); r += 1
    ws.freeze_panes = f"F{header_row+1}"  # Cash 컬럼부터 스크롤

    # === Multiples_Trend Sheet generation ===
    ws_trend = wb.create_sheet('Multiples_Trend')
    ws_trend.merge_cells('A1:M1'); ws_trend['A1'] = "Multiples Trend (PER, PBR, PSR, EV/EBIT)"; sc(ws_trend['A1'], fo=fT)
    ws_trend.merge_cells('A2:M2'); ws_trend['A2'] = "모든 타겟 기간별 Valuation Multiples 흐름 요약 (Formula 기반)"; sc(ws_trend['A2'], fo=fS)
    
    # 0:Comp, 1:Tick, 2:Per, 3:MC, 4:EV, 5:Rev, 6:EBIT, 7:NI, 8:Eq, 9:EV/EB, 10:PER, 11:PSR, 12:PBR
    cols_t = [('Company',15), ('Ticker',10), ('Period',10), ('Market_Cap',15), ('EV',15), ('Revenue(LTM)',15), ('EBIT(LTM)',15), ('NI(LTM)',15), ('Equity', 15), ('EV/EBIT',12), ('PER',10), ('PSR',10), ('PBR', 10)]
    header_row_t = 4
    ws_trend.append([]); ws_trend.append([c[0] for c in cols_t])
    for i, (_, w) in enumerate(cols_t): ws_trend.column_dimensions[get_column_letter(i+1)].width = w; sc(ws_trend.cell(header_row_t, i+1), fo=fH, fi=pH, al=aC, bd=BD)
    
    rt = header_row_t + 1
    if df_screen is not None and not df_screen.empty:
        # Ticker x Period 순회 (df_screen 기반)
        for _, row_data in df_screen.iterrows():
            ticker = row_data.get('Ticker')
            period = row_data.get('Period')
            comp_name = row_data.get('Company')
            
            # Static basic info
            ws_trend.cell(rt, 1, comp_name); sc(ws_trend.cell(rt, 1), fo=fA, fi=pW, al=aL, bd=BD)
            ws_trend.cell(rt, 2, ticker);    sc(ws_trend.cell(rt, 2), fo=fA, fi=pW, al=aL, bd=BD)
            ws_trend.cell(rt, 3, period);    sc(ws_trend.cell(rt, 3), fo=fA, fi=pW, al=aL, bd=BD)
            
            # MC & EV from Market_Cap sheet
            ws_trend.cell(rt, 4).value = f'=SUMIFS(Market_Cap!G:G, Market_Cap!B:B, B{rt}, Market_Cap!C:C, C{rt})'; sc(ws_trend.cell(rt, 4), fo=fLINK, nf=NB, bd=BD)
            ws_trend.cell(rt, 5).value = f'=D{rt}+SUMIFS(BS_Full!H:H, BS_Full!B:B, B{rt}, BS_Full!C:C, C{rt}, BS_Full!G:G, "IBD")-SUMIFS(BS_Full!H:H, BS_Full!B:B, B{rt}, BS_Full!C:C, C{rt}, BS_Full!G:G, "Cash")+SUMIFS(BS_Full!H:H, BS_Full!B:B, B{rt}, BS_Full!C:C, C{rt}, BS_Full!G:G, "NCI")-SUMIFS(BS_Full!H:H, BS_Full!B:B, B{rt}, BS_Full!C:C, C{rt}, BS_Full!G:G, "NOA")'
            sc(ws_trend.cell(rt, 5), fo=fFRM, nf=NB, bd=BD)
            
            # LTM Figures from LTM_Calc (Sumifs calc_key)
            ws_trend.cell(rt, 6).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{rt}, LTM_Calc!C:C, C{rt}, LTM_Calc!D:D, "Revenue")'; sc(ws_trend.cell(rt, 6), fo=fLINK, nf=NB, bd=BD)
            ws_trend.cell(rt, 7).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{rt}, LTM_Calc!C:C, C{rt}, LTM_Calc!D:D, "EBIT")'; sc(ws_trend.cell(rt, 7), fo=fLINK, nf=NB, bd=BD)
            ws_trend.cell(rt, 8).value = f'=SUMIFS(LTM_Calc!H:H, LTM_Calc!B:B, B{rt}, LTM_Calc!C:C, C{rt}, LTM_Calc!D:D, "NI")'; sc(ws_trend.cell(rt, 8), fo=fLINK, nf=NB, bd=BD)
            
            # Equity from BS_Full
            ws_trend.cell(rt, 9).value = f'=SUMIFS(BS_Full!H:H, BS_Full!B:B, B{rt}, BS_Full!C:C, C{rt}, BS_Full!G:G, "Equity_Total")'; sc(ws_trend.cell(rt, 9), fo=fLINK, nf=NB, bd=BD)
            
            # Multiples by Formula
            pMULT = PatternFill('solid', fgColor=C_PB)
            ws_trend.cell(rt, 10).value = f'=IF(G{rt}>0, E{rt}/G{rt}, "N/M")'; sc(ws_trend.cell(rt, 10), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
            ws_trend.cell(rt, 11).value = f'=IF(H{rt}>0, D{rt}/H{rt}, "N/M")'; sc(ws_trend.cell(rt, 11), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
            ws_trend.cell(rt, 12).value = f'=IF(F{rt}>0, D{rt}/F{rt}, "N/M")'; sc(ws_trend.cell(rt, 12), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
            ws_trend.cell(rt, 13).value = f'=IF(I{rt}>0, D{rt}/I{rt}, "N/M")'; sc(ws_trend.cell(rt, 13), fo=fMUL, fi=pMULT, nf=NF_X, bd=BD)
            
            rt += 1
            
    ws_trend.auto_filter.ref = f"A{header_row_t}:M{rt-1}"
    ws_trend.freeze_panes = f"D{header_row_t+1}" # Scroll from Market_Cap

    # === Data_Quality Sheet — 자동 수집이 채우지 못한 자리 ===
    # 못 가져온 값은 0으로 남고 나머지 계산은 계속 돌아간다. 파일만 받아 본 사람이
    # 그 0을 '정말 0'으로 읽지 않도록, 어디가 비었는지 여기에 모아 둔다.
    ws_dq = wb.create_sheet('Data_Quality')
    ws_dq.sheet_properties.tabColor = 'EF6C00'

    ws_dq.merge_cells('A1:E1'); sc(ws_dq.cell(1, 1, 'Data Quality Check (확인 필요 항목)'), fo=fT)
    ws_dq.merge_cells('A2:E2')
    sc(ws_dq.cell(2, 1,
                  'ERROR = 그 값을 쓰면 안 됨 / WARN = 값은 나왔으나 왜곡 가능 / INFO = 참고. '
                  '자동 수집의 한계를 표시한 것이므로 공시자료와 대조 검증하십시오.'), fo=fS)

    dq_cols = [('Level', 10), ('Ticker', 12), ('Company', 20), ('Item', 20), ('Message', 100)]
    dq_hdr = 4
    for ci, (h, w) in enumerate(dq_cols, 1):
        ws_dq.column_dimensions[get_column_letter(ci)].width = w
        sc(ws_dq.cell(dq_hdr, ci, h), fo=fH, fi=pH, al=aC, bd=BD)

    level_fill = {SEV_ERROR: PatternFill('solid', fgColor='FFCDD2'),
                  SEV_WARN: PatternFill('solid', fgColor='FFE0B2'),
                  SEV_INFO: PatternFill('solid', fgColor='E3F2FD')}
    level_order = {SEV_ERROR: 0, SEV_WARN: 1, SEV_INFO: 2}
    dq_rows = sorted(quality.rows,
                     key=lambda x: (level_order.get(x.get('Level'), 9),
                                    str(x.get('Ticker', '')), str(x.get('Item', ''))))

    r_dq = dq_hdr + 1
    if dq_rows:
        for fl in dq_rows:
            lv = fl.get('Level', SEV_INFO)
            vals = [lv, fl.get('Ticker', ''), fl.get('Company', ''),
                    fl.get('Item', ''), fl.get('Message', '')]
            for ci, v in enumerate(vals, 1):
                sc(ws_dq.cell(r_dq, ci, v), fo=fA, fi=level_fill.get(lv, pW), bd=BD,
                   al=aC if ci == 1 else aL)
            r_dq += 1
    else:
        ws_dq.merge_cells(start_row=r_dq, start_column=1, end_row=r_dq, end_column=5)
        sc(ws_dq.cell(r_dq, 1, '✅ 자동 점검에서 특이사항이 발견되지 않았습니다. '
                               '점검 대상이 아닌 항목은 여전히 직접 확인이 필요합니다.'), fo=fA)
        r_dq += 1

    ws_dq.auto_filter.ref = f"A{dq_hdr}:E{r_dq-1}"
    ws_dq.freeze_panes = f'A{dq_hdr+1}'
    # GPCM 바로 뒤 — 숫자를 보기 전에 눈에 걸리도록 맨 앞쪽에 둔다
    wb.move_sheet('Data_Quality', offset=1 - wb.sheetnames.index('Data_Quality'))

    # === Peer_Selection Sheet — 업종에서 고른 경우, 무엇을 보고 골랐는지 ===
    # 후보 목록은 실행 시점의 상장 현황이라 나중에 다시 만들 수 없다. "Peer 를 왜
    # 이 회사들로 했나"에 답하려면 그때 무엇이 떠 있었는지가 파일에 남아 있어야 한다.
    if peer_selection:
        ws_ps = wb.create_sheet('Peer_Selection')
        ws_ps.sheet_properties.tabColor = '00695C'
        picked_set = set(peer_selection.get('picked') or [])

        ws_ps.merge_cells('A1:F1'); sc(ws_ps.cell(1, 1, 'Peer Selection (모집단과 선택 근거)'), fo=fT)
        ws_ps.merge_cells('A2:F2')
        sc(ws_ps.cell(2, 1,
                      f"업종: {peer_selection.get('sector', '')} | "
                      f"후보 {len(peer_selection.get('candidates') or [])}개 중 {len(picked_set)}개 선택 | "
                      f"조회 시점: {datetime.now().strftime('%Y-%m-%d %H:%M')} — "
                      f"상장·폐지에 따라 후보 목록은 시점마다 달라집니다."), fo=fS)

        ps_cols = [('선택', 8), ('Code', 12), ('Name', 24), ('Market', 10),
                   ('주요제품', 60), ('결산월', 10)]
        ps_hdr = 4
        for ci, (h, w) in enumerate(ps_cols, 1):
            ws_ps.column_dimensions[get_column_letter(ci)].width = w
            sc(ws_ps.cell(ps_hdr, ci, h), fo=fH, fi=pH, al=aC, bd=BD)

        pPICK = PatternFill('solid', fgColor='C8E6C9')
        r_ps = ps_hdr + 1
        for c in (peer_selection.get('candidates') or []):
            chosen = c['Code'] in picked_set
            vals = ['●' if chosen else '', c['Code'], c['Name'], c['Market'],
                    c['Product'], c['SettleMonth']]
            for ci, v in enumerate(vals, 1):
                sc(ws_ps.cell(r_ps, ci, v), fo=fA,
                   fi=pPICK if chosen else (PatternFill('solid', fgColor='FFE0B2') if c['FiscalNot12'] else pW),
                   al=aC if ci in (1, 2, 4, 6) else aL, bd=BD)
            r_ps += 1

        ws_ps.auto_filter.ref = f"A{ps_hdr}:F{r_ps-1}"
        ws_ps.freeze_panes = f'A{ps_hdr+1}'
        wb.move_sheet('Peer_Selection', offset=2 - wb.sheetnames.index('Peer_Selection'))

    wb.save(output)
    output.seek(0)

    return output
