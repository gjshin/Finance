import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from io import BytesIO
import plotly.colors
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- 1. 디자인 시스템 & 설정 ---
st.set_page_config(layout="wide", page_title="Integrated Sales Dashboard (MBR)", page_icon="📈")
COLORS = {
    'primary': '#00338D',    # KPMG Blue
    'secondary': '#0091DA',  # Light Blue
    'accent': '#00A3A1',     # Teal
    'negative': '#D32929',   # Red
    'positive': '#009944',   # Green
    'neutral': '#6D6E71',    # Grey
    'sequence': plotly.colors.qualitative.Prism
}

# --- 2. 데이터 로드 및 통합 전처리 ---
@st.cache_data(ttl=3600)
def load_all_data(file_obj):
    try:
        all_dfs = []
        
        # 1) Bumper
        cols_bumper = "A, C, R, Y, AA, AB, AI"
        df_b = pd.read_excel(file_obj, sheet_name='bumper', usecols=cols_bumper)
        df_b.columns = ['Year_Raw', 'Date', 'Product_Group', 'Sub_Product', 'Customer', 'Quantity', 'Revenue']
        df_b['Category'] = 'Bumper'
        all_dfs.append(df_b)
        
        # 2) Sill side
        cols_sill = "A, E, I, N, U"
        df_s = pd.read_excel(file_obj, sheet_name='Sill side', usecols=cols_sill)
        df_s.columns = ['Year_Raw', 'Date', 'Product_Group', 'Quantity', 'Revenue']
        df_s['Category'] = 'Sill side'
        df_s['Sub_Product'] = '-'
        df_s['Customer'] = '-'
        all_dfs.append(df_s)
        
        # 3) Carrier
        cols_carrier = "A, E, I, N, U"
        df_c = pd.read_excel(file_obj, sheet_name='Carrier', usecols=cols_carrier)
        df_c.columns = ['Year_Raw', 'Date', 'Product_Group', 'Quantity', 'Revenue']
        df_c['Category'] = 'Carrier'
        df_c['Sub_Product'] = '-'
        df_c['Customer'] = '-'
        all_dfs.append(df_c)
        
        # 통합
        df = pd.concat(all_dfs, ignore_index=True)
        
        # 공통 전처리
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date'])
        df['Year'] = df['Date'].dt.year
        
        df['Quantity'] = df['Quantity'].fillna(0).astype(float)
        df['Revenue'] = df['Revenue'].fillna(0).astype(float)
        df['Product_Group'] = df['Product_Group'].fillna("Unknown")
        
        # Clean Data
        df = df[~((df['Revenue'] == 0) & (df['Quantity'] == 0))]
        
        # 파생 변수
        df['Month_Dt'] = df['Date'].dt.to_period('M').dt.to_timestamp()
        df['Quarter_Str'] = df['Date'].dt.to_period('Q').astype(str)
        
        # ASP
        df['ASP'] = np.where(df['Quantity'] > 0, df['Revenue'] / df['Quantity'], 0)
        
        return df
        
    except ValueError as ve:
        st.error(f"❌ 엑셀 구조 오류: {ve}")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"❌ 데이터 로드 오류: {e}")
        return pd.DataFrame()

# --- 3. 분석 로직 (P/V/M Bridge) ---
def calculate_bridge_mix_steps(df, years, group_col):
    all_steps = []
    if len(years) < 2: return []
    years = sorted(years)
    
    for i in range(len(years) - 1):
        y0, y1 = years[i], years[i+1]
        
        d0 = df[df['Year'] == y0].groupby(group_col).agg({'Revenue':'sum', 'Quantity':'sum'}).reset_index()
        d1 = df[df['Year'] == y1].groupby(group_col).agg({'Revenue':'sum', 'Quantity':'sum'}).reset_index()
        
        total_q0 = d0['Quantity'].sum()
        total_q1 = d1['Quantity'].sum()
        total_r0 = d0['Revenue'].sum()
        total_r1 = d1['Revenue'].sum()
        
        avg_p0 = total_r0 / total_q0 if total_q0 > 0 else 0
        
        merged = pd.merge(d0, d1, on=group_col, how='outer', suffixes=('_0', '_1')).fillna(0)
        
        merged['P0'] = np.where(merged['Quantity_0'] > 0, merged['Revenue_0'] / merged['Quantity_0'], 0)
        merged['P1'] = np.where(merged['Quantity_1'] > 0, merged['Revenue_1'] / merged['Quantity_1'], 0)
        
        vol_effect = (total_q1 - total_q0) * avg_p0
        merged['Price_Impact'] = (merged['P1'] - merged['P0']) * merged['Quantity_1']
        price_effect = merged['Price_Impact'].sum()
        
        total_variance = total_r1 - total_r0
        mix_effect = total_variance - vol_effect - price_effect
        
        all_steps.append({
            'label': f"{y0}→{y1}",
            'start_year': str(y0),
            'end_year': str(y1),
            'volume': vol_effect,
            'price': price_effect,
            'mix': mix_effect,
            'start_val': total_r0,
            'end_val': total_r1
        })
        
    return all_steps

def create_waterfall_fig_multi(steps):
    if not steps: return go.Figure()
    
    first_step = steps[0]
    
    x_labels = [first_step['start_year']]
    y_vals = [first_step['start_val']]
    measures = ["absolute"]
    text_vals = [f"R$ {first_step['start_val']/1000000:,.1f}M"]
    
    for step in steps:
        for eff_name, eff_val in [("Volume", step['volume']), ("Mix", step['mix']), ("Price", step['price'])]:
            x_labels.append(f"{eff_name}<br>({step['label']})")
            y_vals.append(eff_val)
            measures.append("relative")
            text_vals.append(f"{eff_val/1000000:+,.1f}M")
        
        x_labels.append(step['end_year'])
        y_vals.append(step['end_val'])
        measures.append("absolute")
        text_vals.append(f"R$ {step['end_val']/1000000:,.1f}M")
    
    fig = go.Figure(go.Waterfall(
        orientation="v",
        measure=measures,
        x=x_labels,
        y=y_vals,
        text=text_vals,
        textposition="outside",
        decreasing={"marker":{"color": COLORS['negative']}},
        increasing={"marker":{"color": COLORS['accent']}},
        totals={"marker":{"color": COLORS['primary']}},
        connector={"line":{"color":"#666"}}
    ))
    
    fig.update_layout(title_text="Multi-Year Price-Volume-Mix Bridge", xaxis_type='category')
    return fig

def detect_outliers_iqr(df, col='ASP', factor=1.5):
    if df.empty: return df
    q1 = df[col].quantile(0.25)
    q3 = df[col].quantile(0.75)
    iqr = q3 - q1
    lower, upper = q1 - (factor * iqr), q3 + (factor * iqr)
    return df[(df[col] < lower) | (df[col] > upper)]


# --- 4. Excel 리포트 생성 함수 ---
@st.cache_data(ttl=3600, show_spinner="📊 전체 엑셀 리포트 생성 중...")
def generate_excel_report(df, all_years_list):
    """Generate comprehensive KPMG-styled Excel report with all categories"""
    wb = Workbook()
    wb.remove(wb.active)
    
    KPMG_BLUE = "00338D"
    LIGHT_BLUE = "E8EEF7"
    WHITE = "FFFFFF"
    
    header_fill = PatternFill(fill_type="solid", fgColor=KPMG_BLUE)
    header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    bold_font = Font(name="Calibri", bold=True, size=10)
    title_font = Font(name="Calibri", bold=True, color=KPMG_BLUE, size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    
    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
    
    def add_section_title(ws, title, row, num_cols):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row, 1, title)
        cell.font = title_font
        cell.alignment = Alignment(horizontal='left')
        ws.row_dimensions[row].height = 20
    
    def create_growth_share_sheet(ws_name, data_df, category_name):
        ws = wb.create_sheet(ws_name)
        
        # Section 1: Yearly Revenue + YoY
        y_df = data_df.groupby('Year')['Revenue'].sum().reset_index()
        y_df['YoY'] = y_df['Revenue'].pct_change()
        
        add_section_title(ws, f"1. [{category_name}] Yearly Revenue & YoY Growth", 1, 3)
        for col, h in enumerate(['Year', 'Revenue', 'YoY %'], 1):
            cell = ws.cell(2, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for i, (_, row) in enumerate(y_df.iterrows()):
            dr = 3 + i
            ws.cell(dr, 1, int(row['Year']))
            c = ws.cell(dr, 2, row['Revenue']); c.number_format = '#,##0'
            yoy_val = row['YoY']
            if pd.notnull(yoy_val):
                c = ws.cell(dr, 3, yoy_val); c.number_format = '0.0%'
            else:
                ws.cell(dr, 3, '-')
        
        # Section 2: Revenue Share by Product Group
        comp_df = data_df.groupby(['Year', 'Product_Group'])['Revenue'].sum().reset_index()
        comp_df['Total'] = comp_df.groupby('Year')['Revenue'].transform('sum')
        comp_df['Share'] = comp_df['Revenue'] / comp_df['Total']
        
        s2_start = len(y_df) + 4
        add_section_title(ws, f"2. [{category_name}] Revenue Share by Product Group", s2_start, 4)
        for col, h in enumerate(['Year', 'Product_Group', 'Revenue', 'Share'], 1):
            cell = ws.cell(s2_start+1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for i, (_, row) in enumerate(comp_df.iterrows()):
            dr = s2_start + 2 + i
            ws.cell(dr, 1, int(row['Year']))
            ws.cell(dr, 2, row['Product_Group'])
            c = ws.cell(dr, 3, row['Revenue']); c.number_format = '#,##0'
            c = ws.cell(dr, 4, row['Share']); c.number_format = '0.0%'
        
        # Section 3: Volume & ASP Trend by Product Group
        trend_det = data_df.groupby(['Year', 'Product_Group']).agg({'Quantity': 'sum', 'Revenue': 'sum'}).reset_index()
        trend_det['ASP'] = trend_det['Revenue'] / trend_det['Quantity']
        
        s3_start = s2_start + len(comp_df) + 3
        add_section_title(ws, f"3. [{category_name}] Volume & ASP Trend by Product Group", s3_start, 4)
        for col, h in enumerate(['Year', 'Product_Group', 'Quantity', 'ASP'], 1):
            cell = ws.cell(s3_start+1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for i, (_, row) in enumerate(trend_det.iterrows()):
            dr = s3_start + 2 + i
            ws.cell(dr, 1, int(row['Year']))
            ws.cell(dr, 2, row['Product_Group'])
            c = ws.cell(dr, 3, row['Quantity']); c.number_format = '#,##0'
            c = ws.cell(dr, 4, row['ASP']); c.number_format = '#,##0'
        
        auto_width(ws)
    
    def create_timeseries_sheet(ws_name, data_df, category_name):
        ws = wb.create_sheet(ws_name)
        
        mix_col_ts = 'Category' if category_name == "All Categories" else 'Product_Group'
        
        # Monthly
        t_df_m = data_df.groupby(['Month_Dt', mix_col_ts])['Revenue'].sum().reset_index()
        t_df_m['Total'] = t_df_m.groupby('Month_Dt')['Revenue'].transform('sum')
        t_df_m['Share'] = t_df_m['Revenue'] / t_df_m['Total']
        t_df_m['Period'] = t_df_m['Month_Dt'].dt.strftime('%Y-%m')
        
        agg_m = data_df.groupby('Month_Dt').agg({'Revenue': 'sum', 'Quantity': 'sum'}).reset_index()
        agg_m['ASP'] = np.where(agg_m['Quantity'] > 0, agg_m['Revenue'] / agg_m['Quantity'], 0)
        agg_m['Period'] = agg_m['Month_Dt'].dt.strftime('%Y-%m')
        
        add_section_title(ws, f"1. [{category_name}] Monthly Revenue Mix by {mix_col_ts}", 1, 4)
        for col, h in enumerate(['Period', mix_col_ts, 'Revenue', 'Share'], 1):
            cell = ws.cell(2, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for i, (_, row) in enumerate(t_df_m.iterrows()):
            dr = 3 + i
            ws.cell(dr, 1, row['Period'])
            ws.cell(dr, 2, row[mix_col_ts])
            c = ws.cell(dr, 3, row['Revenue']); c.number_format = '#,##0'
            c = ws.cell(dr, 4, row['Share']); c.number_format = '0.0%'
        
        m_agg_start = len(t_df_m) + 4
        add_section_title(ws, f"2. [{category_name}] Monthly Quantity & ASP Trend", m_agg_start, 3)
        for col, h in enumerate(['Period', 'Quantity', 'ASP'], 1):
            cell = ws.cell(m_agg_start+1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for i, (_, row) in enumerate(agg_m.iterrows()):
            dr = m_agg_start + 2 + i
            ws.cell(dr, 1, row['Period'])
            c = ws.cell(dr, 2, row['Quantity']); c.number_format = '#,##0'
            c = ws.cell(dr, 3, row['ASP']); c.number_format = '#,##0'
        
        # Quarterly
        t_df_q = data_df.groupby(['Quarter_Str', mix_col_ts])['Revenue'].sum().reset_index().sort_values('Quarter_Str')
        t_df_q['Total'] = t_df_q.groupby('Quarter_Str')['Revenue'].transform('sum')
        t_df_q['Share'] = t_df_q['Revenue'] / t_df_q['Total']
        
        agg_q = data_df.groupby('Quarter_Str').agg({'Revenue': 'sum', 'Quantity': 'sum'}).reset_index().sort_values('Quarter_Str')
        agg_q['ASP'] = np.where(agg_q['Quantity'] > 0, agg_q['Revenue'] / agg_q['Quantity'], 0)
        
        q_start = m_agg_start + len(agg_m) + 3
        add_section_title(ws, f"3. [{category_name}] Quarterly Revenue Mix by {mix_col_ts}", q_start, 4)
        for col, h in enumerate(['Quarter', mix_col_ts, 'Revenue', 'Share'], 1):
            cell = ws.cell(q_start+1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for i, (_, row) in enumerate(t_df_q.iterrows()):
            dr = q_start + 2 + i
            ws.cell(dr, 1, row['Quarter_Str'])
            ws.cell(dr, 2, row[mix_col_ts])
            c = ws.cell(dr, 3, row['Revenue']); c.number_format = '#,##0'
            c = ws.cell(dr, 4, row['Share']); c.number_format = '0.0%'
        
        q_agg_start = q_start + len(t_df_q) + 3
        add_section_title(ws, f"4. [{category_name}] Quarterly Quantity & ASP Trend", q_agg_start, 3)
        for col, h in enumerate(['Quarter', 'Quantity', 'ASP'], 1):
            cell = ws.cell(q_agg_start+1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for i, (_, row) in enumerate(agg_q.iterrows()):
            dr = q_agg_start + 2 + i
            ws.cell(dr, 1, row['Quarter_Str'])
            c = ws.cell(dr, 2, row['Quantity']); c.number_format = '#,##0'
            c = ws.cell(dr, 3, row['ASP']); c.number_format = '#,##0'
        
        auto_width(ws)
    
    def create_outlier_sheet(ws_name, data_df, category_name):
        ws = wb.create_sheet(ws_name)

        stats = data_df.groupby('Product_Group')['ASP'].agg(['count', 'mean', 'min', 'max']).reset_index()
        stats.columns = ['Product_Group', 'Count', 'Mean ASP', 'Min ASP', 'Max ASP']

        add_section_title(ws, f"1. [{category_name}] ASP Statistics by Product Group", 1, 5)
        for col, h in enumerate(stats.columns.tolist(), 1):
            cell = ws.cell(2, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for i, (_, row) in enumerate(stats.iterrows()):
            dr = 3 + i
            ws.cell(dr, 1, row['Product_Group'])
            ws.cell(dr, 2, int(row['Count']))
            for col, key in [(3, 'Mean ASP'), (4, 'Min ASP'), (5, 'Max ASP')]:
                c = ws.cell(dr, col, row[key]); c.number_format = '#,##0'

        outliers = data_df.groupby('Product_Group', group_keys=False).apply(lambda x: detect_outliers_iqr(x))
        outlier_view = outliers[['Date', 'Category', 'Product_Group', 'Sub_Product', 'Quantity', 'Revenue', 'ASP']].sort_values('ASP', ascending=False).head(50) if not outliers.empty else pd.DataFrame()

        out_start = len(stats) + 4
        add_section_title(ws, f"2. [{category_name}] Top 50 ASP Outliers (IQR Factor=1.5)", out_start, 7)
        if not outlier_view.empty:
            out_headers = ['Date', 'Category', 'Product_Group', 'Sub_Product', 'Quantity', 'Revenue', 'ASP']
            for col, h in enumerate(out_headers, 1):
                cell = ws.cell(out_start+1, col, h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            for i, (_, row) in enumerate(outlier_view.iterrows()):
                dr = out_start + 2 + i
                ws.cell(dr, 1, row['Date'].strftime('%Y-%m-%d') if hasattr(row['Date'], 'strftime') else str(row['Date']))
                ws.cell(dr, 2, row['Category'])
                ws.cell(dr, 3, row['Product_Group'])
                ws.cell(dr, 4, row['Sub_Product'])
                c = ws.cell(dr, 5, row['Quantity']); c.number_format = '#,##0'
                c = ws.cell(dr, 6, row['Revenue']); c.number_format = '#,##0'
                c = ws.cell(dr, 7, row['ASP']); c.number_format = '#,##0'
        else:
            ws.cell(out_start+1, 1, "이상치 없음")

        auto_width(ws)

    def create_validation_sheet(data_df):
        """Create validation guide sheet with calculation methodology and sample validations"""
        ws = wb.create_sheet("Validation_Guide")

        # Styling for explanation text
        explain_font = Font(name="Calibri", size=10, italic=True)
        formula_font = Font(name="Calibri", size=10, color="0066CC")

        # Section 1: Calculation Methodology
        add_section_title(ws, "📋 1. 계산 방법론 (Calculation Methodology)", 1, 5)

        methodologies = [
            ("ASP (Average Selling Price)", "Revenue ÷ Quantity", "개별 거래의 단가"),
            ("Revenue Share", "개별 Revenue ÷ Total Revenue × 100%", "매출 점유율"),
            ("Quantity Share", "개별 Quantity ÷ Total Quantity × 100%", "수량 점유율"),
            ("YoY Growth", "(현재년도 - 전년도) ÷ 전년도 × 100%", "전년 대비 성장률"),
            ("Volume Effect", "(현재 Qty - 이전 Qty) × 이전 ASP", "수량 변화 영향"),
            ("Price Effect", "(현재 ASP - 이전 ASP) × 현재 Qty", "가격 변화 영향"),
            ("Mix Effect", "Total Change - Volume Effect - Price Effect", "믹스 변화 영향"),
        ]

        headers = ["Metric", "Formula", "Description"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(2, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for i, (metric, formula, desc) in enumerate(methodologies):
            dr = 3 + i
            ws.cell(dr, 1, metric).font = bold_font
            cell = ws.cell(dr, 2, formula)
            cell.font = formula_font
            ws.cell(dr, 3, desc).font = explain_font

        # Section 2: Sample Validation with Formulas
        sample_start = len(methodologies) + 5
        add_section_title(ws, "✅ 2. 샘플 검증 예시 (Sample Validation)", sample_start, 6)

        # Get sample data for validation
        sample_df = data_df.head(10).copy()

        val_headers = ["Date", "Category", "Product_Group", "Quantity", "Revenue", "ASP (Formula)"]
        for col, h in enumerate(val_headers, 1):
            cell = ws.cell(sample_start + 1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for i, (_, row) in enumerate(sample_df.iterrows()):
            dr = sample_start + 2 + i
            ws.cell(dr, 1, row['Date'].strftime('%Y-%m-%d') if hasattr(row['Date'], 'strftime') else str(row['Date']))
            ws.cell(dr, 2, row['Category'])
            ws.cell(dr, 3, row['Product_Group'])
            c = ws.cell(dr, 4, row['Quantity']); c.number_format = '#,##0'
            c = ws.cell(dr, 5, row['Revenue']); c.number_format = '#,##0'
            # ASP with formula
            asp_cell = ws.cell(dr, 6)
            asp_cell.value = f"=E{dr}/D{dr}"
            asp_cell.number_format = '#,##0'
            asp_cell.font = formula_font

        # Add a summary calculation example
        summary_start = sample_start + len(sample_df) + 4
        add_section_title(ws, "📊 3. 집계 계산 예시 (Aggregation Example)", summary_start, 4)

        ws.cell(summary_start + 1, 1, "Metric").fill = header_fill
        ws.cell(summary_start + 1, 1).font = header_font
        ws.cell(summary_start + 1, 2, "Formula").fill = header_fill
        ws.cell(summary_start + 1, 2).font = header_font
        ws.cell(summary_start + 1, 3, "Value").fill = header_fill
        ws.cell(summary_start + 1, 3).font = header_font
        ws.cell(summary_start + 1, 4, "Description").fill = header_fill
        ws.cell(summary_start + 1, 4).font = header_font

        first_data_row = sample_start + 2
        last_data_row = sample_start + 1 + len(sample_df)

        agg_examples = [
            ("Total Quantity", f"=SUM(D{first_data_row}:D{last_data_row})", "수량 합계"),
            ("Total Revenue", f"=SUM(E{first_data_row}:E{last_data_row})", "매출 합계"),
            ("Average ASP", f"=AVERAGE(F{first_data_row}:F{last_data_row})", "평균 단가"),
            ("Weighted ASP", f"=E{summary_start+2}/D{summary_start+2}", "매출÷수량 방식의 가중평균 단가"),
        ]

        for i, (metric, formula, desc) in enumerate(agg_examples):
            dr = summary_start + 2 + i
            ws.cell(dr, 1, metric).font = bold_font
            formula_cell = ws.cell(dr, 2, formula)
            formula_cell.font = formula_font
            value_cell = ws.cell(dr, 3)
            value_cell.value = formula
            value_cell.number_format = '#,##0'
            ws.cell(dr, 4, desc).font = explain_font

        # Section 3: Data Source Mapping
        mapping_start = summary_start + len(agg_examples) + 4
        add_section_title(ws, "🗂️ 4. 데이터 소스 매핑 (Data Source Mapping)", mapping_start, 3)

        map_headers = ["Sheet Name", "Data Source", "Aggregation Level"]
        for col, h in enumerate(map_headers, 1):
            cell = ws.cell(mapping_start + 1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        mappings = [
            ("Overview", "Raw_Data (전체)", "Category"),
            ("All_Growth_Share", "Raw_Data (전체)", "Year, Product_Group"),
            ("All_TimeSeries", "Raw_Data (전체)", "Month/Quarter, Product_Group"),
            ("Bumper_Growth", "Raw_Data (Category=Bumper)", "Year, Product_Group"),
            ("SillSide_Growth", "Raw_Data (Category=Sill side)", "Year, Product_Group"),
            ("Carrier_Growth", "Raw_Data (Category=Carrier)", "Year, Product_Group"),
            ("PQ_Bridge", "Raw_Data (연도별)", "Bridge 분석 (Volume/Price/Mix)"),
            ("Outlier_*", "Raw_Data (카테고리별)", "IQR Method (Factor=1.5)"),
        ]

        for i, (sheet, source, agg) in enumerate(mappings):
            dr = mapping_start + 2 + i
            ws.cell(dr, 1, sheet).font = bold_font
            ws.cell(dr, 2, source)
            ws.cell(dr, 3, agg).font = explain_font

        # Add note
        note_start = mapping_start + len(mappings) + 3
        note_cell = ws.cell(note_start, 1)
        note_cell.value = "💡 참고: 모든 집계는 Python pandas를 통해 계산되어 값으로 저장됩니다. 이 시트의 샘플 수식은 검증 목적으로만 제공됩니다."
        note_cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
        ws.merge_cells(start_row=note_start, start_column=1, end_row=note_start, end_column=6)

        auto_width(ws)
    
    # ============ Sheet Creation ============
    
    # Sheet 1: Overview (Company-Wide)
    ws_ov = wb.create_sheet("Overview")
    df_overview = df[df['Year'].isin(all_years_list)]
    ov_grp = df_overview.groupby('Category').agg({'Revenue': 'sum', 'Quantity': 'sum'}).reset_index()
    ov_grp['ASP'] = ov_grp['Revenue'] / ov_grp['Quantity']
    total_rev = ov_grp['Revenue'].sum()
    total_qty = ov_grp['Quantity'].sum()
    ov_grp['Rev_Share'] = ov_grp['Revenue'] / total_rev
    ov_grp['Qty_Share'] = ov_grp['Quantity'] / total_qty
    
    add_section_title(ws_ov, "1. Category Overview Summary", 1, 6)
    headers = ['Category', 'Revenue', 'Rev Share', 'Quantity', 'Qty Share', 'ASP']
    for col, h in enumerate(headers, 1):
        cell = ws_ov.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    for r_idx, row in ov_grp.iterrows():
        data_row = r_idx + 3
        ws_ov.cell(data_row, 1, row['Category'])
        c = ws_ov.cell(data_row, 2, row['Revenue']); c.number_format = '#,##0'
        c = ws_ov.cell(data_row, 3, row['Rev_Share']); c.number_format = '0.0%'
        c = ws_ov.cell(data_row, 4, row['Quantity']); c.number_format = '#,##0'
        c = ws_ov.cell(data_row, 5, row['Qty_Share']); c.number_format = '0.0%'
        c = ws_ov.cell(data_row, 6, row['ASP']); c.number_format = '#,##0'
    
    total_row = len(ov_grp) + 3
    ws_ov.cell(total_row, 1, 'Total').font = bold_font
    c = ws_ov.cell(total_row, 2, total_rev); c.number_format = '#,##0'; c.font = bold_font
    c = ws_ov.cell(total_row, 4, total_qty); c.number_format = '#,##0'; c.font = bold_font
    
    trend_ov = df_overview.groupby(['Year', 'Category']).agg({'Revenue': 'sum', 'Quantity': 'sum'}).reset_index()
    trend_ov['ASP'] = trend_ov['Revenue'] / trend_ov['Quantity']
    
    section2_start = total_row + 2
    add_section_title(ws_ov, "2. Yearly Trend by Category", section2_start, 5)
    headers2 = ['Year', 'Category', 'Revenue', 'Quantity', 'ASP']
    for col, h in enumerate(headers2, 1):
        cell = ws_ov.cell(section2_start + 1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    for i, (_, row) in enumerate(trend_ov.iterrows()):
        dr = section2_start + 2 + i
        ws_ov.cell(dr, 1, int(row['Year']))
        ws_ov.cell(dr, 2, row['Category'])
        c = ws_ov.cell(dr, 3, row['Revenue']); c.number_format = '#,##0'
        c = ws_ov.cell(dr, 4, row['Quantity']); c.number_format = '#,##0'
        c = ws_ov.cell(dr, 5, row['ASP']); c.number_format = '#,##0'
    
    auto_width(ws_ov)
    
    # Sheets 2-3: All Categories
    df_all = df[df['Year'].isin(all_years_list)]
    create_growth_share_sheet("All_Growth_Share", df_all, "All Categories")
    create_timeseries_sheet("All_TimeSeries", df_all, "All Categories")
    
    # Sheets for each category
    categories = sorted(df['Category'].unique())
    for cat in categories:
        cat_df = df[df['Category'] == cat]
        cat_df = cat_df[cat_df['Year'].isin(all_years_list)]
        sheet_prefix = cat.replace(' ', '')
        create_growth_share_sheet(f"{sheet_prefix}_Growth", cat_df, cat)
        create_timeseries_sheet(f"{sheet_prefix}_TimeSeries", cat_df, cat)
    
    # PQ_Bridge Sheet
    ws_br = wb.create_sheet("PQ_Bridge")
    target_years = sorted(all_years_list)
    
    def write_bridge_section(ws, title, steps, start_row):
        add_section_title(ws, title, start_row, 7)
        headers = ['Period', 'Start Revenue', 'Volume Effect', 'Price Effect', 'Mix Effect', 'End Revenue', 'Total Change']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(start_row+1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for i, step in enumerate(steps):
            dr = start_row + 2 + i
            ws.cell(dr, 1, step['label'])
            for col, key in [(2, 'start_val'), (3, 'volume'), (4, 'price'), (5, 'mix'), (6, 'end_val')]:
                c = ws.cell(dr, col, step[key]); c.number_format = '#,##0'
            c = ws.cell(dr, 7, step['end_val'] - step['start_val']); c.number_format = '+#,##0;-#,##0;0'
        
        return start_row + len(steps) + 3
    
    if len(target_years) >= 2:
        df_company = df[df['Year'].isin(target_years)]
        res1 = calculate_bridge_mix_steps(df_company, target_years, 'Category')
        next_row = write_bridge_section(ws_br, "1. Company-Wide Bridge (by Major Category)", res1, 1)
        
        res2 = calculate_bridge_mix_steps(df_company, target_years, 'Product_Group')
        next_row = write_bridge_section(ws_br, "2. All Categories Bridge (by Product Group)", res2, next_row)
        
        # Add category-specific bridges
        for cat in categories:
            df_cat_bridge = df_company[df_company['Category'] == cat]
            res_cat = calculate_bridge_mix_steps(df_cat_bridge, target_years, 'Product_Group')
            if res_cat:
                next_row = write_bridge_section(ws_br, f"3. {cat} Bridge (by Product Group)", res_cat, next_row)
    else:
        ws_br.cell(1, 1, "최소 2개 연도를 선택해야 Bridge 분석이 가능합니다.")
    
    auto_width(ws_br)
    
    # Outlier Sheets
    create_outlier_sheet("Outlier_All", df_all, "All Categories")
    for cat in categories:
        cat_df = df[df['Category'] == cat]
        cat_df = cat_df[cat_df['Year'].isin(all_years_list)]
        sheet_prefix = cat.replace(' ', '')
        create_outlier_sheet(f"Outlier_{sheet_prefix}", cat_df, cat)

    # Validation Guide Sheet
    create_validation_sheet(df_all)

    # Raw Data Sheet
    ws_raw = wb.create_sheet("Raw_Data")
    raw_cols = ['Date', 'Year', 'Category', 'Product_Group', 'Sub_Product', 'Customer', 'Quantity', 'Revenue', 'ASP']
    raw_export = df[raw_cols].copy()
    
    add_section_title(ws_raw, "Raw Transaction Data (All Data)", 1, len(raw_cols))
    for col, h in enumerate(raw_cols, 1):
        cell = ws_raw.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    for i, (_, row) in enumerate(raw_export.iterrows()):
        dr = 3 + i
        ws_raw.cell(dr, 1, row['Date'].strftime('%Y-%m-%d') if hasattr(row['Date'], 'strftime') else str(row['Date']))
        ws_raw.cell(dr, 2, int(row['Year']))
        ws_raw.cell(dr, 3, row['Category'])
        ws_raw.cell(dr, 4, row['Product_Group'])
        ws_raw.cell(dr, 5, row['Sub_Product'])
        ws_raw.cell(dr, 6, row['Customer'])
        c = ws_raw.cell(dr, 7, row['Quantity']); c.number_format = '#,##0'
        c = ws_raw.cell(dr, 8, row['Revenue']); c.number_format = '#,##0'
        c = ws_raw.cell(dr, 9, row['ASP']); c.number_format = '#,##0'
    
    auto_width(ws_raw)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
# --- 5. 메인 UI ---
def main():
    st.sidebar.markdown(f"<h1 style='color: {COLORS['primary']};'>KPMG Analysis (MBR)</h1>", unsafe_allow_html=True)
    
    st.sidebar.markdown("### 📂 Data Upload")
    uploaded_file = st.sidebar.file_uploader("업로드: 매출데이터_MBR.xlsx", type=['xlsx'])
    
    if uploaded_file is None:
        st.info("👈 좌측 사이드바에서 엑셀 파일을 업로드해주세요.")
        return 
    df = load_all_data(uploaded_file)
    if df.empty: return

    # [엑셀 리포트 생성 - 전체 데이터 기준, 1회만 실행]
    all_years_list = sorted(df['Year'].unique())
    excel_data_full = generate_excel_report(df, all_years_list)

    # [사이드바 필터]
    st.sidebar.header("🔍 Global Filters")
    
    all_years = sorted(df['Year'].unique())
    selected_years = st.sidebar.multiselect(
        "Analysis Years", all_years, 
        default=all_years if len(all_years) <= 3 else all_years[-3:],
        key="year_multi"
    )
    
    st.sidebar.markdown("---")
    st.sidebar.header("📂 Category Filter")
    
    cat_options = ["All Categories"] + sorted(df['Category'].unique().tolist())
    sel_cat = st.sidebar.selectbox("Major Category", cat_options, key="cat_select")
    
    if sel_cat == "All Categories":
        temp_df = df
    else:
        temp_df = df[df['Category'] == sel_cat]
        
    prod_options = ["All Products"] + sorted(temp_df['Product_Group'].unique().tolist())
    sel_prod = st.sidebar.selectbox("Product Group", prod_options, key="prod_select")
    
    # [참조표]
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ℹ️ Product Group Reference")
    ref_data = {
        'Product': ['GH', 'JB', 'CV', 'IJ'],
        'Car Model': ['HB20 (소형)', '크레타', '크레타(소형)', 'HB12']
    }
    st.sidebar.dataframe(pd.DataFrame(ref_data), hide_index=True, use_container_width=True)
    
    # 필터링 데이터
    filtered_df = df[df['Year'].isin(selected_years)]
    if sel_cat != "All Categories":
        filtered_df = filtered_df[filtered_df['Category'] == sel_cat]
    if sel_prod != "All Products":
        filtered_df = filtered_df[filtered_df['Product_Group'] == sel_prod]
    
    # Excel Export Button (전체 데이터 리포트, 필터 무관)
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📥 Excel Export")
    st.sidebar.info("💡 리포트는 필터와 무관하게 **전체 데이터**를 포함합니다.")
    filename = f"Sales_Report_Full_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.sidebar.download_button(
        label="📥 Download Full Excel Report",
        data=excel_data_full,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # KPI Summary
    st.markdown("### 📊 Executive Summary (Filtered Scope)")
    kpi1, kpi2, kpi3, kpi4 = st.columns(4)
    
    f_rev = filtered_df['Revenue'].sum()
    f_qty = filtered_df['Quantity'].sum()
    f_asp = f_rev / f_qty if f_qty > 0 else 0
    
    if len(selected_years) > 0:
        max_y = max(selected_years)
        curr = filtered_df[filtered_df['Year'] == max_y]['Revenue'].sum()
        prev = filtered_df[filtered_df['Year'] == (max_y - 1)]['Revenue'].sum()
        yoy = ((curr - prev) / prev * 100) if prev > 0 else 0
    else:
        max_y, yoy = "-", 0
        
    kpi1.metric("Revenue", f"R$ {f_rev:,.0f}")
    kpi2.metric("Quantity", f"{f_qty:,.0f}")
    kpi3.metric("Avg ASP", f"R$ {f_asp:,.0f}")
    kpi4.metric(f"YoY Growth ('{max_y})", f"{yoy:,.1f}%", delta_color="normal")
    
    # 탭 구성
    tab0, tab1, tab2, tab3, tab4 = st.tabs(["📑 전사 요약", "📈 Growth & Share", "📅 Time Series", "💰 P/Q/Mix Bridge", "🚨 Outlier"])
    
    # === Tab 0: 전사 요약 ===
    with tab0:
        st.subheader("🏢 Major Category Overview")
        df_overview = df[df['Year'].isin(selected_years)]
        ov_grp = df_overview.groupby('Category').agg({'Revenue': 'sum', 'Quantity': 'sum'}).reset_index()
        ov_grp['ASP'] = ov_grp['Revenue'] / ov_grp['Quantity']
        
        total_rev_ov = ov_grp['Revenue'].sum()
        total_qty_ov = ov_grp['Quantity'].sum()
        ov_grp['Rev_Label'] = ov_grp.apply(lambda x: f"R$ {x['Revenue']:,.0f}<br>({x['Revenue']/total_rev_ov:.1%})", axis=1)
        ov_grp['Qty_Label'] = ov_grp.apply(lambda x: f"{x['Quantity']:,.0f}<br>({x['Quantity']/total_qty_ov:.1%})", axis=1)
        
        c1, c2, c3 = st.columns(3)
        with c1:
            fig_rev = px.bar(ov_grp, x='Category', y='Revenue', text='Rev_Label', color='Category', 
                             title="Revenue & Share", color_discrete_sequence=COLORS['sequence'])
            fig_rev.update_traces(textposition='outside')
            st.plotly_chart(fig_rev, use_container_width=True)
        with c2:
            fig_qty = px.bar(ov_grp, x='Category', y='Quantity', text='Qty_Label', color='Category', 
                             title="Sales Volume & Share", color_discrete_sequence=COLORS['sequence'])
            fig_qty.update_traces(textposition='outside')
            st.plotly_chart(fig_qty, use_container_width=True)
        with c3:
            st.plotly_chart(px.bar(ov_grp, x='Category', y='ASP', color='Category', title="Avg ASP (R$)", text_auto='.0f', color_discrete_sequence=COLORS['sequence']), use_container_width=True)
        
        st.markdown("---")
        st.subheader("📈 Yearly Trend by Category")
        trend_ov = df_overview.groupby(['Year', 'Category']).agg({'Revenue':'sum', 'Quantity':'sum'}).reset_index()
        trend_ov['ASP'] = trend_ov['Revenue'] / trend_ov['Quantity']
        
        c_tr1, c_tr2, c_tr3 = st.columns(3)
        with c_tr1:
            fig_l1 = px.line(trend_ov, x='Year', y='Revenue', color='Category', markers=True, title="Revenue Trend", color_discrete_sequence=COLORS['sequence'])
            fig_l1.update_xaxes(dtick=1)
            st.plotly_chart(fig_l1, use_container_width=True)
        with c_tr2:
            fig_l2 = px.line(trend_ov, x='Year', y='Quantity', color='Category', markers=True, title="Quantity Trend", color_discrete_sequence=COLORS['sequence'])
            fig_l2.update_xaxes(dtick=1)
            st.plotly_chart(fig_l2, use_container_width=True)
        with c_tr3:
            fig_l3 = px.line(trend_ov, x='Year', y='ASP', color='Category', markers=True, title="ASP Trend (R$)", color_discrete_sequence=COLORS['sequence'])
            fig_l3.update_xaxes(dtick=1)
            st.plotly_chart(fig_l3, use_container_width=True)
    
    # === Tab 1: Growth & Share ===
    with tab1:
        c1, c2 = st.columns(2)
        with c1:
            st.subheader(f"[{sel_cat}] Yearly Revenue")
            y_df = filtered_df.groupby('Year')['Revenue'].sum().reset_index()
            y_df['YoY'] = y_df['Revenue'].pct_change() * 100
            
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(go.Bar(x=y_df['Year'], y=y_df['Revenue'], name="Revenue", marker_color=COLORS['primary'],
                                 text=y_df['Revenue'].apply(lambda x: f"R$ {x:,.0f}"), textposition='auto'), secondary_y=False)
            fig.add_trace(go.Scatter(x=y_df['Year'], y=y_df['YoY'], name="YoY %", mode='lines+markers+text', 
                                     line=dict(color=COLORS['secondary'], width=3), text=y_df['YoY'].apply(lambda x: f"{x:.1f}%" if pd.notnull(x) else ""), textposition="top center"), secondary_y=True)
            fig.update_xaxes(dtick=1)
            st.plotly_chart(fig, use_container_width=True)
            
        with c2:
            st.subheader(f"[{sel_cat}] Revenue Share")
            comp_df = filtered_df.groupby(['Year', 'Product_Group'])['Revenue'].sum().reset_index()
            comp_df['Total'] = comp_df.groupby('Year')['Revenue'].transform('sum')
            comp_df['Share'] = (comp_df['Revenue'] / comp_df['Total'] * 100)
            comp_df['Label'] = comp_df.apply(lambda x: f"R$ {x['Revenue']:,.0f}<br>({x['Share']:.1f}%)", axis=1)
            fig = px.bar(comp_df, x='Year', y='Revenue', color='Product_Group', 
                         text='Label', color_discrete_sequence=COLORS['sequence'])
            fig.update_traces(textposition='inside')
            fig.update_xaxes(dtick=1)
            st.plotly_chart(fig, use_container_width=True)
        
        st.markdown("---")
        st.subheader(f"[{sel_cat}] Volume & Price Trend Analysis")
        group_col = 'Product_Group' if sel_prod == "All Products" else 'Sub_Product'
        trend_detailed = filtered_df.groupby(['Year', group_col]).agg({'Quantity': 'sum','Revenue': 'sum'}).reset_index()
        trend_detailed['ASP'] = trend_detailed['Revenue'] / trend_detailed['Quantity']
        trend_detailed['Total_Qty_Year'] = trend_detailed.groupby('Year')['Quantity'].transform('sum')
        trend_detailed['Qty_Label'] = trend_detailed.apply(lambda x: f"{x['Quantity']:,.0f}<br>({x['Quantity']/x['Total_Qty_Year']:.1%})", axis=1)
        
        c3, c4 = st.columns(2)
        with c3:
            fig_qty = px.bar(trend_detailed, x='Year', y='Quantity', color=group_col, text='Qty_Label',
                             title="Quantity by Product Group", color_discrete_sequence=COLORS['sequence'])
            fig_qty.update_traces(textposition='inside')
            fig_qty.update_xaxes(dtick=1)
            st.plotly_chart(fig_qty, use_container_width=True)
        with c4:
            fig_asp = px.line(trend_detailed, x='Year', y='ASP', color=group_col,
                              title="ASP by Product Group", markers=True, color_discrete_sequence=COLORS['sequence'])
            fig_asp.update_yaxes(tickprefix="R$ ")
            fig_asp.update_xaxes(dtick=1)
            st.plotly_chart(fig_asp, use_container_width=True)
    
    # === Tab 2: 시계열 ===
    with tab2:
        mix_col = 'Category' if sel_cat == "All Categories" else 'Product_Group'
        st.subheader(f"Revenue Mix by {mix_col}")
        
        mode = st.radio("Time Unit", ["Monthly", "Quarterly"], horizontal=True, key="ts_mode")
        t_col = 'Month_Dt' if mode == "Monthly" else 'Quarter_Str'
        
        t_df = filtered_df.groupby([t_col, mix_col])['Revenue'].sum().reset_index()
        if mode == "Quarterly": t_df = t_df.sort_values(t_col)
        t_df['Total'] = t_df.groupby(t_col)['Revenue'].transform('sum')
        t_df['Share'] = t_df['Revenue'] / t_df['Total']
        
        c1, c2 = st.columns([2, 1])
        with c1:
            fig = px.area(t_df, x=t_col, y='Revenue', color=mix_col, title="Revenue Trend", color_discrete_sequence=COLORS['sequence'])
            if mode == "Monthly": fig.update_xaxes(tickformat="%Y-%m")
            st.plotly_chart(fig, use_container_width=True)
        with c2:
            fig = px.bar(t_df, x=t_col, y='Share', color=mix_col, title="Revenue Mix", color_discrete_sequence=COLORS['sequence'])
            fig.update_yaxes(tickformat=".0%")
            st.plotly_chart(fig, use_container_width=True)
        
        st.markdown("---")
        st.subheader(f"Quantity vs ASP Trend ({mode})")
        if mode == "Monthly":
            agg_df = filtered_df.groupby('Month_Dt').agg({'Revenue':'sum', 'Quantity':'sum'}).reset_index()
            x_ax = 'Month_Dt'
        else:
            agg_df = filtered_df.groupby('Quarter_Str').agg({'Revenue':'sum', 'Quantity':'sum'}).reset_index().sort_values('Quarter_Str')
            x_ax = 'Quarter_Str'
        
        agg_df['ASP'] = np.where(agg_df['Quantity']>0, agg_df['Revenue']/agg_df['Quantity'], 0)
        
        fig_dual = make_subplots(specs=[[{"secondary_y": True}]])
        fig_dual.add_trace(go.Bar(x=agg_df[x_ax], y=agg_df['Quantity'], name="Quantity", marker_color='#B0BEC5', opacity=0.6), secondary_y=False)
        fig_dual.add_trace(go.Scatter(x=agg_df[x_ax], y=agg_df['ASP'], name="ASP", mode='lines+markers', line=dict(color=COLORS['primary'], width=3)), secondary_y=True)
        fig_dual.update_layout(hovermode="x unified")
        fig_dual.update_yaxes(title_text="Quantity", showgrid=False, secondary_y=False)
        fig_dual.update_yaxes(title_text="ASP (R$)", showgrid=False, secondary_y=True)
        if mode == "Monthly": fig_dual.update_xaxes(tickformat="%Y-%m")
        st.plotly_chart(fig_dual, use_container_width=True)
    
    # === Tab 3: P/Q/Mix Bridge ===
    with tab3:
        st.subheader("💰 Price-Volume-Mix Bridge Analysis")
        st.info("매출 변동을 **물량(Volume)**, **가격(Price)**, **믹스(Mix)** 효과로 분해합니다.")
        
        target_years = sorted(list(set(selected_years)))
        
        if len(target_years) >= 2:
            st.markdown("#### 1. Company-Wide Bridge (by Major Category)")
            st.caption("전사 기준: 대분류(Category) 믹스 효과 분석")
            df_company = df[df['Year'].isin(selected_years)]
            res1 = calculate_bridge_mix_steps(df_company, target_years, 'Category')
            if res1:
                st.plotly_chart(create_waterfall_fig_multi(res1), use_container_width=True)
            
            st.markdown("---")
            
            if sel_cat == "All Categories":
                st.markdown("#### 2. Major Category Bridge (All -> Product Group)")
                st.caption("전체 카테고리 내 제품군(Product Group) 믹스 효과 분석")
                res2 = calculate_bridge_mix_steps(df_company, target_years, 'Product_Group')
            else:
                st.markdown(f"#### 2. Major Category Bridge: {sel_cat} (by Product Group)")
                st.caption(f"선택된 {sel_cat} 내 제품군 믹스 효과 분석")
                df_cat_bridge = df_company[df_company['Category'] == sel_cat]
                res2 = calculate_bridge_mix_steps(df_cat_bridge, target_years, 'Product_Group')
                
            if res2:
                st.plotly_chart(create_waterfall_fig_multi(res2), use_container_width=True)
                
            st.markdown("---")
            
            st.markdown("#### 3. Detailed Bridge (Specific Product Group)")
            if sel_cat != "All Categories" and sel_prod != "All Products":
                st.success(f"Selected: {sel_cat} > {sel_prod}")
                st.caption(f"선택된 제품군 내 거래처(Customer) 믹스 효과 상세 분석")
                
                is_bumper = (sel_cat == 'Bumper')
                detail_col = 'Sub_Product' if is_bumper else 'Customer'
                
                res3 = calculate_bridge_mix_steps(filtered_df, target_years, detail_col)
                if res3:
                    st.plotly_chart(create_waterfall_fig_multi(res3), use_container_width=True)
                else:
                    st.warning("분석할 데이터가 부족합니다.")
            else:
                st.info("💡 **Major Category**와 **Product Group**을 모두 선택해야 최하단 상세 브릿지가 표시됩니다.")
                
        else:
            st.info("비교할 연도를 2개 이상 선택해주세요.")
    
    # === Tab 4: Outlier ===
    with tab4:
        st.subheader("Statistical Outlier Analysis")
        c1, c2 = st.columns([2, 1])
        with c1:
            fig_box = px.box(filtered_df, x='Product_Group', y='ASP', color='Product_Group', 
                             points="outliers", notched=True, color_discrete_sequence=COLORS['sequence'])
            st.plotly_chart(fig_box, use_container_width=True)
        with c2:
            stats = filtered_df.groupby('Product_Group')['ASP'].agg(['count', 'mean', 'min', 'max']).reset_index()
            st.dataframe(stats.style.format({'mean': 'R$ {:,.0f}', 'min': 'R$ {:,.0f}', 'max': 'R$ {:,.0f}'}), use_container_width=True)
        
        st.markdown("---")
        st.markdown("#### 🚨 Top 50 Outliers")
        outliers = filtered_df.groupby('Product_Group', group_keys=False).apply(lambda x: detect_outliers_iqr(x))
        if not outliers.empty:
            outlier_view = outliers[['Date', 'Category', 'Product_Group', 'Sub_Product', 'Quantity', 'Revenue', 'ASP']].sort_values('ASP', ascending=False).head(50)
            st.dataframe(outlier_view.style.format({'Revenue': 'R$ {:,.0f}', 'ASP': 'R$ {:,.0f}', 'Quantity': '{:,.0f}'}).bar(subset=['ASP'], color='#ffcccc'), use_container_width=True)
        else:
            st.success("이상치 없음")

if __name__ == "__main__":
    main()
