import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- 1. 디자인 시스템 & 설정 ---
st.set_page_config(layout="wide", page_title="Sales Dashboard (MSK)", page_icon="💶")
COLORS = {
    'primary': '#00338D',    # KPMG Blue
    'secondary': '#0091DA',  # Light Blue
    'accent': '#00A3A1',     # Teal
    'negative': '#D32929',   # Red
    'positive': '#009944',   # Green
    'neutral': '#6D6E71',    # Grey
    'sequence': px.colors.qualitative.Prism
}

# --- 2. 데이터 로드 및 전처리 ---
@st.cache_data(ttl=3600)
def load_data(file_obj):
    try:
        cols = "A, B, I, M, N, O, P, S"
        df = pd.read_excel(file_obj, sheet_name='raw', usecols=cols)
        df.columns = ['Year', 'Date', 'Category', 'Quantity', 'Revenue', 'Customer', 'Region', 'Car_Type']

        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df.dropna(subset=['Date'])

        df['Revenue'] = pd.to_numeric(df['Revenue'], errors='coerce').fillna(0)
        df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(0)

        df['Category'] = df['Category'].fillna("Unknown")
        df['Car_Type'] = df['Car_Type'].fillna("-")
        df['Customer'] = df['Customer'].fillna("Unknown")
        df['Region'] = df['Region'].fillna("Unknown")

        # 매출 0 제외
        df = df[df['Revenue'] != 0]

        df['Month_Dt'] = df['Date'].dt.to_period('M').dt.to_timestamp()
        df['Quarter_Str'] = df['Date'].dt.to_period('Q').astype(str)

        df['ASP'] = np.where(df['Quantity'] > 0, df['Revenue'] / df['Quantity'], 0)

        return df
    except ValueError as ve:
        st.error(f"❌ 엑셀 구조 오류: 'raw' 시트 확인 필요.\n({ve})")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"❌ 데이터 로드 오류: {e}")
        return pd.DataFrame()

# --- 3. 데이터 분리 ---
def split_data(df):
    mask_exception = (df['Quantity'] == 0)
    df_exception = df[mask_exception].copy()
    df_normal = df[~mask_exception].copy()
    return df_normal, df_exception

# --- 4. 분석 로직 (Bridge with Mix) ---
def calculate_bridge_mix_steps(df, years, group_col):
    """
    Price / Volume / Mix Bridge 계산 (Multi-year)
    """
    all_steps = []
    if len(years) < 2: return all_steps

    years = sorted(years)

    for i in range(len(years) - 1):
        y0, y1 = years[i], years[i+1]

        d0 = df[df['Year'] == y0].groupby(group_col).agg({'Revenue':'sum', 'Quantity':'sum'}).reset_index()
        d1 = df[df['Year'] == y1].groupby(group_col).agg({'Revenue':'sum', 'Quantity':'sum'}).reset_index()

        total_q0 = d0['Quantity'].sum()
        total_q1 = d1['Quantity'].sum()
        total_r0 = d0['Revenue'].sum()
        total_r1 = d1['Revenue'].sum()

        avg_p0 = total_r0 / total_q0 if total_q0 > 0 else 0

        merged = pd.merge(d0, d1, on=group_col, how='outer', suffixes=('_0', '_1')).fillna(0)

        merged['P0'] = np.where(merged['Quantity_0'] > 0, merged['Revenue_0'] / merged['Quantity_0'], 0)
        merged['P1'] = np.where(merged['Quantity_1'] > 0, merged['Revenue_1'] / merged['Quantity_1'], 0)

        vol_effect = (total_q1 - total_q0) * avg_p0
        merged['Price_Impact'] = (merged['P1'] - merged['P0']) * merged['Quantity_1']
        price_effect = merged['Price_Impact'].sum()

        total_variance = total_r1 - total_r0
        mix_effect = total_variance - vol_effect - price_effect

        all_steps.append({
            'label': f"{y0}→{y1}",
            'start_year': str(y0),
            'end_year': str(y1),
            'volume': vol_effect,
            'price': price_effect,
            'mix': mix_effect,
            'start_val': total_r0,
            'end_val': total_r1
        })

    return all_steps

def create_waterfall_fig_multi(steps, title="Price-Volume-Mix Bridge"):
    """Multi-year waterfall chart"""
    if not steps: return go.Figure()

    first_step = steps[0]

    x_labels = [first_step['start_year']]
    y_vals = [first_step['start_val']]
    measures = ["absolute"]
    text_vals = [f"€ {first_step['start_val']/1000000:,.1f}M"]

    for step in steps:
        for eff_name, eff_val in [("Volume", step['volume']), ("Mix", step['mix']), ("Price", step['price'])]:
            x_labels.append(f"{eff_name}<br>({step['label']})")
            y_vals.append(eff_val)
            measures.append("relative")
            text_vals.append(f"{eff_val/1000000:+,.1f}M")

        x_labels.append(step['end_year'])
        y_vals.append(step['end_val'])
        measures.append("absolute")
        text_vals.append(f"€ {step['end_val']/1000000:,.1f}M")

    fig = go.Figure(go.Waterfall(
        orientation="v",
        measure=measures,
        x=x_labels,
        y=y_vals,
        text=text_vals,
        textposition="outside",
        decreasing={"marker":{"color": COLORS['negative']}},
        increasing={"marker":{"color": COLORS['accent']}},
        totals={"marker":{"color": COLORS['primary']}},
        connector={"line":{"color":"#666"}}
    ))

    fig.update_layout(title_text=title, xaxis_type='category')
    return fig

def detect_outliers_iqr(df, col='ASP', factor=1.5):
    if df.empty: return df
    q1 = df[col].quantile(0.25)
    q3 = df[col].quantile(0.75)
    iqr = q3 - q1
    lower, upper = q1 - (factor * iqr), q3 + (factor * iqr)
    return df[(df[col] < lower) | (df[col] > upper)]

# --- 5. Excel 리포트 생성 함수 ---
@st.cache_data(ttl=3600, show_spinner="📊 전체 엑셀 리포트 생성 중...")
def generate_excel_report(df_normal, df_exception, all_years_list):
    """Generate comprehensive KPMG-styled Excel report for MSK data"""
    wb = Workbook()
    wb.remove(wb.active)

    KPMG_BLUE = "00338D"
    LIGHT_BLUE = "E8EEF7"
    WHITE = "FFFFFF"

    header_fill = PatternFill(fill_type="solid", fgColor=KPMG_BLUE)
    header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    bold_font = Font(name="Calibri", bold=True, size=10)
    title_font = Font(name="Calibri", bold=True, color=KPMG_BLUE, size=12)
    center_align = Alignment(horizontal='center', vertical='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    def add_section_title(ws, title, row, num_cols):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row, 1, title)
        cell.font = title_font
        cell.alignment = Alignment(horizontal='left')
        ws.row_dimensions[row].height = 20

    # === Sheet 1: Validation Guide ===
    ws_val = wb.create_sheet("Validation_Guide")

    explain_font = Font(name="Calibri", size=10, italic=True)
    formula_font = Font(name="Calibri", size=10, color="0066CC")

    add_section_title(ws_val, "📋 1. 계산 방법론 (Calculation Methodology)", 1, 3)

    methodologies = [
        ("ASP (Average Selling Price)", "Revenue ÷ Quantity", "개별 거래의 단가"),
        ("Revenue Share", "개별 Revenue ÷ Total Revenue × 100%", "매출 점유율"),
        ("YoY Growth", "(현재년도 - 전년도) ÷ 전년도 × 100%", "전년 대비 성장률"),
        ("Volume Effect", "(현재 Qty - 이전 Qty) × 이전 ASP", "수량 변화 영향"),
        ("Price Effect", "(현재 ASP - 이전 ASP) × 현재 Qty", "가격 변화 영향"),
        ("Mix Effect", "Total Change - Volume Effect - Price Effect", "믹스 변화 영향"),
    ]

    headers = ["Metric", "Formula", "Description"]
    for col, h in enumerate(headers, 1):
        cell = ws_val.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for i, (metric, formula, desc) in enumerate(methodologies):
        dr = 3 + i
        ws_val.cell(dr, 1, metric).font = bold_font
        cell = ws_val.cell(dr, 2, formula)
        cell.font = formula_font
        ws_val.cell(dr, 3, desc).font = explain_font

    auto_width(ws_val)

    # === Sheet 2: Overview ===
    ws_ov = wb.create_sheet("Overview")
    df_overview = df_normal[df_normal['Year'].isin(all_years_list)]

    # Section 1: Category Overview
    ov_cat = df_overview.groupby('Category').agg({'Revenue': 'sum', 'Quantity': 'sum'}).reset_index()
    ov_cat['ASP'] = ov_cat['Revenue'] / ov_cat['Quantity']
    total_rev = ov_cat['Revenue'].sum()
    total_qty = ov_cat['Quantity'].sum()
    ov_cat['Rev_Share'] = ov_cat['Revenue'] / total_rev
    ov_cat['Qty_Share'] = ov_cat['Quantity'] / total_qty

    add_section_title(ws_ov, "1. Category Overview Summary", 1, 6)
    headers = ['Category', 'Revenue', 'Rev Share', 'Quantity', 'Qty Share', 'ASP']
    for col, h in enumerate(headers, 1):
        cell = ws_ov.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for r_idx, row in ov_cat.iterrows():
        data_row = r_idx + 3
        ws_ov.cell(data_row, 1, row['Category'])
        c = ws_ov.cell(data_row, 2, row['Revenue']); c.number_format = '#,##0'
        c = ws_ov.cell(data_row, 3, row['Rev_Share']); c.number_format = '0.0%'
        c = ws_ov.cell(data_row, 4, row['Quantity']); c.number_format = '#,##0'
        c = ws_ov.cell(data_row, 5, row['Qty_Share']); c.number_format = '0.0%'
        c = ws_ov.cell(data_row, 6, row['ASP']); c.number_format = '#,##0'

    total_row = len(ov_cat) + 3
    ws_ov.cell(total_row, 1, 'Total').font = bold_font
    c = ws_ov.cell(total_row, 2, total_rev); c.number_format = '#,##0'; c.font = bold_font
    c = ws_ov.cell(total_row, 4, total_qty); c.number_format = '#,##0'; c.font = bold_font

    # Section 2: Customer Revenue (Top 10)
    section2_start = total_row + 2
    ov_cust = df_overview.groupby('Customer')['Revenue'].sum().reset_index().sort_values('Revenue', ascending=False).head(10)
    ov_cust['Share'] = ov_cust['Revenue'] / total_rev

    add_section_title(ws_ov, "2. Top 10 Customers by Revenue", section2_start, 3)
    for col, h in enumerate(['Customer', 'Revenue', 'Share'], 1):
        cell = ws_ov.cell(section2_start+1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for i, (_, row) in enumerate(ov_cust.iterrows()):
        dr = section2_start + 2 + i
        ws_ov.cell(dr, 1, row['Customer'])
        c = ws_ov.cell(dr, 2, row['Revenue']); c.number_format = '#,##0'
        c = ws_ov.cell(dr, 3, row['Share']); c.number_format = '0.0%'

    auto_width(ws_ov)

    # === Sheet 3: Growth & Share Analysis ===
    ws_growth = wb.create_sheet("Growth_Share")

    # Yearly Revenue + YoY
    y_df = df_overview.groupby('Year')['Revenue'].sum().reset_index()
    y_df['YoY'] = y_df['Revenue'].pct_change()

    add_section_title(ws_growth, "1. Yearly Revenue & YoY Growth", 1, 3)
    for col, h in enumerate(['Year', 'Revenue', 'YoY %'], 1):
        cell = ws_growth.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for i, (_, row) in enumerate(y_df.iterrows()):
        dr = 3 + i
        ws_growth.cell(dr, 1, int(row['Year']))
        c = ws_growth.cell(dr, 2, row['Revenue']); c.number_format = '#,##0'
        yoy_val = row['YoY']
        if pd.notnull(yoy_val):
            c = ws_growth.cell(dr, 3, yoy_val); c.number_format = '0.0%'
        else:
            ws_growth.cell(dr, 3, '-')

    # Revenue Share by Category
    s2_start = len(y_df) + 4
    comp_df = df_overview.groupby(['Year', 'Category'])['Revenue'].sum().reset_index()
    comp_df['Total'] = comp_df.groupby('Year')['Revenue'].transform('sum')
    comp_df['Share'] = comp_df['Revenue'] / comp_df['Total']

    add_section_title(ws_growth, "2. Revenue Share by Category", s2_start, 4)
    for col, h in enumerate(['Year', 'Category', 'Revenue', 'Share'], 1):
        cell = ws_growth.cell(s2_start+1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for i, (_, row) in enumerate(comp_df.iterrows()):
        dr = s2_start + 2 + i
        ws_growth.cell(dr, 1, int(row['Year']))
        ws_growth.cell(dr, 2, row['Category'])
        c = ws_growth.cell(dr, 3, row['Revenue']); c.number_format = '#,##0'
        c = ws_growth.cell(dr, 4, row['Share']); c.number_format = '0.0%'

    auto_width(ws_growth)

    # === Sheet 4: Time Series ===
    ws_ts = wb.create_sheet("TimeSeries")

    # Monthly Revenue by Category
    t_df_m = df_overview.groupby(['Month_Dt', 'Category'])['Revenue'].sum().reset_index()
    t_df_m['Total'] = t_df_m.groupby('Month_Dt')['Revenue'].transform('sum')
    t_df_m['Share'] = t_df_m['Revenue'] / t_df_m['Total']
    t_df_m['Period'] = t_df_m['Month_Dt'].dt.strftime('%Y-%m')

    add_section_title(ws_ts, "1. Monthly Revenue Mix by Category", 1, 4)
    for col, h in enumerate(['Period', 'Category', 'Revenue', 'Share'], 1):
        cell = ws_ts.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for i, (_, row) in enumerate(t_df_m.iterrows()):
        dr = 3 + i
        ws_ts.cell(dr, 1, row['Period'])
        ws_ts.cell(dr, 2, row['Category'])
        c = ws_ts.cell(dr, 3, row['Revenue']); c.number_format = '#,##0'
        c = ws_ts.cell(dr, 4, row['Share']); c.number_format = '0.0%'

    auto_width(ws_ts)

    # === Sheet 5: PQ_Bridge ===
    ws_br = wb.create_sheet("PQ_Bridge")
    target_years = sorted(all_years_list)

    def write_bridge_section(ws, title, steps, start_row):
        add_section_title(ws, title, start_row, 7)
        headers = ['Period', 'Start Revenue', 'Volume Effect', 'Price Effect', 'Mix Effect', 'End Revenue', 'Total Change']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(start_row+1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for i, step in enumerate(steps):
            dr = start_row + 2 + i
            ws.cell(dr, 1, step['label'])
            for col, key in [(2, 'start_val'), (3, 'volume'), (4, 'price'), (5, 'mix'), (6, 'end_val')]:
                c = ws.cell(dr, col, step[key]); c.number_format = '#,##0'
            c = ws.cell(dr, 7, step['end_val'] - step['start_val']); c.number_format = '+#,##0;-#,##0;0'

        return start_row + len(steps) + 3

    if len(target_years) >= 2:
        df_bridge = df_normal[df_normal['Year'].isin(target_years)]

        # Company-wide bridge (by Category)
        res1 = calculate_bridge_mix_steps(df_bridge, target_years, 'Category')
        next_row = write_bridge_section(ws_br, "1. Company-Wide Bridge (by Category)", res1, 1)

        # All Categories bridge (by Car Type)
        res2 = calculate_bridge_mix_steps(df_bridge, target_years, 'Car_Type')
        next_row = write_bridge_section(ws_br, "2. All Categories Bridge (by Car Type)", res2, next_row)

        # Category-specific bridges
        categories = sorted(df_normal['Category'].unique())
        for cat in categories:
            df_cat = df_bridge[df_bridge['Category'] == cat]
            res_cat = calculate_bridge_mix_steps(df_cat, target_years, 'Car_Type')
            if res_cat:
                next_row = write_bridge_section(ws_br, f"3. {cat} Bridge (by Car Type)", res_cat, next_row)
    else:
        ws_br.cell(1, 1, "최소 2개 연도를 선택해야 Bridge 분석이 가능합니다.")

    auto_width(ws_br)

    # === Sheet 6: Outlier Analysis ===
    ws_out = wb.create_sheet("Outlier_Analysis")

    stats = df_overview.groupby('Category')['ASP'].agg(['count', 'mean', 'min', 'max']).reset_index()
    stats.columns = ['Category', 'Count', 'Mean ASP', 'Min ASP', 'Max ASP']

    add_section_title(ws_out, "1. ASP Statistics by Category", 1, 5)
    for col, h in enumerate(stats.columns.tolist(), 1):
        cell = ws_out.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for i, (_, row) in enumerate(stats.iterrows()):
        dr = 3 + i
        ws_out.cell(dr, 1, row['Category'])
        ws_out.cell(dr, 2, int(row['Count']))
        for col, key in [(3, 'Mean ASP'), (4, 'Min ASP'), (5, 'Max ASP')]:
            c = ws_out.cell(dr, col, row[key]); c.number_format = '#,##0'

    # Top 50 Outliers
    outliers = df_overview.groupby('Category', group_keys=False).apply(lambda x: detect_outliers_iqr(x))
    outlier_view = outliers[['Date', 'Category', 'Car_Type', 'Customer', 'Quantity', 'Revenue', 'ASP']].sort_values('ASP', ascending=False).head(50) if not outliers.empty else pd.DataFrame()

    out_start = len(stats) + 4
    add_section_title(ws_out, "2. Top 50 ASP Outliers (IQR Factor=1.5)", out_start, 7)
    if not outlier_view.empty:
        out_headers = ['Date', 'Category', 'Car_Type', 'Customer', 'Quantity', 'Revenue', 'ASP']
        for col, h in enumerate(out_headers, 1):
            cell = ws_out.cell(out_start+1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for i, (_, row) in enumerate(outlier_view.iterrows()):
            dr = out_start + 2 + i
            ws_out.cell(dr, 1, row['Date'].strftime('%Y-%m-%d') if hasattr(row['Date'], 'strftime') else str(row['Date']))
            ws_out.cell(dr, 2, row['Category'])
            ws_out.cell(dr, 3, row['Car_Type'])
            ws_out.cell(dr, 4, row['Customer'])
            c = ws_out.cell(dr, 5, row['Quantity']); c.number_format = '#,##0'
            c = ws_out.cell(dr, 6, row['Revenue']); c.number_format = '#,##0'
            c = ws_out.cell(dr, 7, row['ASP']); c.number_format = '#,##0'
    else:
        ws_out.cell(out_start+1, 1, "이상치 없음")

    auto_width(ws_out)

    # === Sheet 7: Exception Data ===
    ws_exc = wb.create_sheet("Exception_Data")

    if not df_exception.empty:
        exc_summary = df_exception.groupby(['Year', 'Category', 'Customer'])['Revenue'].sum().reset_index()

        add_section_title(ws_exc, "Exception Sales (Quantity=0) Summary", 1, 4)
        for col, h in enumerate(['Year', 'Category', 'Customer', 'Revenue'], 1):
            cell = ws_exc.cell(2, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for i, (_, row) in enumerate(exc_summary.iterrows()):
            dr = 3 + i
            ws_exc.cell(dr, 1, int(row['Year']))
            ws_exc.cell(dr, 2, row['Category'])
            ws_exc.cell(dr, 3, row['Customer'])
            c = ws_exc.cell(dr, 4, row['Revenue']); c.number_format = '#,##0'
    else:
        ws_exc.cell(1, 1, "예외 데이터 없음")

    auto_width(ws_exc)

    # === Sheet 8: Raw Data ===
    ws_raw = wb.create_sheet("Raw_Data")
    raw_cols = ['Date', 'Year', 'Category', 'Car_Type', 'Customer', 'Region', 'Quantity', 'Revenue', 'ASP']
    raw_export = df_normal[raw_cols].copy()

    add_section_title(ws_raw, "Raw Transaction Data (Normal Sales)", 1, len(raw_cols))
    for col, h in enumerate(raw_cols, 1):
        cell = ws_raw.cell(2, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for i, (_, row) in enumerate(raw_export.head(10000).iterrows()):  # Limit to 10k rows for performance
        dr = 3 + i
        ws_raw.cell(dr, 1, row['Date'].strftime('%Y-%m-%d') if hasattr(row['Date'], 'strftime') else str(row['Date']))
        ws_raw.cell(dr, 2, int(row['Year']))
        ws_raw.cell(dr, 3, row['Category'])
        ws_raw.cell(dr, 4, row['Car_Type'])
        ws_raw.cell(dr, 5, row['Customer'])
        ws_raw.cell(dr, 6, row['Region'])
        c = ws_raw.cell(dr, 7, row['Quantity']); c.number_format = '#,##0'
        c = ws_raw.cell(dr, 8, row['Revenue']); c.number_format = '#,##0'
        c = ws_raw.cell(dr, 9, row['ASP']); c.number_format = '#,##0'

    auto_width(ws_raw)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# --- Helper: 차트 레이블 생성 ---
def create_label_col(df, val_col, total_val=None, prefix="€ "):
    if total_val is None:
        total_val = df[val_col].sum()

    if total_val == 0: total_val = 1

    return df[val_col].apply(lambda x: f"{prefix}{x/1000000:,.1f}M<br>({x/total_val*100:.1f}%)")

# --- 6. 메인 UI ---
def main():
    st.sidebar.markdown(f"<h1 style='color: {COLORS['primary']};'>Sales Dashboard (MSK)</h1>", unsafe_allow_html=True)

    # 1. 파일 업로드
    st.sidebar.markdown("### 📂 Data Upload")
    uploaded_file = st.sidebar.file_uploader("업로드: 매출데이터_MSK.xlsx", type=['xlsx'])

    if uploaded_file is None:
        st.info("👈 좌측 사이드바에서 엑셀 파일을 업로드해주세요.")
        return

    raw_df = load_data(uploaded_file)
    if raw_df.empty: return

    # 2. 데이터 분리
    df_normal, df_exception = split_data(raw_df)

    # 엑셀 리포트 생성 (전체 데이터 기준, 1회만 실행)
    all_years_list = sorted(df_normal['Year'].unique())
    excel_data_full = generate_excel_report(df_normal, df_exception, all_years_list)

    # 3. 사이드바 필터
    st.sidebar.header("🔍 Global Filters (for Tabs 2~7)")

    all_years = sorted(df_normal['Year'].unique())
    selected_years = st.sidebar.multiselect(
        "Analysis Years", all_years,
        default=all_years[-2:] if len(all_years) >= 2 else all_years
    )

    cat_opts = ["All"] + sorted(df_normal['Category'].unique().tolist())
    sel_cat = st.sidebar.selectbox("Category", cat_opts)

    if sel_cat != "All":
        temp_df = df_normal[df_normal['Category'] == sel_cat]
    else:
        temp_df = df_normal

    car_opts = ["All"] + sorted(temp_df['Car_Type'].unique().tolist())
    sel_car = st.sidebar.selectbox("Car Type", car_opts)

    cust_opts = ["All"] + sorted(temp_df['Customer'].unique().tolist())
    sel_cust = st.sidebar.selectbox("Customer", cust_opts)

    # Excel Export Button
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📥 Excel Export")
    st.sidebar.info("💡 리포트는 필터와 무관하게 **전체 데이터**를 포함합니다.")
    filename = f"Sales_Report_MSK_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.sidebar.download_button(
        label="📥 Download Full Excel Report",
        data=excel_data_full,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # 차종 설명 (좌측 하단)
    st.sidebar.markdown("---")
    st.sidebar.markdown("### ℹ️ 차종(Code) Reference")
    code_map_data = {
        '차종 (Code)': ['F1', 'HW', 'J7', 'JM', 'JY', '7F', 'GT', 'G4', 'IT', 'QL', 'CD'],
        'Model': ['SPORTAGE', 'SPORTAGE', 'CEED', 'CEED', 'SPORTAGE', 'KONA', 'TUCSON', 'I30', 'TUCSON', 'SPORTAGE', 'CEED']
    }
    st.sidebar.dataframe(pd.DataFrame(code_map_data), hide_index=True, use_container_width=True)

    # 4. 필터 적용
    filtered_normal = df_normal[df_normal['Year'].isin(selected_years)]
    if sel_cat != "All": filtered_normal = filtered_normal[filtered_normal['Category'] == sel_cat]
    if sel_car != "All": filtered_normal = filtered_normal[filtered_normal['Car_Type'] == sel_car]
    if sel_cust != "All": filtered_normal = filtered_normal[filtered_normal['Customer'] == sel_cust]

    if filtered_normal.empty:
        st.warning("⚠️ 선택된 필터 조건에 해당하는 데이터가 없습니다.")
        return

    # KPI Summary
    st.markdown("### 📊 Executive Summary (Filtered Scope)")
    k1, k2, k3, k4 = st.columns(4)
    f_rev = filtered_normal['Revenue'].sum()
    f_qty = filtered_normal['Quantity'].sum()
    f_asp = f_rev / f_qty if f_qty > 0 else 0

    if len(selected_years) > 0:
        max_y = max(selected_years)
        curr = filtered_normal[filtered_normal['Year'] == max_y]['Revenue'].sum()
        prev = filtered_normal[filtered_normal['Year'] == (max_y - 1)]['Revenue'].sum()
        yoy = ((curr - prev) / prev * 100) if prev > 0 else 0
    else:
        max_y, yoy = "-", 0

    k1.metric("Revenue", f"€ {f_rev:,.0f}")
    k2.metric("Quantity", f"{f_qty:,.0f}")
    k3.metric("Avg ASP", f"€ {f_asp:,.0f}")
    k4.metric(f"YoY Growth ('{max_y})", f"{yoy:,.1f}%")

    tabs = st.tabs([
        "🏢 Main Overview (Fixed)",
        "🚀 Growth & Share",
        "🤝 Customer Analysis",
        "📅 Time Series",
        "💰 P/Q/Mix Bridge",
        "🚨 Outlier",
        "⚠️ Exception (Qty=0)"
    ])

    # === Tab 1: Main Overview (Fixed) ===
    with tabs[0]:
        st.subheader("🏢 Overall Business Overview (Unfiltered)")

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("##### 1. Revenue by Category")
            df_cat = df_normal.groupby('Category')['Revenue'].sum().reset_index()
            df_cat['Label'] = create_label_col(df_cat, 'Revenue')

            fig1 = px.bar(df_cat, x='Category', y='Revenue', text='Label',
                          color='Category', color_discrete_sequence=COLORS['sequence'])
            st.plotly_chart(fig1, use_container_width=True)

            st.markdown("##### 3. Revenue by Customer (Top 10)")
            df_cust = df_normal.groupby('Customer')['Revenue'].sum().reset_index().sort_values('Revenue', ascending=False).head(10)
            df_cust['Label'] = create_label_col(df_cust, 'Revenue', total_val=df_normal['Revenue'].sum())

            fig3 = px.bar(df_cust, x='Revenue', y='Customer', orientation='h', text='Label',
                          color_discrete_sequence=[COLORS['secondary']])
            fig3.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig3, use_container_width=True)

        with c2:
            st.markdown("##### 2. Revenue by Car Type (Top 10)")
            df_car = df_normal.groupby('Car_Type')['Revenue'].sum().reset_index().sort_values('Revenue', ascending=False).head(10)
            df_car['Label'] = create_label_col(df_car, 'Revenue', total_val=df_normal['Revenue'].sum())

            fig2 = px.bar(df_car, x='Car_Type', y='Revenue', text='Label',
                          color_discrete_sequence=[COLORS['accent']])
            st.plotly_chart(fig2, use_container_width=True)

            st.markdown("##### 4. Revenue by Region")
            df_reg = df_normal.groupby('Region')['Revenue'].sum().reset_index()
            fig4 = px.pie(df_reg, values='Revenue', names='Region', hole=0.4,
                          color_discrete_sequence=px.colors.qualitative.Pastel)
            fig4.update_traces(textinfo='percent+label', texttemplate='%{label}<br>%{value:,.0f}€<br>(%{percent})')
            st.plotly_chart(fig4, use_container_width=True)

        st.markdown("---")
        st.subheader("📈 Yearly Trend by Category (Line Charts)")
        trend_cat = df_normal.groupby(['Year', 'Category']).agg({'Revenue':'sum', 'Quantity':'sum'}).reset_index()
        trend_cat['ASP'] = trend_cat['Revenue'] / trend_cat['Quantity']

        l1, l2, l3 = st.columns(3)
        with l1:
            fig_l1 = px.line(trend_cat, x='Year', y='Revenue', color='Category', markers=True, title="Revenue Trend", color_discrete_sequence=COLORS['sequence'])
            fig_l1.update_traces(texttemplate='%{y:,.2s}', textposition='top center')
            fig_l1.update_xaxes(dtick=1)
            st.plotly_chart(fig_l1, use_container_width=True)
        with l2:
            fig_l2 = px.line(trend_cat, x='Year', y='Quantity', color='Category', markers=True, title="Quantity Trend", color_discrete_sequence=COLORS['sequence'])
            fig_l2.update_traces(texttemplate='%{y:,.2s}', textposition='top center')
            fig_l2.update_xaxes(dtick=1)
            st.plotly_chart(fig_l2, use_container_width=True)
        with l3:
            fig_l3 = px.line(trend_cat, x='Year', y='ASP', color='Category', markers=True, title="ASP Trend (€)", color_discrete_sequence=COLORS['sequence'])
            fig_l3.update_traces(texttemplate='%{y:,.0f}', textposition='top center')
            fig_l3.update_xaxes(dtick=1)
            st.plotly_chart(fig_l3, use_container_width=True)

    # === Tab 2: Growth & Share ===
    with tabs[1]:
        st.subheader("🚀 Growth & Share Analysis")
        g1, g2 = st.columns(2)
        with g1:
            st.markdown("##### Yearly Revenue & Growth")
            y_df = filtered_normal.groupby('Year')['Revenue'].sum().reset_index()
            y_df['YoY'] = y_df['Revenue'].pct_change() * 100

            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(go.Bar(x=y_df['Year'], y=y_df['Revenue'], name="Revenue", marker_color=COLORS['primary'],
                                 text=y_df['Revenue'].apply(lambda x: f"€ {x/1000000:.1f}M"), textposition='auto'), secondary_y=False)
            fig.add_trace(go.Scatter(x=y_df['Year'], y=y_df['YoY'], name="YoY %", mode='lines+markers+text',
                                     line=dict(color=COLORS['secondary'], width=3),
                                     text=y_df['YoY'].apply(lambda x: f"{x:.1f}%" if pd.notnull(x) else ""), textposition="top center"), secondary_y=True)
            fig.update_xaxes(dtick=1)
            st.plotly_chart(fig, use_container_width=True)

        with g2:
            st.markdown("##### Revenue Share by Car Type")
            comp_df = filtered_normal.groupby(['Year', 'Car_Type'])['Revenue'].sum().reset_index()
            comp_df['Total'] = comp_df.groupby('Year')['Revenue'].transform('sum')
            comp_df['Share'] = (comp_df['Revenue'] / comp_df['Total'] * 100)
            comp_df['Label'] = comp_df.apply(lambda r: f"{r['Revenue']/1000000:.1f}M<br>({r['Share']:.1f}%)", axis=1)

            fig = px.bar(comp_df, x='Year', y='Revenue', color='Car_Type',
                         text='Label', color_discrete_sequence=COLORS['sequence'])
            fig.update_traces(textposition='inside')
            fig.update_xaxes(dtick=1)
            st.plotly_chart(fig, use_container_width=True)

        st.markdown("---")
        st.subheader("📦 Volume & Price Trend by Car Type")
        trend_car = filtered_normal.groupby(['Year', 'Car_Type']).agg({'Quantity': 'sum','Revenue': 'sum'}).reset_index()
        trend_car['ASP'] = trend_car['Revenue'] / trend_car['Quantity']
        trend_total = filtered_normal.groupby('Year').agg({'Quantity': 'sum','Revenue': 'sum'}).reset_index()
        trend_total['ASP'] = trend_total['Revenue'] / trend_total['Quantity']

        g3, g4 = st.columns(2)
        with g3:
            trend_car['Total_Qty'] = trend_car.groupby('Year')['Quantity'].transform('sum')
            trend_car['Qty_Label'] = trend_car.apply(lambda r: f"{r['Quantity']/1000:.0f}k<br>({r['Quantity']/r['Total_Qty']*100:.0f}%)", axis=1)

            fig_qty = px.bar(trend_car, x='Year', y='Quantity', color='Car_Type',
                             text='Qty_Label', title="Quantity Trend & Total", color_discrete_sequence=COLORS['sequence'])
            fig_qty.add_trace(go.Scatter(x=trend_total['Year'], y=trend_total['Quantity'], name='Total Qty', mode='lines+markers+text',
                line=dict(color='#333333', width=3, dash='dot'), text=trend_total['Quantity'].apply(lambda x: f'{x:,.0f}'), textposition="top center"))
            fig_qty.update_xaxes(dtick=1)
            st.plotly_chart(fig_qty, use_container_width=True)

        with g4:
            fig_asp = px.line(trend_car, x='Year', y='ASP', color='Car_Type',
                              title="ASP Trend & Average", markers=True, color_discrete_sequence=COLORS['sequence'])
            fig_asp.add_trace(go.Scatter(x=trend_total['Year'], y=trend_total['ASP'], name='Avg ASP', mode='lines+markers+text',
                line=dict(color='#D32929', width=4, dash='dot'), text=trend_total['ASP'].apply(lambda x: f'{x:,.0f}'), textposition="top center"))
            fig_asp.update_yaxes(tickprefix="€ ")
            fig_asp.update_xaxes(dtick=1)
            st.plotly_chart(fig_asp, use_container_width=True)

    # === Tab 3: Customer Analysis ===
    with tabs[2]:
        st.subheader("🤝 Customer-Centric Analysis")
        st.markdown("#### 1. Customer Revenue Trend (Fixed, All Data)")

        cust_trend = df_normal.groupby(['Year', 'Customer'])['Revenue'].sum().reset_index()
        cust_trend['Total'] = cust_trend.groupby('Year')['Revenue'].transform('sum')
        cust_trend['Share'] = (cust_trend['Revenue'] / cust_trend['Total'] * 100)
        cust_trend['Label'] = cust_trend.apply(lambda r: f"€{r['Revenue']/1000000:.1f}M<br>({r['Share']:.1f}%)", axis=1)

        fig_cust = px.bar(cust_trend, x='Year', y='Revenue', color='Customer',
                          text='Label', title="Yearly Revenue by Customer",
                          color_discrete_sequence=COLORS['sequence'])
        fig_cust.update_traces(textposition='inside')
        fig_cust.update_xaxes(dtick=1)
        st.plotly_chart(fig_cust, use_container_width=True)

        st.markdown("---")
        st.markdown("#### 2. Detailed Breakdown by Customer")

        if sel_cust == "All":
            st.info("👈 좌측 사이드바에서 **특정 거래처(Customer)**를 선택하면 상세 분석 차트가 표시됩니다.")
        else:
            st.success(f"Selected Customer: **{sel_cust}**")
            target_cust_df = df_normal[df_normal['Customer'] == sel_cust]
            categories = sorted(target_cust_df['Category'].unique())

            if not categories:
                st.warning("데이터 없음")
            else:
                cols = st.columns(len(categories))
                for idx, cat in enumerate(categories):
                    with cols[idx]:
                        st.markdown(f"##### {cat}")
                        cat_df = target_cust_df[target_cust_df['Category'] == cat]
                        cat_grp = cat_df.groupby(['Year', 'Car_Type'])['Revenue'].sum().reset_index()
                        cat_grp['Total'] = cat_grp.groupby('Year')['Revenue'].transform('sum')
                        cat_grp['Label'] = cat_grp.apply(lambda r: f"{r['Revenue']/1000:.0f}k ({r['Revenue']/r['Total']*100:.0f}%)", axis=1)

                        fig_cat = px.bar(cat_grp, x='Year', y='Revenue', color='Car_Type',
                                         title=f"{cat} Revenue", text='Label',
                                         color_discrete_sequence=COLORS['sequence'])
                        fig_cat.update_xaxes(dtick=1)
                        st.plotly_chart(fig_cat, use_container_width=True)

    # === Tab 4: Time Series ===
    with tabs[3]:
        st.subheader("📅 Time Series Analysis")
        mode = st.radio("Time Unit", ["Monthly", "Quarterly"], horizontal=True)
        t_col = 'Month_Dt' if mode == "Monthly" else 'Quarter_Str'

        if sel_cat == "All":
            mix_col = 'Category'
        else:
            mix_col = 'Car_Type'

        t_df = filtered_normal.groupby([t_col, mix_col])['Revenue'].sum().reset_index()
        if mode == "Quarterly": t_df = t_df.sort_values(t_col)

        t_df['Total'] = t_df.groupby(t_col)['Revenue'].transform('sum')
        t_df['Share'] = t_df['Revenue'] / t_df['Total']

        col_t1, col_t2 = st.columns([2, 1])
        with col_t1:
            fig_area = px.area(t_df, x=t_col, y='Revenue', color=mix_col,
                               title=f"Revenue Trend", color_discrete_sequence=COLORS['sequence'])
            if mode == "Monthly": fig_area.update_xaxes(tickformat="%Y-%m")
            st.plotly_chart(fig_area, use_container_width=True)

        with col_t2:
            fig_mix = px.bar(t_df, x=t_col, y='Share', color=mix_col,
                             title="Revenue Mix", color_discrete_sequence=COLORS['sequence'])
            fig_mix.update_traces(texttemplate='%{y:.0%}', textposition='inside')
            fig_mix.update_yaxes(tickformat=".0%")
            st.plotly_chart(fig_mix, use_container_width=True)

        st.markdown("---")
        st.markdown("##### Quantity vs ASP Trend")
        if mode == "Monthly":
            agg_ts = filtered_normal.groupby('Month_Dt').agg({'Revenue':'sum', 'Quantity':'sum'}).reset_index()
            x_ax = 'Month_Dt'
        else:
            agg_ts = filtered_normal.groupby('Quarter_Str').agg({'Revenue':'sum', 'Quantity':'sum'}).reset_index().sort_values('Quarter_Str')
            x_ax = 'Quarter_Str'

        agg_ts['ASP'] = np.where(agg_ts['Quantity']>0, agg_ts['Revenue']/agg_ts['Quantity'], 0)

        fig_dual = make_subplots(specs=[[{"secondary_y": True}]])
        fig_dual.add_trace(go.Bar(x=agg_ts[x_ax], y=agg_ts['Quantity'], name="Quantity",
                                  text=agg_ts['Quantity'].apply(lambda x: f"{x:,.0f}"), textposition='auto',
                                  marker_color='#B0BEC5', opacity=0.6), secondary_y=False)
        fig_dual.add_trace(go.Scatter(x=agg_ts[x_ax], y=agg_ts['ASP'], name="ASP", mode='lines+markers+text',
                                      text=agg_ts['ASP'].apply(lambda x: f"{x:,.0f}"), textposition='top center',
                                      line=dict(color=COLORS['primary'], width=3)), secondary_y=True)
        if mode == "Monthly": fig_dual.update_xaxes(tickformat="%Y-%m")
        st.plotly_chart(fig_dual, use_container_width=True)

    # === Tab 5: Bridge (3-Level Analysis) ===
    with tabs[4]:
        st.subheader("💰 Price / Volume / Mix Bridge (3-Level Analysis)")
        st.info("매출 변동을 **물량(Volume)**, **가격(Price)**, **제품 구성(Mix)** 효과로 분해하여 3단계로 분석합니다.")

        target_years = sorted(list(set(selected_years)))

        if len(target_years) >= 2:
            # Customer 필터 적용한 베이스 데이터
            df_bridge_base = df_normal[df_normal['Year'].isin(target_years)]
            if sel_cust != "All":
                df_bridge_base = df_bridge_base[df_bridge_base['Customer'] == sel_cust]

            # === Bridge 1: Company-Wide (전체 데이터, Customer 필터만) ===
            st.markdown("#### 🌐 Bridge 1: Company-Wide Analysis (by Category)")
            st.caption("전체 데이터 기준 (Customer 필터만 적용) - Category별 집계")

            steps1 = calculate_bridge_mix_steps(df_bridge_base, target_years, 'Category')
            if steps1:
                fig1 = create_waterfall_fig_multi(steps1, "Company-Wide Bridge (by Category)")
                st.plotly_chart(fig1, use_container_width=True)
            else:
                st.warning("분석할 데이터가 부족합니다.")

            st.markdown("---")

            # === Bridge 2: Category-Level (Category 필터 적용) ===
            if sel_cat != "All":
                st.markdown(f"#### 🏭 Bridge 2: Category-Level Analysis - {sel_cat} (by Car Type)")
                st.caption(f"선택된 Category '{sel_cat}' + Customer 필터 - Car_Type별 집계")

                df_bridge2 = df_bridge_base[df_bridge_base['Category'] == sel_cat]
                steps2 = calculate_bridge_mix_steps(df_bridge2, target_years, 'Car_Type')

                if steps2:
                    fig2 = create_waterfall_fig_multi(steps2, f"{sel_cat} Bridge (by Car Type)")
                    st.plotly_chart(fig2, use_container_width=True)
                else:
                    st.warning("분석할 데이터가 부족합니다.")

                st.markdown("---")

                # === Bridge 3: Detailed (Category + Car Type 필터) ===
                if sel_car != "All":
                    st.markdown(f"#### 🔍 Bridge 3: Detailed Analysis - {sel_cat} > {sel_car} (by Customer)")
                    st.caption(f"선택된 Category + Car Type + Customer 필터 - Customer별 집계")

                    df_bridge3 = df_bridge2[df_bridge2['Car_Type'] == sel_car]
                    steps3 = calculate_bridge_mix_steps(df_bridge3, target_years, 'Customer')

                    if steps3:
                        fig3 = create_waterfall_fig_multi(steps3, f"{sel_cat} > {sel_car} Bridge (by Customer)")
                        st.plotly_chart(fig3, use_container_width=True)
                    else:
                        st.warning("분석할 데이터가 부족합니다.")
                else:
                    st.info("💡 **Car Type**을 선택하면 Bridge 3 (Customer별 상세 분석)이 표시됩니다.")
            else:
                st.info("💡 **Category**를 선택하면 Bridge 2 (Car Type별 분석)와 Bridge 3 (Customer별 분석)이 표시됩니다.")
        else:
            st.info("비교할 연도를 2개 이상 선택해주세요.")

    # === Tab 6: Outlier ===
    with tabs[5]:
        st.subheader("🚨 Outlier Analysis (ASP)")
        c_o1, c_o2 = st.columns([2, 1])
        with c_o1:
            fig_box = px.box(filtered_normal, x='Category', y='ASP', color='Category', points="outliers", notched=True)
            st.plotly_chart(fig_box, use_container_width=True)
        with c_o2:
            st.dataframe(filtered_normal.groupby('Category')['ASP'].agg(['count', 'mean', 'min', 'max']), use_container_width=True)

        st.markdown("#### Top Outliers")
        outliers = filtered_normal.groupby('Category', group_keys=False).apply(lambda x: detect_outliers_iqr(x))
        if not outliers.empty:
            st.dataframe(outliers[['Date','Customer','Category','Car_Type','Quantity','Revenue','ASP']].sort_values('ASP', ascending=False).head(50), use_container_width=True)
        else:
            st.success("이상치 없음")

    # === Tab 7: Exception (Fixed) ===
    with tabs[6]:
        st.subheader("⚠️ Exception Sales Analysis (Quantity = 0)")
        if df_exception.empty:
            st.success("예외 데이터 없음")
        else:
            exc_year = df_exception.groupby('Year')['Revenue'].sum().reset_index()
            fig_exc = px.bar(exc_year, x='Year', y='Revenue', text_auto=',.0f', color_discrete_sequence=[COLORS['negative']])
            fig_exc.update_yaxes(tickprefix="€ ")
            st.plotly_chart(fig_exc, use_container_width=True)

            ec1, ec2 = st.columns(2)
            with ec1:
                st.markdown("##### By Customer")
                exc_cust = df_exception.groupby('Customer')['Revenue'].sum().reset_index().sort_values('Revenue', ascending=False).head(10)
                fig_ec = px.bar(exc_cust, y='Customer', x='Revenue', orientation='h', color_discrete_sequence=[COLORS['secondary']])
                fig_ec.update_layout(yaxis={'categoryorder':'total ascending'})
                st.plotly_chart(fig_ec, use_container_width=True)

            with ec2:
                st.markdown("##### By Category & Car Type")
                tree_df = df_exception.groupby(['Category', 'Car_Type'])['Revenue'].sum().reset_index()
                tree_df = tree_df[tree_df['Revenue'] > 0]

                if tree_df.empty:
                    st.warning("표시할 양수 매출 데이터가 없습니다.")
                else:
                    fig_tree = px.treemap(tree_df, path=['Category', 'Car_Type'], values='Revenue',
                                          color='Revenue', color_continuous_scale='Reds')
                    st.plotly_chart(fig_tree, use_container_width=True)

            with st.expander("데이터 원본 보기"):
                st.dataframe(df_exception)

if __name__ == "__main__":
    main()
