"""gpcm_mcp.py 회귀 테스트 — Streamlit 없이 MCP 도구 경로를 검증한다.

이 환경은 DART·KRX가 차단돼 있어 test_gpcm_kr_quality.py 와 같은 스텁 패턴을 쓴다:
gpcm_kr 모듈 전역을 원숭이 패칭해 파이프라인을 통째로 돌리고, 산출 엑셀까지 연다.

따로 확인하는 것 하나 — stdout 위생. MCP 는 stdout 이 프로토콜 통로라,
gpcm_kr 임포트(사이드바 실행)가 한 글자라도 흘리면 클라이언트 파서가 깨진다.
"""
import json, os, subprocess, sys, tempfile, threading
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

os.environ["DART_API_KEY"] = "TESTKEY-" + "0" * 32

import gpcm_mcp as W

M = W._load()

# --- 파이프라인 스텁 (test_gpcm_kr_quality.py 와 동일한 대역) -----------------

BS = pd.DataFrame([
    {'sj_nm': '연결재무상태표', 'account_nm': '현금및현금성자산', 'account_id': 'ifrs-full_CashAndCashEquivalents', 'thstrm_amount': '100000000000'},
    {'sj_nm': '연결재무상태표', 'account_nm': '자본총계', 'account_id': 'ifrs-full_Equity', 'thstrm_amount': '600000000000'},
])
PL = pd.DataFrame([
    {'sj_nm': '연결포괄손익계산서', 'account_nm': '매출액', 'account_id': 'ifrs-full_Revenue',
     'thstrm_amount': '800000000000', 'thstrm_add_amount': '800000000000'},
    {'sj_nm': '연결포괄손익계산서', 'account_nm': '영업이익', 'account_id': 'dart_OperatingIncomeLoss',
     'thstrm_amount': '80000000000', 'thstrm_add_amount': '80000000000'},
])


class Dart:
    corp_codes = pd.DataFrame({'corp_code': ['00126380'], 'corp_name': ['테스트'], 'stock_code': ['005930']})
    def find_corp_code(self, c): return '00126380'
    def company(self, c): return {'corp_name': '테스트'}
    def finstate_all(self, corp, y, reprt_code='11011', fs_div='CFS'):
        return BS.copy() if fs_div == 'CFS' else pd.DataFrame()


def stub_pipeline():
    M.get_krx_listing = lambda: pd.DataFrame({'Code': ['005930'], 'Name': ['테스트'], 'Stocks': [100]})
    M.get_stock_price = lambda t, d: (70000, d)
    M.get_outstanding_shares = lambda *a, **k: (1_000_000, 'KRX', {})
    M.fetch_pl_df = lambda d, c, y, r: (PL.copy(), 'CFS', 'OK')
    M.check_dart_reachable = lambda timeout=10: (True, None)
    M.get_dart_reader = lambda key: Dart()
    M.time.sleep = lambda s: None  # 티커당 0.5초 대기 생략


fails = []


def check(label, ok, detail=''):
    print(f"{'OK  ' if ok else '실패'}  {label}{('  — ' + str(detail)) if detail and not ok else ''}")
    if not ok:
        fails.append(label)


def expect_error(label, fn, fragment):
    try:
        fn()
    except Exception as e:
        check(label, fragment in str(e), str(e)[:120])
    else:
        check(label, False, '에러가 나야 하는데 성공했다')


# --- 1. 기간 조립 --------------------------------------------------------------
check('단일 기간', W._build_periods('2025.4Q') == ['2025.4Q'])
check('연도 걸친 범위', W._build_periods('2024.3Q', '2025.2Q') ==
      ['2024.3Q', '2024.4Q', '2025.1Q', '2025.2Q'])
expect_error('역순 입력 거부', lambda: W._build_periods('2025.4Q', '2025.1Q'), '빠릅니다')
expect_error('형식 오류 거부', lambda: W._build_periods('2025-4Q'), '형식')

# --- 2. 시작 전 검증 -----------------------------------------------------------
stub_pipeline()
expect_error('빈 티커 거부', lambda: W.run_gpcm([], '2025.4Q'), '비어')
expect_error('6자리 아닌 티커 거부', lambda: W.run_gpcm(['5930'], '2025.4Q'), '6자리')
expect_error('beta_type 검증', lambda: W.run_gpcm(['005930'], '2025.4Q', beta_type='10Y'), '5Y')

M.check_dart_reachable = lambda timeout=10: (False, 'timeout')
expect_error('DART 불통이면 시작 전에 거부', lambda: W.run_gpcm(['005930'], '2025.4Q'), '연결할 수 없습니다')
M.check_dart_reachable = lambda timeout=10: (True, None)

saved_key = os.environ.pop('DART_API_KEY')
expect_error('인증키 없으면 거부', lambda: W.run_gpcm(['005930'], '2025.4Q'), '인증키가 없습니다')
# 확장(.mcpb)에서 입력칸이 비면 플레이스홀더 리터럴이 그대로 온다 — 없음으로 취급해야 한다
os.environ['DART_API_KEY'] = '${user_config.dart_api_key}'
expect_error('플레이스홀더 리터럴도 없음 취급', lambda: W.run_gpcm(['005930'], '2025.4Q'), '인증키가 없습니다')
os.environ['DART_API_KEY'] = saved_key

# --- 3. 실행 → 완료 → 엑셀 ----------------------------------------------------
W.OUTPUT_DIR = Path(tempfile.mkdtemp()) / 'GPCM'

r = W.run_gpcm(['005930'], '2024.1Q')
check('시작 즉시 running 으로 돌아온다', r['state'] == 'running', r)
check('출력 경로를 미리 알려준다', r['output_file'].endswith('.xlsx'), r['output_file'])

job = W._jobs[r['job_id']]
job['thread'].join(timeout=60)

s = W.gpcm_status(r['job_id'])
check('완료 상태 전이', s['state'] == 'done', s.get('error', s['state']))
check('진행률 100', s['progress_pct'] == 100, s['progress_pct'])
check('Data_Quality 요약이 실린다', 'errors' in s.get('data_quality', {}), s.get('data_quality'))
check('Target WACC 가 실린다', isinstance(s.get('target_wacc_pct'), float), s.get('target_wacc_pct'))
check('Data_Quality 확인 안내문', 'Data_Quality' in s.get('note', ''))

out = Path(s['file'])
check('엑셀 파일이 실제로 쓰였다', out.exists(), s['file'])
wb = load_workbook(out)
check('GPCM·Data_Quality 시트 존재', {'GPCM', 'Data_Quality'} <= set(wb.sheetnames), wb.sheetnames)
check('Notes 의 Base Date 가 요청 기간', any(
    '2024.1Q' in str(c.value) for row in wb['GPCM'].iter_rows() for c in row if c.value))

# --- 3-b. GPCM 시트 열 배치 (우선주를 EV 구성요소로 옮긴 뒤의 등가성) ----------
gp = wb['GPCM']
hdr = [c.value for c in gp[5][:len(M.GPCM_COL_DEFS)]]
check('헤더 순서가 열 정의와 일치', hdr == [h for _, h, _ in M.GPCM_COL_DEFS], hdr[:14])
check('우선주(장부)가 11열(K) — NCI와 Equity 사이', hdr[10] == '우선주(장부)', hdr[8:13])
check('BS & EV Components 섹션이 F4:M4', 'F4:M4' in {str(rng) for rng in gp.merged_cells.ranges}
      and gp.cell(4, 6).value == 'BS & EV Components')
check('Equity Adj 섹션은 사라짐', all(gp.cell(4, c).value != 'Equity Adj' for c in range(1, 40)))
check('EV 수식이 새 열 문자를 참조 (U=MktCap, K=우선주)',
      gp.cell(6, 13).value == '=U6+K6+G6-F6+J6-H6', gp.cell(6, 13).value)
check('우선주 셀은 BS_Full의 Preferred 를 집계', 'Preferred' in (gp.cell(6, 11).value or ''), gp.cell(6, 11).value)
check('D/E 수식이 우선주 새 위치(K) 참조', 'K6' in (gp.cell(6, 33).value or ''), gp.cell(6, 33).value)
wacc_refs = {c.value for row in wb['WACC_Calculation'].iter_rows() for c in row
             if isinstance(c.value, str) and c.value.startswith('=GPCM!')}
check('WACC 시트의 GPCM 참조가 이동된 열(AI·AH)로', 
      any('=GPCM!AI' in v for v in wacc_refs) and any('=GPCM!AH' in v for v in wacc_refs), wacc_refs)
check('Cash 고정틀 유지', gp.freeze_panes == 'F6', gp.freeze_panes)

check('job_id 생략 시 최근 작업', W.gpcm_status()['job_id'] == r['job_id'])
expect_error('없는 job_id', lambda: W.gpcm_status('job-999'), '찾을 수 없습니다')

# --- 4. 실행 중 중복 시작 거부 -------------------------------------------------
import threading
gate = threading.Event()
M.fetch_financial_data_orig = M.fetch_financial_data
def slow_fetch(*a, **k):
    gate.wait(timeout=30)
    return M.fetch_financial_data_orig(*a, **k)
M.fetch_financial_data = slow_fetch
r2 = W.run_gpcm(['005930'], '2024.1Q')
expect_error('실행 중이면 새 작업 거부', lambda: W.run_gpcm(['005930'], '2024.1Q'), '이미 실행 중')
gate.set()
W._jobs[r2['job_id']]['thread'].join(timeout=60)
M.fetch_financial_data = M.fetch_financial_data_orig

# --- 5. 실패 경로 --------------------------------------------------------------
def boom(*a, **k):
    raise RuntimeError('의도한 폭발')
M.fetch_financial_data = boom
r3 = W.run_gpcm(['005930'], '2024.1Q')
W._jobs[r3['job_id']]['thread'].join(timeout=30)
s3 = W.gpcm_status(r3['job_id'])
check('실패 상태와 원인이 남는다', s3['state'] == 'failed' and '의도한 폭발' in s3['error'], s3.get('error', '')[:80])
M.fetch_financial_data = M.fetch_financial_data_orig

# --- 6. stdout 위생 (별도 프로세스에서 임포트만) --------------------------------
proc = subprocess.run(
    [sys.executable, '-c', 'import gpcm_mcp; gpcm_mcp._load()'],
    cwd=os.path.dirname(os.path.abspath(__file__)),
    env={**os.environ, 'DART_API_KEY': 'x' * 40},
    capture_output=True, timeout=300)
check('임포트·로드가 stdout 에 아무것도 안 쓴다', proc.stdout == b'',
      proc.stdout[:200])

# 6-b. 로드 중에도 프로토콜 통로(stdout)가 살아 있어야 한다.
# 예전엔 contextlib.redirect_stdout 으로 임포트 출력을 막았는데, 그건 프로세스
# 전역이라 로드하는 60초 동안 서버 응답까지 통째로 stderr 로 샜다 — 클라이언트는
# initialize 응답을 못 받고 끊었다(데스크톱 로그로 확인). 그 상황을 그대로 세운다.
LEAK = 'IMPORT-CHATTER'
PROTO = 'PROTOCOL-FRAME'
script = (
    'import sys, threading, time\n'
    'import gpcm_mcp as W\n'
    'started, done = threading.Event(), threading.Event()\n'
    'def loader():\n'
    '    with W._muted_stdout():\n'
    f'        print("{LEAK}")\n'          # 임포트가 흘리는 출력을 흉내
    '        started.set(); time.sleep(1.0)\n'
    '    done.set()\n'
    'threading.Thread(target=loader, daemon=True).start()\n'
    'started.wait(5)\n'
    f'print("{PROTO}"); sys.stdout.flush()\n'   # 그 사이 서버가 보내는 응답
    'done.wait(5)\n'
)
proc2 = subprocess.run(
    [sys.executable, '-c', script],
    cwd=os.path.dirname(os.path.abspath(__file__)),
    env={**os.environ, 'DART_API_KEY': 'x' * 40},
    capture_output=True, text=True, timeout=120)
check('로드 중에도 프로토콜 출력은 stdout 으로 나간다', PROTO in proc2.stdout, proc2.stdout[:200])
check('로드가 흘린 출력만 stderr 로 빠진다',
      LEAK not in proc2.stdout and LEAK in proc2.stderr, proc2.stdout[:200])

# 6-c. 종단 — 실제로 서버를 띄워 initialize 응답이 stdout 으로 돌아오는지 본다.
server = subprocess.Popen(
    [sys.executable, 'gpcm_mcp.py'],
    cwd=os.path.dirname(os.path.abspath(__file__)),
    env={**os.environ, 'DART_API_KEY': 'x' * 40},
    stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, text=True)
try:
    server.stdin.write(json.dumps({
        "jsonrpc": "2.0", "id": 0, "method": "initialize",
        "params": {"protocolVersion": "2025-11-25", "capabilities": {},
                   "clientInfo": {"name": "test", "version": "0"}},
    }) + '\n')
    server.stdin.flush()
    box: list[str] = []
    reader = threading.Thread(target=lambda: box.append(server.stdout.readline()), daemon=True)
    reader.start()
    reader.join(timeout=30)
    reply = json.loads(box[0]) if box and box[0].strip() else {}
    check('initialize 응답이 30초 안에 stdout 으로 온다', reply.get('id') == 0, box[:1])
    check('응답에 도구 능력이 실려 있다',
          'tools' in reply.get('result', {}).get('capabilities', {}), reply.get('result'))
finally:
    server.kill()
    server.wait(timeout=10)

# --- 7. 오늘 상장사 명단 (KRX 차단 환경 — 응답을 흉내 내 검증) --------------------
IND = pd.DataFrame([
    {'Code': '005930', 'Name': '삼성전자', 'Market': 'KOSPI', 'Sector': '반도체 제조업',
     'Industry': 'DRAM, NAND Flash', 'SettleMonth': '12월'},
    {'Code': '000660', 'Name': 'SK하이닉스', 'Market': 'KOSPI', 'Sector': '반도체 제조업',
     'Industry': 'DRAM', 'SettleMonth': '12월'},
    {'Code': '111111', 'Name': '삼월결산㈜', 'Market': 'KOSDAQ', 'Sector': '반도체 제조업',
     'Industry': '반도체 장비', 'SettleMonth': '3월'},
    {'Code': '222222', 'Name': '농심', 'Market': 'KOSPI', 'Sector': '음식료품',
     'Industry': '라면', 'SettleMonth': '12월'},
])
M.get_krx_industry_listing = lambda: IND.copy()
W._roster_memo.update(at=0.0, df=None)  # 메모 비우기

r = W.list_krx_companies()
check('빈 query 는 업종 집계만', 'sectors' in r and 'companies' not in r, r)
check('업종별 종목수가 맞는다',
      {s['sector']: s['count'] for s in r['sectors']} == {'반도체 제조업': 3, '음식료품': 1}, r['sectors'])

r = W.list_krx_companies('반도체')
check('업종명 매칭 3개 (장비 주요제품 포함)', r['meta']['count'] == 3, r['meta'])
check('주요제품이 실린다', r['companies'][1]['product'] == 'DRAM, NAND Flash', r['companies'][1])
check('3월 결산 플래그', [c['fiscalMonthNot12'] for c in r['companies']] == [False, False, True])

r = W.list_krx_companies('반도체', december_only=True)
check('12월 결산 필터', r['meta']['count'] == 2, r['meta'])

r = W.list_krx_companies('라면')
check('주요제품으로도 찾는다', r['meta']['count'] == 1 and r['companies'][0]['name'] == '농심', r)

W._roster_memo.update(at=0.0, df=None)
M.get_krx_industry_listing = lambda: pd.DataFrame()
expect_error('KRX 실패는 명확한 에러 (조용한 폴백 없음)', lambda: W.list_krx_companies('반도체'), 'KRX')
W._roster_memo.update(at=0.0, df=None)

# --- 8. 거래정지 이력 점검 (fdr 응답을 흉내 내 검증) -----------------------------
days = pd.bdate_range('2024-01-01', periods=100)
def _px(idx): 
    return pd.DataFrame({'Close': [100.0]*len(idx)}, index=idx)

class FakeFdr:
    def DataReader(self, symbol, start=None, end=None):
        if symbol == 'KS11':                 return _px(days)
        if symbol == '000001':               return _px(days)                          # 정상
        if symbol == '000002':               return _px(days[:40].append(days[50:]))   # 중간 10거래일 결측
        if symbol == '000003':               return _px(days[50:])                     # 늦은 상장
        if symbol == '000004':               return _px(days[:-6])                     # 최근 6거래일 결측
        if symbol == '000005':               return pd.DataFrame()                     # 조회 실패
        raise RuntimeError(symbol)

M.fdr = FakeFdr()
g = W.check_trading_gaps(['000001', '000002', '000003', '000004', '000005'])
by = {x['code']: x for x in g['results']}
check('정상 종목은 무표시', by['000001']['flag'] is False and not by['000001']['suspectedHalts'])
check('중간 10거래일 결측을 정지 의심으로 잡는다',
      by['000002']['flag'] and by['000002']['suspectedHalts'][0]['tradingDays'] == 10, by['000002'])
check('늦은 상장은 정지로 오인하지 않는다', by['000003']['flag'] is False
      and by['000003']['observedFrom'] == days[50].strftime('%Y-%m-%d'), by['000003'])
check('최근 결측은 현재 정지 중으로 표시', by['000004']['currentlySuspended'] is True, by['000004'])
check('조회 실패는 failed 로 드러난다', g.get('failed') == ['000005'], g.get('failed'))
check('자동 배제 금지 안내가 실린다', '자동 배제' in g['meta']['note'])

# --- 8. 시장금리 조회 (WACC 입력값의 근거) -------------------------------------
class FakeRateFdr:
    """국고채만 답하는 FDR 대역 — 회사채는 FDR 에 없다."""
    def __init__(self, rows=None, symbol='KR5YT=RR'):
        self.rows, self.symbol, self.asked = rows, symbol, []
    def DataReader(self, symbol, start=None, end=None):
        self.asked.append(symbol)
        if symbol != self.symbol:
            raise RuntimeError(symbol)
        if self.rows is None:
            return pd.DataFrame()
        idx = pd.to_datetime(['2026-08-13', '2026-08-14'])
        return pd.DataFrame({'Close': self.rows}, index=idx)

M.fdr = FakeRateFdr([3.28, 3.31])
os.environ.pop('ECOS_API_KEY', None)
w = W.get_wacc_inputs(as_of='2026-08-18')
check('rf 는 조회한 값과 금리기준일을 함께 준다',
      w['rf']['value'] == 3.31 and w['rf']['rateDate'] == '2026-08-14', w['rf'])
check('출처가 값에 붙어 나온다', 'FinanceDataReader' in w['rf']['source'], w['rf'])
check('citation 한 줄로 Notes 에 넣을 수 있다', '3.31%' in w['citation'], w['citation'])
check('못 구한 회사채는 value 를 지어내지 않는다',
      w['kd_pretax']['value'] is None and 'failed' in w['kd_pretax'], w['kd_pretax'])
check('키 미설정을 실패 사유에 밝힌다', '인증키 미설정' in w['kd_pretax']['failed'], w['kd_pretax'])
check('mrp·size_premium 은 판단 항목이라 값이 안 나온다',
      'mrp' in w['judgment'] and 'mrp' not in w, list(w))

M.fdr = FakeRateFdr(None)          # 전부 실패 — 조용한 기본값이 있으면 안 된다
try:
    W.get_wacc_inputs()
    check('전부 실패하면 값을 만들지 않고 실패를 알린다', False, '에러가 안 났다')
except RuntimeError as exc:
    check('전부 실패하면 값을 만들지 않고 실패를 알린다', '직접 넣어야' in str(exc), str(exc)[:80])

os.environ['ECOS_API_KEY'] = '${user_config.ecos_api_key}'
check('미입력 확장 리터럴은 키 없음으로 본다', W._ecos_key() == '', W._ecos_key())
os.environ.pop('ECOS_API_KEY', None)

try:
    W.get_wacc_inputs(bond_grade='AAA')
    check('없는 등급은 거절한다', False, '통과했다')
except ValueError as exc:
    check('없는 등급은 거절한다', 'AA-' in str(exc), str(exc)[:60])

# ECOS 경로 — 항목코드를 박지 않고 이름으로 찾는지
os.environ['ECOS_API_KEY'] = 'ECOSKEY'
calls = []
def fake_ecos(path):
    calls.append(path)
    if path.startswith('StatisticItemList'):
        return {'StatisticItemList': {'row': [
            {'ITEM_CODE': '010200000', 'ITEM_NAME': '국고채(5년)'},
            {'ITEM_CODE': '010210000', 'ITEM_NAME': '회사채(3년, AA-)'},
            {'ITEM_CODE': '010220000', 'ITEM_NAME': '회사채(3년, BBB-)'},
        ]}}
    return {'StatisticSearch': {'row': [
        {'TIME': '20260813', 'DATA_VALUE': '4.10'},
        {'TIME': '20260814', 'DATA_VALUE': '4.12'},
    ]}}
W._ecos_get = fake_ecos
M.fdr = FakeRateFdr(None)          # 무키 경로 실패 → ECOS 로 넘어가야 한다
w2 = W.get_wacc_inputs(as_of='2026-08-18', bond_grade='BBB-')
check('무키 경로가 막히면 ECOS 로 넘어간다', w2['kd_pretax']['value'] == 4.12, w2['kd_pretax'])
check('등급에 맞는 항목코드를 이름으로 찾는다',
      any('010220000' in c for c in calls) and not any('010210000' in c for c in calls), calls)
check('ECOS 출처와 항목명이 남는다', '회사채(3년, BBB-)' in w2['kd_pretax']['source'], w2['kd_pretax'])
os.environ.pop('ECOS_API_KEY', None)

# 만기는 고정이 아니다 — 평가 대상 현금흐름 기간에 맞춰 바뀐다
check('만기 표기를 "10"·"10Y"·"10년" 모두 같게 읽는다',
      {W._maturity(x, 'rf') for x in ('10', '10Y', '10년', ' 10 년 ')} == {'10년'},
      [W._maturity(x, 'rf') for x in ('10', '10Y', '10년')])
check('빈 값이면 종목별 기본 만기', (W._maturity('', 'rf'), W._maturity('', 'kd')) == ('5년', '3년'))
expect_error('만기 형식 오류는 거절', lambda: W._maturity('오년', 'rf'), '만기는')

f10 = FakeRateFdr([4.05, 4.11], symbol='KR10YT=RR')
M.fdr = f10
w3 = W.get_wacc_inputs(as_of='2026-08-18', rf_maturity='10Y')
check('10년을 요청하면 10년 심볼로 조회한다', f10.asked == ['KR10YT=RR'], f10.asked)
check('라벨·citation 에 요청한 만기가 찍힌다',
      w3['rf']['label'] == '국고채 10년' and '국고채 10년' in w3['citation'], w3['rf']['label'])
check('만기 선택이 판단 항목으로 남는다', '10년' in w3['judgment']['maturity'], w3['judgment']['maturity'])

# ECOS 경로도 만기를 이름으로 찾고, 없는 만기면 있는 항목을 알려준다
os.environ['ECOS_API_KEY'] = 'ECOSKEY'
W._ecos_get = fake_ecos
M.fdr = FakeRateFdr(None, symbol='없음')
w4 = W.get_wacc_inputs(as_of='2026-08-18', rf_maturity='5년')
check('ECOS 도 요청한 만기의 항목을 찾는다', '국고채(5년)' in w4['rf']['source'], w4['rf']['source'])
w5 = W.get_wacc_inputs(as_of='2026-08-18', rf_maturity='20년')
check('없는 만기는 실제로 있는 항목을 알려준다',
      '국고채(5년)' in w5['rf']['failed'], w5['rf'].get('failed'))
os.environ.pop('ECOS_API_KEY', None)

# 근거 문장이 실제로 엑셀 Notes 에 실리는지
seen = {}
export_orig = M.export_gpcm_excel
def spy_export(*a, **k):
    seen['notes'] = a[10]
    return export_orig(*a, **k)
M.export_gpcm_excel = spy_export
r4 = W.run_gpcm(['005930'], '2024.1Q', rate_source='국고채 5년 3.31% (2026-08-14, 한국은행 ECOS)')
W._jobs[r4['job_id']]['thread'].join(timeout=60)
M.export_gpcm_excel = export_orig
check('Notes 에 Rf/Kd 출처 줄이 Base Date 바로 아래 붙는다',
      seen.get('notes', ['', ''])[1].startswith('• Rf/Kd 출처: 국고채 5년 3.31%'),
      seen.get('notes', [])[:2])

# --- 9. 확장 manifest — 파이썬 버전을 못 박았는지 ------------------------------
# uv 는 지정이 없으면 시스템에서 제일 새 파이썬을 집는다. 실제로 Python 3.14 에서
# pydantic 이 깨져(MCP SDK 모델 생성 실패) 도구가 1개만 등록되는 사고가 났다.
# 데스크톱마다 다른 파이썬이 잡히지 않도록 manifest 가 버전을 지정해야 한다.
manifest = json.loads(Path(__file__).with_name('manifest.json').read_text(encoding='utf-8'))
args = manifest['server']['mcp_config']['args']
check('manifest 가 파이썬 버전을 지정한다', '--python' in args, args)
pinned = args[args.index('--python') + 1] if '--python' in args else ''
check('지정한 버전이 검증된 범위(3.10~3.13)', pinned in ('3.10', '3.11', '3.12', '3.13'), pinned)
check('선언한 호환 범위가 3.14 를 제외한다',
      '<3.14' in manifest['compatibility']['runtimes']['python'],
      manifest['compatibility']['runtimes']['python'])
# uv 캐시를 확장 폴더 안에 두면 안 된다. 그 안에 만들어지는 가상환경의 .pyd 가
# 실행 중 메모리에 매핑돼 잠기고, 재설치 때 설치 프로그램이 폴더를 못 지워
# EBUSY 로 실패한다(실제로 세 번 겪었다). 기본 위치(%LOCALAPPDATA%)를 쓴다.
env = manifest['server']['mcp_config'].get('env', {})
check('uv 캐시·임시폴더를 확장 폴더로 돌리지 않는다',
      not {'UV_CACHE_DIR', 'TMP', 'TEMP'} & set(env), list(env))
check('필요한 환경변수만 남긴다', set(env) == {'DART_API_KEY', 'ECOS_API_KEY'}, list(env))

check('엔트리와 요구사항 경로가 args 에 그대로 있다',
      any(a.endswith('gpcm_mcp.py') for a in args) and
      any(a.endswith('requirements-mcp.txt') for a in args), args)

# --- 10. 자가진단 · 베타 신뢰도 · 결과 검토 ------------------------------------
d = W.gpcm_doctor()
names = {c['항목']: c for c in d['점검']}
check('점검에 파이썬·도구·키·출력폴더가 다 있다',
      {'파이썬', '등록된 도구', 'DART 인증키', '출력 폴더'} <= set(names), list(names))
check('선언한 도구가 전부 등록돼 있다', names['등록된 도구']['결과'] == '정상',
      names['등록된 도구']['내용'])
check('인증키는 값이 아니라 길이만 노출', 'TESTKEY' not in json.dumps(d, ensure_ascii=False),
      names['DART 인증키']['내용'])
check('실패 항목에는 조치가 붙는다',
      all('조치' in c for c in d['점검'] if c['결과'] == '실패'),
      [c for c in d['점검'] if c['결과'] == '실패'])

saved = os.environ.pop('DART_API_KEY')
d2 = W.gpcm_doctor()
n2 = {c['항목']: c for c in d2['점검']}
check('키가 없으면 실패로 잡고 발급처를 알려준다',
      n2['DART 인증키']['결과'] == '실패' and 'opendart' in n2['DART 인증키']['조치'], n2['DART 인증키'])
check('판정에 실패 건수가 실린다', '실패' in d2['판정'], d2['판정'])
os.environ['DART_API_KEY'] = saved

# 베타 신뢰도 — 시장과 정확히 2배로 움직이는 계열이면 beta=2, R²=1
import numpy as np
idx = pd.date_range('2020-01-31', periods=40, freq='ME')
mkt = pd.Series(np.linspace(100, 200, 40) + np.sin(np.arange(40)) * 5, index=idx)
mret = mkt.pct_change().dropna()
stk = pd.Series([100.0] * 40, index=idx)
for i in range(1, 40):
    stk.iloc[i] = stk.iloc[i - 1] * (1 + 2 * mret.iloc[i - 1])
summary = [{'Ticker': '005930', 'Company': '테스트', 'Market_Cap': 1000, 'IBD': 0, 'NCI': 0,
            'Preferred': 0, 'Equity': 500, 'Pretax_Income': 100, 'Market_Index': 'KS11',
            'Stock_Monthly_Prices_5Y': stk, 'Market_Monthly_Prices_5Y': mkt,
            'Stock_Weekly_Prices_2Y': None, 'Market_Weekly_Prices_2Y': None}]
q = M.QualityLog()
wacc, dr = M.calculate_wacc_and_beta(['005930'], summary, 0.264, 0.033, 0.08, 0.0402,
                                     0.035, '5Y', fiscal_year=2025, quality=q)
check('R² 를 계산해 남긴다', round(summary[0]['Beta_5Y_R2'], 3) == 1.0, summary[0].get('Beta_5Y_R2'))
check('관측치 수 n 을 남긴다', summary[0]['Beta_5Y_N'] == 39, summary[0].get('Beta_5Y_N'))
check('Raw 베타는 종전 산식 그대로', abs(summary[0]['Beta_5Y_Raw'] - 2.0) < 1e-6, summary[0].get('Beta_5Y_Raw'))

# 베타를 하나도 못 구하면 0.8 을 쓰되 반드시 기록한다
q2 = M.QualityLog()
bare = [{'Ticker': '005930', 'Company': '테스트', 'Market_Cap': 1000, 'IBD': 0, 'NCI': 0,
         'Preferred': 0, 'Equity': 500, 'Pretax_Income': 100, 'Market_Index': 'KS11'}]
w2, _ = M.calculate_wacc_and_beta(['005930'], bare, 0.264, 0.033, 0.08, 0.0402,
                                  0.035, '5Y', fiscal_year=2025, quality=q2)
msgs = ' '.join(r['Message'] for r in q2.rows)
check('기본값 0.8 이 쓰이면 ERROR 로 남는다',
      any(r['Level'] == M.SEV_ERROR and '0.8' in r['Message'] for r in q2.rows), msgs[:120])
check('quality 를 안 넘겨도 종전처럼 동작한다',
      M.calculate_wacc_and_beta(['005930'], bare, 0.264, 0.033, 0.08, 0.0402, 0.035, '5Y',
                                fiscal_year=2025)[0]['Target_WACC'] == w2['Target_WACC'])

# 결과 검토 — 이상치는 표시하되 배제하지 않는다
mult = [
    {'Ticker': 'A', 'Period': '2025.4Q', 'EV/EBITDA': 10.0, 'PER': 12.0},
    {'Ticker': 'B', 'Period': '2025.4Q', 'EV/EBITDA': 11.0, 'PER': 13.0},
    {'Ticker': 'C', 'Period': '2025.4Q', 'EV/EBITDA': 90.0, 'PER': -5.0},
    {'Ticker': 'A', 'Period': '2024.4Q', 'EV/EBITDA': 99.0},
]
summ = [{'Ticker': 'A', 'Company': '가', 'Beta_5Y_R2': 0.45, 'Beta_5Y_N': 58},
        {'Ticker': 'B', 'Company': '나', 'Beta_5Y_R2': 0.02, 'Beta_5Y_N': 55},
        {'Ticker': 'C', 'Company': '다', 'Beta_5Y_R2': 0.50, 'Beta_5Y_N': 20}]
rv = W._summarize(mult, summ, ['A', 'B', 'C'], '2025.4Q', '5Y')
check('기준기간만 본다 (다른 분기 섞이지 않음)', rv['multiples']['EV/EBITDA']['n'] == 3,
      rv['multiples']['EV/EBITDA'])
check('중앙값에서 크게 벗어난 종목을 짚는다',
      rv['multiples']['EV/EBITDA'].get('farFromMedian') == ['C'], rv['multiples']['EV/EBITDA'])
check('0 이하 배수는 못 쓴다고 표시', rv['multiples']['PER'].get('nonPositive') == ['C'],
      rv['multiples']['PER'])
by_code = {b['code']: b for b in rv['beta']}
check('R² 낮은 종목에 사유가 붙는다', any('설명력' in r for r in by_code['B']['caution']), by_code['B'])
check('관측치 적은 종목에 사유가 붙는다', any('관측치' in r for r in by_code['C']['caution']), by_code['C'])
check('R²·n 이 충분하면 주의가 없다', 'caution' not in by_code['A'], by_code['A'])
check('재검토 목록에 사유가 함께 나온다',
      {t['code'] for t in rv['toReview']} == {'B', 'C'}, rv['toReview'])
check('자동 배제 금지 안내가 실린다', '배제 기준이 아닙니다' in rv['note'])

# --- 11. 손익 계정 매칭 2단계 (파이프라인 수준) --------------------------------
# 조이시티에서 2025.1Q·2Q 매출이 빈 사고: 손익계산서는 받았는데 계정을 못 골랐다.
def run_with_pl(pl_df):
    """주어진 손익계산서로 한 기간을 돌려 (수집값, 품질기록) 을 준다."""
    M.fetch_pl_df = lambda d, c, y, r: (pl_df.copy(), 'CFS', 'OK')
    res = M.fetch_financial_data(os.environ['DART_API_KEY'], ['005930'], ['2024.1Q'],
                                 Dart(), Recorder(), Recorder())
    summary, quality = res[4], res[9]
    return summary[0], quality


class Recorder:
    def write(self, *a, **k): pass
    def update(self, *a, **k): pass
    def progress(self, *a, **k): pass


def pl(rows):
    return pd.DataFrame([{'sj_nm': '연결포괄손익계산서', 'sj_div': 'IS',
                          'thstrm_amount': amt, 'thstrm_add_amount': amt,
                          'account_nm': nm, 'account_id': aid} for nm, aid, amt in rows])


# 번호가 붙은 표기 — 종전에는 0 이 되던 경우
s, q = run_with_pl(pl([('Ⅰ. 영업수익', '', '500000000000'),
                       ('Ⅱ. 영업이익', '', '50000000000'),
                       ('당기순이익', '', '40000000000')]))
check('번호 붙은 매출 표기를 2단계에서 줍는다', s['Revenue'] == 5000.0, s.get('Revenue'))
check('영업이익도 함께 잡힌다', s['EBIT'] == 500.0, s.get('EBIT'))

# 이름이 낯설어도 표준태그로
s2, _ = run_with_pl(pl([('게임서비스수익', 'ifrs-full_Revenue', '300000000000'),
                        ('당기순이익', '', '10000000000')]))
check('표준태그만으로도 매출을 잡는다', s2['Revenue'] == 3000.0, s2.get('Revenue'))

# 1단계 우선 — 완전 일치 행이 있으면 그것이 이긴다 (기존 회사 값이 밀리지 않음)
s3, _ = run_with_pl(pl([('Ⅰ. 영업수익', '', '999000000000'),
                        ('매출액', '', '100000000000'),
                        ('당기순이익', '', '10000000000')]))
check('완전 일치 행이 관대 매칭보다 먼저다', s3['Revenue'] == 1000.0, s3.get('Revenue'))

# 못 찾으면 무엇을 봤는지 보여준다 (없는 시트를 가리키지 않는다)
s4, q4 = run_with_pl(pl([('알수없는수익항목', '', '100000000000'),
                         ('당기순이익', '', '10000000000')]))
warn = [r for r in q4.rows if '계정 매칭' in r['Item']]
check('매칭 실패는 경고로 남는다', bool(warn), [r['Item'] for r in q4.rows])
check('경고에 실제 계정과목명이 실린다', '알수없는수익항목' in warn[0]['Message'], warn[0]['Message'][:120])
check('없는 시트를 가리키지 않는다', 'PL_Data 시트에서' not in warn[0]['Message'], warn[0]['Message'][:120])

# --- 12. 베타 평균 모집단이 엑셀과 같은가 --------------------------------------
# 엑셀 UB5 = IF(조정베타>0, 조정베타/(1+(1-세율)*D/E), "") → 빈칸은 AVERAGE 에서 제외.
# 파이썬이 다른 조건으로 걸러내면 조서(엑셀)와 채팅 보고값이 갈린다.
import numpy as np

def series(vals, freq='ME'):
    return pd.Series(vals, index=pd.date_range('2021-01-31', periods=len(vals), freq=freq))

def wacc_with(stock_px, mkt_px, **over):
    comp = {'Ticker': '005930', 'Company': '테스트', 'Market_Cap': 1000, 'IBD': 200,
            'NCI': 0, 'Preferred': 0, 'Equity': 500, 'Pretax_Income': 100,
            'Market_Index': 'KS11',
            'Stock_Monthly_Prices_5Y': stock_px, 'Market_Monthly_Prices_5Y': mkt_px,
            'Stock_Weekly_Prices_2Y': None, 'Market_Weekly_Prices_2Y': None}
    comp.update(over)
    q = M.QualityLog()
    w, _ = M.calculate_wacc_and_beta(['005930'], [comp], 0.264, 0.033, 0.08, 0.0402,
                                     0.035, '5Y', fiscal_year=2025, quality=q)
    return comp, w, q

mkt = series([100 + i for i in range(40)])
up2 = series([100 * (1 + 2 * (i / 100.0)) for i in range(40)])

# 관측치가 적어도 평균에 들어간다 (엑셀 SLOPE 는 계산하므로)
short_m = series([100, 103, 106, 110, 115])
short_s = series([100, 106, 112, 120, 130])
comp, w, q = wacc_with(short_s, short_m)
check('관측치가 적어도 평균에서 빼지 않는다 (엑셀과 동일)',
      w['Avg_Unlevered_Beta'] != 0.8, w['Avg_Unlevered_Beta'])
check('대신 관측치 부족을 경고한다',
      any('관측치가' in r['Message'] for r in q.rows), [r['Message'][:40] for r in q.rows])

# 자기자본(장부)이 0 이하여도 엑셀은 계산한다 → 파이썬도 포함
comp2, w2, q2 = wacc_with(up2, mkt, Equity=-100)
check('자본잠식이어도 평균에 넣는다 (엑셀에 없는 조건이었음)',
      w2['Avg_Unlevered_Beta'] != 0.8, w2['Avg_Unlevered_Beta'])

# 조정베타가 0 이하면 엑셀이 빈칸으로 두므로 파이썬도 뺀다
# 시장과 반대로 움직이게 만든다 (수익률을 -2배로 되짚어 가격을 만든다)
mret = mkt.pct_change().dropna()
inv = [100.0]
for x in mret:
    inv.append(inv[-1] * (1 - 2 * x))
down = series(inv)
comp3, w3, q3 = wacc_with(down, mkt)
check('조정베타 0 이하는 엑셀처럼 평균에서 뺀다',
      comp3['Beta_5Y_Raw'] < 0 and w3['Avg_Unlevered_Beta'] == 0.8, comp3.get('Beta_5Y_Raw'))
check('뺀 사유를 남긴다', any('조정베타가 0 이하' in r['Message'] for r in q3.rows),
      [r['Message'][:40] for r in q3.rows])

# 극단 베타는 버리지 않고 경고만
wild = series([100 * (1 + 8 * (i / 100.0)) for i in range(40)])
comp4, w4, q4 = wacc_with(wild, mkt)
check('극단 베타도 버리지 않는다', comp4['Beta_5Y_Raw'] > M.BETA_SANITY_LIMIT
      and w4['Avg_Unlevered_Beta'] != 0.8, comp4.get('Beta_5Y_Raw'))
check('극단 베타는 경고로 남는다', any('통상 범위' in r['Message'] for r in q4.rows),
      [r['Message'][:40] for r in q4.rows])

# --- 13. 법인세율 자동 산출 · 기준일 표기 · 버전 --------------------------------
stub_pipeline()
M.fetch_pl_df = lambda d, c, y, r: (PL.copy(), 'CFS', 'OK')

# 세율을 안 주면 피평가회사 세전이익으로 정한다
r5 = W.run_gpcm(['005930'], '2024.1Q')
W._jobs[r5['job_id']]['thread'].join(timeout=60)
s5 = W.gpcm_status(r5['job_id'])
basis = s5.get('tax_rate_basis')
check('세율을 비우면 자동 산출한다', basis is not None, s5.get('state'))
check('근거(종목·세전이익·사업연도·세율)를 함께 보고한다',
      basis and {'ticker', 'pretaxIncome100M', 'fiscalYear', 'rate_pct'} <= set(basis), basis)
check('사업연도는 기준일 연도', basis and basis['fiscalYear'] == 2024, basis)
check('세율이 그 해 세율표와 일치',
      basis and abs(basis['rate_pct'] / 100 -
                    M.get_korean_marginal_tax_rate(basis['pretaxIncome100M'], 2024)) < 1e-9, basis)
dq_msgs = ' '.join(r['Message'] for r in W._jobs[r5['job_id']].get('quality_rows', []))

# 직접 지정하면 그 값을 쓴다 (자동 산출하지 않는다)
r6 = W.run_gpcm(['005930'], '2024.1Q', tax_rate=22.0)
W._jobs[r6['job_id']]['thread'].join(timeout=60)
s6 = W.gpcm_status(r6['job_id'])
check('직접 지정하면 자동 산출하지 않는다', 'tax_rate_basis' not in s6, s6.get('tax_rate_basis'))

# 피평가회사가 목록에 없으면 거절
expect_error('target_ticker 가 목록에 없으면 거절',
             lambda: W.run_gpcm(['005930'], '2024.1Q', target_ticker='000660'), 'tickers 안에 없습니다')

# 기준일 표기 — 분기 표기도 받는다
check('as_of 가 분기 표기를 분기말로 읽는다',
      W._as_of_date('2025.4Q').strftime('%Y-%m-%d') == '2025-12-31')
check('as_of 가 날짜 표기도 받는다',
      W._as_of_date('2026-06-30').strftime('%Y-%m-%d') == '2026-06-30')
expect_error('as_of 형식 오류는 거절', lambda: W._as_of_date('2026/06/30'), '형식입니다')

# 버전 — 두 설치 경로가 같은 판인지 확인할 근거
d3 = W.gpcm_doctor()
check('진단이 버전과 실행 위치를 알려준다',
      d3.get('버전') not in (None, 'unknown') and d3.get('실행 위치'), (d3.get('버전'), d3.get('실행 위치')))
check('버전이 manifest 와 일치', d3['버전'] == manifest['version'], (d3['버전'], manifest['version']))

print()
print(f"잘못된 항목 {len(fails)}건")
sys.exit(1 if fails else 0)
