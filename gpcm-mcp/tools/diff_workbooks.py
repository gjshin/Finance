"""두 엑셀 워크북을 셀 단위로 비교한다.

국내 PC 에서 Streamlit 앱과 MCP 서버의 결과가 같은지 확인할 때 쓴다.
CHECKLIST_KR.md 3번의 합격 판정 도구다.

    python tools/diff_workbooks.py 앱.xlsx MCP.xlsx

수식은 문자열 그대로, 숫자는 1e-9 까지 같으면 같은 것으로 본다.
"""

import sys

from openpyxl import load_workbook

TOLERANCE = 1e-9


def cells(path):
    wb = load_workbook(path)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    out[(name, cell.row, cell.column)] = cell.value
    return wb.sheetnames, out


def same(a, b):
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        if isinstance(a, bool) or isinstance(b, bool):
            return a == b
        return abs(float(a) - float(b)) <= TOLERANCE
    return a == b


def main(left_path, right_path):
    left_sheets, left = cells(left_path)
    right_sheets, right = cells(right_path)

    problems = 0

    if left_sheets != right_sheets:
        print('시트 구성이 다릅니다.')
        print(f'  {left_path}: {left_sheets}')
        print(f'  {right_path}: {right_sheets}')
        problems += 1

    only_left = sorted(set(left) - set(right))
    only_right = sorted(set(right) - set(left))
    for label, missing in (('왼쪽에만', only_left), ('오른쪽에만', only_right)):
        if missing:
            print(f'{label} 있는 셀 {len(missing)}개 (앞 10개):')
            for key in missing[:10]:
                print(f'  {key[0]}!R{key[1]}C{key[2]}')
            problems += 1

    diffs = [(k, left[k], right[k]) for k in sorted(set(left) & set(right))
             if not same(left[k], right[k])]
    if diffs:
        by_sheet = {}
        for (sheet, row, col), a, b in diffs:
            by_sheet.setdefault(sheet, []).append((row, col, a, b))
        print(f'값이 다른 셀 {len(diffs)}개:')
        for sheet, items in by_sheet.items():
            cols = sorted({c for _, c, _, _ in items})
            print(f'  [{sheet}] {len(items)}개 — 열 {cols}')
            for row, col, a, b in items[:5]:
                print(f'    R{row}C{col}: {a!r}  vs  {b!r}')
        problems += 1

    if problems == 0:
        print(f'완전히 같습니다. (셀 {len(left)}개, 시트 {len(left_sheets)}개)')
        return 0

    print()
    print('GPCM 시트의 31열(Tax Rate)과 34·35열(Unlevered β)만 다르고 기준일이 '
          'FY2026 이후라면, 의도한 수정입니다. README 의 "원본과 다른 곳" 을 보십시오.')
    return 1


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)
    sys.exit(main(sys.argv[1], sys.argv[2]))
