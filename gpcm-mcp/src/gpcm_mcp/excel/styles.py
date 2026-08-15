"""엑셀 스타일과 셀 유틸. gpcm_kr.py 에서 그대로 옮겼다."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


C_BL='00338D'; C_DB='1E2A5E'; C_LB='C3D7EE'; C_PB='E8EFF8'
C_DG='333333'; C_MG='666666'; C_LG='F5F5F5'; C_BG='B0B0B0'; C_W='FFFFFF'
C_GR='E2EFDA'; C_YL='FFF8E1'; C_NOA='FCE4EC'

S1=Side(style='thin',color=C_BG); BD=Border(left=S1,right=S1,top=S1,bottom=S1)
fT=Font(name='KoPub돋움체 Medium',bold=True,size=14,color=C_BL)
fS=Font(name='KoPub돋움체 Medium',size=9,color=C_MG,italic=True)
fH=Font(name='KoPub돋움체 Medium',bold=True,size=9,color=C_W)
fA=Font(name='KoPub돋움체 Medium',size=9,color=C_DG)
fHL=Font(name='KoPub돋움체 Medium',bold=True,size=9,color=C_DB)
fMUL=Font(name='KoPub돋움체 Medium',bold=True,size=10,color=C_BL)
fNOTE=Font(name='KoPub돋움체 Medium',size=8,color=C_MG,italic=True)
fSTAT=Font(name='KoPub돋움체 Medium',bold=True,size=9,color=C_DB)
fFRM=Font(name='KoPub돋움체 Medium',size=9,color='000000')  # 검은색 (수식/계산값)
fHARD=Font(name='KoPub돋움체 Medium',size=9,color='FF0000') # 빨간색 (하드코딩/외부데이터)
fASSUM=Font(name='KoPub돋움체 Medium',size=9,color='FFC000') # 노란색 (주요 가정사항)
fLINK=Font(name='KoPub돋움체 Medium',size=9,color='008000') # 초록색 (내부시트 링크)
fSEC = Font(name='KoPub돋움체 Medium', bold=True, size=10, color=C_W)

pH=PatternFill('solid',fgColor=C_BL); pW=PatternFill('solid',fgColor=C_W)
pST=PatternFill('solid',fgColor=C_LG); pSTAT=PatternFill('solid',fgColor=C_LB)
pSEC1 = PatternFill('solid', fgColor=C_DB); pSEC2 = PatternFill('solid', fgColor=C_BL)
pSEC3 = PatternFill('solid', fgColor='2E7D32'); pSEC4 = PatternFill('solid', fgColor='6A1B9A')
pSEC5 = PatternFill('solid', fgColor='C62828'); pSEC6 = PatternFill('solid', fgColor='455A64')

ev_fills = {
    'Cash': PatternFill('solid',fgColor=C_GR), 'IBD': PatternFill('solid',fgColor=C_YL),
    'IBD(Option)': PatternFill('solid',fgColor=C_YL), 'NOA(Option)': PatternFill('solid',fgColor=C_NOA),
    'NOA': PatternFill('solid',fgColor=C_NOA), 'NCI': PatternFill('solid',fgColor=C_PB),
    'Equity': PatternFill('solid',fgColor=C_LB), 'PL_HL': PatternFill('solid',fgColor=C_YL),
}

aC=Alignment(horizontal='center',vertical='center',wrap_text=True)
aL=Alignment(horizontal='left',vertical='center',indent=1)
aR=Alignment(horizontal='right',vertical='center')

NB='#,##0;(#,##0);"-"'; NB1='#,##0.0;(#,##0.0);"-"'; NI_FMT='#,##0;(#,##0);"-"'
NP='₩#,##0;(₩#,##0);"-"'; NF_X='#,##0.0x;(#,##0.0x);"-"'

def sc(c,fo=None,fi=None,al=None,bd=None,nf=None):
    if fo: c.font=fo
    if fi: c.fill=fi
    if al: c.alignment=al
    if bd: c.border=bd
    if nf: c.number_format=nf

def style_range(ws, r1, c1, r2, c2, fo=None, fi=None, al=None, bd=None, nf=None):
    for rr in range(r1, r2+1):
        for cc in range(c1, c2+1):
            sc(ws.cell(rr, cc), fo=fo, fi=fi, al=al, bd=bd, nf=nf)



def add_gpcm_section_row(ws):
    sec_row = 4
    sections = [
        (1, 2,  "Company Info",       pSEC1), (3, 5,  "Other Info",         pSEC2),
        (6, 12, "BS & EV Components", pSEC3), (13,17, "PL(Annual & LTM)",   pSEC4),
        (18,20, "Market Data",        pSEC5), (21,25, "Valuation Multiples", pSEC6),
        (26,35, "Beta & Risk Analysis", PatternFill('solid', fgColor='6A1B9A')),
        (36,36, "Equity Adj",           pSEC3),
    ]
    for c1, c2, label, fill in sections:
        ws.merge_cells(start_row=sec_row, start_column=c1, end_row=sec_row, end_column=c2)
        ws.cell(sec_row, c1).value = label
        style_range(ws, sec_row, c1, sec_row, c2, fo=fSEC, fi=fill, al=aC, bd=BD)
