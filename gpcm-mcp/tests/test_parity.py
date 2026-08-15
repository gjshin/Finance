"""이식본이 원본 gpcm_kr.py 와 같은 숫자를 내는지 대조한다.

이 패키지는 gpcm_kr.py 의 복사본이다. 복사본이 갈라지는 것은 시간 문제이고,
갈라졌다는 사실을 아무도 모르는 채로 조서에 숫자가 들어가는 것이 최악이다.
그래서 같은 가짜 입력을 원본과 이식본에 각각 먹여 결과를 통째로 비교한다.

DART 가 해외 IP 를 막아 실제 조회로는 검증할 수 없으므로, **이 테스트가
"숫자가 안 바뀌었다" 의 유일한 기계적 증거**다.

누군가 gpcm_kr.py 를 고치면 여기서 깨진다. 그때 이식본도 함께 고쳐야 한다.
"""

import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parents[1] / 'src'))

import fakes  # noqa: E402
import reference  # noqa: E402

import gpcm_mcp.excel.gpcm_book as port_gpcm_book  # noqa: E402
import gpcm_mcp.excel.historical_book as port_hist_book  # noqa: E402
import gpcm_mcp.gpcm as port_gpcm  # noqa: E402
import gpcm_mcp.historical as port_hist  # noqa: E402
import gpcm_mcp.listings as port_listings  # noqa: E402

TICKERS = ['005930', '000660']
NAMES = {'005930': '테스트전자', '000660': '테스트반도체'}
PERIODS = ['2024.3Q', '2024.4Q', '2025.1Q']
END = '2025-03-31'

WACC_ARGS = dict(target_tax_rate_input=0.264, rf_input=0.033, mrp_input=0.08,
                 size_premium_input=0.0402, kd_pretax_input=0.035, beta_type_input='5Y')


@pytest.fixture(scope='module')
def original():
    return reference.load()


def _run_mode1(module, listings_module, dart_names=NAMES, **kw):
    """모드 1 계산을 끝까지 돌린다. 원본/이식본 어느 쪽이든 같은 방식으로."""
    modules = [module] if listings_module is None else [module, listings_module]
    restore = fakes.install(*modules, names=dart_names, end_date=END, **kw)
    try:
        dart = fakes.FakeDart(names=dart_names)
        out = module.fetch_financial_data(
            'KEY', TICKERS, PERIODS, dart, fakes.Silent(), fakes.Silent())
        summary = out[4]
        base_year = out[5]
        wacc, adr = module.calculate_wacc_and_beta(
            TICKERS, summary, WACC_ARGS['target_tax_rate_input'], WACC_ARGS['rf_input'],
            WACC_ARGS['mrp_input'], WACC_ARGS['size_premium_input'],
            WACC_ARGS['kd_pretax_input'], WACC_ARGS['beta_type_input'],
            fiscal_year=base_year)
        return out, wacc, adr
    finally:
        restore()


def test_mode1_compute_matches(original):
    ref_out, ref_wacc, ref_adr = _run_mode1(original, None)
    port_out, port_wacc, port_adr = _run_mode1(port_gpcm, port_listings)

    (r_bs, r_pl, r_mkt, r_names, r_summary,
     r_year, r_qtr, r_date, r_mult, r_quality) = ref_out
    (p_bs, p_pl, p_mkt, p_names, p_summary,
     p_year, p_qtr, p_date, p_mult, p_quality) = port_out

    assert r_year == p_year and r_qtr == p_qtr and r_date == p_date
    assert r_names == p_names
    assert r_bs == p_bs, '재무상태표 원본 행이 다르다'
    assert r_pl == p_pl, '손익계산서 원본 행이 다르다'
    assert r_mkt == p_mkt, '시가총액 행이 다르다'
    assert r_mult == p_mult, '배수 계산의 입력값이 다르다'

    # 품질 기록은 순서까지 같아야 한다 — Data_Quality 시트 행 순서가 여기서 결정된다.
    assert r_quality.rows == p_quality.rows

    assert len(r_summary) == len(p_summary)
    for r_item, p_item in zip(r_summary, p_summary):
        for key in ('Company', 'Ticker', 'Exchange', 'Market_Index'):
            assert r_item[key] == p_item[key]
        for key in ('Market_Cap', 'Cash', 'IBD', 'NCI', 'NOA', 'Equity',
                    'Preferred', 'Revenue', 'EBIT', 'NI', 'Pretax_Income'):
            assert r_item[key] == p_item[key], f'{key} 가 다르다'
        for key in ('Stock_Monthly_Prices_5Y', 'Market_Monthly_Prices_5Y',
                    'Stock_Weekly_Prices_2Y', 'Market_Weekly_Prices_2Y'):
            r_series, p_series = r_item[key], p_item[key]
            assert (r_series is None) == (p_series is None)
            if r_series is not None:
                pd.testing.assert_series_equal(r_series, p_series)

    assert set(ref_wacc) == set(port_wacc)
    for key in ref_wacc:
        assert ref_wacc[key] == pytest.approx(port_wacc[key], rel=0, abs=0), \
            f'WACC 항목 {key} 가 다르다'
    assert ref_adr == port_adr


def _cells(book):
    """워크북을 (시트, 행, 열) -> 값 으로 펼친다. 수식은 문자열 그대로 비교된다."""
    wb = load_workbook(book)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    out[(name, cell.row, cell.column)] = cell.value
    return wb.sheetnames, out


def _reference_orphans(original, all_multiples, base_period_str, quality):
    """원본 UI 블록 안에 있는 계산을 그대로 재현한다 (L2549-2676).

    잘라낸 뒤쪽이라 reference.load() 로는 안 딸려온다. 이식본의 orphans.py 가
    이것과 같은 것을 만드는지 보려면 여기에 원본 코드를 그대로 두어야 한다.
    """
    import numpy as np

    base_year, base_qtr = original.parse_period(base_period_str)
    base_date_display = original.get_base_date_str(base_year, base_qtr)
    notes_list = [
        f'• Base Date: {base_period_str} ({base_date_display}) | Unit: 억원 (KRW 100M)',
        '• 공통: 연결재무제표 작성 시 CFS 우선, 미존재 시 OFS 기준으로 수집',
        '• PL: 요약 손익계산서에서 매출액/영업이익/당기순이익 3개 계정만 엄격 추출',
        '• PL Fetch: finstate(요약) → finstate_all(CFS/OFS) fallback',
        '• Shares: DART(stockTotqySttus) 유통주식수(distb_stock_co) 우선, 미공시 시 DART 과거보고서 fallback',
        '• EV = Market Cap + 우선주(장부) + IBD − Cash + NCI − NOA',
        '• Net Debt = IBD − Cash − NOA',
        '• IBD(Option): CB/EB/BW 등 메자닌은 기본적으로 IBD(Option)으로 태깅되어 EV/NetDebt에서 제외됨',
        '• NOA(Option): 투자자산/관계기업 등은 기본적으로 NOA(Option)으로 태깅되어 EV/NetDebt에서 제외됨',
        '• LTM = Current Cumulative + Prior Annual − Prior Same Quarter Cumulative (단, 4Q는 Annual)',
        '• Beta: 5년 월간 & 2년 주간 수익률 기준 (FinanceDataReader 사용)',
        '• Adjusted Beta = 2/3 × Raw Beta + 1/3 × 1',
        '• D/E Ratio = IBD / (Market Cap + 우선주 + NCI)',
        '• Debt Ratio (D/V) = IBD / (Market Cap + 우선주 + IBD + NCI)',
        '• 우선주: BS의 우선주자본금(액면) 기준. 시가총액은 보통주만 반영하므로 자기자본가치에 가산',
        '• Unlevered Beta = Levered Beta / (1 + (1 - Tax Rate) × D/E Ratio)',
        '• Tax Rate: 한국 법인세 한계세율 (지방소득세 포함, 세전순이익 기준, 사업연도별 세율표 적용)',
        '   - FY2023~2025: 2억 이하 9.9% | 2~200억 20.9% | 200~3,000억 23.1% | 3,000억 초과 26.4%',
        '   - FY2026~    : 2억 이하 11.0% | 2~200억 22.0% | 200~3,000억 24.2% | 3,000억 초과 27.5% (2025년 세법개정)',
    ]

    df_screen = pd.DataFrame(all_multiples)

    problems = []
    base_rows = [m for m in all_multiples if m['Period'] == base_period_str]
    if not base_rows:
        problems.append(f"기준 기간({base_period_str}) 데이터를 한 건도 수집하지 못했습니다.")
    for m in base_rows:
        empty_fields = [k for k in ('Market_Cap', 'Revenue', 'Equity') if not m.get(k)]
        if empty_fields:
            problems.append(f"[{m['Company']}] {base_period_str} — {', '.join(empty_fields)} 값이 없습니다.")
    if problems:
        for p in problems:
            quality.add(original.SEV_ERROR, '', '', f'기준기간 {base_period_str}', p)

    if not df_screen.empty:
        df_screen['EV'] = df_screen['Market_Cap'] + df_screen['Preferred'] + df_screen['IBD'] - df_screen['Cash'] + df_screen['NCI'] - df_screen['NOA']
        df_screen['EV/EBIT'] = np.where(df_screen['EBIT'] > 0, df_screen['EV'] / df_screen['EBIT'], np.nan)
        df_screen['PER'] = np.where(df_screen['NI'] > 0, df_screen['Market_Cap'] / df_screen['NI'], np.nan)
        df_screen['PSR'] = np.where(df_screen['Revenue'] > 0, df_screen['Market_Cap'] / df_screen['Revenue'], np.nan)

    return notes_list, df_screen


def test_ui_block_orphans_match(original):
    """화면 코드 안에 있던 네 토막이 orphans.py 와 같은 것을 만드는가."""
    import gpcm_mcp.orphans as orphans

    ref_out, _, _ = _run_mode1(original, None)
    all_multiples, ref_quality = ref_out[8], ref_out[9]
    base = PERIODS[-1]

    ref_notes, ref_screen = _reference_orphans(
        original, all_multiples, base, ref_quality)

    port_quality = port_gpcm.QualityLog()
    port_notes = orphans.build_notes(base)
    port_screen = orphans.build_screen_frame(all_multiples)
    port_problems = orphans.diagnose_base_period(all_multiples, base, port_quality)

    assert ref_notes == port_notes, '방법론 주석이 다르다'
    pd.testing.assert_frame_equal(ref_screen, port_screen)
    # 원본은 진단 결과를 quality 에 붙이기만 하고 목록을 남기지 않으므로,
    # 붙은 행으로 비교한다.
    ref_added = [r for r in ref_quality.rows if r['Item'].startswith('기준기간')]
    assert [r['Message'] for r in ref_added] == port_problems
    assert [r['Message'] for r in port_quality.rows] == port_problems


def test_mode1_workbook_matches(original):
    ref_out, ref_wacc, ref_adr = _run_mode1(original, None)
    port_out, port_wacc, port_adr = _run_mode1(port_gpcm, port_listings)

    import gpcm_mcp.orphans as orphans

    base = PERIODS[-1]
    ref_notes, ref_screen = _reference_orphans(
        original, ref_out[8], base, ref_out[9])
    port_notes = orphans.build_notes(base)
    port_screen = orphans.build_screen_frame(port_out[8])
    orphans.diagnose_base_period(port_out[8], base, port_out[9])

    def build(module, out, wacc, adr, notes, screen):
        (bs, pl, mkt, names, summary, year, qtr, date, mult, quality) = out
        return module.export_gpcm_excel(
            base, qtr, TICKERS, summary, bs, pl, mkt, names, wacc,
            '5Y', notes, adr, date, screen, PERIODS, quality)

    ref_sheets, ref_cells = _cells(
        build(original, ref_out, ref_wacc, ref_adr, ref_notes, ref_screen))
    port_sheets, port_cells = _cells(
        build(port_gpcm_book, port_out, port_wacc, port_adr, port_notes, port_screen))

    assert ref_sheets == port_sheets, '시트 구성이 다르다'

    only_ref = set(ref_cells) - set(port_cells)
    only_port = set(port_cells) - set(ref_cells)
    assert not only_ref, f'원본에만 있는 셀: {sorted(only_ref)[:5]}'
    assert not only_port, f'이식본에만 있는 셀: {sorted(only_port)[:5]}'

    diffs = [(k, ref_cells[k], port_cells[k])
             for k in ref_cells if ref_cells[k] != port_cells[k]]
    assert not diffs, f'값이 다른 셀 {len(diffs)}개: {diffs[:5]}'


# --- 모드 2 -----------------------------------------------------------------

HIST_PERIODS = [
    {'year': 2023, 'qtr': None, 'label': '2023년'},
    {'year': 2024, 'qtr': None, 'label': '2024년'},
]


def _run_mode2(module, listings_module):
    modules = [module] if listings_module is None else [module, listings_module]
    restore = fakes.install(*modules, names=NAMES, end_date=END)
    try:
        dart = fakes.FakeDart(names=NAMES)
        df_summ, df_details = module.fetch_historical_financials(
            'KEY', TICKERS, HIST_PERIODS, dart, fakes.Silent(), fakes.Silent(), None)
        return module.calculate_historical_metrics(df_summ), df_details
    finally:
        restore()


def test_mode2_compute_matches(original):
    ref_summ, ref_details = _run_mode2(original, None)
    port_summ, port_details = _run_mode2(port_hist, port_listings)

    pd.testing.assert_frame_equal(ref_summ, port_summ)
    pd.testing.assert_frame_equal(ref_details, port_details)


def test_mode2_workbook_matches(original):
    ref_summ, ref_details = _run_mode2(original, None)
    port_summ, port_details = _run_mode2(port_hist, port_listings)

    ref_sheets, ref_cells = _cells(
        original.export_historical_excel(ref_summ, ref_details, HIST_PERIODS))
    port_sheets, port_cells = _cells(
        port_hist_book.export_historical_excel(port_summ, port_details, HIST_PERIODS))

    assert ref_sheets == port_sheets
    diffs = [(k, ref_cells[k], port_cells[k])
             for k in ref_cells if ref_cells.get(k) != port_cells.get(k)]
    assert not diffs, f'값이 다른 셀 {len(diffs)}개: {diffs[:5]}'
